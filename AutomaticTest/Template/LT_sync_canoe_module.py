#!/usr/bin/env python3
"""
按 LT 用例表结构同步 CANoe XML 与 CAPL 脚本。

Excel 指定 sheet 要求包含列：
- 分类
- 一级测试用例
- 二级测试用例
- 用例ID
- 前置条件
- 测试步骤
- 期望结果
"""
import argparse
import re
import sys
import xml.etree.ElementTree as ET
from collections import OrderedDict
from pathlib import Path

from customrule_sync_canoe_module import (
    NS,
    backup_file,
    create_empty_can,
    create_empty_xml,
    update_can,
    update_xml,
)


def is_placeholder_text(value) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    return text in {"", "/", "／", "-", "--", "N/A", "n/a", "NA", "na"}


def normalize_header(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).rstrip("*").strip()


HEADER_ALIASES = {
    "group": {"分类", "testgroup"},
    "level1_case": {"一级测试用例"},
    "level2_case": {"二级测试用例"},
    "func_name": {"用例ID"},
    "precondition": {"前置条件"},
    "steps": {"测试步骤", "步骤"},
    "expected_result": {"期望结果", "预期结果"},
}


def find_header_row_and_cols(ws):
    search_rows = min(30, ws.max_row)
    for row_idx in range(1, search_rows + 1):
        row_values = [normalize_header(ws.cell(row=row_idx, column=c).value) for c in range(1, ws.max_column + 1)]
        col_map = {}
        for col_idx, header in enumerate(row_values, start=1):
            for key, aliases in HEADER_ALIASES.items():
                if header in aliases:
                    col_map[key] = col_idx
        required = {"group", "level1_case", "level2_case", "func_name", "precondition", "steps", "expected_result"}
        if required.issubset(set(col_map.keys())):
            return row_idx, col_map

    raise ValueError(
        "未识别到 LT 用例表头，请确认 sheet 包含列："
        "分类、一级测试用例、二级测试用例、用例ID、前置条件、测试步骤、期望结果"
    )


def build_case_name(level1: str, level2: str, fallback: str) -> str:
    a = (level1 or "").strip()
    b = (level2 or "").strip()
    if a and b:
        return f"{a}-{b}"
    if a:
        return a
    if b:
        return b
    return fallback


def derive_group_prefix(func_name: str, group_name: str) -> str:
    words = [w for w in (func_name or "").split("_") if w]
    if words:
        return "_".join(words[:3])
    safe_group = re.sub(r"[^A-Za-z0-9_]+", "_", group_name).strip("_")
    return safe_group if safe_group else "LT_Group"


STEP_SEQ_RE = re.compile(r"^\s*(\d+)\s*[\.、，,\)\]]\s*(.*)$")


def merge_multiline_numbered_text(text: str) -> str:
    """
    将多行步骤/期望结果整理为“每个序号一行”：
    - 以序号开头的行，作为新条目
    - 非序号开头行，拼接到上一条序号项
    """
    if not text:
        return ""
    raw_lines = str(text).replace("\r\n", "\n").replace("\r", "\n").split("\n")
    lines = [ln.strip() for ln in raw_lines if ln and ln.strip()]
    if not lines:
        return ""

    merged_items = []
    current_seq = None
    current_parts = []

    def flush_current():
        if current_seq is None:
            return
        merged_text = " ".join(part for part in current_parts if part).strip()
        if merged_text:
            merged_items.append(f"{current_seq}. {merged_text}")

    for line in lines:
        m = STEP_SEQ_RE.match(line)
        if m:
            flush_current()
            current_seq = m.group(1)
            first_part = (m.group(2) or "").strip()
            current_parts = [first_part] if first_part else []
            continue

        if current_seq is not None:
            current_parts.append(line)
        else:
            # 没有显式序号时保持原样，避免丢失信息
            merged_items.append(line)

    flush_current()
    return "\n".join(merged_items)


def parse_lt_rows(excel_path: Path, sheet_name: str):
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] 需要安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excel 中不存在指定 sheet: {sheet_name}")
    ws = wb[sheet_name]

    header_row, col_map = find_header_row_and_cols(ws)
    print(f"[INFO] 使用 Excel 页面: {ws.title}")
    print(f"[INFO] 表头行: {header_row}, 列映射: {col_map}")

    grouped_cases = OrderedDict()
    for row_idx in range(header_row + 1, ws.max_row + 1):
        def get_cell(key):
            col = col_map[key]
            value = ws.cell(row=row_idx, column=col).value
            if is_placeholder_text(value):
                return ""
            return str(value).strip()

        group_name = get_cell("group")
        func_name = get_cell("func_name")
        if not group_name or not func_name:
            continue

        level1_case = get_cell("level1_case")
        level2_case = get_cell("level2_case")
        precondition = get_cell("precondition")
        steps = merge_multiline_numbered_text(get_cell("steps"))
        expected_result = merge_multiline_numbered_text(get_cell("expected_result"))
        case_name = build_case_name(level1_case, level2_case, func_name)

        grouped_cases.setdefault(group_name, []).append(
            {
                "name": func_name,
                "case_id": func_name,
                "case_name": case_name,
                "precondition": precondition,
                "steps": steps,
                "expected_result": expected_result,
            }
        )

    rows = [{"node_type": "testmodule", "name": ws.title, "variant": "", "level": 0, "case_id": "", "case_name": ws.title, "precondition": "", "steps": "", "expected_result": ""}]
    # testmodule 下固定公共 preparation: testModuleInit
    rows.append(
        {
            "node_type": "preparation",
            "name": "",
            "variant": "",
            "level": 1,
            "case_id": "",
            "case_name": "",
            "precondition": "",
            "steps": "",
            "expected_result": "",
        }
    )
    rows.append(
        {
            "node_type": "capltestfunction",
            "name": "testModuleInit",
            "variant": "",
            "level": 2,
            "case_id": "",
            "case_name": "testModuleInit",
            "precondition": "",
            "steps": "",
            "expected_result": "",
        }
    )
    for group_name, cases in grouped_cases.items():
        if not cases:
            continue

        prefix = derive_group_prefix(cases[0]["name"], group_name)
        pre_func = f"{prefix}_Pre"
        post_func = f"{prefix}_Post"

        rows.append(
            {
                "node_type": "testgroup",
                "name": group_name,
                "variant": "",
                "level": 1,
                "case_id": "",
                "case_name": group_name,
                "precondition": "",
                "steps": "",
                "expected_result": "",
            }
        )
        rows.append(
            {
                "node_type": "preparation",
                "name": "",
                "variant": "",
                "level": 2,
                "case_id": "",
                "case_name": "",
                "precondition": "",
                "steps": "",
                "expected_result": "",
            }
        )
        rows.append(
            {
                "node_type": "capltestfunction",
                "name": pre_func,
                "variant": "",
                "level": 3,
                "case_id": "",
                "case_name": f"{group_name} Preparation",
                "precondition": "",
                "steps": "",
                "expected_result": "",
            }
        )

        for case in cases:
            rows.append(
                {
                    "node_type": "capltestcase",
                    "name": case["name"],
                    "variant": "",
                    "level": 2,
                    "case_id": case["case_id"],
                    "case_name": case["case_name"],
                    "precondition": case["precondition"],
                    "steps": case["steps"],
                    "expected_result": case["expected_result"],
                }
            )

        rows.append(
            {
                "node_type": "completion",
                "name": "",
                "variant": "",
                "level": 2,
                "case_id": "",
                "case_name": "",
                "precondition": "",
                "steps": "",
                "expected_result": "",
            }
        )
        rows.append(
            {
                "node_type": "capltestfunction",
                "name": post_func,
                "variant": "",
                "level": 3,
                "case_id": "",
                "case_name": f"{group_name} Completion",
                "precondition": "",
                "steps": "",
                "expected_result": "",
            }
        )

    # testmodule 下固定公共 completion: testModuleEnd
    rows.append(
        {
            "node_type": "completion",
            "name": "",
            "variant": "",
            "level": 1,
            "case_id": "",
            "case_name": "",
            "precondition": "",
            "steps": "",
            "expected_result": "",
        }
    )
    rows.append(
        {
            "node_type": "capltestfunction",
            "name": "testModuleEnd",
            "variant": "",
            "level": 2,
            "case_id": "",
            "case_name": "testModuleEnd",
            "precondition": "",
            "steps": "",
            "expected_result": "",
        }
    )

    return rows


def build_default_output_paths(excel_path: Path, sheet_name: str):
    safe_sheet = re.sub(r"[\\/:*?\"<>|]+", "_", sheet_name).strip() or "Sheet"
    xml_path = excel_path.with_name(f"{excel_path.stem}_{safe_sheet}.xml")
    can_path = xml_path.with_suffix(".can")
    return xml_path, can_path


def localname(tag: str) -> str:
    if tag.startswith("{"):
        return tag.split("}", 1)[1]
    return tag


def _is_named_section(elem: ET.Element, section_tag: str, func_name: str) -> bool:
    if localname(elem.tag) != section_tag:
        return False
    for child in elem:
        if localname(child.tag) != "capltestfunction":
            continue
        if (child.get("name") or "").strip() == func_name:
            return True
    return False


def reorder_module_init_end(xml_path: Path):
    """
    固定根级顺序：
    1) testModuleInit 所在 preparation 在第一个 testgroup 之前
    2) testModuleEnd 所在 completion 在最后一个 testgroup 之后
    """
    tree = ET.parse(xml_path)
    root = tree.getroot()
    children = list(root)

    module_prep = None
    module_comp = None
    for child in children:
        if module_prep is None and _is_named_section(child, "preparation", "testModuleInit"):
            module_prep = child
        elif module_comp is None and _is_named_section(child, "completion", "testModuleEnd"):
            module_comp = child

    changed = False

    if module_prep is not None:
        root.remove(module_prep)
        remaining = list(root)
        first_group_idx = next((i for i, e in enumerate(remaining) if localname(e.tag) == "testgroup"), len(remaining))
        root.insert(first_group_idx, module_prep)
        changed = True

    if module_comp is not None:
        root.remove(module_comp)
        remaining = list(root)
        group_indices = [i for i, e in enumerate(remaining) if localname(e.tag) == "testgroup"]
        insert_idx = group_indices[-1] + 1 if group_indices else len(remaining)
        root.insert(insert_idx, module_comp)
        changed = True

    if not changed:
        return

    ET.register_namespace("", NS)
    ET.register_namespace("xsi", "http://www.w3.org/2001/XMLSchema-instance")
    ET.indent(tree, space="  ")
    tree.write(xml_path, encoding="utf-8", xml_declaration=True)


def main():
    parser = argparse.ArgumentParser(
        description="按 LT 用例 sheet 同步 CANoe XML 与 CAPL 脚本"
    )
    parser.add_argument("excel", type=Path, help="Excel 测试用例文件路径（.xlsx 或 .xls）")
    parser.add_argument("sheet", help="要解析的 sheet 名称")
    parser.add_argument("xml", type=Path, nargs="?", default=None, help="XML 输出路径（可选）")
    parser.add_argument("can", type=Path, nargs="?", default=None, help="CAN 输出路径（可选）")
    parser.add_argument("--xml-only", action="store_true", help="仅生成/同步 XML，不创建和更新 CAN")
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"[ERROR] Excel 文件不存在: {args.excel}")
        sys.exit(1)

    try:
        rows = parse_lt_rows(args.excel, args.sheet)
    except Exception as exc:
        print(f"[ERROR] 解析失败: {exc}")
        sys.exit(1)

    if len(rows) <= 1:
        print("[WARN] 指定 sheet 未发现可用的测试用例（需至少包含分类 + 用例ID）。")
        return
    print(f"[INFO] 解析到 {len(rows)} 条自动化测试项")

    default_xml, default_can = build_default_output_paths(args.excel, args.sheet)
    xml_path = args.xml if args.xml else default_xml
    can_path = args.can if args.can else default_can

    xml_path.parent.mkdir(parents=True, exist_ok=True)
    if not args.xml_only:
        can_path.parent.mkdir(parents=True, exist_ok=True)

    if not xml_path.exists():
        create_empty_xml(xml_path, args.sheet)
    if (not args.xml_only) and (not can_path.exists()):
        create_empty_can(can_path)

    backup_file(xml_path)
    if not args.xml_only:
        backup_file(can_path)

    update_xml(rows, xml_path)
    reorder_module_init_end(xml_path)
    if not args.xml_only:
        update_can(rows, can_path, xml_path)

    print("[OK] 同步完成：")
    print(f"     XML: {xml_path}")
    if not args.xml_only:
        print(f"     CAN: {can_path}")


if __name__ == "__main__":
    main()

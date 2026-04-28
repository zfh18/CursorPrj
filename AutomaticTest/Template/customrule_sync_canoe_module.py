#!/usr/bin/env python3
"""
根据 Excel 测试用例文件增量同步 CANoe XML 与 CAPL 脚本
"""
import argparse
import shutil
import re
import sys
import xml.etree.ElementTree as ET
from pathlib import Path


NS = "http://www.vector-informatik.de/CANoe/TestModule/1.8"

# Excel 中关键列的默认索引（1-based，用于找不到表头时的备选）
EXCEL_COL_TEST_TYPE = 11
EXCEL_COL_NODE_TYPE = 12
EXCEL_COL_NAME = 13
EXCEL_COL_VARIANT = 14
EXCEL_COL_LEVEL = 15
# 注释相关列的默认索引（常见测试用例表结构）
EXCEL_COL_CASE_ID = 6
EXCEL_COL_CASE_NAME = 7
EXCEL_COL_PRECONDITION = 16
EXCEL_COL_STEPS = 17
EXCEL_COL_EXPECTED_RESULT = 18


def parse_variants(cell_value: str) -> str:
    """
    解析变体单元格：支持空格或换行分隔的多个变体，
    返回 'variant1 variant2' 格式，用于 variants 属性
    """
    if not cell_value or not str(cell_value).strip():
        return ""
    parts = re.split(r"[\s\n\r]+", str(cell_value).strip())
    return " ".join(p for p in parts if p)


def is_placeholder_text(s: str) -> bool:
    """判定 Excel 占位符文本（表示空值）"""
    if s is None:
        return True
    t = str(s).strip()
    return t in {"", "/", "／", "-", "--", "N/A", "n/a", "NA", "na"}


def read_text_with_fallback(path: Path):
    """尝试多种编码读取文本文件"""
    encodings = ["utf-8-sig", "utf-8", "gbk", "cp936"]
    last_err = None
    for enc in encodings:
        try:
            return path.read_text(encoding=enc), enc
        except UnicodeDecodeError as exc:
            last_err = exc
    raise RuntimeError(f"无法读取文件编码: {path}") from last_err


def _build_col_mapping(row_values: list) -> dict:
    """根据表头单元格内容建立列名到索引的映射"""
    col_mapping = {}
    for col_idx, val in enumerate(row_values, start=1):
        val_clean = (val or "").strip().rstrip("*").strip()
        if "测试类型" in val_clean and ("自动" in val_clean or "手动" in val_clean):
            col_mapping["test_type"] = col_idx
        elif "CANoe" in val_clean and "类型" in val_clean:
            col_mapping["node_type"] = col_idx
        elif "自动化测试函数" in val_clean or val_clean == "自动化测试函数":
            col_mapping["name"] = col_idx
        elif val_clean == "变体":
            col_mapping["variant"] = col_idx
        elif "层级" in val_clean:
            col_mapping["level"] = col_idx
        elif ("用例" in val_clean and ("id" in val_clean.lower() or "编号" in val_clean)) or val_clean in ("ID", "id", "编号", "用例编号", "规范ID", "编码"):
            col_mapping["case_id"] = col_idx
        elif val_clean in ("名称", "用例名称", "测试名称", "标题", "用例标题") or (val_clean.endswith("名称") and "函数" not in val_clean):
            col_mapping["case_name"] = col_idx
        elif "前置条件" in val_clean or "预置条件" in val_clean or "前提条件" in val_clean or val_clean.lower() == "precondition":
            col_mapping["precondition"] = col_idx
        elif ("步骤" in val_clean or val_clean.lower() in ("steps", "step")) and "结果" not in val_clean:
            col_mapping["steps"] = col_idx
        elif "预期结果" in val_clean or "期望结果" in val_clean or val_clean in ("预期", "期望") or val_clean.lower() in ("expected", "expectedresult"):
            col_mapping["expected_result"] = col_idx
    return col_mapping


def _clean_cell_for_match(v) -> str:
    if v is None:
        return ""
    return str(v).strip().replace("\r", " ").replace("\n", " ").replace("\t", " ").strip().lower()


def _infer_columns_by_data(ws):
    """
    当表头检测失败时，通过数据本身推断列：
    1) 测试类型列：Automatic / 自动
    2) 节点类型列：testmodule/testgroup/preparation/completion/capltestfunction/capltestcase
    3) 层级列：数字（如 0/1/2/3）
    4) 名称列：位于节点类型列后一列（常见结构：测试类型/节点类型/名称/变体/层级）
    5) 变体列：位于名称列后一列（如果存在）
    """
    node_tokens = {
        "testmodule",
        "testgroup",
        "preparation",
        "completion",
        "capltestfunction",
        "capltestcase",
    }

    max_col = ws.max_column
    search_rows = min(30, ws.max_row)

    test_type_scores = [0] * (max_col + 1)  # 1-based
    node_type_scores = [0] * (max_col + 1)
    level_scores = [0] * (max_col + 1)

    def is_automatic(s: str) -> bool:
        return ("automatic" in s) or ("auto" in s) or ("自动" in s) or (s == "自动化")

    for r in range(1, search_rows + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c).value
            s = _clean_cell_for_match(cell)
            if not s:
                continue
            if is_automatic(s):
                test_type_scores[c] += 1
            if s in node_tokens:
                node_type_scores[c] += 1
            if s.isdigit():
                try:
                    n = int(s)
                except Exception:
                    continue
                if 0 <= n <= 50:
                    level_scores[c] += 1

    best_test_col = max(range(1, max_col + 1), key=lambda c: test_type_scores[c], default=None)
    best_node_col = max(range(1, max_col + 1), key=lambda c: node_type_scores[c], default=None)
    best_level_col = max(range(1, max_col + 1), key=lambda c: level_scores[c], default=None)

    if test_type_scores[best_test_col] == 0 or node_type_scores[best_node_col] == 0:
        return None

    # 找到数据首行：同时满足测试类型 + 节点类型
    first_data_row = None
    for r in range(1, search_rows + 1):
        test_s = _clean_cell_for_match(ws.cell(row=r, column=best_test_col).value)
        node_s = _clean_cell_for_match(ws.cell(row=r, column=best_node_col).value)
        if is_automatic(test_s) and node_s in node_tokens:
            first_data_row = r
            break

    if first_data_row is None:
        first_data_row = 2

    # 常见结构：名称在节点类型列后
    name_col = best_node_col + 1 if best_node_col + 1 <= max_col else None
    variant_col = name_col + 1 if name_col and (name_col + 1) <= max_col else None
    level_col = best_level_col

    # 若层级列太不可信，退化：取距离变体后/末尾那列中的数字列
    if level_scores[level_col] == 0:
        # fallback：在 1..max_col 中找最像“层级”的数字列
        candidates = [c for c in range(1, max_col + 1) if level_scores[c] > 0]
        if candidates:
            level_col = max(candidates, key=lambda c: level_scores[c])

    header_row = max(1, first_data_row - 1)
    col_map = {
        "test_type": best_test_col,
        "node_type": best_node_col,
        "name": name_col,
        "variant": variant_col if variant_col else None,
        "level": level_col if level_col else EXCEL_COL_LEVEL,
    }

    # 清理 None
    col_map = {k: v for k, v in col_map.items() if v is not None}

    # 从第 1 行表头补充注释列（用例 ID、名称、前置条件、步骤、预期结果）
    for hdr_row in (1, 2):
        if hdr_row > ws.max_row:
            continue
        row_values = [str(ws.cell(row=hdr_row, column=c).value or "") for c in range(1, max_col + 1)]
        extra = _build_col_mapping(row_values)
        for key in ("case_id", "case_name", "precondition", "steps", "expected_result"):
            if key not in col_map and key in extra and extra[key] <= max_col:
                col_map[key] = extra[key]

    return header_row, col_map


def find_header_row_and_cols(ws):
    """
    自动检测表头行：扫描前 20 行，选取包含「node_type」且匹配列数最多的行作为表头
    """
    search_rows = min(20, ws.max_row)
    best_row, best_mapping, best_score = None, None, -1

    for row_idx in range(1, search_rows + 1):
        row_values = [str(ws.cell(row=row_idx, column=c).value or "") for c in range(1, ws.max_column + 1)]
        col_mapping = _build_col_mapping(row_values)

        # 必须包含 node_type（CANoe自动化测试类型）才是有效表头
        if "node_type" not in col_mapping:
            continue

        # 按匹配列数评分，优先选匹配更多的行
        score = len(col_mapping)
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_mapping = col_mapping

    if best_row is not None:
        return best_row, best_mapping

    # 表头检测不到：尝试按数据推断列
    inferred = _infer_columns_by_data(ws)
    if inferred is not None:
        header_row, col_map = inferred
        return header_row, col_map

    # 最后兜底：使用默认
    return 2, {
        "test_type": EXCEL_COL_TEST_TYPE,
        "node_type": EXCEL_COL_NODE_TYPE,
        "name": EXCEL_COL_NAME,
        "variant": EXCEL_COL_VARIANT,
        "level": EXCEL_COL_LEVEL,
    }


def find_test_case_sheet(wb):
    """在 Excel 工作簿中查找测试用例页面"""
    # 优先查找包含关键字的 sheet
    keywords = ["测试用例", "测试记录", "Test"]
    for name in wb.sheetnames:
        ws = wb[name]
        # 检查是否有足够的行和列
        if ws.max_row < 5 or ws.max_column < 10:
            continue
        # 检查是否包含 CANoe 相关内容
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_text = "".join(str(ws.cell(row=row_idx, column=c).value or "") for c in range(1, min(20, ws.max_column + 1)))
            if "CANoe" in row_text or "testmodule" in row_text.lower():
                return ws
    # 返回行数最多的 sheet
    best_ws = None
    best_rows = 0
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row > best_rows and ws.max_column >= 10:
            best_rows = ws.max_row
            best_ws = ws
    return best_ws


def parse_excel_rows(excel_path: Path, sheet_name: str = None):
    """解析 Excel 文件中的测试用例"""
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] 需要安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # 查找目标 sheet
    if sheet_name:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
    else:
        ws = find_test_case_sheet(wb)
    
    if ws is None:
        print(f"[ERROR] 未找到测试用例页面")
        return []

    print(f"[INFO] 使用 Excel 页面: {ws.title}")
    
    # 查找表头行和列映射
    header_row, col_map = find_header_row_and_cols(ws)
    # 确保注释列存在：列少时用 1–5，列多时用 6/7/16/17/18
    if ws.max_column >= EXCEL_COL_EXPECTED_RESULT:
        defaults = (("case_id", EXCEL_COL_CASE_ID), ("case_name", EXCEL_COL_CASE_NAME),
                    ("precondition", EXCEL_COL_PRECONDITION), ("steps", EXCEL_COL_STEPS),
                    ("expected_result", EXCEL_COL_EXPECTED_RESULT))
    else:
        defaults = (("case_id", 1), ("case_name", 2), ("precondition", 3), ("steps", 4), ("expected_result", 5))
    for k, c in defaults:
        if k not in col_map and c <= ws.max_column:
            col_map[k] = c
    print(f"[INFO] 表头行: {header_row}, 列映射: {col_map}")
    
    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        def get_cell(key):
            col = col_map.get(key)
            if col:
                val = ws.cell(row=row_idx, column=col).value
                if is_placeholder_text(val):
                    return ""
                return str(val).strip()
            return ""
        
        test_type_raw = get_cell("test_type")
        test_type = str(test_type_raw).strip().lower().replace(" ", "")
        node_type = get_cell("node_type").lower()
        name = get_cell("name")
        variant = get_cell("variant")
        level_str = get_cell("level")
        level = int(level_str) if level_str.isdigit() else -1

        case_id = get_cell("case_id")
        case_name = get_cell("case_name")
        precondition = get_cell("precondition")
        steps = get_cell("steps")
        expected_result = get_cell("expected_result")
        
        # 兼容不同 Excel 写法：Automatic / 自动 / Auto
        if not test_type:
            continue
        if ("automatic" not in test_type) and ("auto" not in test_type) and ("自动" not in test_type) and (test_type != "自动化"):
            continue
        if not node_type or node_type in {"null", "none", ""}:
            continue
        
        rows.append({
            "node_type": node_type,
            "name": name,
            "variant": variant,
            "level": level,
            "case_id": case_id,
            "case_name": case_name,
            "precondition": precondition,
            "steps": steps,
            "expected_result": expected_result,
        })
    
    return rows




def qname(tag: str):
    return f"{{{NS}}}{tag}"


def localname(tag: str):
    if tag.startswith("{"):
        return tag.split("}", 1)[1]
    return tag


def find_first_by_attr(root: ET.Element, tag: str, attr: str, value: str):
    for elem in root.iter(qname(tag)):
        if (elem.get(attr) or "") == value:
            return elem
    return None


def ensure_child(parent: ET.Element, tag: str):
    for child in parent:
        if localname(child.tag) == tag:
            return child
    child = ET.SubElement(parent, qname(tag))
    return child


def find_child_by_attr(parent: ET.Element, tag: str, attr: str, value: str):
    """仅在当前父节点的直接子节点中查找匹配元素。"""
    for child in parent:
        if localname(child.tag) != tag:
            continue
        if (child.get(attr) or "") == value:
            return child
    return None


def get_row_title(row: dict) -> str:
    """节点 title 统一取 Excel“名称”列，空时回退到函数名列"""
    case_name = (row.get("case_name") or "").strip()
    if case_name:
        return case_name
    return (row.get("name") or "").strip()


def get_module_title_from_rows(rows):
    """从解析结果中提取 testmodule 标题"""
    for row in rows:
        if row["node_type"] == "testmodule":
            title = get_row_title(row)
            if title:
                return title
    return "Test Module"


def create_empty_xml(xml_path: Path, title: str):
    """创建空的 CANoe XML 测试模块文件"""
    xml_content = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<testmodule xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
 xmlns="http://www.vector-informatik.de/CANoe/TestModule/1.8"
 xsi:schemaLocation="http://www.vector-informatik.de/CANoe/TestModule/1.8 testmodule.xsd"
 title="{title}" version="1.0">
  <variants>
  </variants>
  <description>Generated from Excel</description>
</testmodule>'''
    xml_path.write_text(xml_content, encoding="utf-8")
    print(f"[INFO] 已创建 XML 文件: {xml_path}")


def create_empty_can(can_path: Path):
    can_content = '''/*@!Encoding:936*/
includes
{
  
}

variables
{  

}

'''
    can_path.write_text(can_content, encoding="gbk")
    print(f"[INFO] 已创建 CAN 文件: {can_path}")


def backup_file(path: Path):
    """创建单份备份：<原文件名><后缀>.bak（仅保留一个，重复运行覆盖）"""
    if not path.exists():
        return
    backup_path = path.with_suffix(path.suffix + ".bak")
    shutil.copy2(path, backup_path)
    print(f"[INFO] 已备份: {backup_path}")


def update_xml(rows, xml_path: Path):
    ET.register_namespace("", NS)
    ET.register_namespace("xsi", "http://www.w3.org/2001/XMLSchema-instance")

    tree = ET.parse(xml_path)
    root = tree.getroot()
    # 根节点 title 也与 Excel 保持一致（取“名称”列）
    root.set("title", get_module_title_from_rows(rows))

    # 先收集 Excel 中所有使用的变体（单元格可能包含多个变体，空格或换行分隔）
    used_variants = set()
    for row in rows:
        variant_str = parse_variants(row.get("variant", ""))
        for var_name in variant_str.split():
            if var_name:
                used_variants.add(var_name)

    # 重建 variants 列表：清空并只添加实际使用的变体
    variants_parent = ensure_child(root, "variants")
    # 清空现有变体
    for child in list(variants_parent):
        variants_parent.remove(child)
    # 添加实际使用的变体（按名称排序）
    variant_names = set()
    for var_name in sorted(used_variants):
        v = ET.SubElement(variants_parent, qname("variant"))
        v.set("name", var_name)
        v.text = var_name
        variant_names.add(var_name)

    # 层级栈: [(level, element), ...]，level 0 对应 root
    level_stack = [(0, root)]
    current_section = None
    current_section_level = -1

    def pop_to_parent_level(target_level):
        """弹出栈直到栈顶层级 < target_level（用于 testgroup 找父节点）"""
        while len(level_stack) > 1 and level_stack[-1][0] >= target_level:
            level_stack.pop()

    def pop_to_current_level(target_level):
        """弹出栈直到栈顶层级 <= target_level（用于 preparation/completion/testcase 找所属节点）"""
        while len(level_stack) > 1 and level_stack[-1][0] > target_level:
            level_stack.pop()

    def get_current_parent():
        """获取当前栈顶元素作为父节点"""
        return level_stack[-1][1]

    for row in rows:
        node_type = row["node_type"]
        name = row["name"]
        title = get_row_title(row)
        variant = row["variant"]
        level = row["level"]

        if node_type == "testmodule":
            level_stack = [(0, root)]
            current_section = None
            current_section_level = -1
            continue

        if node_type == "testgroup":
            if not name and not title:
                continue
            # testgroup: 弹出到父层级（level < 当前），再创建/复用
            pop_to_parent_level(level)
            desired_title = title if title else name
            parent = get_current_parent()
            # 仅在当前父节点下复用，避免同名 testgroup 跨层级误复用
            found = find_child_by_attr(parent, "testgroup", "title", desired_title)
            # 兼容老策略：若历史上 title 用的是 name，则在当前父节点下按 name 再查一次
            if found is None and name and desired_title != name:
                found = find_child_by_attr(parent, "testgroup", "title", name)
            if found is not None:
                if desired_title:
                    found.set("title", desired_title)
                level_stack.append((level, found))
            else:
                new_group = ET.SubElement(parent, qname("testgroup"))
                new_group.set("title", desired_title)
                level_stack.append((level, new_group))
            current_section = None
            current_section_level = -1
            continue

        if node_type in {"preparation", "completion"}:
            # preparation/completion: 层级 N 表示它是层级 N-1 父节点的子节点（和 testgroup 同逻辑）
            pop_to_parent_level(level)
            parent = get_current_parent()
            current_section = ensure_child(parent, node_type)
            current_section_level = level
            continue

        if node_type == "capltestfunction":
            if not name:
                continue
            variant_str = parse_variants(variant)
            # 确保每个变体被添加到 variants 列表
            for var_name in variant_str.split():
                if var_name and var_name not in variant_names:
                    v = ET.SubElement(variants_parent, qname("variant"))
                    v.set("name", var_name)
                    v.text = var_name
                    variant_names.add(var_name)
            # 查找已存在的节点
            existing = find_first_by_attr(root, "capltestfunction", "name", name)
            if existing is not None:
                if title:
                    existing.set("title", title)
                # 始终同步 variants 属性：有值则设置，空值则删除（格式: "variant1 variant2"）
                if variant_str:
                    existing.set("variants", variant_str)
                elif "variants" in existing.attrib:
                    del existing.attrib["variants"]
                continue
            # 创建新节点
            if current_section is not None and level > current_section_level:
                parent = current_section
            else:
                if level <= current_section_level:
                    current_section = None
                    current_section_level = -1
                pop_to_current_level(level)
                parent = get_current_parent()
            node = ET.SubElement(parent, qname("capltestfunction"))
            node.set("name", name)
            node.set("title", title if title else name)
            if variant_str:
                node.set("variants", variant_str)
            continue

        if node_type == "capltestcase":
            if not name:
                continue
            variant_str = parse_variants(variant)
            # 确保每个变体被添加到 variants 列表
            for var_name in variant_str.split():
                if var_name and var_name not in variant_names:
                    v = ET.SubElement(variants_parent, qname("variant"))
                    v.set("name", var_name)
                    v.text = var_name
                    variant_names.add(var_name)
            # 查找已存在的节点
            existing = find_first_by_attr(root, "capltestcase", "name", name)
            if existing is not None:
                if title:
                    existing.set("title", title)
                # 始终同步 variants 属性：有值则设置，空值则删除（格式: "variant1 variant2"）
                if variant_str:
                    existing.set("variants", variant_str)
                elif "variants" in existing.attrib:
                    del existing.attrib["variants"]
                continue
            # 创建新节点
            if current_section is not None and level > current_section_level:
                parent = current_section
            else:
                if level <= current_section_level:
                    current_section = None
                    current_section_level = -1
                pop_to_current_level(level)
                parent = get_current_parent()
            node = ET.SubElement(parent, qname("capltestcase"))
            node.set("name", name)
            node.set("title", title if title else name)
            if variant_str:
                node.set("variants", variant_str)

    # Excel 已删除的用例：从 XML 中移除对应节点，与 Excel 保持一致
    excel_capl_names = set()
    for row in rows:
        nt = row["node_type"]
        nm = row.get("name", "").strip()
        if nt in ("capltestfunction", "capltestcase") and nm:
            excel_capl_names.add((nt, nm))
    parent_map = {c: p for p in root.iter() for c in p}
    to_remove = []
    for elem in root.iter():
        if localname(elem.tag) == "capltestfunction":
            n = (elem.get("name") or "").strip()
            if n and ("capltestfunction", n) not in excel_capl_names:
                to_remove.append(elem)
        elif localname(elem.tag) == "capltestcase":
            n = (elem.get("name") or "").strip()
            if n and ("capltestcase", n) not in excel_capl_names:
                to_remove.append(elem)
    for elem in to_remove:
        parent = parent_map.get(elem)
        if parent is not None:
            parent.remove(elem)

    ET.indent(tree, space="  ")
    tree.write(xml_path, encoding="utf-8", xml_declaration=True)


def has_capl_definition(text: str, node_type: str, name: str):
    if node_type == "capltestfunction":
        pattern = rf"\btestfunction\s+{re.escape(name)}\s*\("
    else:
        pattern = rf"\btestcase\s+{re.escape(name)}\s*\("
    return re.search(pattern, text) is not None


def _format_comment_lines(text: str) -> list:
    """将文本转为注释行列表，每行以 // 开头，避免 */ 破坏块注释"""
    if not text or not str(text).strip():
        return []
    lines = str(text).strip().replace("\r\n", "\n").replace("\r", "\n").split("\n")
    return [f"// {line.strip()}" for line in lines if line.strip()]


def build_comment_block(row: dict) -> str:
    """生成注释块（仅用例 ID、名称）"""
    name = row["name"]
    case_id = (row.get("case_id") or "").strip()
    case_name = (row.get("case_name") or "").strip() or name

    comment_lines = ["/// <Auto Generated>"]
    if case_id:
        comment_lines.extend(_format_comment_lines(f"用例 ID: {case_id}"))
    if case_name:
        comment_lines.extend(_format_comment_lines(f"名称: {case_name}"))

    return "\n".join(comment_lines)


def _escape_capl_string(s: str) -> str:
    return s.replace("\\", "\\\\").replace("\"", "\\\"")


def _split_nonempty_lines(text: str) -> list:
    if not text:
        return []
    lines = str(text).replace("\r\n", "\n").replace("\r", "\n").split("\n")
    out = []
    for ln in lines:
        t = ln.strip()
        if not t:
            continue
        if is_placeholder_text(t):
            continue
        out.append(t)
    return out


TBD_REBUILD_MARKER = "// TBD，实现时请先删除本行，否则下次重新生成时将重建。"


def _extract_step_seq_and_text(line: str, fallback_seq: int):
    m = re.match(r"^\s*(\d+)\s*[\.、，,\)\]]?\s*(.*)$", line)
    if m:
        seq = m.group(1)
        content = m.group(2).strip() if m.group(2) else line.strip()
        return seq, content
    return str(fallback_seq), line.strip()


def _extract_optional_seq_and_text(line: str):
    """提取可选序号；若无序号则返回 (None, 原文)。"""
    m = re.match(r"^\s*(\d+)\s*[\.、，,\)\]]?\s*(.*)$", line)
    if m:
        seq = m.group(1)
        content = (m.group(2) or "").strip()
        return seq, content
    return None, line.strip()


def _strip_leading_seq(text: str) -> str:
    """去掉文本前缀序号（如 1. / 2、），避免在 testStep 第二参数重复展示"""
    if not text:
        return ""
    m = re.match(r"^\s*\d+\s*[\.、，,\)\]]?\s*(.*)$", text.strip())
    if m:
        return (m.group(1) or "").strip()
    return text.strip()


def build_test_step_calls(row: dict, func_name: str) -> list:
    """将步骤/预期结果解析为 testStep 调用行"""
    step_lines = _split_nonempty_lines(row.get("steps", ""))
    expected_lines = _split_nonempty_lines(row.get("expected_result", ""))

    # 预期结果优先按“编号”匹配步骤；无编号项按顺序兜底
    expected_by_seq = {}
    expected_no_seq = []
    for exp_line in expected_lines:
        exp_seq, exp_text = _extract_optional_seq_and_text(exp_line)
        exp_text = _strip_leading_seq(exp_text)
        if not exp_text:
            continue
        if exp_seq:
            expected_by_seq[exp_seq] = exp_text
        else:
            expected_no_seq.append(exp_text)
    no_seq_idx = 0

    calls = []
    for idx, step_line in enumerate(step_lines, start=1):
        seq, step_text = _extract_step_seq_and_text(step_line, idx)
        step_text = _strip_leading_seq(step_text)
        expected_text = expected_by_seq.get(seq, "")
        if not expected_text and no_seq_idx < len(expected_no_seq):
            expected_text = expected_no_seq[no_seq_idx]
            no_seq_idx += 1
        if expected_text:
            if step_text and step_text[-1] in "；;，,。.":
                msg = f"步骤：{step_text}预期结果：{expected_text}"
            else:
                msg = f"步骤：{step_text}，预期结果：{expected_text}"
        else:
            msg = f"步骤：{step_text}"
        calls.append(f"  testStep(\"{_escape_capl_string(seq)}\", \"{_escape_capl_string(msg)}\");")

    if not calls:
        calls.append(f"  testStep(\"{_escape_capl_string(func_name)}\", \"Auto-generated placeholder.\");")
    return calls


def build_capl_stub(row: dict):
    """生成 CAPL 函数代码（注释: 用例 ID/名称；正文: 步骤与预期结果）"""
    node_type = row["node_type"]
    name = row["name"]
    comment_block = build_comment_block(row)
    step_calls = "\n".join(build_test_step_calls(row, name))

    if node_type == "capltestfunction":
        return (
            f"{comment_block}\n"
            f"testfunction {name}()\n"
            "{\n"
            f"{step_calls}\n"
            "}\n"
        )
    return (
        f"{comment_block}\n"
        f"testcase {name}()\n"
        "{\n"
        f"  {TBD_REBUILD_MARKER}\n"
        "  testCaseStartLogging();\n"
        f"{step_calls}\n"
        "  resetEnv();\n"
        "  testCaseStopLogging();\n"
        "}\n"
    )


def _update_function_comment(text: str, node_type: str, name: str, new_comment: str) -> str:
    """替换目标函数正上方的注释块，避免误删其他函数"""
    kw = "testfunction" if node_type == "capltestfunction" else "testcase"
    func_pattern = rf"({re.escape(kw)}\s+{re.escape(name)}\s*\()"
    m = re.search(func_pattern, text)
    if not m:
        return text
    func_start = m.start()
    before = text[:func_start]
    # 在目标函数之前找最近的 /// 或 // <Auto Generated>
    comment_markers = list(re.finditer(r"^(///? <Auto Generated>)", before, re.MULTILINE))
    if not comment_markers:
        return text
    # 取最后一个，即紧挨目标函数的那个注释块起点
    block_start = comment_markers[-1].start()
    return text[:block_start] + new_comment + "\n" + text[m.start():]


def _ensure_reset_env_call(text: str, testcase_name: str) -> str:
    """确保指定 testcase 在 testCaseStopLogging 前调用 resetEnv。"""
    pattern = rf"(testcase\s+{re.escape(testcase_name)}\s*\(\)\s*\{{[\s\S]*?)(\n\s*testCaseStopLogging\(\);)"
    m = re.search(pattern, text)
    if not m:
        return text
    before_stop = m.group(1)
    if re.search(r"\n\s*resetEnv\(\);\s*$", before_stop):
        return text
    replacement = before_stop + "\n  resetEnv();" + m.group(2)
    return text[:m.start()] + replacement + text[m.end():]


def _is_empty_testcase(text: str, testcase_name: str) -> bool:
    """
    判断 testcase 是否为空实现：
    - 不包含 testStep(...)
    - 函数体内除 logging/reset 调用外无其他有效语句
    """
    pattern = rf"testcase\s+{re.escape(testcase_name)}\s*\(\)\s*\{{([\s\S]*?)\n\}}"
    m = re.search(pattern, text)
    if not m:
        return False
    body = m.group(1)
    if re.search(r"\btestStep\s*\(", body):
        return False

    # 去掉注释与空白后，检查是否只剩允许的占位调用
    lines = []
    for ln in body.replace("\r\n", "\n").split("\n"):
        s = ln.strip()
        if not s or s.startswith("//"):
            continue
        lines.append(s)

    allowed = {
        TBD_REBUILD_MARKER,
        "testCaseStartLogging();",
        "testCaseStopLogging();",
        "resetEnv();",
    }
    return all(ln in allowed for ln in lines)


def _remove_testcase_if_tbd(text: str, testcase_name: str):
    """若 testcase 内存在 TBD 标记，则删除该函数（及其自动生成注释块）"""
    func_match = re.search(rf"testcase\s+{re.escape(testcase_name)}\s*\(\)\s*\{{", text)
    if not func_match:
        return text, False

    func_start = func_match.start()
    brace_start = text.find("{", func_match.start())
    if brace_start < 0:
        return text, False

    depth = 0
    end_idx = -1
    for i in range(brace_start, len(text)):
        ch = text[i]
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                end_idx = i
                break
    if end_idx < 0:
        return text, False

    block = text[func_start:end_idx + 1]
    if TBD_REBUILD_MARKER not in block:
        return text, False

    # 尝试同时删除紧邻的自动生成注释块
    remove_start = func_start
    before = text[:func_start]
    marker_pos = max(before.rfind("/// <Auto Generated>"), before.rfind("// <Auto Generated>"))
    if marker_pos >= 0:
        seg = before[marker_pos:func_start]
        lines = seg.splitlines()
        if all((not ln.strip()) or ln.lstrip().startswith("//") for ln in lines):
            remove_start = marker_pos

    new_text = text[:remove_start] + text[end_idx + 1:]
    return new_text, True


def _cleanup_can_artifacts(text: str) -> str:
    """清理历史生成残留并收敛空行格式。"""
    cleaned_lines = []
    for ln in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        if ln.strip() == "/":
            continue
        cleaned_lines.append(ln)
    cleaned = "\n".join(cleaned_lines)
    # 将 3 个及以上连续空行压缩为 2 个换行（即最多保留 1 个空白行）
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    # 去掉文件尾多余空行，保留一个结尾换行
    return cleaned.rstrip() + "\n"


def _find_function_span(text: str, node_type: str, name: str):
    """查找函数定义在文本中的 [start, end) 区间（含函数体）。"""
    kw = "testfunction" if node_type == "capltestfunction" else "testcase"
    m = re.search(rf"{re.escape(kw)}\s+{re.escape(name)}\s*\(\)\s*\{{", text)
    if not m:
        return None
    start = m.start()
    brace_start = text.find("{", m.start())
    if brace_start < 0:
        return None
    depth = 0
    end_idx = -1
    for i in range(brace_start, len(text)):
        ch = text[i]
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                end_idx = i
                break
    if end_idx < 0:
        return None
    return start, end_idx + 1


def _insert_after_prev_function(text: str, prev_row: dict, new_block: str) -> str:
    """把新函数块插入到前一个函数块之后。"""
    span = _find_function_span(text, prev_row["node_type"], prev_row["name"])
    if not span:
        return text
    _, end = span
    insertion = "\n\n" + new_block.rstrip() + "\n"
    return text[:end] + insertion + text[end:]


def _collect_capl_order_from_xml(xml_path: Path):
    """按 XML 出现顺序收集 capltestfunction/capltestcase。"""
    if not xml_path or (not xml_path.exists()):
        return []
    tree = ET.parse(xml_path)
    root = tree.getroot()
    ordered = []
    for elem in root.iter():
        tag = localname(elem.tag)
        if tag not in {"capltestfunction", "capltestcase"}:
            continue
        name = (elem.get("name") or "").strip()
        if not name:
            continue
        ordered.append((tag, name))
    return ordered


def update_can(rows, can_path: Path, xml_path: Path = None):
    if can_path.exists():
        text, _ = read_text_with_fallback(can_path)
    else:
        text = ""

    # 以 XML 节点顺序驱动 CAN 生成顺序
    row_map = {}
    for row in rows:
        nt = row.get("node_type", "")
        nm = (row.get("name") or "").strip()
        if nt in {"capltestfunction", "capltestcase"} and nm:
            row_map[(nt, nm)] = row

    ordered_rows = []
    seen = set()
    for key in _collect_capl_order_from_xml(xml_path):
        nt, nm = key
        row = row_map.get(key)
        if row is None:
            # XML 存在但 Excel 行缺失时，兜底生成最小信息
            row = {
                "node_type": nt,
                "name": nm,
                "case_id": "",
                "case_name": nm,
                "steps": "",
                "expected_result": "",
            }
        ordered_rows.append(row)
        seen.add(key)
    # XML 中没有但 rows 里有的，追加在最后
    for key, row in row_map.items():
        if key not in seen:
            ordered_rows.append(row)

    append_blocks = []
    updated_text = text
    for idx, row in enumerate(ordered_rows):
        node_type = row["node_type"]
        name = row["name"]
        if node_type not in {"capltestfunction", "capltestcase"}:
            continue
        if not name:
            continue
        if has_capl_definition(updated_text, node_type, name):
            if node_type == "capltestcase":
                updated_text, removed = _remove_testcase_if_tbd(updated_text, name)
                if removed:
                    block = build_capl_stub(row)
                    inserted = False
                    for p in range(idx - 1, -1, -1):
                        prev_row = ordered_rows[p]
                        if has_capl_definition(updated_text, prev_row["node_type"], prev_row["name"]):
                            updated_text = _insert_after_prev_function(updated_text, prev_row, block)
                            inserted = True
                            break
                    if not inserted:
                        append_blocks.append(block)
                    continue
                # testcase 已实现（无 TBD 标记）时，保留原注释与函数体，不做覆盖更新。
                if _is_empty_testcase(updated_text, name):
                    updated_text = _ensure_reset_env_call(updated_text, name)
                continue
            # 已存在：刷新注释
            new_comment = build_comment_block(row)
            updated_text = _update_function_comment(updated_text, node_type, name, new_comment)
            if node_type == "capltestcase" and _is_empty_testcase(updated_text, name):
                updated_text = _ensure_reset_env_call(updated_text, name)
        else:
            block = build_capl_stub(row)
            inserted = False
            # 参考前一个函数排序：优先插在前一个已存在函数之后
            for p in range(idx - 1, -1, -1):
                prev_row = ordered_rows[p]
                if has_capl_definition(updated_text, prev_row["node_type"], prev_row["name"]):
                    updated_text = _insert_after_prev_function(updated_text, prev_row, block)
                    inserted = True
                    break
            if not inserted:
                append_blocks.append(block)

    # 追加新函数
    if append_blocks:
        insertion = "\n" + "\n".join(append_blocks) + "\n"
        marker = "/**************************************"
        marker_pos = updated_text.find(marker)
        if marker_pos >= 0:
            updated_text = updated_text[:marker_pos] + insertion + updated_text[marker_pos:]
        else:
            updated_text = updated_text.rstrip() + insertion

    updated_text = _cleanup_can_artifacts(updated_text)

    if append_blocks or updated_text != text:
        can_path.write_text(updated_text, encoding="gbk")


def main():
    parser = argparse.ArgumentParser(
        description="根据 Excel 测试用例文件增量同步 CANoe XML 与 CAPL 脚本（仅补缺，不覆盖已有内容；文件不存在则自动创建）"
    )
    parser.add_argument("excel", type=Path, help="Excel 测试用例文件路径（.xlsx 或 .xls）")
    parser.add_argument("xml", type=Path, nargs="?", default=None, help="XML 文件路径（可选，默认与 Excel 同名）")
    parser.add_argument("can", type=Path, nargs="?", default=None, help="CAN 文件路径（可选，默认与 XML 同名）")
    parser.add_argument("--xml-only", action="store_true", help="仅生成/同步 XML，不创建和更新 CAN")
    args = parser.parse_args()

    # XML 文件路径：如果未指定，则与 Excel 同名但扩展名为 .xml
    xml_path = args.xml if args.xml else args.excel.with_suffix(".xml")
    # CAN 文件路径：如果未指定，则与 XML 同名但扩展名为 .can
    can_path = args.can if args.can else xml_path.with_suffix(".can")

    if not args.excel.exists():
        print(f"[ERROR] Excel 文件不存在: {args.excel}")
        sys.exit(1)

    rows = parse_excel_rows(args.excel)
    if not rows:
        print("[WARN] 未发现可用的 Automatic 行，未做修改。")
        return

    print(f"[INFO] 解析到 {len(rows)} 条自动化测试项")

    xml_path.parent.mkdir(parents=True, exist_ok=True)
    if not args.xml_only:
        can_path.parent.mkdir(parents=True, exist_ok=True)

    if not xml_path.exists():
        title = get_module_title_from_rows(rows)
        create_empty_xml(xml_path, title)

    if (not args.xml_only) and (not can_path.exists()):
        create_empty_can(can_path)

    # 修改前备份（最多保留一个备份文件，重复运行覆盖）
    backup_file(xml_path)
    if not args.xml_only:
        backup_file(can_path)

    update_xml(rows, xml_path)
    if not args.xml_only:
        update_can(rows, can_path, xml_path)
    print(f"[OK] 同步完成：")
    print(f"     XML: {xml_path}")
    if not args.xml_only:
        print(f"     CAN: {can_path}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Generate a CANdelaStudio-compatible JAC PDX from a JAC survey workbook.

The JAC workbook adapter below maps the OEM-specific 01/02/03/04 sheets into
the shared canonical diagnostic model. The ODX writer reuses the mature flat
service and CANdela compatibility logic from the VF generator base.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import importlib.util
import re
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Iterable

from lxml import etree
from openpyxl import load_workbook


def load_vf_writer_base() -> Any:
    here = Path(__file__).resolve().parent
    candidates = [
        here / "pdxBase_VF.py",
        here.parent / "VF" / "pdxBase_VF.py",
    ]
    for path in candidates:
        if path.exists():
            spec = importlib.util.spec_from_file_location("pdxBase_VF", path)
            if spec is None or spec.loader is None:
                continue
            module = importlib.util.module_from_spec(spec)
            sys.modules[spec.name] = module
            spec.loader.exec_module(module)
            return module
    searched = ", ".join(str(path) for path in candidates)
    raise FileNotFoundError(f"Cannot find the reusable ODX writer base. Searched: {searched}")


BASE = load_vf_writer_base()
XSI_TYPE = BASE.XSI_TYPE

Conversion = BASE.Conversion
ParamDef = BASE.ParamDef
DidDef = BASE.DidDef
IoDidDef = BASE.IoDidDef
RoutineDef = BASE.RoutineDef
RoutineSubFunction = BASE.RoutineSubFunction
DtcDef = BASE.DtcDef
SnapshotDef = BASE.SnapshotDef
ExtendedRecordDef = BASE.ExtendedRecordDef
CoverInfo = BASE.CoverInfo
SurveyData = BASE.SurveyData


JAC_KEEP_FILES = {
    "ISO_11898_2_DWCAN.odx-cs",
    "ISO_11898_3_DWFTCAN.odx-cs",
    "ISO_15765_2.odx-cs",
    "ISO_15765_3.odx-cs",
    "ISO_15765_3_on_ISO_15765_2.odx-c",
    "SAE_J2411_SWCAN.odx-cs",
    "JAC_ECU_CAN_v15.odx-d",
    "index.xml",
}

CANDELA_SHORT_NAME_MAX_LEN = 64
UUDT_DISABLED_CAN_ID = 0xFFFFFFFF


def cell_text(value: Any) -> str:
    return BASE.cell_text(value)


def compact_text(value: Any) -> str:
    return BASE.compact_text(value)


def normalize_access(value: Any) -> str:
    return BASE.normalize_access(value)


def parse_int_cell(value: Any, default: int = 0) -> int:
    return BASE.parse_int_cell(value, default)


def parse_hex_cell(value: Any, *, max_value: int = 0xFFFFFFFF) -> int | None:
    return BASE.parse_hex_cell(value, max_value=max_value)


def parse_hex_in_text(value: Any, *, max_value: int = 0xFFFFFFFF) -> int | None:
    text = compact_text(value)
    match = re.search(r"0[xX]([0-9A-Fa-f]{1,8})|(?<![A-Za-z0-9])([0-9A-Fa-f]{2,8})(?![A-Za-z0-9])", text)
    if not match:
        return None
    raw = match.group(1) or match.group(2)
    result = int(raw, 16)
    return result if result <= max_value else None


def parse_hex_loose(value: Any, *, max_value: int = 0xFFFFFFFF) -> int | None:
    text = compact_text(value).replace("`", "").strip()
    if not text or text.upper() in {"/", "N/A", "NA", "ALL"}:
        return None
    parsed = parse_hex_cell(text, max_value=max_value)
    if parsed is not None:
        return parsed
    return parse_hex_in_text(text, max_value=max_value)


def find_sheet(workbook: Any, predicate: Any) -> Any | None:
    return BASE.find_sheet(workbook, predicate)


def sanitize_short_name(value: str, fallback: str, used: set[str] | None = None, max_len: int = 120) -> str:
    return BASE.sanitize_short_name(value, fallback, used, max_len)


def hex_short(prefix: str, value: int, width: int = 4) -> str:
    return BASE.hex_short(prefix, value, width)


def usable_text(value: Any) -> str:
    text = compact_text(value)
    if text.upper() in {"", "/", "N/A", "NA", "NONE", "NULL"}:
        return ""
    return text


def normalize_unit_text(value: Any) -> str:
    text = usable_text(value)
    normalized = text.replace("（", "(").replace("）", ")")
    if re.fullmatch(r"-?\s*\(?\s*no\s*unit\s*\)?", normalized, flags=re.IGNORECASE):
        return ""
    return text


def dual_name(english_value: Any, chinese_value: Any = "", fallback: str = "") -> tuple[str, str]:
    english = usable_text(english_value)
    chinese = usable_text(chinese_value)
    if not english and not chinese:
        return fallback, fallback
    if english and chinese and english != chinese:
        return english, f"{english} / {chinese}"
    return english or chinese or fallback, english or chinese or fallback


def normalize_conversion_text(value: Any) -> str:
    text = cell_text(value)
    replacements = {
        "\u3000": " ",
        "?": ";",
        "?": ":",
        "?": ",",
        "?": "~",
        "?": "-",
        "?": "-",
        "\r": "\n",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.strip()


LINEAR_NUMBER_PATTERN = r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)"


def decimal_precision_from_token(token: str | None) -> int:
    if not token:
        return 0
    match = re.search(r"\.(\d+)", token)
    return len(match.group(1)) if match else 0


def parse_float_assignment_token(text: str, key: str, default: float) -> tuple[float, str | None]:
    match = re.search(rf"\b{re.escape(key)}\s*=\s*({LINEAR_NUMBER_PATTERN})", text, flags=re.IGNORECASE)
    if not match:
        return default, None
    token = match.group(1)
    return float(token), token


def is_conversion_directive_line(line: str) -> bool:
    return re.fullmatch(r"\s*(?:type|value|a|b|precision)\s*=.*", line, flags=re.IGNORECASE) is not None


def parse_conversion(value: Any) -> Conversion:
    text = normalize_conversion_text(value)
    if not text or text.upper() in {"/", "N/A", "NA"}:
        return Conversion()
    if re.search(r"\btype\s*=\s*identical\b", text, flags=re.IGNORECASE):
        return Conversion()
    if re.search(r"\btype\s*=\s*bcd\b", text, flags=re.IGNORECASE):
        return Conversion(kind="bcd")

    linear = re.search(
        rf"(?:phy|phys|physical|y)\s*=\s*(?:raw|XX|xx|X|x)"
        rf"(?:\s*\*\s*(?P<a>{LINEAR_NUMBER_PATTERN}))?"
        rf"(?:\s*(?P<sign>[+-])\s*(?P<b>{LINEAR_NUMBER_PATTERN}))?",
        text,
        flags=re.IGNORECASE,
    )
    if linear:
        a_token = linear.group("a")
        b_token = linear.group("b")
        a = float(a_token) if a_token else 1.0
        b = float(b_token) if b_token else 0.0
        if linear.group("sign") == "-":
            b = -b
        precision = max(decimal_precision_from_token(a_token), decimal_precision_from_token(b_token))
        return Conversion(kind="linear", a=a, b=b, precision=precision if precision > 0 else None)

    if re.search(r"\btype\s*=\s*a\s*x\s*\+\s*b\b|\by\s*=\s*a\s*\*?\s*x\s*\+\s*b\b", text, flags=re.IGNORECASE):
        a, a_token = parse_float_assignment_token(text, "a", 1.0)
        b, b_token = parse_float_assignment_token(text, "b", 0.0)
        precision_match = re.search(r"\bprecision\s*=\s*(-?\d+)", text, flags=re.IGNORECASE)
        precision = (
            int(precision_match.group(1))
            if precision_match
            else max(decimal_precision_from_token(a_token), decimal_precision_from_token(b_token))
        )
        return Conversion(kind="linear", a=a, b=b, precision=precision if precision > 0 else None)

    if re.search(r"\ba\s*=", text, flags=re.IGNORECASE) and re.search(r"\bb\s*=", text, flags=re.IGNORECASE):
        a, a_token = parse_float_assignment_token(text, "a", 1.0)
        b, b_token = parse_float_assignment_token(text, "b", 0.0)
        c, c_token = parse_float_assignment_token(text, "c", 1.0)
        if c == 0:
            c = 1.0
        precision = max(
            decimal_precision_from_token(a_token),
            decimal_precision_from_token(b_token),
            decimal_precision_from_token(c_token),
        )
        return Conversion(kind="linear", a=a / c, b=b / c, precision=precision if precision > 0 else None)

    enum_entries: list[tuple[int, int, str]] = []
    enum_pattern = re.compile(
        r"(?:0[xX])?([0-9A-Fa-f]+)\s*(?:[-~]\s*(?:0[xX])?([0-9A-Fa-f]+))?\s*[:=]\s*([^;\n\r]+)"
    )
    for line in re.split(r"[\r\n;]+", text):
        if is_conversion_directive_line(line):
            continue
        for match in enum_pattern.finditer(line):
            label = match.group(3).strip(" ;,")
            if not label:
                continue
            lo = int(match.group(1), 16)
            hi = int(match.group(2), 16) if match.group(2) else lo
            enum_entries.append((lo, hi, label))
    if enum_entries:
        return Conversion(kind="enum", enum=enum_entries)

    inline_type_match = re.search(r"\btype\s*=\s*\w+\b", text, flags=re.IGNORECASE)
    data_text = text[inline_type_match.end():] if inline_type_match else text
    for line in re.split(r"[\r\n;]+", data_text):
        if is_conversion_directive_line(line):
            continue
        for match in enum_pattern.finditer(line):
            label = match.group(3).strip(" ;,")
            if not label:
                continue
            lo = int(match.group(1), 16)
            hi = int(match.group(2), 16) if match.group(2) else lo
            enum_entries.append((lo, hi, label))
    if enum_entries:
        return Conversion(kind="enum", enum=enum_entries)

    bare_enum_pattern = re.compile(
        r"(?<![0-9A-Za-z])(?:0[xX])?([0-9A-Fa-f]{1,2})\s+([^/;\n\r]+)"
    )
    for line in re.split(r"[/;\n\r]+", data_text):
        if is_conversion_directive_line(line):
            continue
        match = bare_enum_pattern.search(line.strip())
        if not match:
            continue
        label = match.group(2).strip(" ;,")
        if label:
            value = int(match.group(1), 16)
            enum_entries.append((value, value, label))
    if enum_entries:
        return Conversion(kind="enum", enum=enum_entries)

    return Conversion()


def choose_conversion(primary_value: Any, fallback_value: Any = "") -> Conversion:
    """Parse both conversion columns and keep the richer physical meaning.

    Legacy survey sheets normally put English expressions in column L and Chinese text
    in column M. In a few rows one column contains a pass-through formula while
    the other carries the value table, so enum > linear > identity gives CANdela
    the more useful data type without hard-coding row numbers.
    """

    primary = parse_conversion(primary_value)
    fallback = parse_conversion(fallback_value)
    rank = {"enum": 4, "linear": 3, "bcd": 2, "identity": 1}
    if rank.get(fallback.kind, 0) > rank.get(primary.kind, 0):
        return fallback
    return primary


def parse_index_range(value: Any, *, size: int = 0, default: int = 0) -> tuple[int, int]:
    text = compact_text(value).upper()
    if not text or text in {"/", "N/A", "NA"}:
        return default, default
    if text == "ALL":
        return 0, max(0, size - 1)
    nums = [int(num) for num in re.findall(r"\d+", text)]
    if not nums:
        return default, default
    start = nums[0]
    end = nums[1] if len(nums) > 1 else start
    if end < start:
        end = start
    return start, end


def parse_bit_range(value: Any, byte_start: int, byte_end: int, *, size: int = 0) -> tuple[int, int]:
    text = compact_text(value).upper()
    byte_count = max(1, byte_end - byte_start + 1)
    if not text or text in {"/", "N/A", "NA", "ALL"}:
        return 0, byte_count * 8
    nums = [int(num) for num in re.findall(r"\d+", text)]
    if not nums:
        return 0, byte_count * 8
    if len(nums) > 1 and ("-" in text or "~" in text):
        start, end = nums[0], nums[1]
        if end < start:
            end = start
        return start % 8, min(max(1, end - start + 1), byte_count * 8)
    return nums[0] % 8, 1


def make_param_from_cells(
    *,
    name_value: Any,
    chinese_name_value: Any = "",
    byte_value: Any,
    bit_value: Any,
    data_type_value: Any,
    unit_value: Any = "",
    conversion_value: Any = "",
    conversion_fallback_value: Any = "",
    min_value: Any = "",
    max_value: Any = "",
    fallback_name: str = "Data",
    size: int = 0,
) -> ParamDef | None:
    english, long_name = dual_name(name_value, chinese_name_value, fallback_name)
    if not usable_text(english) and not usable_text(long_name):
        english = fallback_name
        long_name = fallback_name
    byte_start, byte_end = parse_index_range(byte_value, size=size, default=0)
    bit_pos, bit_len = parse_bit_range(bit_value, byte_start, byte_end, size=size)
    data_type = usable_text(data_type_value) or "Hex(Unsigned)"
    conversion = choose_conversion(
        usable_text(conversion_value),
        usable_text(conversion_fallback_value),
    )
    if bit_len > 32 and conversion.kind == "enum":
        conversion = Conversion()
    return ParamDef(
        name=english or fallback_name,
        long_name=long_name or english or fallback_name,
        byte_pos=byte_start,
        bit_pos=bit_pos,
        bit_len=bit_len,
        data_type=data_type,
        unit=normalize_unit_text(unit_value),
        conversion=conversion,
        min_value=usable_text(min_value),
        max_value=usable_text(max_value),
    )


def jac_make_param(
    *,
    name_value: Any,
    chinese_name_value: Any = "",
    byte_value: Any,
    bit_value: Any,
    data_type_value: Any,
    unit_value: Any = "",
    conversion_value: Any = "",
    conversion_fallback_value: Any = "",
    min_value: Any = "",
    max_value: Any = "",
    fallback_name: str = "Data",
    size: int = 0,
) -> ParamDef | None:
    param = make_param_from_cells(
        name_value=name_value,
        chinese_name_value=chinese_name_value,
        byte_value=byte_value,
        bit_value=bit_value,
        data_type_value=data_type_value,
        unit_value=unit_value,
        conversion_value=conversion_value,
        conversion_fallback_value=conversion_fallback_value,
        min_value=min_value,
        max_value=max_value,
        fallback_name=fallback_name,
        size=size,
    )
    if param is not None and size > 0 and param.byte_pos >= size:
        param.byte_pos = max(0, size - 1)
    return param


def merge_session_access(read_values: Iterable[Any], write_values: Iterable[Any]) -> list[str]:
    sessions: list[str] = []
    for read_value, write_value in zip(read_values, write_values, strict=False):
        flags = ""
        if usable_text(read_value).upper() != "N":
            flags += "R"
        if usable_text(write_value).upper() != "N":
            flags += "W"
        sessions.append(flags)
    return sessions


def first_security_level(values: Iterable[Any]) -> str:
    for value in values:
        text = usable_text(value)
        if text and text.upper() != "N":
            return text
    return "N"


def is_supported_flag(value: Any, *, default: bool = False) -> bool:
    text = compact_text(value).upper()
    if text in {"Y", "YES", "TBD"}:
        return True
    if text in {"N", "NO", "-", "/", "N/A", "NA"}:
        return False
    return default


def merge_security_text(old_value: str, new_value: str) -> str:
    old_text = usable_text(old_value) or "N"
    new_text = usable_text(new_value) or "N"
    if old_text.upper() == "N":
        return new_text
    if new_text.upper() == "N" or new_text in old_text.split("&"):
        return old_text
    parts = [part for part in old_text.split("&") if part]
    for part in new_text.split("&"):
        if part and part not in parts:
            parts.append(part)
    return "&".join(parts) if parts else "N"


def vf_security_text(value: Any) -> str:
    text = usable_text(value) or "N"
    normalized = normalize_access(text)
    if normalized in {"", "N", "NO", "NONE", "LEVEL_0", "LEVEL0"}:
        return "N"
    return text


def vf_hex_tokens(text: str) -> list[int]:
    values: list[int] = []
    token_pattern = r"(?i)(?<![0-9a-z])(?:0x[0-9a-f]{1,2}|[0-9][0-9a-f]?|[a-f][0-9a-f])(?=$|[^0-9a-z])"
    for token in re.findall(token_pattern, text or ""):
        try:
            values.append(int(token, 16) if token.lower().startswith("0x") else int(token, 16))
        except ValueError:
            continue
    return values


def vf_subfunction_values(value_text: str, meaning_text: str) -> list[int]:
    """Return base sub-functions, normalizing suppress-positive-response values."""

    search_text = meaning_text.split("=", 1)[0] if "=" in meaning_text else value_text
    values = []
    for value in vf_hex_tokens(search_text):
        if value <= 0xFF:
            values.append(value & 0x7F)
    return list(dict.fromkeys(values))


def vf_parse_diagnostics_services_access(workbook: Any) -> dict[tuple[int, int | None], dict[str, Any]]:
    """Parse VF's Diagnostics Services support matrix.

    VF keeps the core UDS service support flags in one worksheet:
    Application Default/Extended are columns G/H, Bootloader
    Default/Programming are columns I/J, and each request sub-function row
    carries the supported sessions for that sub-function.
    """

    sheet = find_sheet(workbook, lambda name: name.strip() == "Diagnostics Services")
    if sheet is None:
        return {}

    access: dict[tuple[int, int | None], dict[str, Any]] = {}
    current_service_id: int | None = None
    current_service_name = ""
    current_security = "N"
    in_request_block = False
    in_subfunction_rows = False

    for row in range(1, sheet.max_row + 1):
        first_col = usable_text(sheet.cell(row, 1).value) or ""
        first_norm = first_col.strip().lower()
        if first_col and "$" in first_col and "service" in first_norm:
            current_service_name = first_col
            current_service_id = None
            current_security = "N"
            in_request_block = False
            in_subfunction_rows = False
            continue

        if first_norm.startswith("block"):
            in_request_block = "request" in first_norm
            if in_request_block:
                current_service_id = None
                current_security = "N"
                in_subfunction_rows = False
            continue

        if not in_request_block:
            continue

        description = usable_text(sheet.cell(row, 3).value) or ""
        value_text = usable_text(sheet.cell(row, 4).value) or ""
        meaning_text = usable_text(sheet.cell(row, 5).value) or ""
        description_norm = description.strip().lower()

        if "request sid" in description_norm:
            service_id = parse_hex_in_text(value_text, max_value=0xFF)
            if service_id is not None:
                current_service_id = service_id
                current_security = vf_security_text(sheet.cell(row, 13).value)
            continue

        if current_service_id is None:
            continue
        if current_service_id not in {0x10, 0x27, 0x28}:
            continue

        is_subfunction_header = any(
            marker in description_norm
            for marker in ("session type", "sub-function", "sub function", "control type")
        )
        if description_norm and not is_subfunction_header:
            in_subfunction_rows = False
        if is_subfunction_header:
            in_subfunction_rows = True
        elif not in_subfunction_rows or "=" not in meaning_text:
            continue

        requirement = compact_text(sheet.cell(row, 6).value).upper()
        if requirement == "U":
            continue

        subfunctions = vf_subfunction_values(value_text, meaning_text)
        if not subfunctions:
            continue

        sessions = {
            state_name
            for col, state_name in ((7, "Default"), (8, "Extended"), (9, "Default"), (10, "Programming"))
            if is_supported_flag(sheet.cell(row, col).value)
        }
        if not sessions:
            continue

        sources = {
            source
            for col, source in ((7, "Application"), (8, "Application"), (9, "Boot"), (10, "Boot"))
            if is_supported_flag(sheet.cell(row, col).value)
        }
        if not sources:
            continue

        security = vf_security_text(sheet.cell(row, 13).value) if usable_text(sheet.cell(row, 13).value) else current_security
        subfunction_name = meaning_text.split("=", 1)[1].split("/", 1)[0].strip() if "=" in meaning_text else description
        for subfunction in subfunctions:
            key = (current_service_id, subfunction)
            item = access.setdefault(
                key,
                {
                    "service_id": current_service_id,
                    "subfunction": subfunction,
                    "service_name": current_service_name,
                    "subfunction_name": subfunction_name,
                    "sessions": set(),
                    "security": "N",
                    "sources": set(),
                },
            )
            item["sessions"].update(sessions)
            item["security"] = merge_security_text(item["security"], security)
            item["sources"].update(sources)

    return access


# ---------------------------------------------------------------------------
# VF workbook adapter
# ---------------------------------------------------------------------------


def vf_is_end(value: Any) -> bool:
    return compact_text(value).upper().startswith("#END")


def vf_yes(value: Any, default: bool = False) -> bool:
    text = compact_text(value).upper()
    if text in {"Y", "YES", "TRUE", "SUPPORTED", "SUPPORT", "TBD"}:
        return True
    if text in {"N", "NO", "FALSE", "-", "/", "N/A", "NA", "NOT SUPPORTED", "NOTSUPPORTED"}:
        return False
    return default


def vf_find_sheet(workbook: Any, name: str) -> Any | None:
    return find_sheet(workbook, lambda sheet_name: sheet_name.strip().casefold() == name.casefold())


def vf_parse_baudrate(value: Any, default: int = 500000) -> int:
    text = compact_text(value).lower()
    number = parse_int_cell(text, default)
    if "kbps" in text or "kbit" in text or text.endswith("k"):
        return number * 1000
    return number


def vf_record_number(value: Any) -> int | None:
    text = compact_text(value).replace("0x", "").replace("0X", "")
    match = re.search(r"[0-9A-Fa-f]{1,2}", text)
    if not match:
        return None
    return int(match.group(0), 16)


def vf_make_param(
    *,
    name_value: Any,
    byte_value: Any,
    bit_value: Any,
    bit_len_value: Any,
    data_type_value: Any,
    method_value: Any = "",
    unit_value: Any = "",
    min_value: Any = "",
    max_value: Any = "",
    fallback_name: str = "Data",
    byte_offset: int = 0,
) -> ParamDef | None:
    name = usable_text(name_value) or fallback_name
    byte_abs = parse_int_cell(byte_value, default=-999999)
    bit_len = parse_int_cell(bit_len_value, default=0)
    if byte_abs == -999999 or bit_len <= 0:
        return None
    bit_pos = parse_int_cell(bit_value, default=0)
    conversion = parse_conversion(method_value)
    if conversion.kind == "identity" and re.fullmatch(r"(?i)\s*BCD\s*", usable_text(data_type_value)):
        conversion = Conversion(kind="bcd")
    if bit_len > 32 and conversion.kind == "enum":
        conversion = Conversion()
    return ParamDef(
        name=name,
        long_name=name,
        byte_pos=max(0, byte_abs - byte_offset),
        bit_pos=bit_pos,
        bit_len=bit_len,
        data_type=usable_text(data_type_value) or "Unsigned",
        unit=normalize_unit_text(unit_value),
        conversion=conversion,
        min_value=usable_text(min_value),
        max_value=usable_text(max_value),
    )


def vf_param_size(params: Iterable[ParamDef], default: int = 1) -> int:
    max_end = 0
    for param in params:
        max_end = max(max_end, param.byte_pos + max(1, (param.bit_pos + max(1, param.bit_len) + 7) // 8))
    return max(default, max_end)


def vf_parse_cover(workbook: Any) -> CoverInfo:
    cover = CoverInfo(ecu_name="VF_ECU_CAN")
    sheet = vf_find_sheet(workbook, "Change Log & General Info")
    rows: dict[str, str] = {}
    if sheet is not None:
        for row in range(1, sheet.max_row + 1):
            key = compact_text(sheet.cell(row, 1).value)
            if key:
                rows[key.casefold()] = compact_text(sheet.cell(row, 2).value)

    cover.ecu_name = rows.get("ecu-name", "") or "VF_ECU_CAN"
    cover.vehicle = rows.get("vehicle program - project name", "")
    cover.supplier = rows.get("ecu supplier", "")
    cover.bus_type = "CAN"
    cover.rx_fun_id = parse_hex_in_text(rows.get("funcreqid", ""), max_value=0x1FFFFFFF)
    cover.rx_phy_id = parse_hex_in_text(rows.get("physreqid", ""), max_value=0x1FFFFFFF)
    cover.tx_id = parse_hex_in_text(rows.get("physrespid", ""), max_value=0x1FFFFFFF)
    baudrate = vf_parse_baudrate(rows.get("baudrate", "500kbps"))

    cover.comm_params = {
        "ISO_15765_2.CP_CanFuncReqId": cover.rx_fun_id if cover.rx_fun_id is not None else 0x6FF,
        "ISO_15765_3.CP_P2Max": 150_000,
        "ISO_15765_3.CP_P2Star": 5_100_000,
        "ISO_15765_3.CP_TesterPresentTime": 2_000_000,
        "ISO_15765_3.CP_P3Phys": 100_000,
        "ISO_15765_3.CP_P3Func": 100_000,
        "ISO_15765_2.CP_StMin": 20_000,
        "ISO_15765_2.CP_BlockSize": 0,
        "ISO_15765_2.CP_As": 70_000,
        "ISO_15765_2.CP_Ar": 70_000,
        "ISO_15765_2.CP_Bs": 150_000,
        "ISO_15765_2.CP_Br": 70_000,
        "ISO_15765_2.CP_Cs": 70_000,
        "ISO_15765_2.CP_Cr": 150_000,
        "ISO_11898_2_DWCAN.CP_Baudrate": baudrate,
    }
    cover.session_timing = {"P2": 50, "P2Ex": 2000}
    return cover


def vf_parse_did_sheet(sheet: Any, *, system_sheet: bool) -> list[DidDef]:
    dids: dict[int, DidDef] = {}
    current: DidDef | None = None
    current_supported = True

    for row in range(3, sheet.max_row + 1):
        if vf_is_end(sheet.cell(row, 1).value):
            break
        did_value = parse_hex_cell(sheet.cell(row, 1).value, max_value=0xFFFF)
        if did_value is not None:
            current_supported = vf_yes(sheet.cell(row, 15).value, default=False) if system_sheet else True
            if not current_supported:
                current = None
                continue
            desc = usable_text(sheet.cell(row, 2).value) or hex_short("DID", did_value)
            rw_state = usable_text(sheet.cell(row, 3).value) or "R"
            app_access = usable_text(sheet.cell(row, 11).value) or rw_state
            boot_access = usable_text(sheet.cell(row, 12).value)
            security = usable_text(sheet.cell(row, 13).value)
            sessions = [app_access, app_access, boot_access, boot_access, app_access]
            current = dids.setdefault(
                did_value,
                DidDef(
                    did=did_value,
                    desc=desc,
                    size=0,
                    write_security=security or "N",
                    sessions=sessions,
                ),
            )
            current.desc = current.desc or desc
            current.write_security = current.write_security if current.write_security != "N" else (security or "N")
            current.sessions = current.sessions or sessions

        if current is None or not current_supported or not usable_text(sheet.cell(row, 4).value):
            continue

        param = vf_make_param(
            name_value=sheet.cell(row, 4).value,
            byte_value=sheet.cell(row, 5).value,
            bit_value=sheet.cell(row, 6).value,
            bit_len_value=sheet.cell(row, 7).value,
            data_type_value=sheet.cell(row, 8).value,
            method_value=sheet.cell(row, 9).value,
            unit_value=sheet.cell(row, 10).value,
            min_value=sheet.cell(row, 14).value if not system_sheet else "",
            max_value=sheet.cell(row, 15).value if not system_sheet else "",
            fallback_name=current.desc,
            byte_offset=3,
        )
        if param is not None:
            current.params.append(param)

    for did in dids.values():
        did.size = vf_param_size(did.params)
    return list(dids.values())


def vf_parse_dids(workbook: Any) -> list[DidDef]:
    result: dict[int, DidDef] = {}
    for sheet_name, system_sheet in (("System DID", True), ("ECU DID", False)):
        sheet = vf_find_sheet(workbook, sheet_name)
        if sheet is None:
            continue
        for did in vf_parse_did_sheet(sheet, system_sheet=system_sheet):
            existing = result.get(did.did)
            if existing is None:
                result[did.did] = did
            else:
                existing.params.extend(did.params)
                existing.size = max(existing.size, did.size)
                existing.write_security = existing.write_security if existing.write_security != "N" else did.write_security
                existing.sessions = existing.sessions or did.sessions
    return list(result.values())


def vf_parse_io_dids(workbook: Any) -> list[IoDidDef]:
    sheet = vf_find_sheet(workbook, "IO Control (2F)")
    if sheet is None:
        return []
    by_id: dict[int, IoDidDef] = {}
    current: IoDidDef | None = None
    current_control: int | None = None
    current_direction = ""
    for row in range(24, sheet.max_row + 1):
        if vf_is_end(sheet.cell(row, 1).value):
            break
        did_value = parse_hex_cell(sheet.cell(row, 2).value, max_value=0xFFFF)
        if did_value is not None:
            desc = usable_text(sheet.cell(row, 3).value) or hex_short("IODID", did_value)
            current = by_id.setdefault(did_value, IoDidDef(did=did_value, desc=desc, size=0))
        control = parse_hex_cell(sheet.cell(row, 1).value, max_value=0xFF)
        if control is not None:
            current_control = control
            if current is not None:
                current.controls.add(control)
        direction = usable_text(sheet.cell(row, 4).value)
        if direction:
            current_direction = direction
        if current is None or not usable_text(sheet.cell(row, 5).value):
            continue
        param = vf_make_param(
            name_value=sheet.cell(row, 5).value,
            byte_value=sheet.cell(row, 6).value,
            bit_value=sheet.cell(row, 7).value,
            bit_len_value=sheet.cell(row, 8).value,
            data_type_value=sheet.cell(row, 9).value,
            method_value=sheet.cell(row, 10).value,
            unit_value=sheet.cell(row, 11).value,
            fallback_name=current.desc,
            byte_offset=4,
        )
        if param is not None and not normalize_access(current_direction).startswith("RESP"):
            current.params.append(param)
        if current_control is not None:
            current.controls.add(current_control)
    for io_did in by_id.values():
        io_did.size = vf_param_size(io_did.params)
        if not io_did.controls:
            io_did.controls.add(3)
    return list(by_id.values())


def vf_parse_routine_control_type(value: Any) -> int | None:
    parsed = parse_hex_cell(value, max_value=0xFF)
    if parsed is not None and parsed in {1, 2, 3}:
        return parsed
    match = re.search(r"\b([123])\b", compact_text(value))
    return int(match.group(1)) if match else None


def vf_parse_routines(workbook: Any) -> list[RoutineDef]:
    sheet = vf_find_sheet(workbook, "Routines Data (0x31)")
    if sheet is None:
        return []
    routines: dict[int, RoutineDef] = {}
    current_routine: RoutineDef | None = None
    current_subfn: RoutineSubFunction | None = None
    current_direction = ""
    current_supported = True
    for row in range(3, sheet.max_row + 1):
        if vf_is_end(sheet.cell(row, 1).value):
            break
        rid = parse_hex_cell(sheet.cell(row, 3).value, max_value=0xFFFF)
        control_type = vf_parse_routine_control_type(sheet.cell(row, 2).value)
        if rid is not None:
            current_supported = vf_yes(sheet.cell(row, 15).value, default=True)
            desc = usable_text(sheet.cell(row, 4).value) or hex_short("RID", rid)
            current_routine = routines.setdefault(
                rid,
                RoutineDef(
                    rid=rid,
                    desc=desc,
                    security=usable_text(sheet.cell(row, 13).value) or "N",
                    sessions=[usable_text(sheet.cell(row, 1).value)],
                ),
            )
        if current_routine is not None and control_type is not None:
            current_subfn = current_routine.subfunctions.setdefault(
                control_type, RoutineSubFunction(control_type=control_type, supported=current_supported)
            )
            current_subfn.supported = current_subfn.supported or current_supported
        direction = usable_text(sheet.cell(row, 5).value)
        if direction:
            current_direction = direction
        if current_routine is None or current_subfn is None or not current_supported or not usable_text(sheet.cell(row, 6).value):
            continue
        param = vf_make_param(
            name_value=sheet.cell(row, 6).value,
            byte_value=sheet.cell(row, 7).value,
            bit_value=sheet.cell(row, 8).value,
            bit_len_value=sheet.cell(row, 9).value,
            data_type_value=sheet.cell(row, 10).value,
            method_value=sheet.cell(row, 11).value,
            unit_value=sheet.cell(row, 12).value,
            fallback_name=current_routine.desc,
            byte_offset=4,
        )
        if param is None:
            continue
        if normalize_access(current_direction).startswith("RESP"):
            current_subfn.status_params.append(param)
        else:
            current_subfn.option_params.append(param)
    return list(routines.values())


def vf_encode_dtc(value: Any) -> tuple[int, str] | None:
    text = compact_text(value).upper().replace(" ", "").replace("_", "")
    text = re.sub(r"^0X", "", text)
    match = re.fullmatch(r"([PCBU])([0-3A-F][0-9A-F]{3})([0-9A-F]{2})", text)
    if match:
        category, base, ftb = match.groups()
        category_base = {"P": 0x0, "C": 0x4, "B": 0x8, "U": 0xC}[category]
        first = int(base[0], 16) & 0x3
        encoded = f"{category_base + first:X}{base[1:]}{ftb}"
        return int(encoded, 16), f"{category}{base}{ftb}"
    match = re.fullmatch(r"([0-9A-F]{4})-?([0-9A-F]{2})", text)
    if match:
        encoded = "".join(match.groups())
        return int(encoded, 16), f"{match.group(1)}-{match.group(2)}"
    match = re.search(r"([0-9A-F]{6})", text)
    if match:
        return int(match.group(1), 16), match.group(1)
    return None


def vf_parse_dtcs(workbook: Any) -> list[DtcDef]:
    sheet = vf_find_sheet(workbook, "DTC-List")
    if sheet is None:
        return []
    result: list[DtcDef] = []
    seen: set[int] = set()
    for row in range(3, sheet.max_row + 1):
        encoded = vf_encode_dtc(sheet.cell(row, 2).value) or vf_encode_dtc(sheet.cell(row, 3).value)
        if encoded is None:
            continue
        byte_code, display = encoded
        if byte_code in seen:
            continue
        seen.add(byte_code)
        priority = re.search(r"\d+", compact_text(sheet.cell(row, 6).value))
        result.append(
            DtcDef(
                display_code=display,
                byte_code=byte_code,
                text=usable_text(sheet.cell(row, 5).value) or display,
                priority=priority.group(0) if priority else usable_text(sheet.cell(row, 6).value),
            )
        )
    return result


def vf_snapshot_record_nums(value: Any) -> list[int]:
    return sorted({int(match, 16) for match in re.findall(r"0[xX]([0-9A-Fa-f]{1,2})", compact_text(value))})


def vf_snapshot_record_names(value: Any) -> dict[int, str]:
    names: dict[int, str] = {}
    for line in re.split(r"[\r\n]+", cell_text(value)):
        match = re.search(r"0[xX]([0-9A-Fa-f]{1,2})\s*[-–—]\s*(.+)", line)
        if not match:
            continue
        record_num = int(match.group(1), 16)
        label = compact_text(match.group(2))
        if label:
            names[record_num] = label
    return names


def vf_parse_snapshots(workbook: Any) -> tuple[list[SnapshotDef], list[int], dict[int, str]]:
    sheet = vf_find_sheet(workbook, "Snapshot DIDs")
    if sheet is None:
        return [], [], {}
    record_header = sheet.cell(3, 1).value
    record_nums = vf_snapshot_record_nums(record_header) or [1, 2]
    record_names = vf_snapshot_record_names(record_header)
    snapshots: list[SnapshotDef] = []
    for row in range(5, sheet.max_row + 1):
        if vf_is_end(sheet.cell(row, 1).value):
            break
        did = parse_hex_cell(sheet.cell(row, 1).value, max_value=0xFFFF)
        if did is None or not vf_yes(sheet.cell(row, 11).value, default=False):
            continue
        size = parse_int_cell(sheet.cell(row, 6).value, default=0)
        if size <= 0:
            continue
        desc = usable_text(sheet.cell(row, 4).value) or hex_short("Snapshot", did)
        snapshots.append(
            SnapshotDef(
                record_num=None,
                did=did,
                desc=desc,
                size=size,
                params=[
                    ParamDef(
                        name=desc,
                        long_name=desc,
                        byte_pos=0,
                        bit_pos=0,
                        bit_len=max(8, size * 8),
                        data_type=usable_text(sheet.cell(row, 5).value) or "Unsigned",
                    )
                ],
            )
        )
    return snapshots, record_nums, record_names


def vf_parse_extended_records(workbook: Any) -> list[ExtendedRecordDef]:
    sheet = vf_find_sheet(workbook, "Extended Data")
    if sheet is None:
        return []
    result: list[ExtendedRecordDef] = []
    for row in range(6, sheet.max_row + 1):
        if vf_is_end(sheet.cell(row, 1).value):
            break
        if not vf_yes(sheet.cell(row, 8).value, default=False):
            continue
        record_num = vf_record_number(sheet.cell(row, 2).value)
        size = parse_int_cell(sheet.cell(row, 3).value, default=0)
        if record_num is None or size <= 0:
            continue
        desc = usable_text(sheet.cell(row, 1).value) or f"Extended Record 0x{record_num:02X}"
        result.append(
            ExtendedRecordDef(
                record_num=record_num,
                desc=desc,
                size=size,
                params=[
                    ParamDef(
                        name=desc,
                        long_name=desc,
                        byte_pos=0,
                        bit_pos=0,
                        bit_len=max(8, size * 8),
                        data_type=usable_text(sheet.cell(row, 4).value) or "INT",
                    )
                ],
            )
        )
    return result


def jac_is_end(value: Any) -> bool:
    return compact_text(value).upper().startswith("#END")


def jac_yes(value: Any, default: bool = False) -> bool:
    return is_supported_flag(value, default=default)


def jac_find_sheet(workbook: Any, exact_name: str) -> Any | None:
    return find_sheet(workbook, lambda name: name.strip() == exact_name.strip())


def jac_ms_to_us(value: Any, default_us: int) -> int:
    text = compact_text(value)
    if not text:
        return default_us
    number = parse_int_cell(text, default_us // 1000)
    if re.search(r"\bus\b|\bmicro", text, flags=re.IGNORECASE):
        return number
    return number * 1000


def jac_parse_cover(workbook: Any) -> CoverInfo:
    cover = CoverInfo(ecu_name="JAC_ECU_CAN")
    sheet = jac_find_sheet(workbook, "01.General")
    rows: dict[str, str] = {}
    if sheet is not None:
        for row in range(1, sheet.max_row + 1):
            key = compact_text(sheet.cell(row, 1).value)
            if key:
                rows[key.casefold()] = compact_text(sheet.cell(row, 2).value)

    cover.ecu_name = rows.get("ecu-name", "") or "JAC_ECU_CAN"
    cover.supplier = rows.get("supplier id", "")
    cover.bus_type = "CAN"
    cover.rx_fun_id = parse_hex_in_text(rows.get("can functional request id", ""), max_value=0x1FFFFFFF)
    cover.rx_phy_id = parse_hex_in_text(rows.get("can physical request id", ""), max_value=0x1FFFFFFF)
    cover.tx_id = parse_hex_in_text(rows.get("can response id", ""), max_value=0x1FFFFFFF)
    baudrate = parse_int_cell(rows.get("can baudrate", ""), 500000)

    cover.comm_params = {
        "ISO_15765_2.CP_CanFuncReqId": cover.rx_fun_id if cover.rx_fun_id is not None else 0x7DF,
        "ISO_15765_3.CP_P2Max": jac_ms_to_us(rows.get("p2", ""), 150_000),
        "ISO_15765_3.CP_P2Star": jac_ms_to_us(rows.get("p2*", ""), 5_000_000),
        "ISO_15765_3.CP_TesterPresentTime": jac_ms_to_us(rows.get("s3", ""), 2_000_000),
        "ISO_15765_3.CP_P3Phys": 50_000,
        "ISO_15765_3.CP_P3Func": 50_000,
        "ISO_15765_2.CP_StMin": jac_ms_to_us(rows.get("app stmin(ecu)", rows.get("app stmin", "")), 20_000),
        "ISO_15765_2.CP_BlockSize": parse_int_cell(rows.get("bs(ecu)", ""), 0),
        "ISO_15765_2.CP_As": jac_ms_to_us(rows.get("n_as", ""), 70_000),
        "ISO_15765_2.CP_Ar": jac_ms_to_us(rows.get("n_ar", ""), 70_000),
        "ISO_15765_2.CP_Bs": jac_ms_to_us(rows.get("n_bs", ""), 150_000),
        "ISO_15765_2.CP_Br": 50_000,
        "ISO_15765_2.CP_Cs": 50_000,
        "ISO_15765_2.CP_Cr": jac_ms_to_us(rows.get("n_cr", ""), 150_000),
        "ISO_11898_2_DWCAN.CP_Baudrate": baudrate,
    }
    cover.session_timing = {
        "P2": parse_int_cell(rows.get("p2(ecu)", rows.get("p2", "")), 50),
        "P2Ex": parse_int_cell(rows.get("p2*(ecu)", rows.get("p2*", "")), 5000),
    }
    return cover


def jac_access_flags(rw_value: Any, read_values: Iterable[Any], write_values: Iterable[Any]) -> list[str]:
    rw = normalize_access(rw_value)
    result: list[str] = []
    for read_value, write_value in zip(read_values, write_values, strict=False):
        flags = ""
        if "R" in rw and jac_yes(read_value):
            flags += "R"
        if "W" in rw and jac_yes(write_value):
            flags += "W"
        result.append(flags)
    return result


def jac_security_level(values: Iterable[Any]) -> str:
    names = ["Level0", "Level1", "LevelFBL"]
    selected: list[str] = []
    for name, value in zip(names, values, strict=False):
        if jac_yes(value):
            selected.append(name)
    if "LevelFBL" in selected:
        return "LevelFBL"
    if "Level1" in selected:
        return "Level1"
    return "N"


def jac_append_did_param(sheet: Any, row: int, current: DidDef, cols: dict[str, int]) -> None:
    if not usable_text(sheet.cell(row, cols["name_en"]).value) and not usable_text(sheet.cell(row, cols["name_zh"]).value):
        return
    param = jac_make_param(
        name_value=sheet.cell(row, cols["name_en"]).value,
        chinese_name_value=sheet.cell(row, cols["name_zh"]).value,
        byte_value=sheet.cell(row, cols["byte"]).value,
        bit_value=sheet.cell(row, cols["bit"]).value,
        data_type_value=sheet.cell(row, cols["type"]).value,
        unit_value=sheet.cell(row, cols["unit"]).value,
        conversion_value=sheet.cell(row, cols["method_en"]).value,
        conversion_fallback_value=sheet.cell(row, cols["method_zh"]).value,
        min_value=sheet.cell(row, cols["min"]).value,
        max_value=sheet.cell(row, cols["max"]).value,
        fallback_name=current.desc,
        size=current.size,
    )
    if param is not None:
        current.params.append(param)


def jac_parse_did_sheet(sheet: Any, *, system_sheet: bool, start_row: int) -> list[DidDef]:
    if system_sheet:
        cols = {
            "did": 2,
            "desc_en": 3,
            "desc_zh": 4,
            "supported": 6,
            "rw": 7,
            "size": 8,
            "byte": 9,
            "bit": 10,
            "name_en": 11,
            "name_zh": 12,
            "min": 13,
            "max": 14,
            "unit": 15,
            "method_en": 16,
            "method_zh": 17,
            "type": 19,
            "read": (21, 23, 24, 25, 26),
            "write": (30, 32, 33, 34, 35),
            "write_security": (36, 37, 38),
        }
    else:
        cols = {
            "did": 2,
            "desc_en": 3,
            "desc_zh": 4,
            "supported": None,
            "rw": 5,
            "size": 6,
            "byte": 7,
            "bit": 8,
            "name_en": 9,
            "name_zh": 10,
            "min": 11,
            "max": 12,
            "unit": 13,
            "method_en": 14,
            "method_zh": 15,
            "type": 17,
            "read": (19, 21, 22, 23, 24),
            "write": (27, 29, 30, 31, 32),
            "write_security": (33, 34, 35),
        }

    dids: dict[int, DidDef] = {}
    current: DidDef | None = None
    current_supported = True
    for row in range(start_row, sheet.max_row + 1):
        if jac_is_end(sheet.cell(row, 1).value) or jac_is_end(sheet.cell(row, 2).value):
            break
        first_col = compact_text(sheet.cell(row, 1).value)
        if "internal" in first_col.casefold() or "供应商内部" in first_col:
            break

        did_value = parse_hex_loose(sheet.cell(row, cols["did"]).value, max_value=0xFFFF)
        if did_value is not None:
            support_col = cols.get("supported")
            current_supported = jac_yes(sheet.cell(row, support_col).value, default=True) if support_col else True
            if not current_supported:
                current = None
                continue
            desc, long_name = dual_name(
                sheet.cell(row, cols["desc_en"]).value,
                sheet.cell(row, cols["desc_zh"]).value,
                hex_short("DID", did_value),
            )
            size = parse_int_cell(sheet.cell(row, cols["size"]).value, 0)
            sessions = jac_access_flags(
                sheet.cell(row, cols["rw"]).value,
                (sheet.cell(row, col).value for col in cols["read"]),
                (sheet.cell(row, col).value for col in cols["write"]),
            )
            security = jac_security_level(sheet.cell(row, col).value for col in cols["write_security"])
            current = dids.setdefault(
                did_value,
                DidDef(
                    did=did_value,
                    desc=long_name or desc or hex_short("DID", did_value),
                    size=size,
                    write_security=security,
                    sessions=sessions,
                ),
            )
            current.desc = current.desc or long_name or desc or hex_short("DID", did_value)
            current.size = max(current.size, size)
            current.sessions = current.sessions or sessions
            current.write_security = current.write_security if current.write_security != "N" else security

        if current is None or not current_supported:
            continue
        jac_append_did_param(sheet, row, current, cols)

    for did in dids.values():
        if did.size <= 0:
            did.size = vf_param_size(did.params)
    return list(dids.values())


def jac_parse_dids(workbook: Any) -> list[DidDef]:
    result: dict[int, DidDef] = {}
    profiles = [
        ("03.1.System DID", True, 7),
        ("03.2.ECU DID", False, 7),
        ("03.3.Coding DID", False, 8),
    ]
    for sheet_name, system_sheet, start_row in profiles:
        sheet = jac_find_sheet(workbook, sheet_name)
        if sheet is None:
            continue
        for did in jac_parse_did_sheet(sheet, system_sheet=system_sheet, start_row=start_row):
            existing = result.get(did.did)
            if existing is None:
                result[did.did] = did
            else:
                existing.params.extend(did.params)
                existing.size = max(existing.size, did.size)
                existing.write_security = existing.write_security if existing.write_security != "N" else did.write_security
                existing.sessions = existing.sessions or did.sessions
    return list(result.values())


def jac_parse_io_dids(workbook: Any) -> list[IoDidDef]:
    sheet = jac_find_sheet(workbook, "03.4.IOcontrol DID 0x2F")
    if sheet is None:
        return []
    by_id: dict[int, IoDidDef] = {}
    current: IoDidDef | None = None
    current_control: int | None = None
    current_enabled = False
    for row in range(6, sheet.max_row + 1):
        if jac_is_end(sheet.cell(row, 1).value) or jac_is_end(sheet.cell(row, 2).value):
            break
        did_value = parse_hex_loose(sheet.cell(row, 2).value, max_value=0xFFFF)
        if did_value is not None:
            desc, long_name = dual_name(sheet.cell(row, 3).value, sheet.cell(row, 4).value, hex_short("IODID", did_value))
            current = by_id.setdefault(did_value, IoDidDef(did=did_value, desc=long_name or desc, size=0))
        control = parse_hex_loose(sheet.cell(row, 5).value, max_value=0xFF)
        if control is not None:
            current_control = control
            sessions_supported = any(jac_yes(sheet.cell(row, col).value) for col in (18, 19, 20))
            product_supported = jac_yes(sheet.cell(row, 25).value, default=True) if usable_text(sheet.cell(row, 25).value) else True
            current_enabled = sessions_supported and product_supported
        if current is None or current_control is None or not current_enabled:
            continue
        current.controls.add(current_control)
        if current_control != 0x03:
            continue
        param = jac_make_param(
            name_value=sheet.cell(row, 9).value,
            chinese_name_value=sheet.cell(row, 10).value,
            byte_value=sheet.cell(row, 7).value,
            bit_value=sheet.cell(row, 8).value,
            data_type_value=sheet.cell(row, 17).value,
            unit_value=sheet.cell(row, 13).value,
            conversion_value=sheet.cell(row, 14).value,
            conversion_fallback_value=sheet.cell(row, 15).value,
            min_value=sheet.cell(row, 11).value,
            max_value=sheet.cell(row, 12).value,
            fallback_name=current.desc,
            size=parse_int_cell(sheet.cell(row, 6).value, 0),
        )
        if param is not None:
            current.params.append(param)
    for io_did in by_id.values():
        io_did.size = max(io_did.size, vf_param_size(io_did.params))
        if not io_did.controls:
            io_did.controls.add(3)
    return [item for item in by_id.values() if item.controls]


def jac_parse_routine_control_type(value: Any) -> int | None:
    parsed = parse_hex_loose(value, max_value=0xFF)
    if parsed is not None and parsed in {1, 2, 3}:
        return parsed
    match = re.search(r"\b([123])\b", compact_text(value))
    return int(match.group(1)) if match else None


def jac_parse_routines(workbook: Any) -> list[RoutineDef]:
    sheet = jac_find_sheet(workbook, "03.5.Routine DID 0x31")
    if sheet is None:
        return []
    routines: dict[int, RoutineDef] = {}
    current_routine: RoutineDef | None = None
    current_subfn: RoutineSubFunction | None = None
    current_supported = False
    for row in range(7, sheet.max_row + 1):
        if jac_is_end(sheet.cell(row, 1).value) or jac_is_end(sheet.cell(row, 2).value):
            break
        rid = parse_hex_loose(sheet.cell(row, 2).value, max_value=0xFFFF)
        if rid is not None:
            desc, long_name = dual_name(sheet.cell(row, 3).value, sheet.cell(row, 4).value, hex_short("RID", rid))
            sessions = [
                "R" if jac_yes(sheet.cell(row, 33).value) else "",
                "R" if jac_yes(sheet.cell(row, 35).value) else "",
                "R" if jac_yes(sheet.cell(row, 36).value) else "",
                "R" if jac_yes(sheet.cell(row, 37).value) else "",
                "R" if jac_yes(sheet.cell(row, 38).value) else "",
            ]
            security = jac_security_level(sheet.cell(row, col).value for col in (39, 40, 41))
            current_routine = routines.setdefault(
                rid,
                RoutineDef(rid=rid, desc=long_name or desc, security=security, sessions=sessions),
            )
            current_routine.security = current_routine.security if current_routine.security != "N" else security
            current_routine.sessions = current_routine.sessions or sessions

        control_type = jac_parse_routine_control_type(sheet.cell(row, 7).value)
        supported = jac_yes(sheet.cell(row, 6).value, default=False) and jac_yes(sheet.cell(row, 8).value, default=False)
        if current_routine is not None and control_type is not None:
            current_supported = supported
            current_subfn = current_routine.subfunctions.setdefault(
                control_type,
                RoutineSubFunction(control_type=control_type, supported=supported),
            )
            current_subfn.supported = current_subfn.supported or supported
        elif control_type is None:
            supported = current_supported
        if current_routine is None or current_subfn is None or not supported:
            continue

        option_size = parse_int_cell(sheet.cell(row, 9).value, 0)
        option_has_layout = option_size > 0 or usable_text(sheet.cell(row, 10).value) or usable_text(sheet.cell(row, 11).value)
        if option_has_layout:
            option_param = jac_make_param(
                name_value=sheet.cell(row, 12).value,
                chinese_name_value=sheet.cell(row, 13).value,
                byte_value=sheet.cell(row, 10).value,
                bit_value=sheet.cell(row, 11).value,
                data_type_value=sheet.cell(row, 20).value,
                unit_value=sheet.cell(row, 16).value,
                conversion_value=sheet.cell(row, 17).value,
                conversion_fallback_value=sheet.cell(row, 18).value,
                min_value=sheet.cell(row, 14).value,
                max_value=sheet.cell(row, 15).value,
                fallback_name=current_routine.desc,
                size=option_size,
            )
            if option_param is not None:
                current_subfn.option_params.append(option_param)

        status_size = parse_int_cell(sheet.cell(row, 21).value, 0)
        status_has_layout = status_size > 0 or usable_text(sheet.cell(row, 22).value) or usable_text(sheet.cell(row, 23).value)
        if status_has_layout:
            status_param = jac_make_param(
                name_value=sheet.cell(row, 24).value,
                chinese_name_value=sheet.cell(row, 25).value,
                byte_value=sheet.cell(row, 22).value,
                bit_value=sheet.cell(row, 23).value,
                data_type_value=sheet.cell(row, 32).value,
                unit_value=sheet.cell(row, 28).value,
                conversion_value=sheet.cell(row, 29).value,
                conversion_fallback_value=sheet.cell(row, 30).value,
                min_value=sheet.cell(row, 26).value,
                max_value=sheet.cell(row, 27).value,
                fallback_name=f"{current_routine.desc}_Status",
                size=status_size,
            )
            if status_param is not None:
                current_subfn.status_params.append(status_param)
    return [routine for routine in routines.values() if any(subfn.supported for subfn in routine.subfunctions.values())]


def jac_parse_dtcs(workbook: Any) -> list[DtcDef]:
    sheet = jac_find_sheet(workbook, "04.2.DTC list")
    if sheet is None:
        return []
    result: list[DtcDef] = []
    seen: set[int] = set()
    for row in range(6, sheet.max_row + 1):
        if jac_is_end(sheet.cell(row, 1).value) or jac_is_end(sheet.cell(row, 2).value):
            break
        display = usable_text(sheet.cell(row, 2).value).upper()
        bytes_text = "".join(f"{parse_hex_loose(sheet.cell(row, col).value, max_value=0xFF) or 0:02X}" for col in (3, 4, 5))
        if not display or not re.fullmatch(r"[0-9A-F]{6}", bytes_text):
            continue
        byte_code = int(bytes_text, 16)
        if byte_code in seen:
            continue
        seen.add(byte_code)
        _, text = dual_name(sheet.cell(row, 6).value, sheet.cell(row, 7).value, display)
        priority_match = re.search(r"\d+", compact_text(sheet.cell(row, 13).value))
        result.append(
            DtcDef(
                display_code=display,
                byte_code=byte_code,
                text=text or display,
                priority=priority_match.group(0) if priority_match else usable_text(sheet.cell(row, 13).value),
            )
        )
    return result


def jac_record_numbers(value: Any) -> list[int]:
    text = compact_text(value).replace("~", "-")
    if not text:
        return []
    nums = [int(token, 16) for token in re.findall(r"(?:0[xX])?([0-9A-Fa-f]{1,2})", text)]
    if len(nums) >= 2 and "-" in text:
        start, end = nums[0], nums[1]
        return list(range(start, end + 1)) if end >= start else [start]
    return nums[:1]


def jac_parse_snapshots(workbook: Any, dids: list[DidDef]) -> tuple[list[SnapshotDef], list[int], dict[int, str]]:
    sheet = jac_find_sheet(workbook, "04.3.Snapshot")
    if sheet is None:
        return [], [], {}
    did_by_id = {did.did: did for did in dids}
    snapshots: list[SnapshotDef] = []
    record_nums: set[int] = set()
    record_names: dict[int, str] = {}
    current_records: list[int] = []
    current_did: int | None = None
    for row in range(7, sheet.max_row + 1):
        if jac_is_end(sheet.cell(row, 1).value):
            break
        records = jac_record_numbers(sheet.cell(row, 4).value)
        if records:
            current_records = records
            for num in records:
                record_nums.add(num)
                record_names.setdefault(num, f"Snapshot Record 0x{num:02X}")
        did = parse_hex_loose(sheet.cell(row, 5).value, max_value=0xFFFF)
        if did is not None:
            current_did = did
        if current_did is None:
            continue
        desc, long_name = dual_name(sheet.cell(row, 6).value, sheet.cell(row, 7).value, hex_short("Snapshot", current_did))
        source_did = did_by_id.get(current_did)
        if source_did is None:
            continue
        if not any(snapshot.did == current_did for snapshot in snapshots):
            snapshots.append(
                SnapshotDef(
                    record_num=current_records[0] if len(current_records) == 1 else None,
                    did=current_did,
                    desc=long_name or desc or source_did.desc,
                    size=source_did.size,
                    params=[copy.deepcopy(param) for param in source_did.params],
                )
            )
    return snapshots, sorted(record_nums), record_names


def jac_parse_extended_records(workbook: Any) -> list[ExtendedRecordDef]:
    sheet = jac_find_sheet(workbook, "04.4.ExtendedData")
    if sheet is None:
        return []
    result: list[ExtendedRecordDef] = []
    for row in range(8, sheet.max_row + 1):
        if jac_is_end(sheet.cell(row, 1).value) or jac_is_end(sheet.cell(row, 2).value):
            break
        record_num = parse_hex_loose(sheet.cell(row, 2).value, max_value=0xFF)
        size = parse_int_cell(sheet.cell(row, 5).value, 0)
        if record_num is None or size <= 0:
            continue
        desc, long_name = dual_name(sheet.cell(row, 3).value, sheet.cell(row, 4).value, f"Extended Record 0x{record_num:02X}")
        record = ExtendedRecordDef(record_num=record_num, desc=long_name or desc, size=size)
        param = jac_make_param(
            name_value=sheet.cell(row, 3).value,
            chinese_name_value=sheet.cell(row, 4).value,
            byte_value=sheet.cell(row, 6).value,
            bit_value=sheet.cell(row, 7).value,
            data_type_value=sheet.cell(row, 8).value,
            unit_value=sheet.cell(row, 13).value,
            conversion_value=sheet.cell(row, 9).value,
            min_value=sheet.cell(row, 11).value,
            max_value=sheet.cell(row, 12).value,
            fallback_name=record.desc,
            size=size,
        )
        if param is not None:
            record.params.append(param)
        result.append(record)
    return result


def jac_parse_service_access_sheet(
    sheet: Any,
    *,
    start_row: int,
    session_columns: Iterable[tuple[int, str]],
    security_columns: Iterable[tuple[int, str]],
    source: str,
) -> dict[tuple[int, int | None], dict[str, Any]]:
    access: dict[tuple[int, int | None], dict[str, Any]] = {}
    current_service_id: int | None = None
    current_service_name = ""
    for row in range(start_row, sheet.max_row + 1):
        if jac_is_end(sheet.cell(row, 1).value):
            break
        service_id = parse_hex_loose(sheet.cell(row, 1).value, max_value=0xFF)
        if service_id is not None:
            current_service_id = service_id
            current_service_name = usable_text(sheet.cell(row, 2).value)
        if current_service_id is None:
            continue
        if not jac_yes(sheet.cell(row, 6).value):
            continue
        subfunction = parse_hex_loose(sheet.cell(row, 3).value, max_value=0x7F)
        sessions = {name for col, name in session_columns if jac_yes(sheet.cell(row, col).value)}
        if not sessions:
            continue
        security_values = [name for col, name in security_columns if jac_yes(sheet.cell(row, col).value)]
        security = "LevelFBL" if "LevelFBL" in security_values else ("Level1" if "Level1" in security_values else "N")
        key = (current_service_id, subfunction)
        item = access.setdefault(
            key,
            {
                "service_id": current_service_id,
                "subfunction": subfunction,
                "service_name": current_service_name,
                "subfunction_name": usable_text(sheet.cell(row, 4).value),
                "sessions": set(),
                "security": "N",
                "sources": set(),
            },
        )
        item["sessions"].update(sessions)
        item["security"] = merge_security_text(item["security"], security)
        item["sources"].add(source)
    return access


def jac_parse_core_service_access(workbook: Any) -> dict[tuple[int, int | None], dict[str, Any]]:
    result: dict[tuple[int, int | None], dict[str, Any]] = {}
    profiles = [
        (
            jac_find_sheet(workbook, "02.1.ApplicationServices"),
            ((12, "Default"), (13, "Extended")),
            ((14, "Level0"), (15, "Level1")),
            "Application",
        ),
        (
            jac_find_sheet(workbook, "02.2.BootServices"),
            ((12, "Default"), (13, "Programming"), (14, "Extended")),
            ((15, "Level0"), (16, "Level1"), (17, "LevelFBL")),
            "Boot",
        ),
    ]
    for sheet, session_columns, security_columns, source in profiles:
        if sheet is None:
            continue
        parsed = jac_parse_service_access_sheet(
            sheet,
            start_row=7,
            session_columns=session_columns,
            security_columns=security_columns,
            source=source,
        )
        for key, item in parsed.items():
            target = result.setdefault(
                key,
                {
                    "service_id": item["service_id"],
                    "subfunction": item["subfunction"],
                    "service_name": item["service_name"],
                    "subfunction_name": item["subfunction_name"],
                    "sessions": set(),
                    "security": "N",
                    "sources": set(),
                },
            )
            target["sessions"].update(item["sessions"])
            target["security"] = merge_security_text(target["security"], item["security"])
            target["sources"].update(item["sources"])
    return result


def parse_jac_survey(xlsx_path: Path) -> SurveyData:
    workbook = load_workbook(xlsx_path, data_only=True)
    cover = jac_parse_cover(workbook)
    dids = jac_parse_dids(workbook)
    io_dids = jac_parse_io_dids(workbook)
    routines = jac_parse_routines(workbook)
    dtcs = jac_parse_dtcs(workbook)
    snapshots, snapshot_record_nums, snapshot_record_names = jac_parse_snapshots(workbook, dids)
    extended_records = jac_parse_extended_records(workbook)
    survey = SurveyData(cover, dids, io_dids, routines, dtcs, snapshots, extended_records)
    survey.snapshot_record_nums = snapshot_record_nums
    survey.snapshot_record_names = snapshot_record_names
    survey.core_service_access = jac_parse_core_service_access(workbook)
    return survey


def parse_vf_survey(xlsx_path: Path) -> SurveyData:
    workbook = load_workbook(xlsx_path, data_only=True)
    cover = vf_parse_cover(workbook)
    dids = vf_parse_dids(workbook)
    io_dids = vf_parse_io_dids(workbook)
    routines = vf_parse_routines(workbook)
    dtcs = vf_parse_dtcs(workbook)
    snapshots, snapshot_record_nums, snapshot_record_names = vf_parse_snapshots(workbook)
    extended_records = vf_parse_extended_records(workbook)
    survey = SurveyData(cover, dids, io_dids, routines, dtcs, snapshots, extended_records)
    survey.snapshot_record_nums = snapshot_record_nums
    survey.snapshot_record_names = snapshot_record_names
    survey.core_service_access = vf_parse_diagnostics_services_access(workbook)
    return survey


def update_template(template_pdx: Path, output_pdx: Path, survey: SurveyData, validate: bool = True) -> None:
    with tempfile.TemporaryDirectory(prefix="jac_pdx_") as tmp_name:
        tmp_dir = Path(tmp_name)
        with zipfile.ZipFile(template_pdx, "r") as archive:
            archive.extractall(tmp_dir)

        odx_path = tmp_dir / "JAC_ECU_CAN_v15.odx-d"
        if not odx_path.exists():
            candidates = list(tmp_dir.glob("*.odx-d"))
            if not candidates:
                raise FileNotFoundError("No .odx-d file found inside template PDX")
            odx_path = candidates[0]

        parser = etree.XMLParser(remove_blank_text=False)
        tree = etree.parse(str(odx_path), parser)
        root = tree.getroot()
        id_gen = BASE.IdGenerator(root)
        update_odx_vf(root, id_gen, survey)
        validate_diag_service_child_order(root, odx_path.name)
        BASE.validate_can_dela_odx_structure(root, odx_path.name)
        tree.write(str(odx_path), encoding="UTF-8", xml_declaration=True, pretty_print=True, standalone=False)

        keep_files = set(JAC_KEEP_FILES)
        keep_files.add(odx_path.name)
        BASE.patch_pdx_catalog_for_can_only(tmp_dir / "index.xml", keep_files)

        output_pdx.parent.mkdir(parents=True, exist_ok=True)
        if output_pdx.exists():
            output_pdx.unlink()
        with zipfile.ZipFile(output_pdx, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for path in sorted(tmp_dir.rglob("*")):
                rel_path = path.relative_to(tmp_dir).as_posix()
                if path.is_file() and rel_path in keep_files:
                    archive.write(path, rel_path)

    if validate:
        validate_with_odxtools(output_pdx)


def update_odx_vf(root: etree._Element, id_gen: Any, survey: SurveyData) -> None:
    update_layer_names(root, survey.cover)
    ddds = find_dictionary_spec(root)
    if ddds is None:
        raise RuntimeError("Template ODX has no DIAG-DATA-DICTIONARY-SPEC")
    data_object_props = ddds.find("DATA-OBJECT-PROPS")
    structures = ddds.find("STRUCTURES")
    if data_object_props is None or structures is None:
        raise RuntimeError("Template ODX is missing DOP/STRUCTURE containers")

    unit_ids = BASE.ensure_units(ddds, survey)
    generated_dop_cache: dict[tuple[str, int, str, str, str], str] = {}

    for did in survey.dids:
        BASE.prepare_data_structure(
            id_gen=id_gen,
            data_object_props=data_object_props,
            structures=structures,
            item=did,
            prefix="DID",
            unit_ids=unit_ids,
            generated_dop_cache=generated_dop_cache,
        )

    for io_did in survey.io_dids:
        BASE.prepare_data_structure(
            id_gen=id_gen,
            data_object_props=data_object_props,
            structures=structures,
            item=io_did,
            prefix="IODID",
            unit_ids=unit_ids,
            generated_dop_cache=generated_dop_cache,
        )

    for routine in survey.routines:
        english, long_name = BASE.split_name(routine.desc)
        routine.short_name = sanitize_short_name(english or hex_short("RID", routine.rid), hex_short("RID", routine.rid))
        routine.long_name = long_name or routine.short_name
        for subfn in routine.subfunctions.values():
            if subfn.option_params:
                subfn.option_structure_id = BASE.make_param_structure(
                    id_gen,
                    data_object_props,
                    structures,
                    f"STR_RoutineOption_{routine.short_name}_{subfn.control_type:02X}",
                    f"Routine Option {routine.long_name}",
                    subfn.option_params,
                    unit_ids,
                    generated_dop_cache,
                )
            if subfn.status_params:
                subfn.status_structure_id = BASE.make_param_structure(
                    id_gen,
                    data_object_props,
                    structures,
                    f"STR_RoutineStatus_{routine.short_name}_{subfn.control_type:02X}",
                    f"Routine Status {routine.long_name}",
                    subfn.status_params,
                    unit_ids,
                    generated_dop_cache,
                )

    for snapshot in survey.snapshots:
        if snapshot.params:
            snapshot.structure_id = BASE.make_param_structure(
                id_gen,
                data_object_props,
                structures,
                f"STR_Snapshot_{sanitize_short_name(snapshot.desc, hex_short('Snapshot', snapshot.did))}",
                snapshot.desc,
                snapshot.params,
                unit_ids,
                generated_dop_cache,
                byte_size=snapshot.size or None,
            )

    for record in survey.extended_records:
        if record.params:
            record.structure_id = BASE.make_param_structure(
                id_gen,
                data_object_props,
                structures,
                f"STR_ExtendedRecord_{record.record_num:02X}",
                record.desc,
                record.params,
                unit_ids,
                generated_dop_cache,
                byte_size=record.size or None,
            )

    ensure_vf_placeholder_services(root, id_gen)
    generate_flat_did_services(root, id_gen, survey.dids)
    generate_flat_io_services(root, id_gen, survey.io_dids)
    generate_flat_routine_services(root, id_gen, survey.routines)
    for old_name in ("z_7_Read", "z_Read", "z_Control", "z_ReturnControl"):
        BASE.remove_service_and_messages(root, old_name)
    BASE.remove_service_and_messages(root, "Upload_Download_RequestDownload")
    update_vf_dtc_dop(root, id_gen, survey.dtcs)
    update_vf_snapshot_and_extended_data(
        root,
        id_gen,
        survey.snapshots,
        survey.extended_records,
        getattr(survey, "snapshot_record_nums", []),
        getattr(survey, "snapshot_record_names", {}),
    )
    update_base_variant_comparams(root, survey.cover)
    update_session_timing(root, survey.cover)
    ensure_core_response_services(root, id_gen)
    ensure_vf_session_control_services(root, id_gen, survey)
    ensure_vf_communication_control_services(root, id_gen, survey)
    ensure_vf_security_access_subfunction_scales(root, survey)
    prune_vf_unsupported_security_access_services(root, survey)
    generate_boot_security_access_services(root, id_gen, survey)
    update_vf_security_access_preconditions(root, id_gen, survey)
    update_core_service_preconditions(root, survey)
    shorten_dictionary_short_names(root)
    BASE.prefix_doc_revision_labels(root)


def update_layer_names(root: etree._Element, cover: CoverInfo) -> None:
    container = root.find(".//DIAG-LAYER-CONTAINER")
    if container is not None:
        long_name = container.find("LONG-NAME")
        if long_name is not None:
            long_name.text = cover.ecu_name
    base_variant = get_base_variant(root)
    if base_variant is not None:
        long_name = base_variant.find("LONG-NAME")
        if long_name is not None:
            long_name.text = cover.ecu_name


def get_base_variant(root: etree._Element) -> etree._Element | None:
    node = BASE.first_by_short(root, "BASE-VARIANT", "JAC_ECU_CAN")
    if node is not None:
        return node
    nodes = root.xpath("//*[local-name()='BASE-VARIANT']")
    return nodes[0] if nodes else None


def find_dictionary_spec(root: etree._Element) -> etree._Element | None:
    for ddds in root.xpath("//*[local-name()='DIAG-DATA-DICTIONARY-SPEC']"):
        if ddds.find("DATA-OBJECT-PROPS") is not None and ddds.find("STRUCTURES") is not None:
            return ddds
    nodes = root.xpath("//*[local-name()='DIAG-DATA-DICTIONARY-SPEC']")
    return nodes[0] if nodes else None


def candela_short_name(value: str, used: set[str] | None = None, max_len: int = CANDELA_SHORT_NAME_MAX_LEN) -> str:
    if len(value) <= max_len and (used is None or value not in used):
        if used is not None:
            used.add(value)
        return value

    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:8].upper()
    suffix = f"_{digest}"
    head_len = max(1, max_len - len(suffix))
    base = value[:head_len].rstrip("_") or value[:head_len]
    candidate = f"{base}{suffix}"
    if used is not None:
        index = 2
        while candidate in used:
            index_suffix = f"_{index:02d}"
            candidate = f"{candidate[:max_len - len(index_suffix)].rstrip('_')}{index_suffix}"
            index += 1
        used.add(candidate)
    return candidate


def shorten_dictionary_short_names(root: etree._Element) -> None:
    for container_name, child_name in (
        ("DATA-OBJECT-PROPS", "DATA-OBJECT-PROP"),
        ("STRUCTURES", "STRUCTURE"),
    ):
        for container in root.xpath(f"//*[local-name()='{container_name}']"):
            used: set[str] = set()
            for node in container.xpath(f"./*[local-name()='{child_name}']"):
                short = node.find("SHORT-NAME")
                if short is None or not short.text:
                    continue
                short.text = candela_short_name(short.text, used)


def validate_diag_service_child_order(root: etree._Element, source_name: str) -> None:
    order = {
        "SHORT-NAME": 0,
        "LONG-NAME": 1,
        "DESC": 2,
        "ADMIN-DATA": 3,
        "SDGS": 4,
        "FUNCT-CLASS-REFS": 5,
        "AUDIENCE": 6,
        "PROTOCOL-SNREFS": 7,
        "RELATED-DIAG-COMM-REFS": 8,
        "PRE-CONDITION-STATE-REFS": 9,
        "STATE-TRANSITION-REFS": 10,
        "COMPARAM-REFS": 11,
        "REQUEST-REF": 12,
        "POS-RESPONSE-REFS": 13,
        "NEG-RESPONSE-REFS": 14,
        "POS-RESPONSE-SUPPRESSABLE": 15,
    }
    errors: list[str] = []
    for service in root.xpath("//*[local-name()='DIAG-SERVICE']"):
        service_name = service.findtext("SHORT-NAME") or service.get("ID") or "<unnamed>"
        last_index = -1
        for child in service:
            if not isinstance(child.tag, str):
                continue
            child_name = etree.QName(child).localname
            index = order.get(child_name)
            if index is None:
                continue
            if index < last_index:
                errors.append(f"DIAG-SERVICE '{service_name}' has '{child_name}' out of ODX order")
                break
            last_index = index
    if errors:
        details = "\n".join(f"- {message}" for message in errors[:20])
        raise RuntimeError(f"Generated {source_name} has invalid DIAG-SERVICE child order:\n{details}")


def get_short_name(node: etree._Element) -> str:
    return node.findtext("SHORT-NAME") or ""


def find_container(root: etree._Element, name: str) -> etree._Element:
    node = root.find(f".//{name}")
    if node is None:
        raise RuntimeError(f"Template ODX is missing {name}")
    return node


def find_by_id(root: etree._Element, node_id: str | None) -> etree._Element | None:
    if not node_id:
        return None
    nodes = root.xpath("//*[@ID=$node_id]", node_id=node_id)
    return nodes[0] if nodes else None


def service_messages(root: etree._Element, service_short_name: str) -> tuple[etree._Element, etree._Element, etree._Element | None, etree._Element | None]:
    service = BASE.first_by_short(root, "DIAG-SERVICE", service_short_name)
    if service is None:
        raise RuntimeError(f"Template service {service_short_name} was not found")
    request_ref = service.find("REQUEST-REF")
    request = find_by_id(root, request_ref.get("ID-REF") if request_ref is not None else None)
    if request is None:
        raise RuntimeError(f"Template service {service_short_name} has no request")
    pos_ref = service.find(".//POS-RESPONSE-REF")
    neg_ref = service.find(".//NEG-RESPONSE-REF")
    positive = find_by_id(root, pos_ref.get("ID-REF") if pos_ref is not None else None)
    negative = find_by_id(root, neg_ref.get("ID-REF") if neg_ref is not None else None)
    return service, request, positive, negative


def clone_service_bundle(
    root: etree._Element,
    id_gen: Any,
    base_service_short_name: str,
    new_short_name: str,
    long_name: str,
    service_qualifier: str,
    service_name: str,
    instance_qualifier: str,
    instance_name: str,
    precondition_state_ids: Iterable[str] = (),
    static_value: int | None = None,
    positive_response_suppressed: str = "no",
) -> tuple[etree._Element, etree._Element, etree._Element | None, etree._Element | None]:
    base_service, base_request, base_positive, base_negative = service_messages(root, base_service_short_name)
    request = copy.deepcopy(base_request)
    positive = copy.deepcopy(base_positive) if base_positive is not None else None
    negative = copy.deepcopy(base_negative) if base_negative is not None else None
    service = copy.deepcopy(base_service)

    request_id = id_gen.new("RQ")
    positive_id = id_gen.new("PR") if positive is not None else None
    negative_id = id_gen.new("NR") if negative is not None else None
    service_id = id_gen.new("SVC")

    request.set("ID", request_id)
    BASE.refresh_internal_ids(request, id_gen)
    BASE.set_short_long(request, f"RQ_{new_short_name}", f"RQ {long_name}")

    if positive is not None and positive_id is not None:
        positive.set("ID", positive_id)
        BASE.refresh_internal_ids(positive, id_gen)
        BASE.set_short_long(positive, f"PR_{new_short_name}", f"PR {long_name}")

    if negative is not None and negative_id is not None:
        negative.set("ID", negative_id)
        BASE.refresh_internal_ids(negative, id_gen)
        BASE.set_short_long(negative, f"NR_{new_short_name}", f"NR {long_name}")

    service.set("ID", service_id)
    BASE.refresh_internal_ids(service, id_gen)
    BASE.set_short_long(service, new_short_name, long_name)
    request_ref = service.find("REQUEST-REF")
    if request_ref is not None:
        request_ref.set("ID-REF", request_id)
    pos_ref = service.find(".//POS-RESPONSE-REF")
    if pos_ref is not None and positive_id is not None:
        pos_ref.set("ID-REF", positive_id)
    neg_ref = service.find(".//NEG-RESPONSE-REF")
    if neg_ref is not None and negative_id is not None:
        neg_ref.set("ID-REF", negative_id)
    update_service_sdg(
        service,
        id_gen,
        service_qualifier,
        service_name,
        instance_qualifier,
        instance_name,
        static_value,
        positive_response_suppressed=positive_response_suppressed,
    )
    update_flat_preconditions(service, precondition_state_ids)

    find_container(root, "REQUESTS").append(request)
    if positive is not None:
        find_container(root, "POS-RESPONSES").append(positive)
    if negative is not None:
        find_container(root, "NEG-RESPONSES").append(negative)
    find_container(root, "DIAG-COMMS").append(service)
    return service, request, positive, negative


def update_service_sdg(
    service: etree._Element,
    id_gen: Any,
    service_qualifier: str,
    service_name: str,
    instance_qualifier: str,
    instance_name: str,
    static_value: int | None = None,
    positive_response_suppressed: str = "no",
) -> None:
    sdgs = BASE.element("SDGS")
    sdg = BASE.sub(sdgs, "SDG")
    caption = BASE.sub(sdg, "SDG-CAPTION", attrib={"ID": id_gen.new("CAP")})
    BASE.sub(caption, "SHORT-NAME", "CANdelaServiceInformation")
    if static_value is not None:
        BASE.sub(sdg, "SD", static_value, attrib={"SI": "DiagInstanceStaticValue"})
    BASE.sub(sdg, "SD", sanitize_short_name(instance_qualifier, "Instance"), attrib={"SI": "DiagInstanceQualifier"})
    BASE.sub(sdg, "SD", instance_name or instance_qualifier, attrib={"SI": "DiagInstanceName"})
    BASE.sub(sdg, "SD", service_qualifier, attrib={"SI": "ServiceQualifier"})
    BASE.sub(sdg, "SD", service_name, attrib={"SI": "ServiceName"})
    BASE.sub(sdg, "SD", positive_response_suppressed, attrib={"SI": "PositiveResponseSuppressed"})
    BASE.replace_child(service, "SDGS", sdgs, before_tags={"FUNCT-CLASS-REFS", "AUDIENCE", "PRE-CONDITION-STATE-REFS", "REQUEST-REF"})


def update_flat_preconditions(service: etree._Element, state_ids: Iterable[str]) -> None:
    state_ids = list(dict.fromkeys(state_ids))
    if not state_ids:
        existing = service.find("PRE-CONDITION-STATE-REFS")
        if existing is not None:
            service.remove(existing)
        return
    pc = BASE.element("PRE-CONDITION-STATE-REFS")
    for state_id in state_ids:
        BASE.sub(pc, "PRE-CONDITION-STATE-REF", attrib={"ID-REF": state_id})
        BASE.replace_child(
            service,
            "PRE-CONDITION-STATE-REFS",
            pc,
            before_tags={"STATE-TRANSITION-REFS", "COMPARAM-REFS", "REQUEST-REF"},
        )


def state_id_map(root: etree._Element) -> dict[str, str]:
    result: dict[str, str] = {}
    for state in root.xpath("//*[local-name()='STATE']"):
        short_name = state.findtext("SHORT-NAME")
        if short_name and state.get("ID"):
            result[short_name] = state.get("ID")
    return result


def did_state_ids(root: etree._Element, did: DidDef, mode: str) -> list[str]:
    states = state_id_map(root)
    default = states.get("Default")
    programming = states.get("Programming")
    extended = states.get("Extended")
    locked = states.get("Locked")
    unlocked_l1 = states.get("UnlockedL1")
    unlocked_fbl = states.get("Unlocked_FBL")
    level_9 = states.get("Level_9")

    result: list[str] = []
    session_map = [default, extended, default, programming, extended]
    wanted = "R" if mode == "read" else "W"
    for value, state_id in zip(did.sessions, session_map, strict=False):
        if state_id and wanted in normalize_access(value):
            result.append(state_id)
    if mode == "read":
        result.extend(state for state in (locked, unlocked_l1, level_9, unlocked_fbl) if state)
    else:
        security = normalize_access(did.write_security)
        if "FBL" in security and unlocked_fbl:
            result.append(unlocked_fbl)
        elif ("LEVEL_1" in security or "LEVEL1" in security) and unlocked_l1:
            result.append(unlocked_l1)
        elif normalize_access(did.write_security) in {"", "N", "NO", "NONE"}:
            result.extend(state for state in (locked, unlocked_l1, level_9, unlocked_fbl) if state)
        else:
            result.extend(state for state in (unlocked_l1, level_9, unlocked_fbl) if state)
    if not result:
        result = [state for state in (default, extended, unlocked_l1, level_9) if state]
    return list(dict.fromkeys(result))


CORE_RESPONSE_SERVICE_SOURCES: tuple[str, ...] = (
    "DefaultSession_Start_NoResponse",
    "ExtendedDiagnosticSession_Start_NoResponse",
    "Hard_Reset_Reset_NoResponse",
    "EnableRxAndEnableTx_Control_NoResponse",
    "DisableRxAndDisableTx_Control_NoResponse",
    "TesterPresent_Send_NoResponse",
    "ControlDTCSetting_On_NoResponse",
    "ControlDTCSetting_Off_NoResponse",
)

CORE_SERVICE_PRECONDITION_TARGETS: dict[str, tuple[int, int | None, set[str]]] = {
    "DefaultSession_Start": (0x10, 0x01, {"Default", "Programming", "Extended"}),
    "DefaultSession_Start_NoResponse": (0x10, 0x01, {"Default", "Programming", "Extended"}),
    "ProgrammingSession_Start": (0x10, 0x02, {"Default", "Programming", "Extended"}),
    "ProgrammingSession_Start_NoResponse": (0x10, 0x02, {"Default", "Programming", "Extended"}),
    "ExtendedDiagnosticSession_Start": (0x10, 0x03, {"Default", "Programming", "Extended"}),
    "ExtendedDiagnosticSession_Start_NoResponse": (0x10, 0x03, {"Default", "Programming", "Extended"}),
    "CodingSession_Start": (0x10, 0x41, {"Extended"}),
    "CodingSession_Start_NoResponse": (0x10, 0x41, {"Extended"}),
    "Hard_Reset_Reset": (0x11, 0x01, {"Default", "Programming", "Extended"}),
    "Hard_Reset_Reset_NoResponse": (0x11, 0x01, {"Default", "Programming", "Extended"}),
    "EnableRxAndEnableTx_Control": (0x28, 0x00, {"Default", "Programming", "Extended"}),
    "EnableRxAndEnableTx_Control_NoResponse": (0x28, 0x00, {"Default", "Programming", "Extended"}),
    "EnableRxAndDisableTx_Control": (0x28, 0x01, {"Extended", "Programming"}),
    "EnableRxAndDisableTx_Control_NoResponse": (0x28, 0x01, {"Extended", "Programming"}),
    "DisableRxAndDisableTx_Control": (0x28, 0x03, {"Default", "Programming", "Extended"}),
    "DisableRxAndDisableTx_Control_NoResponse": (0x28, 0x03, {"Default", "Programming", "Extended"}),
    "TesterPresent_Send": (0x3E, 0x00, {"Default", "Programming", "Extended"}),
    "TesterPresent_Send_NoResponse": (0x3E, 0x00, {"Default", "Programming", "Extended"}),
    "ControlDTCSetting_On": (0x85, 0x01, {"Extended"}),
    "ControlDTCSetting_On_NoResponse": (0x85, 0x01, {"Extended"}),
    "ControlDTCSetting_Off": (0x85, 0x02, {"Extended"}),
    "ControlDTCSetting_Off_NoResponse": (0x85, 0x02, {"Extended"}),
}


def service_access_state_ids(root: etree._Element, sessions: Iterable[str], security: str) -> list[str]:
    states = state_id_map(root)
    result: list[str] = []
    session_order = ("Default", "Programming", "Extended")
    session_names = {compact_text(session) for session in sessions if compact_text(session)}
    for name in session_order:
        if name in session_names and states.get(name):
            result.append(states[name])

    security_text = normalize_access(security)
    if security_text in {"", "N", "NO", "NONE"}:
        security_names = ("Locked", "UnlockedL1", "Level_9", "Unlocked_FBL")
    else:
        security_names_list: list[str] = []
        if "LEVEL_1" in security_text or "LEVEL1" in security_text:
            security_names_list.append("UnlockedL1")
        if "LEVEL_9" in security_text or "LEVEL9" in security_text:
            security_names_list.append("Level_9")
        if "FBL" in security_text:
            security_names_list.append("Unlocked_FBL")
        security_names = tuple(security_names_list or ["UnlockedL1", "Level_9", "Unlocked_FBL"])

    for name in security_names:
        if states.get(name):
            result.append(states[name])
    return list(dict.fromkeys(result))


def ensure_vf_funct_class(root: etree._Element, id_gen: Any, short_name: str, long_name: str) -> str:
    existing = BASE.first_by_short(root, "FUNCT-CLASS", short_name)
    if existing is not None and existing.get("ID"):
        return existing.get("ID")
    base_variant = get_base_variant(root)
    classes = base_variant.find("FUNCT-CLASSS")
    if classes is None:
        classes = BASE.element("FUNCT-CLASSS")
        BASE.replace_child(base_variant, "FUNCT-CLASSS", classes, before_tags={"STATE-CHARTS", "DIAG-COMMS"})
    node = BASE.sub(classes, "FUNCT-CLASS", attrib={"ID": id_gen.new("FC")})
    BASE.sub(node, "SHORT-NAME", short_name)
    BASE.sub(node, "LONG-NAME", long_name)
    return node.get("ID")


def append_vf_placeholder_service(
    root: etree._Element,
    id_gen: Any,
    short_name: str,
    long_name: str,
    semantic: str,
    funct_class: str,
    funct_class_long: str,
    request: etree._Element,
    positive: etree._Element | None,
    negative: etree._Element | None,
    service_qualifier: str,
    service_name: str,
    instance_qualifier: str = "z",
    instance_name: str = "VF placeholder",
) -> None:
    if BASE.first_by_short(root, "DIAG-SERVICE", short_name) is not None:
        return
    service = BASE.element("DIAG-SERVICE", attrib={"ID": id_gen.new("SVC"), "SEMANTIC": semantic, "ADDRESSING": "FUNCTIONAL-OR-PHYSICAL"})
    BASE.sub(service, "SHORT-NAME", short_name)
    BASE.sub(service, "LONG-NAME", long_name)
    update_service_sdg(service, id_gen, service_qualifier, service_name, instance_qualifier, instance_name)
    class_id = ensure_vf_funct_class(root, id_gen, funct_class, funct_class_long)
    refs = BASE.element("FUNCT-CLASS-REFS")
    BASE.sub(refs, "FUNCT-CLASS-REF", attrib={"ID-REF": class_id})
    BASE.replace_child(service, "FUNCT-CLASS-REFS", refs, before_tags={"AUDIENCE", "PRE-CONDITION-STATE-REFS", "REQUEST-REF"})
    update_flat_preconditions(service, [state for name, state in state_id_map(root).items() if name in {"Default", "Extended", "Programming", "Locked", "UnlockedL1", "Unlocked_FBL"}])
    BASE.sub(service, "REQUEST-REF", attrib={"ID-REF": request.get("ID")})
    if positive is not None:
        refs = BASE.sub(service, "POS-RESPONSE-REFS")
        BASE.sub(refs, "POS-RESPONSE-REF", attrib={"ID-REF": positive.get("ID")})
    if negative is not None:
        refs = BASE.sub(service, "NEG-RESPONSE-REFS")
        BASE.sub(refs, "NEG-RESPONSE-REF", attrib={"ID-REF": negative.get("ID")})
    find_container(root, "REQUESTS").append(request)
    if positive is not None:
        find_container(root, "POS-RESPONSES").append(positive)
    if negative is not None:
        find_container(root, "NEG-RESPONSES").append(negative)
    find_container(root, "DIAG-COMMS").append(service)


def make_vf_read_placeholder_request(id_gen: Any, short_name: str, service_id: int = 0x22, did: int = 0xF190) -> etree._Element:
    request = BASE.element("REQUEST", attrib={"ID": id_gen.new("RQ")})
    BASE.sub(request, "SHORT-NAME", f"RQ_{short_name}")
    BASE.sub(request, "LONG-NAME", f"RQ {short_name}")
    params = BASE.sub(request, "PARAMS")
    params.append(coded_const_param("SID_RQ", "SID-RQ", 0, service_id, "SERVICE-ID"))
    params.append(coded_const_param("RecordDataIdentifier", "RecordDataIdentifier", 1, did, "ID", bit_length=16))
    return request


def make_vf_read_placeholder_positive(id_gen: Any, short_name: str, service_id: int = 0x62, did: int = 0xF190) -> etree._Element:
    response = BASE.element("POS-RESPONSE", attrib={"ID": id_gen.new("PR")})
    BASE.sub(response, "SHORT-NAME", f"PR_{short_name}")
    BASE.sub(response, "LONG-NAME", f"PR {short_name}")
    params = BASE.sub(response, "PARAMS")
    params.append(coded_const_param("SID_PR", "SID-PR", 0, service_id, "SERVICE-ID"))
    params.append(coded_const_param("RecordDataIdentifier", "RecordDataIdentifier", 1, did, "ID", bit_length=16))
    params.append(BASE.make_value_param("Data", "Data", 3, 0, "_1"))
    return response


def make_vf_io_placeholder_request(id_gen: Any, short_name: str, did: int = 0xFE01, control: int = 0x03) -> etree._Element:
    request = BASE.element("REQUEST", attrib={"ID": id_gen.new("RQ")})
    BASE.sub(request, "SHORT-NAME", f"RQ_{short_name}")
    BASE.sub(request, "LONG-NAME", f"RQ {short_name}")
    params = BASE.sub(request, "PARAMS")
    params.append(coded_const_param("SID_RQ", "SID-RQ", 0, 0x2F, "SERVICE-ID"))
    params.append(coded_const_param("DataIdentifier", "DataIdentifier", 1, did, "ID", bit_length=16))
    params.append(coded_const_param("ControlOptionRecord_InputOutputControlParameter", "InputOutputControlParameter", 3, control, "ID"))
    params.append(BASE.make_value_param("ControlOptionRecord", "ControlOptionRecord", 4, 0, "_1"))
    return request


def make_vf_io_placeholder_positive(id_gen: Any, short_name: str, did: int = 0xFE01, control: int = 0x03) -> etree._Element:
    response = BASE.element("POS-RESPONSE", attrib={"ID": id_gen.new("PR")})
    BASE.sub(response, "SHORT-NAME", f"PR_{short_name}")
    BASE.sub(response, "LONG-NAME", f"PR {short_name}")
    params = BASE.sub(response, "PARAMS")
    params.append(coded_const_param("SID_PR", "SID-PR", 0, 0x6F, "SERVICE-ID"))
    params.append(coded_const_param("DataIdentifier", "DataIdentifier", 1, did, "ID", bit_length=16))
    params.append(coded_const_param("ControlStatusRecord_InputOutputControlParameter", "InputOutputControlParameter", 3, control, "ID"))
    params.append(BASE.make_value_param("ControlStatusRecord", "ControlStatusRecord", 4, 0, "_1"))
    return response


def ensure_vf_placeholder_services(root: etree._Element, id_gen: Any) -> None:
    """Create the flat-service seeds expected by the shared flat-service writer.

    They are deleted again after generated DID/IO/Routine instances are cloned.
    Keeping the seed services in-memory lets the shared writer work without
    requiring permanent placeholder services in the VF template.
    """

    append_vf_placeholder_service(
        root,
        id_gen,
        "z_7_Read",
        "VF DID Read placeholder",
        "IDENTIFICATION",
        "ECU_Identification",
        "ECU Identification",
        make_vf_read_placeholder_request(id_gen, "z_7_Read"),
        make_vf_read_placeholder_positive(id_gen, "z_7_Read"),
        make_negative_response(id_gen.new("NR"), "NR_z_7_Read", "NR VF DID Read placeholder", 0x22, "Read"),
        "Read",
        "Read",
    )
    append_vf_placeholder_service(
        root,
        id_gen,
        "z_Read",
        "VF Routine placeholder",
        "ROUTINE",
        "Routine_Control",
        "Routine Control",
        make_vf_read_placeholder_request(id_gen, "z_Read", service_id=0x31, did=0x0203),
        make_vf_read_placeholder_positive(id_gen, "z_Read", service_id=0x71, did=0x0203),
        make_negative_response(id_gen.new("NR"), "NR_z_Read", "NR VF Routine placeholder", 0x31, "Routine"),
        "Read",
        "Read",
    )
    append_vf_placeholder_service(
        root,
        id_gen,
        "z_Control",
        "VF IOControl Control placeholder",
        "CONTROL",
        "IOControl",
        "IO Control",
        make_vf_io_placeholder_request(id_gen, "z_Control", control=0x03),
        make_vf_io_placeholder_positive(id_gen, "z_Control", control=0x03),
        make_negative_response(id_gen.new("NR"), "NR_z_Control", "NR VF IOControl placeholder", 0x2F, "Control"),
        "Control",
        "Control",
    )
    append_vf_placeholder_service(
        root,
        id_gen,
        "z_ReturnControl",
        "VF IOControl ReturnControl placeholder",
        "CONTROL",
        "IOControl",
        "IO Control",
        make_vf_io_placeholder_request(id_gen, "z_ReturnControl", control=0x00),
        make_vf_io_placeholder_positive(id_gen, "z_ReturnControl", control=0x00),
        make_negative_response(id_gen.new("NR"), "NR_z_ReturnControl", "NR VF IOControl placeholder", 0x2F, "ReturnControl"),
        "ReturnControl",
        "ReturnControl",
    )


def sdg_value(service: etree._Element, si: str, default: str = "") -> str:
    node = service.find(f'.//SD[@SI="{si}"]')
    return node.text if node is not None and node.text is not None else default


def first_coded_value_by_semantic(message: etree._Element, semantic: str) -> int | None:
    for param in message.xpath('.//*[local-name()="PARAM" and @SEMANTIC=$semantic]', semantic=semantic):
        value = parse_int_cell(param.findtext("CODED-VALUE"), default=-1)
        if value >= 0:
            return value
    return None


def first_coded_param_by_semantic(message: etree._Element, semantic: str) -> etree._Element | None:
    params = message.xpath('.//*[local-name()="PARAM" and @SEMANTIC=$semantic]', semantic=semantic)
    return params[0] if params else None


def first_coded_value_by_semantics(message: etree._Element, semantics: Iterable[str]) -> int | None:
    for semantic in semantics:
        value = first_coded_value_by_semantic(message, semantic)
        if value is not None:
            return value
    return None


def diag_service_request_key(root: etree._Element, service: etree._Element) -> tuple[int, int | None] | None:
    request_ref = service.find("REQUEST-REF")
    request = find_by_id(root, request_ref.get("ID-REF") if request_ref is not None else None)
    if request is None:
        return None
    service_id = first_coded_value_by_semantic(request, "SERVICE-ID")
    if service_id is None:
        return None
    subfunction = first_coded_value_by_semantics(request, ("SUBFUNCTION", "ACCESSMODE"))
    return service_id, subfunction


def existing_diag_service_keys(root: etree._Element) -> set[tuple[int, int | None]]:
    result: set[tuple[int, int | None]] = set()
    for service in root.xpath("//*[local-name()='DIAG-SERVICE']"):
        key = diag_service_request_key(root, service)
        if key is not None:
            result.add(key)
    return result


def strip_subfunction_prefix(value: Any) -> str:
    text = usable_text(value)
    text = re.sub(r"^\s*(?:0[xX][0-9A-Fa-f]{1,2}|[0-9A-Fa-f]{2})\s*", "", text)
    return text.strip()


def security_access_instance_name(subfunction: int, access: dict[str, Any]) -> str:
    if subfunction == 0x07:
        return "RequestSeedOfSecurityLevelFBL"
    if subfunction == 0x08:
        return "SendKeyOfSecurityLevelFBL"
    raw_name = strip_subfunction_prefix(access.get("subfunction_name", ""))
    if raw_name:
        return sanitize_short_name(raw_name, f"SecurityLevel{subfunction:02X}")
    prefix = "RequestSeed" if subfunction % 2 else "SendKey"
    return sanitize_short_name(f"{prefix}OfSecurityLevel{subfunction:02X}", f"SecurityLevel{subfunction:02X}")


def ensure_state_transition(root: etree._Element, id_gen: Any, source_state: str, target_state: str) -> str | None:
    chart = BASE.first_by_short(root, "STATE-CHART", "SecurityAccess")
    if chart is None:
        return None
    transitions = chart.find("STATE-TRANSITIONS")
    if transitions is None:
        transitions = BASE.element("STATE-TRANSITIONS")
        BASE.replace_child(chart, "STATE-TRANSITIONS", transitions, before_tags={"START-STATE-SNREF", "STATES"})

    for transition in transitions.findall("STATE-TRANSITION"):
        source = transition.find("SOURCE-SNREF")
        target = transition.find("TARGET-SNREF")
        if (
            source is not None
            and target is not None
            and source.get("SHORT-NAME") == source_state
            and target.get("SHORT-NAME") == target_state
            and transition.get("ID")
        ):
            return transition.get("ID")

    transition = BASE.sub(transitions, "STATE-TRANSITION", attrib={"ID": id_gen.new("ST")})
    BASE.sub(transition, "SHORT-NAME", sanitize_short_name(f"{source_state}_{target_state}", "SecurityTransition"))
    BASE.sub(transition, "LONG-NAME", f"{source_state} {target_state}")
    BASE.sub(transition, "SOURCE-SNREF", attrib={"SHORT-NAME": source_state})
    BASE.sub(transition, "TARGET-SNREF", attrib={"SHORT-NAME": target_state})
    return transition.get("ID")


def set_service_transition_refs(service: etree._Element, transition_id: str | None) -> None:
    existing = service.find("STATE-TRANSITION-REFS")
    if transition_id is None:
        if existing is not None:
            service.remove(existing)
        return
    refs = BASE.element("STATE-TRANSITION-REFS")
    BASE.sub(refs, "STATE-TRANSITION-REF", attrib={"ID-REF": transition_id})
    BASE.replace_child(service, "STATE-TRANSITION-REFS", refs, before_tags={"COMPARAM-REFS", "REQUEST-REF"})


def security_access_target_state(subfunction: int, access: dict[str, Any]) -> str:
    text = normalize_access(access.get("subfunction_name", ""))
    if "FBL" in text:
        return "Unlocked_FBL"
    if "LEVEL_9" in text or "LEVEL9" in text:
        return "Level_9"
    if subfunction == 0x08:
        return "Unlocked_FBL"
    return "UnlockedL1"


def generate_boot_security_access_services(root: etree._Element, id_gen: Any, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    existing_keys = existing_diag_service_keys(root)
    for (service_id, subfunction), access in sorted(access_map.items()):
        if service_id != 0x27 or subfunction is None:
            continue
        if "Boot" not in (access.get("sources") or set()):
            continue
        if (service_id, subfunction) in existing_keys:
            continue

        is_seed = bool(subfunction % 2)
        base_service = "SeedLevel1_Request" if is_seed else "KeyLevel1_Send"
        if BASE.first_by_short(root, "DIAG-SERVICE", base_service) is None:
            continue

        instance_name = security_access_instance_name(subfunction, access)
        service_qualifier = "Request" if is_seed else "Send"
        service_name = f"{instance_name}_{service_qualifier}"
        service, request, positive, negative = clone_service_bundle(
            root,
            id_gen,
            base_service,
            service_name,
            f"{instance_name} {service_qualifier}",
            service_qualifier,
            service_qualifier,
            instance_name,
            instance_name,
            service_access_state_ids(root, access.get("sessions") or {"Programming"}, access.get("security") or "N"),
        )
        set_service_semantic_and_funct_class(root, service, "SECURITY", "SecurityAccess")

        subfunction_semantics = ("ACCESSMODE", "SUBFUNCTION")
        for message in (request, positive):
            if message is not None:
                for semantic in subfunction_semantics:
                    set_first_coded_by_semantic(message, semantic, subfunction)

        if negative is not None:
            set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x27)

        if not is_seed:
            target_state = security_access_target_state(subfunction, access)
            transition_id = ensure_state_transition(root, id_gen, "Locked", target_state)
            set_service_transition_refs(service, transition_id)

        existing_keys.add((service_id, subfunction))


def prune_vf_unsupported_security_access_services(root: etree._Element, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    supported = {key for key in access_map if key[0] == 0x27}
    if not supported:
        return
    to_remove: list[str] = []
    for service in root.xpath("//*[local-name()='DIAG-SERVICE']"):
        key = diag_service_request_key(root, service)
        if key is None or key[0] != 0x27:
            continue
        if key not in supported:
            short_name = service.findtext("SHORT-NAME")
            if short_name:
                to_remove.append(short_name)
    for short_name in to_remove:
        BASE.remove_service_and_messages(root, short_name)


def ensure_vf_security_access_subfunction_scales(root: etree._Element, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    security_subfunctions = sorted(subfunction for service_id, subfunction in access_map if service_id == 0x27 and subfunction is not None)
    if not security_subfunctions:
        return
    dop = BASE.first_by_short(root, "DATA-OBJECT-PROP", "Subfunction_SecurityAccess")
    if dop is None:
        return
    scales = dop.find(".//COMPU-SCALES")
    if scales is None:
        internal = dop.find(".//COMPU-INTERNAL-TO-PHYS")
        if internal is None:
            compu = dop.find("COMPU-METHOD")
            if compu is None:
                compu = BASE.sub(dop, "COMPU-METHOD")
                BASE.sub(compu, "CATEGORY", "TEXTTABLE")
            internal = BASE.sub(compu, "COMPU-INTERNAL-TO-PHYS")
        scales = BASE.sub(internal, "COMPU-SCALES")

    existing: set[int] = set()
    for scale in scales.findall("COMPU-SCALE"):
        low = parse_int_cell(scale.findtext("LOWER-LIMIT"))
        high = parse_int_cell(scale.findtext("UPPER-LIMIT"), low)
        if low is not None and high == low:
            existing.add(low)

    for subfunction in security_subfunctions:
        if subfunction in existing:
            continue
        scale = BASE.sub(scales, "COMPU-SCALE")
        BASE.sub(scale, "LOWER-LIMIT", subfunction)
        BASE.sub(scale, "UPPER-LIMIT", subfunction)
        inverse = BASE.sub(scale, "COMPU-INVERSE-VALUE")
        BASE.sub(inverse, "V", 1 if subfunction % 2 else 2)
        const = BASE.sub(scale, "COMPU-CONST")
        BASE.sub(const, "VT", "requestSeed" if subfunction % 2 else "sendKey")


def update_vf_security_access_preconditions(root: etree._Element, id_gen: Any, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    for service in root.xpath("//*[local-name()='DIAG-SERVICE']"):
        key = diag_service_request_key(root, service)
        if key is None or key[0] != 0x27:
            continue
        access = access_map.get(key)
        if not access:
            continue
        update_flat_preconditions(service, service_access_state_ids(root, access.get("sessions") or {"Extended", "Programming"}, access.get("security") or "N"))
        subfunction = key[1]
        if subfunction is not None and not subfunction % 2:
            target_state = security_access_target_state(subfunction, access)
            transition_id = ensure_state_transition(root, id_gen, "Locked", target_state)
            set_service_transition_refs(service, transition_id)


def data_object_prop_id(root: etree._Element, short_name: str) -> str | None:
    dop = BASE.first_by_short(root, "DATA-OBJECT-PROP", short_name)
    return dop.get("ID") if dop is not None else None


def make_core_positive_response(
    root: etree._Element,
    node_id: str,
    short_name: str,
    long_name: str,
    request_sid: int,
    subfunction_param: etree._Element | None,
    subfunction_value: int,
) -> etree._Element:
    response = BASE.element("POS-RESPONSE", attrib={"ID": node_id})
    BASE.sub(response, "SHORT-NAME", sanitize_short_name(short_name, "PR"))
    BASE.sub(response, "LONG-NAME", long_name)
    params = BASE.sub(response, "PARAMS")
    params.append(coded_const_param("SID_PR", "SID-PR", 0, request_sid + 0x40, "SERVICE-ID"))

    sub_short_name = subfunction_param.findtext("SHORT-NAME") if subfunction_param is not None else "SubFunction"
    sub_long_name = subfunction_param.findtext("LONG-NAME") if subfunction_param is not None else sub_short_name
    params.append(
        coded_const_param(
            sub_short_name or "SubFunction",
            sub_long_name or sub_short_name or "SubFunction",
            1,
            subfunction_value,
            "SUBFUNCTION",
        )
    )

    if request_sid == 0x10:
        p2_id = data_object_prop_id(root, "P2")
        p2ex_id = data_object_prop_id(root, "P2Ex")
        if p2_id:
            params.append(BASE.make_value_param("P2", "P2", 2, 0, p2_id))
        if p2ex_id:
            params.append(BASE.make_value_param("P2Ex", "P2Ex", 4, 0, p2ex_id))
    return response


def create_core_response_service(root: etree._Element, id_gen: Any, no_response_name: str) -> None:
    response_name = no_response_name.removesuffix("_NoResponse")
    if response_name == no_response_name or BASE.first_by_short(root, "DIAG-SERVICE", response_name) is not None:
        return

    base_service, base_request, _, base_negative = service_messages(root, no_response_name)
    request_sid = first_coded_value_by_semantic(base_request, "SERVICE-ID")
    no_response_subfunction = first_coded_value_by_semantic(base_request, "SUBFUNCTION")
    if request_sid is None or no_response_subfunction is None:
        return
    response_subfunction = no_response_subfunction & 0x7F

    request = copy.deepcopy(base_request)
    negative = copy.deepcopy(base_negative) if base_negative is not None else None
    service = copy.deepcopy(base_service)

    request_id = id_gen.new("RQ")
    positive_id = id_gen.new("PR")
    negative_id = id_gen.new("NR") if negative is not None else None
    service_id = id_gen.new("SVC")

    request.set("ID", request_id)
    BASE.refresh_internal_ids(request, id_gen)
    BASE.set_short_long(request, f"RQ_{response_name}", f"RQ {response_name}")
    set_first_coded_by_semantic(request, "SUBFUNCTION", response_subfunction)

    subfunction_param = first_coded_param_by_semantic(request, "SUBFUNCTION")
    positive = make_core_positive_response(
        root,
        positive_id,
        f"PR_{response_name}",
        f"PR {response_name}",
        request_sid,
        subfunction_param,
        response_subfunction,
    )

    if negative is not None and negative_id is not None:
        negative.set("ID", negative_id)
        BASE.refresh_internal_ids(negative, id_gen)
        BASE.set_short_long(negative, f"NR_{response_name}", f"NR {response_name}")

    service.set("ID", service_id)
    service.attrib.pop("TRANSMISSION-MODE", None)
    BASE.refresh_internal_ids(service, id_gen)
    BASE.set_short_long(service, response_name, (base_service.findtext("LONG-NAME") or response_name).replace("_NoResponse", ""))

    request_ref = service.find("REQUEST-REF")
    if request_ref is not None:
        request_ref.set("ID-REF", request_id)

    pos_refs = BASE.element("POS-RESPONSE-REFS")
    BASE.sub(pos_refs, "POS-RESPONSE-REF", attrib={"ID-REF": positive_id})
    BASE.replace_child(service, "POS-RESPONSE-REFS", pos_refs, before_tags={"NEG-RESPONSE-REFS"})

    neg_ref = service.find(".//NEG-RESPONSE-REF")
    if neg_ref is not None and negative_id is not None:
        neg_ref.set("ID-REF", negative_id)

    update_service_sdg(
        service,
        id_gen,
        sdg_value(base_service, "ServiceQualifier", "Send"),
        sdg_value(base_service, "ServiceName", sdg_value(base_service, "ServiceQualifier", "Send")),
        sdg_value(base_service, "DiagInstanceQualifier", response_name),
        sdg_value(base_service, "DiagInstanceName", response_name),
    )

    find_container(root, "REQUESTS").append(request)
    find_container(root, "POS-RESPONSES").append(positive)
    if negative is not None:
        find_container(root, "NEG-RESPONSES").append(negative)
    find_container(root, "DIAG-COMMS").append(service)


def ensure_core_response_services(root: etree._Element, id_gen: Any) -> None:
    for no_response_name in CORE_RESPONSE_SERVICE_SOURCES:
        if BASE.first_by_short(root, "DIAG-SERVICE", no_response_name) is not None:
            create_core_response_service(root, id_gen, no_response_name)


def copy_state_transition_refs(source_service: etree._Element | None, target_service: etree._Element) -> None:
    if source_service is None:
        return
    source_refs = source_service.find("STATE-TRANSITION-REFS")
    if source_refs is None:
        existing = target_service.find("STATE-TRANSITION-REFS")
        if existing is not None:
            target_service.remove(existing)
        return
    BASE.replace_child(
        target_service,
        "STATE-TRANSITION-REFS",
        copy.deepcopy(source_refs),
        before_tags={"COMPARAM-REFS", "REQUEST-REF"},
    )


def ensure_vf_session_control_services(root: etree._Element, id_gen: Any, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    session_defs = {
        0x01: ("DefaultSession", "Default Session", "DefaultSession_Start", "DefaultSession_Start_NoResponse"),
        0x02: ("ProgrammingSession", "Programming Session", "ProgrammingSession_Start", "ProgrammingSession_Start_NoResponse"),
        0x03: (
            "ExtendedDiagnosticSession",
            "Extended Diagnostic Session",
            "ExtendedDiagnosticSession_Start",
            "ExtendedDiagnosticSession_Start_NoResponse",
        ),
        0x41: ("CodingSession", "Coding Session"),
    }
    for subfunction, definition in session_defs.items():
        instance_qualifier = definition[0]
        instance_name = definition[1]
        service_short_name = definition[2] if len(definition) > 2 else f"{instance_qualifier}_Start"
        no_response_short_name = definition[3] if len(definition) > 3 else f"{service_short_name}_NoResponse"
        access = access_map.get((0x10, subfunction))
        if not access or not access.get("sessions"):
            continue

        preconditions = service_access_state_ids(root, access.get("sessions") or {"Extended"}, access.get("security") or "N")
        if BASE.first_by_short(root, "DIAG-SERVICE", service_short_name) is None:
            service, request, positive, negative = clone_service_bundle(
                root,
                id_gen,
                "ExtendedDiagnosticSession_Start",
                service_short_name,
                f"{instance_name} Start",
                "Start",
                "Start",
                instance_qualifier,
                instance_name,
                preconditions,
            )
            for message in (request, positive):
                if message is not None:
                    set_first_coded_by_semantic(message, "SUBFUNCTION", subfunction)
            if negative is not None:
                set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x10)

        if BASE.first_by_short(root, "DIAG-SERVICE", no_response_short_name) is None:
            no_response_base = (
                "DefaultSession_Start_NoResponse"
                if subfunction == 0x01
                else "ExtendedDiagnosticSession_Start_NoResponse"
            )
            service, request, _, negative = clone_service_bundle(
                root,
                id_gen,
                no_response_base,
                no_response_short_name,
                f"{instance_name} Start_NoResponse",
                "Start",
                "Start",
                instance_qualifier,
                instance_name,
                preconditions,
                positive_response_suppressed="yes",
            )
            set_first_coded_by_semantic(request, "SUBFUNCTION", subfunction | 0x80)
            if negative is not None:
                set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x10)
            copy_state_transition_refs(BASE.first_by_short(root, "DIAG-SERVICE", service_short_name), service)


def ensure_vf_communication_control_services(root: etree._Element, id_gen: Any, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    service_defs = {
        0x01: ("EnableRxAndDisableTx", "Enable Rx And Disable Tx"),
    }
    for subfunction, (instance_qualifier, instance_name) in service_defs.items():
        access = access_map.get((0x28, subfunction))
        if not access or not access.get("sessions"):
            continue

        preconditions = service_access_state_ids(root, access.get("sessions") or {"Extended", "Programming"}, access.get("security") or "N")
        service_short_name = f"{instance_qualifier}_Control"
        if BASE.first_by_short(root, "DIAG-SERVICE", service_short_name) is None:
            service, request, positive, negative = clone_service_bundle(
                root,
                id_gen,
                "EnableRxAndEnableTx_Control",
                service_short_name,
                f"{instance_name} Control",
                "Control",
                "Control",
                instance_qualifier,
                instance_name,
                preconditions,
            )
            for message in (request, positive):
                if message is not None:
                    set_first_coded_by_semantic(message, "SUBFUNCTION", subfunction)
            if negative is not None:
                set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x28)

        no_response_short_name = f"{service_short_name}_NoResponse"
        if BASE.first_by_short(root, "DIAG-SERVICE", no_response_short_name) is None:
            service, request, _, negative = clone_service_bundle(
                root,
                id_gen,
                "EnableRxAndEnableTx_Control_NoResponse",
                no_response_short_name,
                f"{instance_name} Control_NoResponse",
                "Control",
                "Control",
                instance_qualifier,
                instance_name,
                preconditions,
                positive_response_suppressed="yes",
            )
            set_first_coded_by_semantic(request, "SUBFUNCTION", subfunction | 0x80)
            if negative is not None:
                set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x28)


def update_core_service_preconditions(root: etree._Element, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    for service_name, (service_id, subfunction, fallback_sessions) in CORE_SERVICE_PRECONDITION_TARGETS.items():
        service = BASE.first_by_short(root, "DIAG-SERVICE", service_name)
        if service is None:
            continue
        access = access_map.get((service_id, subfunction)) or access_map.get((service_id, None))
        if access:
            sessions = access.get("sessions") or fallback_sessions
            security = access.get("security") or "N"
        else:
            sessions = fallback_sessions
            security = "N"
        update_flat_preconditions(service, service_access_state_ids(root, sessions, security))


def set_coded_value(message: etree._Element, param_short_name: str, value: int) -> None:
    for param in message.xpath('.//*[local-name()="PARAM" and SHORT-NAME=$name]', name=param_short_name):
        coded = param.find("CODED-VALUE")
        if coded is not None:
            coded.text = str(value)


def set_first_coded_by_semantic(message: etree._Element, semantic: str, value: int) -> None:
    for param in message.xpath('.//*[local-name()="PARAM" and @SEMANTIC=$semantic]', semantic=semantic):
        coded = param.find("CODED-VALUE")
        if coded is not None:
            coded.text = str(value)
            return


def set_value_param_dop(message: etree._Element, dop_id: str, short_name: str, long_name: str, byte_position: int | None = None) -> None:
    value_params = message.xpath('./*[local-name()="PARAMS"]/*[local-name()="PARAM" and @xsi:type="VALUE"]', namespaces={"xsi": BASE.XSI_NS})
    if not value_params:
        return
    param = value_params[-1]
    short = param.find("SHORT-NAME")
    if short is not None:
        short.text = sanitize_short_name(short_name, "Data")
    long = param.find("LONG-NAME")
    if long is not None:
        long.text = long_name or short_name
    if byte_position is not None:
        byte = param.find("BYTE-POSITION")
        if byte is not None:
            byte.text = str(byte_position)
    dop_ref = param.find("DOP-REF")
    if dop_ref is None:
        dop_ref = BASE.sub(param, "DOP-REF")
    dop_ref.set("ID-REF", dop_id)


def remove_param_by_short_name(message: etree._Element, short_name: str) -> None:
    params = message.find("PARAMS")
    if params is None:
        return
    for param in list(params):
        if param.findtext("SHORT-NAME") == short_name:
            params.remove(param)


def generate_flat_did_services(root: etree._Element, id_gen: Any, dids: list[DidDef]) -> None:
    for did in dids:
        if did.readable:
            service_name = f"{did.short_name}_Read"
            service, request, positive, negative = clone_service_bundle(
                root,
                id_gen,
                "z_7_Read",
                service_name,
                f"{did.long_name} Read",
                "Read",
                "Read",
                did.short_name,
                did.long_name,
                did_state_ids(root, did, "read"),
                static_value=did.did,
            )
            set_service_semantic_and_funct_class(root, service, "IDENTIFICATION", "ECU_Identification")
            set_coded_value(request, "RecordDataIdentifier", did.did)
            if positive is not None:
                set_coded_value(positive, "RecordDataIdentifier", did.did)
                set_value_param_dop(positive, did.structure_id, did.short_name, did.long_name, byte_position=3)
            if negative is not None:
                set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x22)

        if did.writable:
            service_name = f"{did.short_name}_Write"
            service, _, _, _ = clone_service_bundle(
                root,
                id_gen,
                "z_7_Read",
                service_name,
                f"{did.long_name} Write",
                "Write",
                "Write",
                did.short_name,
                did.long_name,
                did_state_ids(root, did, "write"),
                static_value=did.did,
            )
            set_service_semantic_and_funct_class(root, service, "IDENTIFICATION", "ECU_Identification")
            request = make_write_request(id_gen.new("RQ"), f"RQ_{service_name}", did)
            positive = make_write_positive(id_gen.new("PR"), f"PR_{service_name}", did)
            negative = make_negative_response(id_gen.new("NR"), f"NR_{service_name}", f"NR {did.long_name} Write", 0x2E, "Write")
            replace_service_refs(root, service, request, positive, negative)


def replace_service_refs(
    root: etree._Element,
    service: etree._Element,
    request: etree._Element,
    positive: etree._Element | None,
    negative: etree._Element | None,
) -> None:
    old_request_ref = service.find("REQUEST-REF")
    old_pos_ref = service.find(".//POS-RESPONSE-REF")
    old_neg_ref = service.find(".//NEG-RESPONSE-REF")
    old_ids = [
        old_request_ref.get("ID-REF") if old_request_ref is not None else None,
        old_pos_ref.get("ID-REF") if old_pos_ref is not None else None,
        old_neg_ref.get("ID-REF") if old_neg_ref is not None else None,
    ]
    for old_id in old_ids:
        old_node = find_by_id(root, old_id)
        if old_node is not None and old_node.getparent() is not None:
            old_node.getparent().remove(old_node)
    find_container(root, "REQUESTS").append(request)
    if positive is not None:
        find_container(root, "POS-RESPONSES").append(positive)
    if negative is not None:
        find_container(root, "NEG-RESPONSES").append(negative)
    if old_request_ref is not None:
        old_request_ref.set("ID-REF", request.get("ID"))
    if old_pos_ref is not None and positive is not None:
        old_pos_ref.set("ID-REF", positive.get("ID"))
    if old_neg_ref is not None and negative is not None:
        old_neg_ref.set("ID-REF", negative.get("ID"))


def coded_const_param(short_name: str, long_name: str, byte_position: int, coded_value: int, semantic: str, bit_length: int = 8) -> etree._Element:
    param = BASE.element("PARAM", attrib={"SEMANTIC": semantic})
    BASE.set_xsi_type(param, "CODED-CONST")
    BASE.sub(param, "SHORT-NAME", short_name)
    BASE.sub(param, "LONG-NAME", long_name)
    BASE.sub(param, "BYTE-POSITION", byte_position)
    BASE.sub(param, "CODED-VALUE", coded_value)
    coded_type = BASE.sub(param, "DIAG-CODED-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32"})
    BASE.set_xsi_type(coded_type, "STANDARD-LENGTH-TYPE")
    BASE.sub(coded_type, "BIT-LENGTH", bit_length)
    return param


def make_write_request(node_id: str, short_name: str, did: DidDef) -> etree._Element:
    request = BASE.element("REQUEST", attrib={"ID": node_id})
    BASE.sub(request, "SHORT-NAME", sanitize_short_name(short_name, "RQ"))
    BASE.sub(request, "LONG-NAME", f"RQ {did.long_name} Write")
    params = BASE.sub(request, "PARAMS")
    params.append(coded_const_param("SID_RQ", "SID-RQ", 0, 0x2E, "SERVICE-ID"))
    params.append(coded_const_param("RecordDataIdentifier", "RecordDataIdentifier", 1, did.did, "ID", bit_length=16))
    params.append(BASE.make_value_param(did.short_name, did.long_name, 3, 0, did.structure_id))
    return request


def make_write_positive(node_id: str, short_name: str, did: DidDef) -> etree._Element:
    response = BASE.element("POS-RESPONSE", attrib={"ID": node_id})
    BASE.sub(response, "SHORT-NAME", sanitize_short_name(short_name, "PR"))
    BASE.sub(response, "LONG-NAME", f"PR {did.long_name} Write")
    params = BASE.sub(response, "PARAMS")
    params.append(coded_const_param("SID_PR", "SID-PR", 0, 0x6E, "SERVICE-ID"))
    params.append(coded_const_param("RecordDataIdentifier", "RecordDataIdentifier", 1, did.did, "ID", bit_length=16))
    return response


def make_negative_response(node_id: str, short_name: str, long_name: str, request_sid: int, qualifier: str) -> etree._Element:
    response = BASE.element("NEG-RESPONSE", attrib={"ID": node_id})
    BASE.sub(response, "SHORT-NAME", sanitize_short_name(short_name, "NR"))
    BASE.sub(response, "LONG-NAME", long_name)
    params = BASE.sub(response, "PARAMS")
    params.append(coded_const_param("SID_NR", "SID-NR", 0, 0x7F, "SERVICE-ID"))
    params.append(coded_const_param("SIDRQ_NR", "SIDRQ-NR", 1, request_sid, "SERVICEIDRQ"))
    value = BASE.make_value_param(qualifier, qualifier, 2, 0, "_1")
    params.append(value)
    nrc = BASE.element("PARAM", attrib={"SEMANTIC": "DATA"})
    BASE.set_xsi_type(nrc, "NRC-CONST")
    BASE.sub(nrc, "SHORT-NAME", sanitize_short_name(f"NRCConst_{qualifier}", "NRCConst"))
    BASE.sub(nrc, "LONG-NAME", qualifier)
    BASE.sub(nrc, "BYTE-POSITION", 2)
    coded_values = BASE.sub(nrc, "CODED-VALUES")
    for code in (0x13, 0x22, 0x31, 0x33, 0x72, 0x7F):
        BASE.sub(coded_values, "CODED-VALUE", code)
    coded_type = BASE.sub(nrc, "DIAG-CODED-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32"})
    BASE.set_xsi_type(coded_type, "STANDARD-LENGTH-TYPE")
    BASE.sub(coded_type, "BIT-LENGTH", 8)
    params.append(nrc)
    return response


def set_service_semantic_and_funct_class(
    root: etree._Element,
    service: etree._Element,
    semantic: str,
    funct_class_short_name: str,
) -> None:
    service.set("SEMANTIC", semantic)
    class_node = BASE.first_by_short(root, "FUNCT-CLASS", funct_class_short_name)
    if class_node is None or not class_node.get("ID"):
        return
    refs = service.find("FUNCT-CLASS-REFS")
    if refs is None:
        refs = BASE.element("FUNCT-CLASS-REFS")
        BASE.replace_child(service, "FUNCT-CLASS-REFS", refs, before_tags={"AUDIENCE", "PRE-CONDITION-STATE-REFS", "REQUEST-REF"})
    for child in list(refs):
        refs.remove(child)
    BASE.sub(refs, "FUNCT-CLASS-REF", attrib={"ID-REF": class_node.get("ID")})


def generate_flat_io_services(root: etree._Element, id_gen: Any, io_dids: list[IoDidDef]) -> None:
    control_templates = {
        0: ("z_ReturnControl", "ReturnControl", "ReturnControl"),
        1: ("z_ReturnControl", "Reset", "Reset"),
        2: ("z_ReturnControl", "Freeze", "Freeze"),
        3: ("z_Control", "Control", "Control"),
    }
    default_states = [state_id for name, state_id in state_id_map(root).items() if name in {"Extended", "UnlockedL1", "Level_9"}]
    for io_did in io_dids:
        for control in sorted(io_did.controls):
            template, qualifier, service_label = control_templates.get(control, control_templates[3])
            service_name = f"{io_did.short_name}_{qualifier}"
            _, request, positive, negative = clone_service_bundle(
                root,
                id_gen,
                template,
                service_name,
                f"{io_did.long_name} {service_label}",
                qualifier,
                service_label,
                io_did.short_name,
                io_did.long_name,
                default_states,
            )
            set_coded_value(request, "DataIdentifier", io_did.did)
            set_coded_value(request, "ControlOptionRecord_InputOutputControlParameter", control)
            if control == 3:
                set_value_param_dop(request, io_did.structure_id, "ControlOptionRecord", "ControlOptionRecord", byte_position=4)
            else:
                remove_param_by_short_name(request, "ControlOptionRecord")
            if positive is not None:
                set_coded_value(positive, "DataIdentifier", io_did.did)
                set_coded_value(positive, "ControlStatusRecord_InputOutputControlParameter", control)
                set_value_param_dop(positive, io_did.structure_id, "ControlStatusRecord", "ControlStatusRecord", byte_position=4)
            if negative is not None:
                set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x2F)

def generate_flat_routine_services(root: etree._Element, id_gen: Any, routines: list[RoutineDef]) -> None:
    states = state_id_map(root)
    default_states = [state for state in (states.get("Extended"), states.get("Programming"), states.get("UnlockedL1"), states.get("Unlocked_FBL")) if state]
    labels = {1: ("Start", "Start"), 2: ("Stop", "Stop"), 3: ("RequestResults", "RequestResults")}
    for routine in routines:
        for control_type, subfn in sorted(routine.subfunctions.items()):
            if not subfn.supported:
                continue
            qualifier, service_label = labels.get(control_type, (f"Control_{control_type:02X}", f"Control {control_type:02X}"))
            service_name = f"{routine.short_name}_{qualifier}"
            service, _, _, _ = clone_service_bundle(
                root,
                id_gen,
                "z_Read",
                service_name,
                f"{routine.long_name} {service_label}",
                qualifier,
                service_label,
                routine.short_name,
                routine.long_name,
                default_states,
            )
            request = make_routine_request(id_gen.new("RQ"), f"RQ_{service_name}", routine, subfn)
            positive = make_routine_positive(id_gen.new("PR"), f"PR_{service_name}", routine, subfn)
            negative = make_negative_response(id_gen.new("NR"), f"NR_{service_name}", f"NR {routine.long_name} {service_label}", 0x31, qualifier)
            replace_service_refs(root, service, request, positive, negative)


def make_routine_request(node_id: str, short_name: str, routine: RoutineDef, subfn: RoutineSubFunction) -> etree._Element:
    request = BASE.element("REQUEST", attrib={"ID": node_id})
    BASE.sub(request, "SHORT-NAME", sanitize_short_name(short_name, "RQ"))
    BASE.sub(request, "LONG-NAME", f"RQ {routine.long_name}")
    params = BASE.sub(request, "PARAMS")
    params.append(coded_const_param("SID_RQ", "SID-RQ", 0, 0x31, "SERVICE-ID"))
    params.append(coded_const_param("RoutineControlType", "RoutineControlType", 1, subfn.control_type, "SUBFUNCTION"))
    params.append(coded_const_param("RoutineIdentifier", "RoutineIdentifier", 2, routine.rid, "ID", bit_length=16))
    if subfn.option_structure_id:
        params.append(BASE.make_value_param("RoutineControlOptionRecord", "RoutineControlOptionRecord", 4, 0, subfn.option_structure_id))
    return request


def make_routine_positive(node_id: str, short_name: str, routine: RoutineDef, subfn: RoutineSubFunction) -> etree._Element:
    response = BASE.element("POS-RESPONSE", attrib={"ID": node_id})
    BASE.sub(response, "SHORT-NAME", sanitize_short_name(short_name, "PR"))
    BASE.sub(response, "LONG-NAME", f"PR {routine.long_name}")
    params = BASE.sub(response, "PARAMS")
    params.append(coded_const_param("SID_PR", "SID-PR", 0, 0x71, "SERVICE-ID"))
    params.append(coded_const_param("RoutineControlType", "RoutineControlType", 1, subfn.control_type, "SUBFUNCTION"))
    params.append(coded_const_param("RoutineIdentifier", "RoutineIdentifier", 2, routine.rid, "ID", bit_length=16))
    if subfn.status_structure_id:
        params.append(BASE.make_value_param("RoutineStatusRecord", "RoutineStatusRecord", 4, 0, subfn.status_structure_id))
    return response


def normalize_vf_texttable_dop(dop: etree._Element, bit_length: int, labels: list[tuple[int, str]]) -> None:
    compu = BASE.element("COMPU-METHOD")
    BASE.sub(compu, "CATEGORY", "TEXTTABLE")
    internal = BASE.sub(compu, "COMPU-INTERNAL-TO-PHYS")
    scales = BASE.sub(internal, "COMPU-SCALES")
    for value, label in labels:
        scale = BASE.sub(scales, "COMPU-SCALE")
        BASE.sub(scale, "LOWER-LIMIT", value)
        BASE.sub(scale, "UPPER-LIMIT", value)
        const = BASE.sub(scale, "COMPU-CONST")
        BASE.sub(const, "VT", label)

    coded = BASE.element(
        "DIAG-CODED-TYPE",
        attrib={"BASE-TYPE-ENCODING": "NONE", "BASE-DATA-TYPE": "A_UINT32"},
    )
    BASE.set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
    BASE.sub(coded, "BIT-LENGTH", bit_length)
    physical = BASE.element("PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UNICODE2STRING"})

    BASE.replace_child(dop, "COMPU-METHOD", compu, before_tags={"DIAG-CODED-TYPE", "PHYSICAL-TYPE"})
    BASE.replace_child(dop, "DIAG-CODED-TYPE", coded, before_tags={"PHYSICAL-TYPE"})
    BASE.replace_child(dop, "PHYSICAL-TYPE", physical)
    internal_constr = dop.find("INTERNAL-CONSTR")
    if internal_constr is not None:
        dop.remove(internal_constr)


def ensure_vf_texttable_dtc(root: etree._Element, id_gen: Any) -> str:
    existing = BASE.first_by_short(root, "DATA-OBJECT-PROP", "TextTable_DTC")
    if existing is not None and existing.get("ID"):
        normalize_vf_texttable_dop(existing, 24, [])
        return existing.get("ID")
    ddds = find_dictionary_spec(root)
    if ddds is None:
        raise RuntimeError("Template ODX has no DIAG-DATA-DICTIONARY-SPEC")
    data_object_props = ddds.find("DATA-OBJECT-PROPS")
    if data_object_props is None:
        data_object_props = BASE.sub(ddds, "DATA-OBJECT-PROPS")
    dop = BASE.sub(data_object_props, "DATA-OBJECT-PROP", attrib={"ID": id_gen.new("DOP")})
    BASE.sub(dop, "SHORT-NAME", "TextTable_DTC")
    BASE.sub(dop, "LONG-NAME", "Text Table DTC")
    normalize_vf_texttable_dop(dop, 24, [])
    return dop.get("ID")


def ensure_vf_tables_container(root: etree._Element) -> etree._Element:
    ddds = find_dictionary_spec(root)
    if ddds is None:
        raise RuntimeError("Template ODX has no DIAG-DATA-DICTIONARY-SPEC")
    tables = ddds.find("TABLES")
    if tables is None:
        tables = BASE.sub(ddds, "TABLES")
    return tables


def make_vf_ext_record_table_structure(root: etree._Element, id_gen: Any) -> str:
    existing = BASE.first_by_short(root, "STRUCTURE", "STRUC_DTCExtendedDataRecordNumbers")
    if existing is not None and existing.get("ID"):
        return existing.get("ID")
    structures = find_container(root, "STRUCTURES")
    structure = BASE.sub(structures, "STRUCTURE", attrib={"ID": id_gen.new("STR"), "IS-VISIBLE": "false"})
    BASE.sub(structure, "SHORT-NAME", "STRUC_DTCExtendedDataRecordNumbers")
    params = BASE.sub(structure, "PARAMS")
    params.append(
        BASE.make_value_param(
            "Record_Numbers",
            "Record Numbers",
            0,
            0,
            dop_id_by_short(root, "DATA-OBJECT-PROP", "DTCExtendedDataRecordNumbers_All") or "_54",
        )
    )
    return structure.get("ID")


def normalize_vf_extended_data_request(root: etree._Element, id_gen: Any, table_id: str) -> None:
    service = BASE.first_by_short(root, "DIAG-SERVICE", "FaultMemory_Read_DTC_extended_data_record_by_DTC_number")
    if service is None:
        return
    request_ref = service.find("REQUEST-REF")
    request = find_by_id(root, request_ref.get("ID-REF") if request_ref is not None else None)
    if request is None:
        return
    params = request.find("PARAMS")
    if params is None:
        params = BASE.sub(request, "PARAMS")
    table_key_id = id_gen.new("TK")
    for child in list(params):
        params.remove(child)
    params.append(coded_const_param("SID_RQ", "SID-RQ", 0, 0x19, "SERVICE-ID"))
    params.append(
        coded_const_param(
            "ReportDTCExtendedDataRecordByDtcNumber",
            "ReportDTCExtendedDataRecordByDtcNumber",
            1,
            0x06,
            "SUBFUNCTION",
        )
    )
    params.append(BASE.make_table_key_param("DTC", "DTC", 2, table_id, table_key_id))
    params.append(BASE.make_table_struct_param("DTCExtendedDataRecordNumber", "DTCExtendedDataRecordNumber", 5, table_key_id))


def normalize_vf_snapshot_request(root: etree._Element) -> None:
    service = BASE.first_by_short(root, "DIAG-SERVICE", "FaultMemory_Read_DTC_snapshot_record_by_DTC_number")
    if service is None:
        return
    request_ref = service.find("REQUEST-REF")
    request = find_by_id(root, request_ref.get("ID-REF") if request_ref is not None else None)
    if request is None:
        return
    for param in request.xpath('./*[local-name()="PARAMS"]/*[local-name()="PARAM"]'):
        if param.findtext("SHORT-NAME") == "Record_Numbers":
            dop_ref = param.find("DOP-REF")
            if dop_ref is not None:
                dop_ref.set("ID-REF", dop_id_by_short(root, "DATA-OBJECT-PROP", "DTCSnapshotRecordNumbers_All") or "_53")


def ensure_vf_dtc_extended_table(root: etree._Element, id_gen: Any, dtcs: list[DtcDef]) -> None:
    texttable_id = ensure_vf_texttable_dtc(root, id_gen)
    tables = ensure_vf_tables_container(root)
    table = BASE.first_by_short(root, "TABLE", "DTCExtendedDataRecordNumber")
    if table is None:
        table = BASE.sub(tables, "TABLE", attrib={"ID": id_gen.new("TAB")})
        BASE.sub(table, "SHORT-NAME", "DTCExtendedDataRecordNumber")
        BASE.sub(table, "LONG-NAME", "DTCExtendedDataRecordNumber")
        BASE.sub(table, "KEY-DOP-REF", attrib={"ID-REF": texttable_id})
    else:
        key_ref = table.find("KEY-DOP-REF")
        if key_ref is None:
            key_ref = BASE.element("KEY-DOP-REF", attrib={"ID-REF": texttable_id})
            table.insert(2, key_ref)
        key_ref.set("ID-REF", texttable_id)
    for row in table.findall("TABLE-ROW"):
        table.remove(row)
    structure_id = make_vf_ext_record_table_structure(root, id_gen)
    for dtc in dtcs:
        label = BASE.dtc_table_key(dtc)
        row = BASE.append_table_row(table, id_gen, f"TR_DTC_{dtc.byte_code:06X}", label, label, structure_id)
        sdgs = BASE.sub(row, "SDGS")
        sdg = BASE.sub(sdgs, "SDG")
        caption = BASE.sub(sdg, "SDG-CAPTION", attrib={"ID": id_gen.new("CAP")})
        BASE.sub(caption, "SHORT-NAME", "IsDefaultCase")
        BASE.sub(sdg, "SD", "Yes")
    normalize_vf_extended_data_request(root, id_gen, table.get("ID"))
    normalize_vf_snapshot_request(root)


def update_vf_dtc_dop(root: etree._Element, id_gen: Any, dtcs: list[DtcDef]) -> None:
    dtc_dop = BASE.first_by_short(root, "DTC-DOP", "RecordDataType")
    if dtc_dop is None:
        data_dop = BASE.first_by_short(root, "DATA-OBJECT-PROP", "RecordDataType")
        record_id = data_dop.get("ID") if data_dop is not None and data_dop.get("ID") else id_gen.new("DTC_DOP")
        if data_dop is not None and data_dop.getparent() is not None:
            data_dop.getparent().remove(data_dop)
        ddds = find_dictionary_spec(root)
        if ddds is None:
            return
        dtc_dops = ddds.find("DTC-DOPS")
        if dtc_dops is None:
            dtc_dops = BASE.element("DTC-DOPS")
            ddds.insert(0, dtc_dops)
        dtc_dop = BASE.sub(dtc_dops, "DTC-DOP", attrib={"ID": record_id})
        BASE.sub(dtc_dop, "SHORT-NAME", "RecordDataType")
        BASE.sub(dtc_dop, "LONG-NAME", "RecordDataType")
        coded = BASE.sub(
            dtc_dop,
            "DIAG-CODED-TYPE",
            attrib={"BASE-TYPE-ENCODING": "NONE", "BASE-DATA-TYPE": "A_UINT32"},
        )
        BASE.set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
        BASE.sub(coded, "BIT-LENGTH", 24)
        BASE.sub(dtc_dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32", "DISPLAY-RADIX": "HEX"})
        compu = BASE.sub(dtc_dop, "COMPU-METHOD")
        BASE.sub(compu, "CATEGORY", "IDENTICAL")
    dtcs_node = dtc_dop.find("DTCS")
    if dtcs_node is None:
        dtcs_node = BASE.sub(dtc_dop, "DTCS")
    for child in list(dtcs_node):
        dtcs_node.remove(child)
    caption_ids: dict[str, str] = {}
    for index, dtc in enumerate(dtcs):
        dtc_node = BASE.sub(dtcs_node, "DTC", attrib={"ID": id_gen.new("DTC")})
        BASE.sub(dtc_node, "SHORT-NAME", sanitize_short_name(f"DTC_{dtc.byte_code:06X}", "DTC"))
        BASE.sub(dtc_node, "TROUBLE-CODE", dtc.byte_code)
        BASE.sub(dtc_node, "DISPLAY-TROUBLE-CODE", dtc.display_code)
        BASE.sub(dtc_node, "TEXT", dtc.text or dtc.display_code)
        sdgs = BASE.sub(dtc_node, "SDGS")
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_SHORTNAME", f"DTC_0X{dtc.byte_code:06X}", first=index == 0)
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_PRIORITY_VALUE", dtc.priority or "2", first=index == 0)
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_SUPPORTED", "supported", first=index == 0)
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_CYCLE", "DEM_POWER", first=index == 0)
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_COUNTER", "40", first=index == 0)
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_SEVERITY_VALUE", "noSeverity", first=index == 0)
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_INIT_MONITOR_REQUIRED", "not required", first=index == 0)
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_FUNCTIONAL_UNIT_VALUE", "0x00", first=index == 0)

    ensure_vf_dtc_extended_table(root, id_gen, dtcs)
    BASE.update_dtc_text_table(root, dtcs)

    ext_table = BASE.first_by_short(root, "TABLE", "DTCExtendedDataRecordNumber")
    if ext_table is None:
        return
    default_structure = first_table_structure_ref(ext_table) or create_extended_record_number_structure(root, id_gen)
    BASE.clear_children(ext_table, "TABLE-ROW")
    for dtc in dtcs:
        label = BASE.dtc_table_key(dtc)
        row = BASE.append_table_row(ext_table, id_gen, f"TR_DTC_{dtc.byte_code:06X}", label, label, default_structure)
        sdgs = BASE.sub(row, "SDGS")
        sdg = BASE.sub(sdgs, "SDG")
        caption = BASE.sub(sdg, "SDG-CAPTION", attrib={"ID": id_gen.new("CAP")})
        BASE.sub(caption, "SHORT-NAME", "IsDefaultCase")
        BASE.sub(sdg, "SD", "Yes")


def first_table_structure_ref(table: etree._Element) -> str | None:
    ref = table.find(".//STRUCTURE-REF")
    return ref.get("ID-REF") if ref is not None else None


def create_extended_record_number_structure(root: etree._Element, id_gen: Any) -> str:
    structures = find_container(root, "STRUCTURES")
    structure_id = id_gen.new("STR")
    structure = BASE.sub(structures, "STRUCTURE", attrib={"ID": structure_id, "IS-VISIBLE": "false"})
    BASE.sub(structure, "SHORT-NAME", "STRUC_DTCExtendedDataRecordNumbers")
    params = BASE.sub(structure, "PARAMS")
    params.append(BASE.make_value_param("Record_Numbers", "Record Numbers", 0, 0, dop_id_by_short(root, "DATA-OBJECT-PROP", "DTCExtendedDataRecordNumbers_All") or "_62"))
    return structure_id


def dop_id_by_short(root: etree._Element, tag: str, short_name: str) -> str | None:
    node = BASE.first_by_short(root, tag, short_name)
    return node.get("ID") if node is not None else None


def snapshot_number_of_ids_param(
    root: etree._Element,
    short_name: str,
    long_name: str,
    byte_position: int,
) -> etree._Element:
    param = BASE.element("PARAM")
    BASE.set_xsi_type(param, "VALUE")
    BASE.sub(param, "SHORT-NAME", short_name)
    BASE.sub(param, "LONG-NAME", long_name)
    BASE.sub(param, "BYTE-POSITION", byte_position)
    BASE.sub(
        param,
        "DOP-REF",
        attrib={
            "ID-REF": (
                dop_id_by_short(root, "DATA-OBJECT-PROP", "DTCSnapshotRecordNumberOfIdentifiers")
                or dop_id_by_short(root, "DATA-OBJECT-PROP", "HexDump_1Byte")
                or "_73"
            )
        },
    )
    return param


def ensure_vf_ddds_container(root: etree._Element, tag: str, before_tags: set[str]) -> etree._Element:
    ddds = find_dictionary_spec(root)
    if ddds is None:
        raise RuntimeError("Template ODX has no DIAG-DATA-DICTIONARY-SPEC")
    existing = ddds.find(tag)
    if existing is not None:
        return existing
    container = BASE.element(tag)
    for index, child in enumerate(ddds):
        if child.tag in before_tags:
            ddds.insert(index, container)
            return container
    ddds.append(container)
    return container


def ensure_vf_snapshot_env_data(root: etree._Element, id_gen: Any) -> etree._Element:
    env_data = BASE.first_by_short(root, "ENV-DATA", "ENVDATA_ALLDTCS")
    if env_data is not None:
        return env_data
    env_datas = ensure_vf_ddds_container(root, "ENV-DATAS", {"UNIT-SPEC", "TABLES"})
    env_data = BASE.sub(env_datas, "ENV-DATA", attrib={"ID": id_gen.new("ENV")})
    BASE.sub(env_data, "SHORT-NAME", "ENVDATA_ALLDTCS")
    BASE.sub(env_data, "LONG-NAME", "ENVDATA_ALLDTCS")
    BASE.sub(env_data, "PARAMS")
    BASE.sub(env_data, "ALL-VALUE")
    return env_data


def ensure_vf_snapshot_env_data_desc(root: etree._Element, id_gen: Any, env_data_id: str) -> str:
    desc = BASE.first_by_short(root, "ENV-DATA-DESC", "DTCSnapshotRecordData")
    if desc is None:
        descs = ensure_vf_ddds_container(root, "ENV-DATA-DESCS", {"DATA-OBJECT-PROPS", "STRUCTURES"})
        desc = BASE.sub(descs, "ENV-DATA-DESC", attrib={"ID": id_gen.new("EDD")})
        BASE.sub(desc, "SHORT-NAME", "DTCSnapshotRecordData")
        BASE.sub(desc, "LONG-NAME", "DTCSnapshotRecordData")
    param_ref = desc.find("PARAM-SNREF")
    if param_ref is None:
        param_ref = BASE.element("PARAM-SNREF", attrib={"SHORT-NAME": "DTC"})
        env_refs = desc.find("ENV-DATA-REFS")
        insert_at = desc.index(env_refs) if env_refs is not None else len(desc)
        desc.insert(insert_at, param_ref)
    else:
        param_ref.set("SHORT-NAME", "DTC")
    env_refs = desc.find("ENV-DATA-REFS")
    if env_refs is None:
        env_refs = BASE.sub(desc, "ENV-DATA-REFS")
    for child in list(env_refs):
        env_refs.remove(child)
    BASE.sub(env_refs, "ENV-DATA-REF", attrib={"ID-REF": env_data_id})
    return desc.get("ID")


def update_vf_snapshot_and_extended_data(
    root: etree._Element,
    id_gen: Any,
    snapshots: list[SnapshotDef],
    extended_records: list[ExtendedRecordDef],
    snapshot_record_nums: list[int],
    snapshot_record_names: dict[int, str],
) -> None:
    update_snapshot_record_number_dop(
        root,
        "DTCSnapshotRecordNumbers_All",
        snapshot_record_nums,
        snapshot_record_names,
        include_all=True,
    )
    update_snapshot_record_number_dop(
        root,
        "DTCSnapshotRecordNumbers_All_except_FF",
        snapshot_record_nums,
        snapshot_record_names,
        include_all=False,
    )

    env_data = BASE.first_by_short(root, "ENV-DATA", "ENVDATA_ALLDTCS")
    if env_data is None and snapshots:
        env_data = ensure_vf_snapshot_env_data(root, id_gen)
    did_dop_id = dop_id_by_short(root, "DATA-OBJECT-PROP", "HexDump_2Byte") or "_2"
    if env_data is not None:
        params = env_data.find("PARAMS")
        if params is None:
            params = BASE.sub(env_data, "PARAMS")
        for child in list(params):
            params.remove(child)
        byte_position = 0
        for snapshot in snapshots:
            if not snapshot.structure_id:
                continue
            did_param = BASE.element("PARAM", attrib={"SEMANTIC": "DATA"})
            BASE.set_xsi_type(did_param, "PHYS-CONST")
            snapshot_short_name = sanitize_short_name(snapshot.desc, hex_short("DID", snapshot.did))
            BASE.sub(did_param, "SHORT-NAME", snapshot_short_name)
            BASE.sub(did_param, "LONG-NAME", snapshot.desc)
            BASE.sub(did_param, "BYTE-POSITION", byte_position)
            BASE.sub(did_param, "PHYS-CONSTANT-VALUE", snapshot.did)
            BASE.sub(did_param, "DOP-REF", attrib={"ID-REF": did_dop_id})
            params.append(did_param)
            byte_position += 2
            value_param = BASE.make_value_param(
                sanitize_short_name(f"{snapshot_short_name}_Data", "SnapshotData"),
                snapshot.desc,
                byte_position,
                0,
                snapshot.structure_id,
            )
            params.append(value_param)
            byte_position += max(1, snapshot.size)
        snapshot_env_desc_id = ensure_vf_snapshot_env_data_desc(root, id_gen, env_data.get("ID"))
        list_struct = BASE.first_by_short(root, "STRUCTURE", "ListOfDTCSnapshotRecord")
        if list_struct is not None and snapshot_env_desc_id:
            params = list_struct.find("PARAMS")
            if params is None:
                params = BASE.sub(list_struct, "PARAMS")
            for child in list(params):
                params.remove(child)
            params.append(
                BASE.make_value_param(
                    "Record_Numbers",
                    "Record Numbers",
                    0,
                    0,
                    dop_id_by_short(root, "DATA-OBJECT-PROP", "DTCSnapshotRecordNumbers_All") or "_53",
                )
            )
            params.append(
                snapshot_number_of_ids_param(
                    root,
                    "DTCSnapshotRecordNumberOfIdentifiers",
                    "DTCSnapshotRecordNumberOfIdentifiers",
                    1,
                )
            )
            params.append(
                BASE.make_value_param(
                    "DTCSnapshotRecordData",
                    "DTCSnapshotRecordData",
                    2,
                    0,
                    snapshot_env_desc_id,
                )
            )

    if extended_records:
        BASE.update_record_number_dop(root, "DTCExtendedDataRecordNumbers_All", extended_records, include_all=True)
        BASE.update_record_number_dop(root, "DTCExtendedDataRecordNumbers_All_except_FF", extended_records, include_all=False)
        mux = BASE.first_by_short(root, "MUX", "DTCExtendedDataRecordData")
        if mux is not None:
            cases = mux.find("CASES")
            if cases is None:
                cases = BASE.sub(mux, "CASES")
            for child in list(cases):
                cases.remove(child)
            for record in extended_records:
                if not record.structure_id:
                    continue
                case = BASE.sub(cases, "CASE")
                label = BASE.extended_record_label(record)
                BASE.sub(case, "SHORT-NAME", sanitize_short_name(f"Case_0x{record.record_num:02X}", "Case"))
                BASE.sub(case, "STRUCTURE-REF", attrib={"ID-REF": record.structure_id})
                BASE.sub(case, "LOWER-LIMIT", record.record_num)
                BASE.sub(case, "UPPER-LIMIT", record.record_num)


def collect_env_data_groups(params: etree._Element) -> dict[int, list[etree._Element]]:
    groups: dict[int, list[etree._Element]] = {}
    current_did: int | None = None
    for param in params.findall("PARAM"):
        if param.get(XSI_TYPE) == "PHYS-CONST":
            try:
                current_did = int(compact_text(param.findtext("PHYS-CONSTANT-VALUE")), 0)
            except ValueError:
                current_did = None
            if current_did is not None:
                groups[current_did] = [param]
            continue
        if current_did is not None:
            groups[current_did].append(param)
    return groups


def existing_env_data_dids(root: etree._Element) -> set[int]:
    env_data = BASE.first_by_short(root, "ENV-DATA", "ENVDATA_ALLDTCS")
    if env_data is None:
        return set()
    params = env_data.find("PARAMS")
    if params is None:
        return set()
    return set(collect_env_data_groups(params))


def env_param_end_byte(root: etree._Element, param: etree._Element) -> int:
    byte_position = parse_int_cell(param.findtext("BYTE-POSITION"), 0)
    dop_ref = param.find("DOP-REF")
    if dop_ref is None:
        return byte_position + 1
    dop = find_by_id(root, dop_ref.get("ID-REF"))
    if dop is None:
        return byte_position + 1
    bit_lengths = dop.xpath('.//*[local-name()="BIT-LENGTH"]/text()')
    if not bit_lengths:
        return byte_position + 1
    try:
        bit_length = int(bit_lengths[0])
    except ValueError:
        return byte_position + 1
    return byte_position + max(1, (bit_length + 7) // 8)


def update_snapshot_record_number_dop(
    root: etree._Element,
    short_name: str,
    record_nums: list[int],
    record_names: dict[int, str],
    *,
    include_all: bool,
) -> None:
    dop = BASE.first_by_short(root, "DATA-OBJECT-PROP", short_name)
    if dop is None:
        return
    labels: list[tuple[int, str]] = []
    seen: set[int] = set()
    for record_num in record_nums:
        if record_num == 0xFF:
            continue
        if record_num in seen:
            continue
        seen.add(record_num)
        labels.append((record_num, record_names.get(record_num, f"Snapshot Record 0x{record_num:02X}")))
    if include_all:
        labels.append((0xFF, record_names.get(0xFF, "All")))
    normalize_vf_texttable_dop(dop, 8, labels)


def update_base_variant_comparams(root: etree._Element, cover: CoverInfo) -> None:
    base_variant = get_base_variant(root)
    if base_variant is None:
        return
    base_short_name = base_variant.findtext("SHORT-NAME") or "JAC_ECU_CAN"
    comp_refs = base_variant.find("COMPARAM-REFS")
    if comp_refs is None:
        comp_refs = BASE.sub(base_variant, "COMPARAM-REFS")
    values = getattr(cover, "comm_params", {})
    for ref in comp_refs.findall("COMPARAM-REF"):
        id_ref = ref.get("ID-REF")
        if id_ref not in values:
            continue
        simple = ref.find("SIMPLE-VALUE")
        if simple is not None:
            simple.text = str(values[id_ref])
    unique_ref = None
    for ref in comp_refs.findall("COMPARAM-REF"):
        if ref.get("ID-REF") == "ISO_15765_2.CP_UniqueRespIdTable":
            unique_ref = ref
            break
    if unique_ref is None:
        unique_ref = BASE.sub(
            comp_refs,
            "COMPARAM-REF",
            attrib={"ID-REF": "ISO_15765_2.CP_UniqueRespIdTable", "DOCREF": "ISO_15765_2", "DOCTYPE": "COMPARAM-SUBSET"},
        )
    for child in list(unique_ref):
        unique_ref.remove(child)
    rx_phy = cover.rx_phy_id if cover.rx_phy_id is not None else 0x705
    tx = cover.tx_id if cover.tx_id is not None else 0x785
    rx_fun = cover.rx_fun_id if cover.rx_fun_id is not None else 0x7DF
    uudt_resp_id = UUDT_DISABLED_CAN_ID
    complex_value = BASE.sub(unique_ref, "COMPLEX-VALUE")
    for value in (
        "0",
        "normal segmented 11-bit transmit with FC",
        str(rx_phy),
        "0",
        "normal segmented 11-bit receive with FC",
        str(tx),
        "0",
        "normal unsegmented 11-bit receive",
        str(uudt_resp_id),
        base_short_name,
    ):
        BASE.sub(complex_value, "SIMPLE-VALUE", value)
    BASE.sub(unique_ref, "PROTOCOL-SNREF", attrib={"SHORT-NAME": "CAN"})


def update_session_timing(root: etree._Element, cover: CoverInfo) -> None:
    timing = getattr(cover, "session_timing", {"P2": 50, "P2Ex": 5000})
    text = f"{{P2={timing.get('P2', 50)}, P2Ex={timing.get('P2Ex', 5000)}}}"
    for state in root.xpath('//*[local-name()="STATE" and (SHORT-NAME="Default" or SHORT-NAME="Programming" or SHORT-NAME="Extended")]'):
        desc = state.find("DESC")
        if desc is None:
            desc = BASE.sub(state, "DESC")
        for child in list(desc):
            desc.remove(child)
        BASE.sub(desc, "p", text)


def validate_with_odxtools(pdx_path: Path) -> None:
    cmd = [sys.executable, "-m", "odxtools", "list", str(pdx_path), "--all"]
    result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
    if result.returncode != 0:
        details = (result.stderr or result.stdout).strip()
        raise RuntimeError(f"odxtools validation failed for {pdx_path}:\n{details}")


def find_default_xlsx(base_dir: Path) -> Path:
    candidates = sorted(path for path in base_dir.glob("*.xlsx") if not path.name.startswith("~$"))
    if not candidates:
        raise FileNotFoundError("No .xlsx survey file found in the current directory")
    return candidates[0]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Generate JAC PDX from a JAC diagnosis survey Excel file.")
    parser.add_argument("xlsx", nargs="?", type=Path, help="Input JAC diagnosis survey .xlsx file")
    parser.add_argument("--template", type=Path, default=Path("templates") / "JAC_ECU_CAN_v15.pdx", help="Template PDX")
    parser.add_argument("--output-dir", type=Path, default=Path("output"), help="Output directory")
    parser.add_argument("--no-validate", action="store_true", help="Skip odxtools validation")
    args = parser.parse_args(argv)

    xlsx_path = args.xlsx or find_default_xlsx(Path.cwd())
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)
    if not args.template.exists():
        raise FileNotFoundError(args.template)

    survey = parse_jac_survey(xlsx_path)
    output_pdx = args.output_dir / f"{xlsx_path.stem}.pdx"
    update_template(args.template, output_pdx, survey, validate=not args.no_validate)
    did_data_objects = sum(len(did.params) for did in survey.dids)
    converted_did_objects = sum(
        1
        for did in survey.dids
        for param in did.params
        if getattr(param.conversion, "kind", "identity") in {"enum", "linear"}
    )
    print(f"Generated: {output_pdx}")
    print(
        "Parsed: "
        f"{len(survey.dids)} DID identifiers, "
        f"{did_data_objects} DID data objects, "
        f"{converted_did_objects} converted DID data objects, "
        f"{len(survey.io_dids)} IO DIDs, "
        f"{len(survey.routines)} routines, "
        f"{len(survey.dtcs)} DTCs, "
        f"{len(survey.snapshots)} snapshot DIDs, "
        f"{len(getattr(survey, 'snapshot_record_nums', []))} snapshot records, "
        f"{len(survey.extended_records)} extended records"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

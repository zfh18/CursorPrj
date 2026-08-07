#!/usr/bin/env python3
"""Generate a CANdelaStudio-compatible SERES PDX from a SERES survey workbook.

The SERES workbook layout is close to the shared diagnostic survey model, but
the SERES CANdela template uses flat DID/IO services. This script keeps the OEM
parser and template writer separated:

    parse_seres_survey(.xlsx) -> shared SurveyData -> update SERES PDX
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import re
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Iterable

from lxml import etree
from openpyxl import load_workbook


XSI_NS = "http://www.w3.org/2001/XMLSchema-instance"
XSI_TYPE = f"{{{XSI_NS}}}type"
etree.register_namespace("xsi", XSI_NS)


@dataclass
class Conversion:
    kind: str = "identity"
    enum: list[tuple[int, int, str]] = field(default_factory=list)
    a: float = 1.0
    b: float = 0.0
    precision: int | None = None


@dataclass
class ParamDef:
    name: str
    long_name: str
    byte_pos: int
    bit_pos: int
    bit_len: int
    data_type: str = "Hex(Unsigned)"
    unit: str = ""
    conversion: Conversion = field(default_factory=Conversion)
    min_value: str = ""
    max_value: str = ""
    dop_id: str = ""


@dataclass
class DidDef:
    did: int
    desc: str
    size: int
    params: list[ParamDef] = field(default_factory=list)
    write_security: str = "N"
    sessions: list[str] = field(default_factory=list)
    structure_id: str = ""
    wrapper_id: str = ""
    short_name: str = ""
    long_name: str = ""

    @property
    def readable(self) -> bool:
        return any("R" in normalize_access(value) for value in self.sessions)

    @property
    def writable(self) -> bool:
        return any("W" in normalize_access(value) for value in self.sessions)


@dataclass
class IoDidDef:
    did: int
    desc: str
    size: int
    controls: set[int] = field(default_factory=set)
    params: list[ParamDef] = field(default_factory=list)
    structure_id: str = ""
    status_wrapper_id: str = ""
    control_request_wrapper_id: str = ""
    short_name: str = ""
    long_name: str = ""


@dataclass
class RoutineSubFunction:
    control_type: int
    supported: bool
    option_params: list[ParamDef] = field(default_factory=list)
    status_params: list[ParamDef] = field(default_factory=list)
    option_structure_id: str = ""
    status_structure_id: str = ""


@dataclass
class RoutineDef:
    rid: int
    desc: str
    security: str = "N"
    sessions: list[str] = field(default_factory=list)
    subfunctions: dict[int, RoutineSubFunction] = field(default_factory=dict)
    short_name: str = ""
    long_name: str = ""


@dataclass
class DtcDef:
    display_code: str
    byte_code: int
    text: str
    priority: str = ""


@dataclass
class SnapshotDef:
    record_num: int | None
    did: int
    desc: str
    size: int
    params: list[ParamDef] = field(default_factory=list)
    structure_id: str = ""


@dataclass
class ExtendedRecordDef:
    record_num: int
    desc: str
    size: int
    params: list[ParamDef] = field(default_factory=list)
    structure_id: str = ""


@dataclass
class CoverInfo:
    ecu_name: str = "SERES_ECU_CAN"
    vehicle: str = ""
    supplier: str = ""
    tx_id: int | None = None
    rx_phy_id: int | None = None
    rx_fun_id: int | None = None
    bus_type: str = "CAN"


@dataclass
class SurveyData:
    cover: CoverInfo
    dids: list[DidDef]
    io_dids: list[IoDidDef]
    routines: list[RoutineDef]
    dtcs: list[DtcDef]
    snapshots: list[SnapshotDef]
    extended_records: list[ExtendedRecordDef]


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip().lstrip("'")


def compact_text(value: Any) -> str:
    return re.sub(r"\s+", " ", cell_text(value).replace("\u3000", " ")).strip()


def normalize_access(value: Any) -> str:
    return compact_text(value).upper().replace(" ", "")


def has_chinese_text(value: Any) -> bool:
    return bool(re.search(r"[\u3400-\u4DBF\u4E00-\u9FFF\uF900-\uFAFF]", cell_text(value)))


def name_parts(value: Any) -> list[str]:
    text = cell_text(value).replace("\u3000", " ")
    parts = [part.strip() for part in re.split(r"[\r\n]+", text) if part.strip()]
    result: list[str] = []
    for part in parts:
        slash_parts = [piece.strip() for piece in re.split(r"\s*/\s*", part) if piece.strip()]
        if len(slash_parts) > 1 and any(has_chinese_text(piece) for piece in slash_parts):
            result.extend(slash_parts)
        else:
            result.append(part)
    return result


def canonical_name(identifier: str, display: str, fallback: str = "") -> str:
    identifier = compact_text(identifier)
    display = compact_text(display)
    if identifier and display and identifier != display:
        return f"{identifier}\n{display}"
    return display or identifier or fallback


def split_name(value: str) -> tuple[str, str]:
    text = cell_text(value).replace("\r\n", "\n").strip()
    parts = name_parts(text)
    if not parts:
        return "", ""
    english = next((part for part in parts if re.search(r"[A-Za-z]", part) and not has_chinese_text(part)), "")
    chinese = next((part for part in parts if has_chinese_text(part)), "")
    identifier = english or ("" if chinese else parts[0])
    return identifier, chinese or text


def parse_hex_cell(value: Any, *, max_value: int = 0xFFFFFFFF) -> int | None:
    text = compact_text(value).replace(" ", "")
    if not re.fullmatch(r"(?:0[xX])?[0-9A-Fa-f]{1,8}", text):
        return None
    result = int(text[2:] if text.lower().startswith("0x") else text, 16)
    return result if result <= max_value else None


def parse_int_cell(value: Any, default: int = 0) -> int:
    text = compact_text(value)
    if not text:
        return default
    match = re.search(r"-?\d+", text)
    return int(match.group(0)) if match else default


def parse_conversion(value: Any) -> Conversion:
    text = cell_text(value)
    normalized = (
        text.replace("锛?, ", ":")
        .replace("锛?, ", ";")
        .replace("鈥?, ", "-")
        .replace("鈥?, ", "-")
    )
    enum_entries: list[tuple[int, int, str]] = []
    enum_pattern = re.compile(
        r"((?:0[xX])?[0-9A-Fa-f]+)\s*(?:-\s*((?:0[xX])?[0-9A-Fa-f]+))?\s*:\s*([^;\n\r]+)"
    )
    for match in enum_pattern.finditer(normalized):
        lo = parse_enum_value(match.group(1))
        hi = parse_enum_value(match.group(2)) if match.group(2) else lo
        label = match.group(3).strip()
        if label:
            enum_entries.append((lo, hi, label))
    if enum_entries:
        return Conversion(kind="enum", enum=enum_entries)

    if re.search(r"y\s*=\s*a\s*x\s*\+\s*b", normalized, flags=re.IGNORECASE):
        a = _parse_float_assignment(normalized, "a", 1.0)
        b = _parse_float_assignment(normalized, "b", 0.0)
        precision_match = re.search(r"precision\s*=\s*(-?\d+)", normalized, flags=re.IGNORECASE)
        precision = int(precision_match.group(1)) if precision_match else None
        return Conversion(kind="linear", a=a, b=b, precision=precision)

    return Conversion()


def _parse_float_assignment(text: str, key: str, default: float) -> float:
    match = re.search(rf"\b{re.escape(key)}\s*=\s*(-?\d+(?:\.\d+)?)", text, flags=re.IGNORECASE)
    return float(match.group(1)) if match else default


def parse_enum_value(value: str) -> int:
    text = value.strip()
    if text.lower().startswith("0x") or re.search(r"[A-Fa-f]", text):
        return int(text[2:] if text.lower().startswith("0x") else text, 16)
    return int(text, 10)


def sanitize_short_name(value: str, fallback: str, used: set[str] | None = None, max_len: int = 120) -> str:
    base = value.strip() if value else fallback
    base = base.encode("ascii", "ignore").decode("ascii")
    base = re.sub(r"[^A-Za-z0-9_]+", "_", base).strip("_")
    if not base:
        base = fallback
    if not re.match(r"[A-Za-z_]", base):
        base = f"X_{base}"
    base = base[:max_len].rstrip("_") or fallback
    if used is None:
        return base
    candidate = base
    index = 2
    while candidate in used:
        suffix = f"_{index}"
        candidate = f"{base[: max_len - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def hex_short(prefix: str, value: int, width: int = 4) -> str:
    return f"{prefix}_{value:0{width}X}"


def find_sheet(workbook: Any, predicate: Any) -> Any | None:
    for name in workbook.sheetnames:
        if predicate(name):
            return workbook[name]
    return None


class IdGenerator:
    def __init__(self, root: etree._Element) -> None:
        self.used = {node.get("ID") for node in root.xpath("//*[@ID]") if node.get("ID")}
        self.index = 1

    def new(self, prefix: str = "SERES") -> str:
        while True:
            candidate = f"_{prefix}_{self.index:06d}"
            self.index += 1
            if candidate not in self.used:
                self.used.add(candidate)
                return candidate


def element(tag: str, text: Any | None = None, attrib: dict[str, str] | None = None) -> etree._Element:
    node = etree.Element(tag, attrib=attrib or {})
    if text is not None:
        node.text = str(text)
    return node


def sub(parent: etree._Element, tag: str, text: Any | None = None, attrib: dict[str, str] | None = None) -> etree._Element:
    child = element(tag, text, attrib)
    parent.append(child)
    return child


def set_xsi_type(node: etree._Element, value: str) -> None:
    node.set(XSI_TYPE, value)


def first_by_short(root: etree._Element, tag: str, short_name: str) -> etree._Element | None:
    nodes = root.xpath(f'//*[local-name()="{tag}" and SHORT-NAME=$name]', name=short_name)
    return nodes[0] if nodes else None


def clear_children(parent: etree._Element, tag: str) -> None:
    for child in list(parent):
        if child.tag == tag:
            parent.remove(child)


def replace_child(parent: etree._Element, old_tag: str, new_child: etree._Element, before_tags: set[str] | None = None) -> None:
    old = parent.find(old_tag)
    if old is not None:
        index = parent.index(old)
        parent.remove(old)
        parent.insert(index, new_child)
        return
    if before_tags:
        for index, child in enumerate(parent):
            if child.tag in before_tags:
                parent.insert(index, new_child)
                return
    parent.append(new_child)


def xml_local_name(node: etree._Element) -> str:
    return etree.QName(node).localname if isinstance(node.tag, str) else ""


def child_text_by_local_name(parent: etree._Element, name: str) -> str | None:
    matches = parent.xpath("./*[local-name()=$name]", name=name)
    return matches[0].text if matches else None


def validate_can_dela_odx_structure(root: etree._Element, source_name: str) -> None:
    errors: list[str] = []
    table_order = {
        "SHORT-NAME": 0,
        "LONG-NAME": 1,
        "DESC": 2,
        "KEY-LABEL": 3,
        "STRUCT-LABEL": 4,
        "ADMIN-DATA": 5,
        "KEY-DOP-REF": 6,
        "TABLE-ROW-REF": 7,
        "TABLE-ROW": 7,
        "TABLE-DIAG-COMM-CONNECTORS": 8,
        "SDGS": 9,
    }

    seen_ids: dict[str, int] = {}
    duplicate_ids: list[tuple[str, int, int]] = []
    for node in root.xpath("//*[@ID]"):
        node_id = node.get("ID")
        if not node_id:
            continue
        if node_id in seen_ids:
            duplicate_ids.append((node_id, seen_ids[node_id], node.sourceline or 0))
        else:
            seen_ids[node_id] = node.sourceline or 0
    for node_id, first_line, second_line in duplicate_ids[:20]:
        errors.append(f"duplicate ID '{node_id}' at line {second_line}, first seen at line {first_line}")

    for ref_node in root.xpath("//*[@ID-REF]"):
        id_ref = ref_node.get("ID-REF")
        if not id_ref or ref_node.get("DOCREF"):
            continue
        if id_ref not in seen_ids:
            errors.append(
                f"dangling local ID-REF '{id_ref}' at line {ref_node.sourceline or '?'} "
                f"({xml_local_name(ref_node)})"
            )

    for table in root.xpath("//*[local-name()='TABLE']"):
        short_name = child_text_by_local_name(table, "SHORT-NAME") or table.get("ID") or "<unnamed>"
        children = [xml_local_name(child) for child in table if isinstance(child.tag, str)]
        if not any(child_name in {"TABLE-ROW", "TABLE-ROW-REF"} for child_name in children):
            errors.append(f"TABLE '{short_name}' at line {table.sourceline or '?'} has no TABLE-ROW/TABLE-ROW-REF")
        if not children or children[0] != "SHORT-NAME":
            errors.append(f"TABLE '{short_name}' at line {table.sourceline or '?'} does not start with SHORT-NAME")
            continue
        last_index = -1
        for child_name in children:
            order_index = table_order.get(child_name)
            if order_index is None:
                errors.append(f"TABLE '{short_name}' at line {table.sourceline or '?'} has unexpected child '{child_name}'")
                continue
            if order_index < last_index:
                errors.append(f"TABLE '{short_name}' at line {table.sourceline or '?'} has '{child_name}' out of ODX order")
                break
            last_index = order_index

    if errors:
        details = "\n".join(f"- {message}" for message in errors[:60])
        if len(errors) > 60:
            details += f"\n- ... {len(errors) - 60} more issue(s)"
        raise RuntimeError(f"Generated {source_name} is not CANdelaStudio-compatible:\n{details}")


def patch_pdx_catalog_for_can_only(index_path: Path, keep_files: set[str]) -> None:
    if not index_path.exists():
        return
    parser = etree.XMLParser(remove_blank_text=False)
    tree = etree.parse(str(index_path), parser)
    root = tree.getroot()
    for file_node in list(root.xpath("//*[local-name()='FILE']")):
        file_name = compact_text(file_node.text)
        if file_name not in keep_files:
            parent = file_node.getparent()
            if parent is not None:
                parent.remove(file_node)
    tree.write(str(index_path), encoding="UTF-8", xml_declaration=False, pretty_print=True, standalone=False)


def prefix_doc_revision_labels(root: etree._Element) -> None:
    for label in root.xpath("//*[local-name()='DOC-REVISION']/*[local-name()='REVISION-LABEL']"):
        value = compact_text(label.text)
        if value and not value.startswith("PDX_"):
            label.text = f"PDX_{value}"


def ensure_units(ddds: etree._Element, survey: SurveyData) -> dict[str, str]:
    unit_spec = ddds.find("UNIT-SPEC")
    if unit_spec is None:
        unit_spec = sub(ddds, "UNIT-SPEC")
    units_node = unit_spec.find("UNITS")
    if units_node is None:
        units_node = sub(unit_spec, "UNITS")

    existing = {
        compact_text(unit.findtext("DISPLAY-NAME") or unit.findtext("SHORT-NAME")): unit.get("ID")
        for unit in units_node.findall("UNIT")
        if unit.get("ID")
    }
    all_params: list[ParamDef] = []
    for item in [*survey.dids, *survey.io_dids, *survey.snapshots, *survey.extended_records]:
        all_params.extend(item.params)
    for routine in survey.routines:
        for subfn in routine.subfunctions.values():
            all_params.extend(subfn.option_params)
            all_params.extend(subfn.status_params)

    for param in all_params:
        unit = compact_text(param.unit)
        if not unit or unit in {"-", "/", "N/A", "NA"} or unit in existing:
            continue
        unit_id = f"_UNIT_{sanitize_short_name(unit, 'UNIT')}"
        index = 2
        while unit_id in existing.values():
            unit_id = f"_UNIT_{sanitize_short_name(unit, 'UNIT')}_{index}"
            index += 1
        unit_node = sub(units_node, "UNIT", attrib={"ID": unit_id})
        sub(unit_node, "SHORT-NAME", sanitize_short_name(unit, "UNIT"))
        sub(unit_node, "DISPLAY-NAME", unit)
        existing[unit] = unit_id
    return {key: value for key, value in existing.items() if value}


def prepare_data_structure(
    *,
    id_gen: IdGenerator,
    data_object_props: etree._Element,
    structures: etree._Element,
    item: DidDef | IoDidDef,
    prefix: str,
    unit_ids: dict[str, str],
    generated_dop_cache: dict[tuple[str, int, str, str, str], str],
) -> None:
    used_param_names: set[str] = set()
    english, long_name = split_name(item.desc)
    item.short_name = sanitize_short_name(english or hex_short(prefix, item.did), hex_short(prefix, item.did))
    item.long_name = long_name or item.short_name
    for param in item.params:
        param.name = sanitize_short_name(param.name, "Data", used_param_names)
        param.dop_id = ensure_param_dop(
            id_gen,
            data_object_props,
            param,
            unit_ids,
            generated_dop_cache,
            f"DOP_{item.short_name}_{param.name}",
        )
    if not item.params:
        raw = ParamDef(
            name="Data",
            long_name="Data",
            byte_pos=0,
            bit_pos=0,
            bit_len=max(8, (item.size or 1) * 8),
            data_type="Hex(Unsigned)",
        )
        raw.name = sanitize_short_name(raw.name, "Data", used_param_names)
        raw.dop_id = ensure_param_dop(
            id_gen, data_object_props, raw, unit_ids, generated_dop_cache, f"DOP_{item.short_name}_Data"
        )
        item.params.append(raw)

    item.structure_id = id_gen.new("DIDSTR")
    structure = sub(structures, "STRUCTURE", attrib={"ID": item.structure_id})
    sub(structure, "SHORT-NAME", item.short_name)
    sub(structure, "LONG-NAME", item.long_name)
    if item.size:
        sub(structure, "BYTE-SIZE", item.size)
    params_node = sub(structure, "PARAMS")
    for param in item.params:
        params_node.append(make_value_param(param.name, param.long_name, param.byte_pos, param.bit_pos, param.dop_id))


def make_param_structure(
    id_gen: IdGenerator,
    data_object_props: etree._Element,
    structures: etree._Element,
    short_name: str,
    long_name: str,
    params: list[ParamDef],
    unit_ids: dict[str, str],
    generated_dop_cache: dict[tuple[str, int, str, str, str], str],
    byte_size: int | None = None,
) -> str:
    used_names: set[str] = set()
    structure_id = id_gen.new("STR")
    structure = sub(structures, "STRUCTURE", attrib={"ID": structure_id})
    short_name = sanitize_short_name(short_name, "Structure")
    sub(structure, "SHORT-NAME", short_name)
    sub(structure, "LONG-NAME", long_name or short_name)
    if byte_size:
        sub(structure, "BYTE-SIZE", byte_size)
    params_node = sub(structure, "PARAMS")
    for param in params:
        param.name = sanitize_short_name(param.name, "Data", used_names)
        param.dop_id = ensure_param_dop(
            id_gen, data_object_props, param, unit_ids, generated_dop_cache, f"DOP_{short_name}_{param.name}"
        )
        params_node.append(make_value_param(param.name, param.long_name, param.byte_pos, param.bit_pos, param.dop_id))
    return structure_id


def ensure_param_dop(
    id_gen: IdGenerator,
    data_object_props: etree._Element,
    param: ParamDef,
    unit_ids: dict[str, str],
    cache: dict[tuple[str, int, str, str, str], str],
    preferred_short_name: str,
) -> str:
    conv_key = conversion_cache_key(param.conversion)
    key = (param.data_type.upper(), param.bit_len, conv_key, compact_text(param.unit), param.name)
    if key in cache:
        return cache[key]
    dop_id = id_gen.new("DOP")
    dop = make_data_object_prop(dop_id, preferred_short_name, param, unit_ids)
    data_object_props.append(dop)
    cache[key] = dop_id
    return dop_id


def conversion_cache_key(conversion: Conversion) -> str:
    if conversion.kind == "enum":
        return "enum:" + "|".join(f"{lo}-{hi}:{label}" for lo, hi, label in conversion.enum)
    if conversion.kind == "linear":
        return f"linear:{conversion.a}:{conversion.b}:{conversion.precision}"
    return "identity"


def clean_float(value: float) -> str:
    return str(int(value)) if float(value).is_integer() else str(value)


def make_data_object_prop(dop_id: str, preferred_short_name: str, param: ParamDef, unit_ids: dict[str, str]) -> etree._Element:
    bit_len = max(1, int(param.bit_len or 8))
    dop = element("DATA-OBJECT-PROP", attrib={"ID": dop_id})
    short_name = sanitize_short_name(preferred_short_name, "DOP")
    sub(dop, "SHORT-NAME", short_name)
    sub(dop, "LONG-NAME", param.long_name or short_name)

    conversion = param.conversion
    if conversion.kind == "enum" and conversion.enum:
        compu = sub(dop, "COMPU-METHOD")
        sub(compu, "CATEGORY", "TEXTTABLE")
        internal_to_phys = sub(compu, "COMPU-INTERNAL-TO-PHYS")
        scales = sub(internal_to_phys, "COMPU-SCALES")
        for lo, hi, label in conversion.enum:
            scale = sub(scales, "COMPU-SCALE")
            sub(scale, "LOWER-LIMIT", lo)
            sub(scale, "UPPER-LIMIT", hi)
            const = sub(scale, "COMPU-CONST")
            sub(const, "VT", label)
        coded = sub(dop, "DIAG-CODED-TYPE", attrib={"BASE-TYPE-ENCODING": "NONE", "BASE-DATA-TYPE": "A_UINT32"})
        set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
        sub(coded, "BIT-LENGTH", bit_len)
        sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UNICODE2STRING"})
    elif conversion.kind == "linear":
        compu = sub(dop, "COMPU-METHOD")
        sub(compu, "CATEGORY", "LINEAR")
        internal_to_phys = sub(compu, "COMPU-INTERNAL-TO-PHYS")
        scales = sub(internal_to_phys, "COMPU-SCALES")
        scale = sub(scales, "COMPU-SCALE")
        coeffs = sub(scale, "COMPU-RATIONAL-COEFFS")
        numerator = sub(coeffs, "COMPU-NUMERATOR")
        sub(numerator, "V", clean_float(conversion.b))
        sub(numerator, "V", clean_float(conversion.a))
        coded = sub(dop, "DIAG-CODED-TYPE", attrib={"BASE-TYPE-ENCODING": "NONE", "BASE-DATA-TYPE": "A_UINT32"})
        set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
        sub(coded, "BIT-LENGTH", min(bit_len, 32))
        physical = sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_FLOAT64"})
        if conversion.precision is not None:
            sub(physical, "PRECISION", conversion.precision)
    else:
        compu = sub(dop, "COMPU-METHOD")
        sub(compu, "CATEGORY", "IDENTICAL")
        data_type = param.data_type.upper()
        if "BCD" in data_type:
            coded = sub(
                dop,
                "DIAG-CODED-TYPE",
                attrib={"BASE-TYPE-ENCODING": "BCD-P", "BASE-DATA-TYPE": "A_UINT32"},
            )
            set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
            sub(coded, "BIT-LENGTH", bit_len)
            sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32", "DISPLAY-RADIX": "DEC"})
        elif "ASCII" in data_type:
            coded = sub(
                dop,
                "DIAG-CODED-TYPE",
                attrib={"BASE-TYPE-ENCODING": "ISO-8859-1", "BASE-DATA-TYPE": "A_ASCIISTRING"},
            )
            set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
            sub(coded, "BIT-LENGTH", bit_len)
            sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UNICODE2STRING"})
        elif bit_len <= 32:
            coded = sub(dop, "DIAG-CODED-TYPE", attrib={"BASE-TYPE-ENCODING": "NONE", "BASE-DATA-TYPE": "A_UINT32"})
            set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
            sub(coded, "BIT-LENGTH", bit_len)
            sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32", "DISPLAY-RADIX": "HEX"})
        else:
            coded = sub(dop, "DIAG-CODED-TYPE", attrib={"BASE-DATA-TYPE": "A_BYTEFIELD"})
            set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
            sub(coded, "BIT-LENGTH", bit_len)
            sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_BYTEFIELD"})

    unit = compact_text(param.unit)
    if unit and unit in unit_ids and conversion.kind != "enum":
        sub(dop, "UNIT-REF", attrib={"ID-REF": unit_ids[unit]})
    return dop


def make_value_param(short_name: str, long_name: str, byte_position: int, bit_position: int, dop_id: str) -> etree._Element:
    param = element("PARAM", attrib={"SEMANTIC": "DATA"})
    set_xsi_type(param, "VALUE")
    sub(param, "SHORT-NAME", short_name)
    sub(param, "LONG-NAME", long_name or short_name)
    sub(param, "BYTE-POSITION", byte_position)
    if bit_position:
        sub(param, "BIT-POSITION", bit_position)
    sub(param, "DOP-REF", attrib={"ID-REF": dop_id})
    return param


def append_table_row(
    table: etree._Element,
    id_gen: IdGenerator,
    short_name: str,
    long_name: str,
    key: int | str,
    structure_ref: str | None = None,
) -> etree._Element:
    row = sub(table, "TABLE-ROW", attrib={"ID": id_gen.new("ROW")})
    sub(row, "SHORT-NAME", sanitize_short_name(short_name, "Row"))
    sub(row, "LONG-NAME", long_name or short_name)
    sub(row, "KEY", key)
    if structure_ref:
        sub(row, "STRUCTURE-REF", attrib={"ID-REF": structure_ref})
    return row


def remove_service_and_messages(root: etree._Element, service_short_name: str) -> None:
    service = first_by_short(root, "DIAG-SERVICE", service_short_name)
    if service is None:
        return
    ids_to_remove: set[str] = set()
    request_ref = service.find("REQUEST-REF")
    if request_ref is not None and request_ref.get("ID-REF"):
        ids_to_remove.add(request_ref.get("ID-REF"))
    ids_to_remove.update(
        ref.get("ID-REF")
        for ref in service.xpath('./*[local-name()="POS-RESPONSE-REFS"]/*[local-name()="POS-RESPONSE-REF"]')
        if ref.get("ID-REF")
    )
    ids_to_remove.update(
        ref.get("ID-REF")
        for ref in service.xpath('./*[local-name()="NEG-RESPONSE-REFS"]/*[local-name()="NEG-RESPONSE-REF"]')
        if ref.get("ID-REF")
    )

    parent = service.getparent()
    if parent is not None:
        parent.remove(service)
    for element_id in ids_to_remove:
        for node in root.xpath(f'//*[@ID="{element_id}"]'):
            node_parent = node.getparent()
            if node_parent is not None:
                node_parent.remove(node)


def set_short_long(node: etree._Element, short_name: str, long_name: str) -> None:
    short = node.find("SHORT-NAME")
    if short is not None:
        short.text = sanitize_short_name(short_name, "Node")
    long = node.find("LONG-NAME")
    if long is not None:
        long.text = long_name


def refresh_internal_ids(node: etree._Element, id_gen: IdGenerator) -> None:
    id_map: dict[str, str] = {}
    for child in node.xpath(".//*[@ID]"):
        old_id = child.get("ID")
        if not old_id:
            continue
        new_id = id_gen.new("IN")
        id_map[old_id] = new_id
        child.set("ID", new_id)
    if not id_map:
        return
    for ref_node in node.xpath(".//*[@ID-REF]"):
        old_ref = ref_node.get("ID-REF")
        if old_ref in id_map and not ref_node.get("DOCREF"):
            ref_node.set("ID-REF", id_map[old_ref])


def dtc_table_key(dtc: DtcDef) -> str:
    text = compact_text(dtc.text or dtc.display_code)
    suffix = f"_{text[:80]}" if text else ""
    return f"0x{dtc.byte_code:06X}{suffix}"


def update_dtc_text_table(root: etree._Element, dtcs: list[DtcDef]) -> None:
    dop = first_by_short(root, "DATA-OBJECT-PROP", "TextTable_DTC")
    if dop is None:
        return
    scales = dop.find(".//COMPU-SCALES")
    if scales is None:
        return
    for child in list(scales):
        scales.remove(child)
    for dtc in dtcs:
        scale = sub(scales, "COMPU-SCALE")
        sub(scale, "LOWER-LIMIT", dtc.byte_code)
        sub(scale, "UPPER-LIMIT", dtc.byte_code)
        const = sub(scale, "COMPU-CONST")
        sub(const, "VT", dtc_table_key(dtc))
    internal_constr = dop.find("INTERNAL-CONSTR")
    if internal_constr is not None:
        dop.remove(internal_constr)


def add_dtc_sdg(
    sdgs: etree._Element,
    id_gen: IdGenerator,
    caption_ids: dict[str, str],
    short_name: str,
    value: str,
    *,
    first: bool,
) -> None:
    sdg = sub(sdgs, "SDG")
    if first or short_name not in caption_ids:
        caption_id = id_gen.new("CAP")
        caption_ids[short_name] = caption_id
        caption = sub(sdg, "SDG-CAPTION", attrib={"ID": caption_id})
        sub(caption, "SHORT-NAME", short_name)
    else:
        sub(sdg, "SDG-CAPTION-REF", attrib={"ID-REF": caption_ids[short_name]})
    sub(sdg, "SD", value)


BASE = SimpleNamespace(
    XSI_NS=XSI_NS,
    XSI_TYPE=XSI_TYPE,
    Conversion=Conversion,
    ParamDef=ParamDef,
    DidDef=DidDef,
    IoDidDef=IoDidDef,
    RoutineDef=RoutineDef,
    RoutineSubFunction=RoutineSubFunction,
    DtcDef=DtcDef,
    SnapshotDef=SnapshotDef,
    ExtendedRecordDef=ExtendedRecordDef,
    CoverInfo=CoverInfo,
    SurveyData=SurveyData,
    cell_text=cell_text,
    compact_text=compact_text,
    normalize_access=normalize_access,
    split_name=split_name,
    parse_hex_cell=parse_hex_cell,
    parse_int_cell=parse_int_cell,
    parse_conversion=parse_conversion,
    sanitize_short_name=sanitize_short_name,
    hex_short=hex_short,
    find_sheet=find_sheet,
    IdGenerator=IdGenerator,
    element=element,
    sub=sub,
    set_xsi_type=set_xsi_type,
    first_by_short=first_by_short,
    clear_children=clear_children,
    replace_child=replace_child,
    validate_can_dela_odx_structure=validate_can_dela_odx_structure,
    patch_pdx_catalog_for_can_only=patch_pdx_catalog_for_can_only,
    prefix_doc_revision_labels=prefix_doc_revision_labels,
    ensure_units=ensure_units,
    prepare_data_structure=prepare_data_structure,
    make_param_structure=make_param_structure,
    make_value_param=make_value_param,
    append_table_row=append_table_row,
    remove_service_and_messages=remove_service_and_messages,
    set_short_long=set_short_long,
    refresh_internal_ids=refresh_internal_ids,
    dtc_table_key=dtc_table_key,
    update_dtc_text_table=update_dtc_text_table,
    add_dtc_sdg=add_dtc_sdg,
)

Conversion = BASE.Conversion
ParamDef = BASE.ParamDef
DidDef = BASE.DidDef
IoDidDef = BASE.IoDidDef
RoutineDef = BASE.RoutineDef
RoutineSubFunction = BASE.RoutineSubFunction
DtcDef = BASE.DtcDef
SnapshotDef = BASE.SnapshotDef
ExtendedRecordDef = BASE.ExtendedRecordDef
CoverInfo = BASE.CoverInfo
SurveyData = BASE.SurveyData


SERES_KEEP_FILES = {
    "ISO_11898_2_DWCAN.odx-cs",
    "ISO_11898_3_DWFTCAN.odx-cs",
    "ISO_15765_2.odx-cs",
    "ISO_15765_3.odx-cs",
    "ISO_15765_3_on_ISO_15765_2.odx-c",
    "SAE_J2411_SWCAN.odx-cs",
    "SERES_ECU_CAN_v15.odx-d",
    "index.xml",
}

SCRIPT_DIR = Path(__file__).resolve().parent
CANDELA_SHORT_NAME_MAX_LEN = 64
UUDT_DISABLED_CAN_ID = 0xFFFFFFFF


def cell_text(value: Any) -> str:
    return BASE.cell_text(value)


def compact_text(value: Any) -> str:
    return BASE.compact_text(value)


def normalize_access(value: Any) -> str:
    return BASE.normalize_access(value)


def parse_int_cell(value: Any, default: int = 0) -> int:
    return BASE.parse_int_cell(value, default)


def parse_hex_cell(value: Any, *, max_value: int = 0xFFFFFFFF) -> int | None:
    return BASE.parse_hex_cell(value, max_value=max_value)


def parse_hex_in_text(value: Any, *, max_value: int = 0xFFFFFFFF) -> int | None:
    text = compact_text(value)
    match = re.search(r"0[xX]([0-9A-Fa-f]{1,8})|(?<![A-Za-z0-9])([0-9A-Fa-f]{2,8})(?![A-Za-z0-9])", text)
    if not match:
        return None
    raw = match.group(1) or match.group(2)
    result = int(raw, 16)
    return result if result <= max_value else None


def find_sheet(workbook: Any, predicate: Any) -> Any | None:
    return BASE.find_sheet(workbook, predicate)


def cell_has_strike(cell: Any) -> bool:
    return bool(getattr(getattr(cell, "font", None), "strike", False))


def row_is_deleted(sheet: Any, row: int, columns: Iterable[int] | None = None, *, threshold: float = 0.75) -> bool:
    cells = [sheet.cell(row, col) for col in columns] if columns is not None else list(sheet[row])
    populated = [cell for cell in cells if compact_text(cell.value)]
    if len(populated) < 2:
        return False
    struck = sum(1 for cell in populated if cell_has_strike(cell))
    return struck == len(populated) or struck / len(populated) >= threshold


def sanitize_short_name(value: str, fallback: str, used: set[str] | None = None, max_len: int = 120) -> str:
    return BASE.sanitize_short_name(value, fallback, used, max_len)


def hex_short(prefix: str, value: int, width: int = 4) -> str:
    return BASE.hex_short(prefix, value, width)


def usable_text(value: Any) -> str:
    text = compact_text(value)
    if text.upper() in {"", "/", "N/A", "NA", "NONE", "NULL"}:
        return ""
    return text


def dual_name(english_value: Any, chinese_value: Any = "", fallback: str = "") -> tuple[str, str]:
    candidates = name_parts(english_value) + name_parts(chinese_value)
    candidates = [candidate for candidate in candidates if usable_text(candidate)]
    if not candidates:
        return fallback, fallback
    english = next((candidate for candidate in candidates if re.search(r"[A-Za-z]", candidate) and not has_chinese_text(candidate)), "")
    chinese = next((candidate for candidate in candidates if has_chinese_text(candidate)), "")
    identifier = english or (fallback if chinese else candidates[0])
    display = chinese or candidates[0]
    return identifier or fallback, display or identifier or fallback


def normalize_conversion_text(value: Any) -> str:
    text = cell_text(value)
    replacements = {
        "\u3000": " ",
        "；": ";",
        "：": ":",
        "，": ",",
        "～": "~",
        "－": "-",
        "—": "-",
        "\r": "\n",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.strip()


def decimal_places_from_literal(value: str | None) -> int | None:
    if not value or "." not in value:
        return None
    return len(value.split(".", 1)[1].rstrip())


def parse_conversion(value: Any) -> Conversion:
    text = normalize_conversion_text(value)
    if not text or text.upper() in {"/", "N/A", "NA"}:
        return Conversion()

    enum_entries: list[tuple[int, int, str]] = []
    enum_pattern = re.compile(
        r"((?:0[xX])?[0-9A-Fa-f]+)\s*(?:[-~]\s*((?:0[xX])?[0-9A-Fa-f]+))?\s*[:=]\s*([^;\n\r]+)"
    )
    for match in enum_pattern.finditer(text):
        label = match.group(3).strip(" ;,")
        if not label:
            continue
        lo = parse_enum_value(match.group(1))
        hi = parse_enum_value(match.group(2)) if match.group(2) else lo
        enum_entries.append((lo, hi, label))
    if enum_entries and re.search(r"0[xX]|[:=]", text):
        return Conversion(kind="enum", enum=enum_entries)

    linear = re.search(
        r"(?:phy|y)\s*=\s*(?:XX|xx|X|x)"
        r"(?:\s*\*\s*(-?\d+(?:\.\d+)?))?"
        r"(?:\s*([+-])\s*(-?\d+(?:\.\d+)?))?",
        text,
    )
    if linear:
        coefficient_text = linear.group(1)
        offset_text = linear.group(3)
        a = float(coefficient_text) if coefficient_text else 1.0
        b = float(offset_text) if offset_text else 0.0
        if linear.group(2) == "-":
            b = -b
        precision_candidates = [
            value
            for value in (
                decimal_places_from_literal(coefficient_text),
                decimal_places_from_literal(offset_text),
            )
            if value is not None
        ]
        precision = max(precision_candidates) if precision_candidates else None
        return Conversion(kind="linear", a=a, b=b, precision=precision)

    yaxb = re.search(r"y\s*=\s*a\s*\*\s*x\s*\+\s*b", text, flags=re.IGNORECASE)
    if yaxb:
        return BASE.parse_conversion(text)
    return Conversion()


def choose_conversion(primary_value: Any, fallback_value: Any = "") -> Conversion:
    """Parse both conversion columns and keep the richer physical meaning.

    SERES sheets normally put English expressions in column L and Chinese text
    in column M. In a few rows one column contains a pass-through formula while
    the other carries the value table, so enum > linear > identity gives CANdela
    the more useful data type without hard-coding row numbers.
    """

    primary = parse_conversion(primary_value)
    fallback = parse_conversion(fallback_value)
    rank = {"enum": 3, "linear": 2, "identity": 1}
    if rank.get(fallback.kind, 0) > rank.get(primary.kind, 0):
        return fallback
    return primary


def parse_index_range(value: Any, *, size: int = 0, default: int = 0) -> tuple[int, int]:
    text = compact_text(value).upper()
    if not text or text in {"/", "N/A", "NA"}:
        return default, default
    if text == "ALL":
        return 0, max(0, size - 1)
    nums = [int(num) for num in re.findall(r"\d+", text)]
    if not nums:
        return default, default
    start = nums[0]
    end = nums[1] if len(nums) > 1 else start
    if end < start:
        end = start
    return start, end


def parse_bit_range(value: Any, byte_start: int, byte_end: int, *, size: int = 0) -> tuple[int, int]:
    text = compact_text(value).upper()
    byte_count = max(1, byte_end - byte_start + 1)
    if not text or text in {"/", "N/A", "NA", "ALL"}:
        return 0, byte_count * 8
    nums = [int(num) for num in re.findall(r"\d+", text)]
    if not nums:
        return 0, byte_count * 8
    if len(nums) > 1 and ("-" in text or "~" in text):
        start, end = nums[0], nums[1]
        if end < start:
            end = start
        return start % 8, min(max(1, end - start + 1), byte_count * 8)
    return nums[0] % 8, 1


def make_param_from_cells(
    *,
    name_value: Any,
    chinese_name_value: Any = "",
    byte_value: Any,
    bit_value: Any,
    data_type_value: Any,
    unit_value: Any = "",
    conversion_value: Any = "",
    conversion_fallback_value: Any = "",
    min_value: Any = "",
    max_value: Any = "",
    fallback_name: str = "Data",
    size: int = 0,
) -> ParamDef | None:
    english, long_name = dual_name(name_value, chinese_name_value, fallback_name)
    if not usable_text(english) and not usable_text(long_name):
        english = fallback_name
        long_name = fallback_name
    byte_start, byte_end = parse_index_range(byte_value, size=size, default=0)
    bit_pos, bit_len = parse_bit_range(bit_value, byte_start, byte_end, size=size)
    data_type = usable_text(data_type_value) or "Hex(Unsigned)"
    conversion = choose_conversion(
        usable_text(conversion_value),
        usable_text(conversion_fallback_value),
    )
    if bit_len > 32 and conversion.kind == "enum":
        conversion = Conversion()
    return ParamDef(
        name=english or fallback_name,
        long_name=long_name or english or fallback_name,
        byte_pos=byte_start,
        bit_pos=bit_pos,
        bit_len=bit_len,
        data_type=data_type,
        unit=usable_text(unit_value),
        conversion=conversion,
        min_value=usable_text(min_value),
        max_value=usable_text(max_value),
    )


def did_param_size(params: Iterable[ParamDef], default: int = 1) -> int:
    max_end = 0
    for param in params:
        max_end = max(max_end, param.byte_pos + max(1, (param.bit_pos + max(1, param.bit_len) + 7) // 8))
    return max(default, max_end)


def normalized_data_type(value: Any) -> str:
    return compact_text(value).upper()


def is_ascii_data_type(value: Any) -> bool:
    return "ASCII" in normalized_data_type(value)


def merge_split_byte_did_params(did: DidDef) -> None:
    if len(did.params) <= 1:
        return

    ordered = sorted(did.params, key=lambda param: (param.byte_pos, param.bit_pos, param.name))
    data_type = normalized_data_type(ordered[0].data_type)
    unit = compact_text(ordered[0].unit)
    if not data_type or not is_ascii_data_type(data_type):
        return

    for expected_byte, param in enumerate(ordered):
        if (
            normalized_data_type(param.data_type) != data_type
            or compact_text(param.unit) != unit
            or param.byte_pos != expected_byte
            or param.bit_pos != 0
            or param.bit_len != 8
            or param.conversion.kind != "identity"
        ):
            return

    identifier, display = split_name(did.desc)
    name = display or identifier or hex_short("DID", did.did)
    did.params = [
        ParamDef(
            name=identifier or name,
            long_name=name,
            byte_pos=0,
            bit_pos=0,
            bit_len=len(ordered) * 8,
            data_type=ordered[0].data_type,
            unit=unit,
        )
    ]


def direct_did_payload_param(did: DidDef) -> ParamDef | None:
    if len(did.params) != 1:
        return None
    param = did.params[0]
    if param.byte_pos != 0 or param.bit_pos != 0 or not param.dop_id:
        return None
    expected_size = did.size if did.size > 0 else did_param_size(did.params)
    expected_bit_len = max(8, expected_size * 8)
    return param if param.bit_len == expected_bit_len else None


def remove_structure_by_id(structures: etree._Element, structure_id: str) -> None:
    if not structure_id:
        return
    for structure in list(structures):
        if structure.get("ID") == structure_id:
            structures.remove(structure)
            return


def did_payload_dop_id(did: DidDef) -> str:
    direct_param = direct_did_payload_param(did)
    if direct_param is not None:
        return direct_param.dop_id
    if not did.structure_id:
        raise RuntimeError(f"DID 0x{did.did:04X} has no payload DOP/STRUCTURE")
    return did.structure_id


def merge_session_access(read_values: Iterable[Any], write_values: Iterable[Any]) -> list[str]:
    sessions: list[str] = []
    for read_value, write_value in zip(read_values, write_values, strict=False):
        flags = ""
        if usable_text(read_value).upper() != "N":
            flags += "R"
        if usable_text(write_value).upper() != "N":
            flags += "W"
        sessions.append(flags)
    return sessions


def first_security_level(values: Iterable[Any]) -> str:
    for value in values:
        text = usable_text(value)
        if text and text.upper() != "N":
            return text
    return "N"


def is_supported_flag(value: Any, *, default: bool = False) -> bool:
    text = compact_text(value).upper()
    if text in {"Y", "YES", "TBD"}:
        return True
    if text in {"N", "NO", "-", "/", "N/A", "NA"}:
        return False
    return default


def merge_security_text(old_value: str, new_value: str) -> str:
    old_text = usable_text(old_value) or "N"
    new_text = usable_text(new_value) or "N"
    if old_text.upper() == "N":
        return new_text
    if new_text.upper() == "N" or new_text in old_text.split("&"):
        return old_text
    parts = [part for part in old_text.split("&") if part]
    for part in new_text.split("&"):
        if part and part not in parts:
            parts.append(part)
    return "&".join(parts) if parts else "N"


def parse_service_access_sheet(
    sheet: Any,
    *,
    start_row: int,
    session_columns: Iterable[tuple[int, str]],
    security_col: int,
    source: str,
) -> dict[tuple[int, int | None], dict[str, Any]]:
    access: dict[tuple[int, int | None], dict[str, Any]] = {}
    for row in range(start_row, sheet.max_row + 1):
        service_marker = compact_text(sheet.cell(row, 2).value)
        if service_marker.startswith("#"):
            break
        if row_is_deleted(sheet, row):
            continue
        service_id = parse_hex_in_text(service_marker, max_value=0xFF)
        if service_id is None:
            continue
        if not is_supported_flag(sheet.cell(row, 5).value):
            continue

        subfunction_text = compact_text(sheet.cell(row, 6).value)
        subfunction = parse_hex_in_text(subfunction_text, max_value=0x7F)
        if subfunction is not None and not is_supported_flag(sheet.cell(row, 8).value, default=True):
            continue

        sessions = {
            state_name
            for col, state_name in session_columns
            if is_supported_flag(sheet.cell(row, col).value)
        }
        if not sessions:
            continue

        key = (service_id, subfunction)
        item = access.setdefault(
            key,
            {
                "service_id": service_id,
                "subfunction": subfunction,
                "service_name": usable_text(sheet.cell(row, 3).value),
                "subfunction_name": usable_text(sheet.cell(row, 6).value),
                "sessions": set(),
                "security": "N",
                "sources": set(),
            },
        )
        item["sessions"].update(sessions)
        item["security"] = merge_security_text(item["security"], usable_text(sheet.cell(row, security_col).value) or "N")
        item["sources"].add(source)
    return access


def parse_core_service_access(workbook: Any) -> dict[tuple[int, int | None], dict[str, Any]]:
    """Parse SERES application/boot service matrices for flat template services."""

    result: dict[tuple[int, int | None], dict[str, Any]] = {}
    sheet_profiles = [
        (
            find_sheet(workbook, lambda name: name.strip().startswith("1_1") and "ApplicationServices" in name),
            7,
            ((10, "Default"), (11, "Extended")),
            13,
            "Application",
        ),
        (
            find_sheet(workbook, lambda name: name.strip().startswith("1_2") and "BootService" in name),
            10,
            ((10, "Default"), (11, "Programming"), (12, "Extended")),
            14,
            "Boot",
        ),
    ]
    for sheet, start_row, session_columns, security_col, source in sheet_profiles:
        if sheet is None:
            continue
        parsed = parse_service_access_sheet(
            sheet,
            start_row=start_row,
            session_columns=session_columns,
            security_col=security_col,
            source=source,
        )
        for key, item in parsed.items():
            target = result.setdefault(
                key,
                {
                    "service_id": item["service_id"],
                    "subfunction": item["subfunction"],
                    "service_name": item["service_name"],
                    "subfunction_name": item["subfunction_name"],
                    "sessions": set(),
                    "security": "N",
                    "sources": set(),
                },
            )
            target["sessions"].update(item["sessions"])
            target["security"] = merge_security_text(target["security"], item["security"])
            target["sources"].update(item["sources"])
    return result


def parse_seres_survey(xlsx_path: Path) -> SurveyData:
    workbook = load_workbook(xlsx_path, data_only=True)
    cover = parse_cover(workbook)
    dids = parse_read_write_dids(workbook)
    io_dids = parse_io_dids(workbook)
    routines = parse_routines(workbook)
    dtcs = parse_dtcs(workbook)
    snapshots, extended_records, snapshot_record_names = parse_snapshot_extended(workbook)
    survey = SurveyData(cover, dids, io_dids, routines, dtcs, snapshots, extended_records)
    survey.snapshot_record_names = snapshot_record_names
    survey.snapshot_record_nums = sorted(snapshot_record_names)
    survey.core_service_access = parse_core_service_access(workbook)
    return survey


def parse_cover(workbook: Any) -> CoverInfo:
    cover = CoverInfo(ecu_name="SERES_ECU_CAN")
    general = find_sheet(workbook, lambda name: name.strip() == "GeneralInfo")
    cover_sheet = find_sheet(workbook, lambda name: name.strip().startswith("0_1"))

    rows: dict[str, tuple[str, str]] = {}
    if general is not None:
        for row in range(1, general.max_row + 1):
            key = compact_text(general.cell(row, 1).value)
            if key:
                rows[key.lower()] = (compact_text(general.cell(row, 2).value), compact_text(general.cell(row, 3).value))

    general_ecu_name = rows.get("ecu-name", ("", ""))[0]
    cover.ecu_name = general_ecu_name or "SERES_ECU_CAN"
    cover.supplier = rows.get("supplier id", ("", ""))[0]
    cover.bus_type = "CAN"
    cover.rx_fun_id = parse_hex_in_text(rows.get("can functional request id", ("", ""))[0], max_value=0x1FFFFFFF)
    cover.rx_phy_id = parse_hex_in_text(rows.get("can physical request id", ("", ""))[0], max_value=0x1FFFFFFF)
    cover.tx_id = parse_hex_in_text(rows.get("can response id", ("", ""))[0], max_value=0x1FFFFFFF)

    if cover_sheet is not None:
        if not general_ecu_name:
            cover.ecu_name = compact_text(cover_sheet["B11"].value) or cover.ecu_name
        cover.tx_id = cover.tx_id if cover.tx_id is not None else parse_hex_in_text(cover_sheet["C16"].value)
        cover.rx_phy_id = cover.rx_phy_id if cover.rx_phy_id is not None else parse_hex_in_text(cover_sheet["C17"].value)
        cover.rx_fun_id = cover.rx_fun_id if cover.rx_fun_id is not None else parse_hex_in_text(cover_sheet["C18"].value)

    def ms_to_us(key: str, default: int) -> int:
        value, unit = rows.get(key.lower(), ("", "ms"))
        number = parse_int_cell(value, default // 1000)
        if unit.lower().startswith("us") or unit.lower().startswith("micro"):
            return number
        return number * 1000

    cover.comm_params = {
        "ISO_15765_2.CP_CanFuncReqId": cover.rx_fun_id if cover.rx_fun_id is not None else 0x7DF,
        "ISO_15765_3.CP_P2Max": ms_to_us("P2", 150_000),
        "ISO_15765_3.CP_P2Star": ms_to_us("P2*", 5_100_000),
        "ISO_15765_3.CP_TesterPresentTime": ms_to_us("S3", 2_000_000),
        "ISO_15765_3.CP_P3Phys": 50_000,
        "ISO_15765_3.CP_P3Func": 50_000,
        "ISO_15765_2.CP_StMin": parse_int_cell(rows.get("app stmin(ecu)", ("20", ""))[0], 20),
        "ISO_15765_2.CP_BlockSize": parse_int_cell(rows.get("bs(ecu)", ("0", ""))[0], 0),
        "ISO_15765_2.CP_As": ms_to_us("N_As", 70_000),
        "ISO_15765_2.CP_Ar": ms_to_us("N_Ar", 70_000),
        "ISO_15765_2.CP_Bs": ms_to_us("N_Bs", 150_000),
        "ISO_15765_2.CP_Br": 50_000,
        "ISO_15765_2.CP_Cs": 50_000,
        "ISO_15765_2.CP_Cr": ms_to_us("N_Cr", 150_000),
        "ISO_11898_2_DWCAN.CP_Baudrate": parse_int_cell(rows.get("can baudrate", ("500000", ""))[0], 500000),
    }
    cover.session_timing = {
        "P2": parse_int_cell(rows.get("p2(ecu)", ("50", ""))[0], 50),
        "P2Ex": parse_int_cell(rows.get("p2*(ecu)", ("5000", ""))[0], 5000),
    }
    return cover


def parse_read_write_dids(workbook: Any) -> list[DidDef]:
    sheet = find_sheet(workbook, lambda name: "DID of read&write" in name)
    if sheet is None:
        return []

    dids: dict[int, DidDef] = {}
    for row in range(12, sheet.max_row + 1):
        if compact_text(sheet.cell(row, 2).value).startswith("#"):
            break
        if row_is_deleted(sheet, row):
            continue
        did_value = parse_hex_cell(sheet.cell(row, 3).value, max_value=0xFFFF)
        if did_value is None:
            continue
        if "DEMO" in compact_text(sheet.cell(row, 2).value).upper() or compact_text(sheet.cell(row, 4).value).lower() == "demo":
            continue

        size = parse_int_cell(sheet.cell(row, 6).value, default=0)
        desc_en, desc_long = dual_name(sheet.cell(row, 4).value, sheet.cell(row, 5).value, hex_short("DID", did_value))
        read_values = [sheet.cell(row, col).value for col in (19, 20, 21, 22, 23)]
        write_values = [sheet.cell(row, col).value for col in (25, 26, 27, 28, 29)]
        did = dids.get(did_value)
        if did is None:
            did = DidDef(
                did=did_value,
                desc=canonical_name(desc_en, desc_long, hex_short("DID", did_value)),
                size=size,
                write_security=first_security_level(write_values),
                sessions=merge_session_access(read_values, write_values),
            )
            dids[did_value] = did
        else:
            did.size = did.size or size
            did.write_security = did.write_security if did.write_security != "N" else first_security_level(write_values)
            did.sessions = [
                "".join(sorted(set((old or "") + (new or ""))))
                for old, new in zip(did.sessions, merge_session_access(read_values, write_values), strict=False)
            ]

        param = make_param_from_cells(
            name_value=sheet.cell(row, 9).value,
            chinese_name_value=sheet.cell(row, 10).value,
            byte_value=sheet.cell(row, 7).value,
            bit_value=sheet.cell(row, 8).value,
            data_type_value=sheet.cell(row, 11).value,
            unit_value=sheet.cell(row, 16).value,
            conversion_value=sheet.cell(row, 12).value,
            conversion_fallback_value=sheet.cell(row, 13).value,
            min_value=sheet.cell(row, 14).value,
            max_value=sheet.cell(row, 15).value,
            fallback_name=desc_en or hex_short("DID", did_value),
            size=size,
        )
        if param:
            did.params.append(param)

    for did in dids.values():
        merge_split_byte_did_params(did)
        did.size = did_param_size(did.params, did.size or 1)
    return list(dids.values())


def parse_io_dids(workbook: Any) -> list[IoDidDef]:
    sheet = find_sheet(workbook, lambda name: "IO DID" in name)
    if sheet is None:
        return []

    io_by_id: dict[int, IoDidDef] = {}
    for row in range(10, sheet.max_row + 1):
        if compact_text(sheet.cell(row, 2).value).startswith("#"):
            break
        if row_is_deleted(sheet, row):
            continue
        raw_did = compact_text(sheet.cell(row, 2).value)
        did_value = parse_hex_in_text(raw_did, max_value=0xFFFF)
        control = parse_hex_in_text(sheet.cell(row, 5).value, max_value=0x03)
        if did_value is None or control is None:
            continue
        if "DEMO" in raw_did.upper() or compact_text(sheet.cell(row, 3).value).lower() == "demo":
            continue

        size = parse_int_cell(sheet.cell(row, 6).value, default=0)
        desc_en, desc_long = dual_name(sheet.cell(row, 3).value, sheet.cell(row, 4).value, hex_short("IODID", did_value))
        io_did = io_by_id.get(did_value)
        if io_did is None:
            io_did = IoDidDef(did=did_value, desc=canonical_name(desc_en, desc_long, hex_short("IODID", did_value)), size=size)
            io_by_id[did_value] = io_did
        io_did.size = max(io_did.size, size)
        io_did.controls.add(control)
        if size <= 0:
            continue

        param = make_param_from_cells(
            name_value=sheet.cell(row, 9).value,
            chinese_name_value=sheet.cell(row, 10).value,
            byte_value=sheet.cell(row, 7).value,
            bit_value=sheet.cell(row, 8).value,
            data_type_value=sheet.cell(row, 11).value,
            unit_value=sheet.cell(row, 16).value,
            conversion_value=sheet.cell(row, 12).value,
            conversion_fallback_value=sheet.cell(row, 13).value,
            min_value=sheet.cell(row, 14).value,
            max_value=sheet.cell(row, 15).value,
            fallback_name=desc_en or hex_short("IODID", did_value),
            size=size,
        )
        if param:
            io_did.params.append(param)

    return list(io_by_id.values())


def parse_routine_control_type(value: Any) -> int | None:
    text = compact_text(value)
    match = re.search(r"0[xX]([0-9A-Fa-f]{1,2})", text)
    if match:
        result = int(match.group(1), 16)
        return result if result in {1, 2, 3} else None
    return parse_hex_cell(text.split()[0] if text else "", max_value=0x03)


def parse_routines(workbook: Any) -> list[RoutineDef]:
    sheet = find_sheet(workbook, lambda name: "Routine DID" in name)
    if sheet is None:
        return []

    routines: dict[int, RoutineDef] = {}
    for row in range(11, sheet.max_row + 1):
        if compact_text(sheet.cell(row, 2).value).startswith("#"):
            break
        if row_is_deleted(sheet, row):
            continue
        rid = parse_hex_in_text(sheet.cell(row, 2).value, max_value=0xFFFF)
        control_type = parse_routine_control_type(sheet.cell(row, 5).value)
        if rid is None or control_type is None:
            continue
        if compact_text(sheet.cell(row, 3).value).lower() == "demo":
            continue
        supported = compact_text(sheet.cell(row, 6).value).upper() == "Y"
        desc_en, desc_long = dual_name(sheet.cell(row, 3).value, sheet.cell(row, 4).value, hex_short("RID", rid))
        routine = routines.get(rid)
        if routine is None:
            routine = RoutineDef(
                rid=rid,
                desc=canonical_name(desc_en, desc_long, hex_short("RID", rid)),
                security=usable_text(sheet.cell(row, 25).value) or "N",
                sessions=[compact_text(sheet.cell(row, col).value) for col in (26, 27, 28, 29, 30)],
            )
            routines[rid] = routine
        elif routine.security == "N" and usable_text(sheet.cell(row, 25).value):
            routine.security = usable_text(sheet.cell(row, 25).value)

        subfn = routine.subfunctions.get(control_type)
        if subfn is None:
            subfn = RoutineSubFunction(control_type=control_type, supported=supported)
            routine.subfunctions[control_type] = subfn
        else:
            subfn.supported = subfn.supported or supported

        option_size = parse_int_cell(sheet.cell(row, 7).value, default=0)
        if option_size > 0:
            param = make_param_from_cells(
                name_value=sheet.cell(row, 10).value,
                byte_value=sheet.cell(row, 8).value,
                bit_value=sheet.cell(row, 9).value,
                data_type_value=sheet.cell(row, 11).value,
                unit_value=sheet.cell(row, 15).value,
                conversion_value=sheet.cell(row, 12).value,
                min_value=sheet.cell(row, 13).value,
                max_value=sheet.cell(row, 14).value,
                fallback_name="RoutineControlOptionRecord",
                size=option_size,
            )
            if param:
                subfn.option_params.append(param)

        status_size = parse_int_cell(sheet.cell(row, 16).value, default=0)
        if status_size > 0:
            param = make_param_from_cells(
                name_value=sheet.cell(row, 19).value,
                byte_value=sheet.cell(row, 17).value,
                bit_value=sheet.cell(row, 18).value,
                data_type_value=sheet.cell(row, 20).value,
                unit_value=sheet.cell(row, 24).value,
                conversion_value=sheet.cell(row, 21).value,
                min_value=sheet.cell(row, 22).value,
                max_value=sheet.cell(row, 23).value,
                fallback_name="RoutineStatusRecord",
                size=status_size,
            )
            if param:
                subfn.status_params.append(param)

    return list(routines.values())


def parse_dtcs(workbook: Any) -> list[DtcDef]:
    sheet = find_sheet(workbook, lambda name: "DTC list" in name)
    if sheet is None:
        return []

    dtcs: list[DtcDef] = []
    for row in range(8, sheet.max_row + 1):
        if compact_text(sheet.cell(row, 2).value).startswith("#"):
            break
        if row_is_deleted(sheet, row):
            continue
        if compact_text(sheet.cell(row, 2).value).lower() == "demo":
            continue
        display = compact_text(sheet.cell(row, 4).value).upper()
        byte_text = compact_text(sheet.cell(row, 5).value).upper().replace("0X", "")
        if not re.fullmatch(r"[PCBU][0-9A-F]{6}", display):
            continue
        if not re.fullmatch(r"[0-9A-F]{6}", byte_text):
            continue
        _, text = dual_name(sheet.cell(row, 6).value, sheet.cell(row, 7).value, display)
        priority = compact_text(sheet.cell(row, 13).value)
        priority_match = re.search(r"\d+", priority)
        dtcs.append(
            DtcDef(
                display_code=display,
                byte_code=int(byte_text, 16),
                text=text or display,
                priority=priority_match.group(0) if priority_match else priority,
            )
        )
    return dtcs


def parse_record_number(value: Any) -> int | None:
    text = compact_text(value)
    hex_match = re.search(r"0[xX]([0-9A-Fa-f]{1,2})", text)
    if hex_match:
        result = int(hex_match.group(1), 16)
        return result if 0 <= result <= 0xFF else None
    dec_match = re.search(r"\d+", text)
    if not dec_match:
        return None
    result = int(dec_match.group(0))
    return result if 0 <= result <= 0xFF else None


def is_record_number_only_text(text: str) -> bool:
    without_number = re.sub(r"0[xX][0-9A-Fa-f]+|\d+", "", text)
    without_punctuation = re.sub(r"[\s()（）\[\]【】{}:：,，.;；/\\_\-]+", "", without_number)
    return not without_punctuation


def record_label_from_number_cell(value: Any, record_num: int, header_text: str, fallback_prefix: str) -> str:
    source = usable_text(value)
    if source and not is_record_number_only_text(source):
        return source
    if source and header_text:
        return f"{header_text} {source}"
    if source:
        return source
    if header_text:
        return f"{header_text} {record_num}"
    return f"{fallback_prefix} 0x{record_num:02X}"


def extended_record_label(record: ExtendedRecordDef) -> str:
    return usable_text(record.desc) or f"Extended Data Record 0x{record.record_num:02X}"


def param_key(param: ParamDef) -> tuple[str, int, int, int]:
    return (compact_text(param.name).lower(), param.byte_pos, param.bit_pos, param.bit_len)


def parse_snapshot_extended(workbook: Any) -> tuple[list[SnapshotDef], list[ExtendedRecordDef], dict[int, str]]:
    sheet = find_sheet(workbook, lambda name: "Snapshot&Extended" in name)
    if sheet is None:
        return [], [], {}

    snapshots_by_did: dict[int, SnapshotDef] = {}
    snapshot_param_keys: dict[int, set[tuple[str, int, int, int]]] = {}
    snapshot_record_names: dict[int, str] = {}
    ext_header = 0
    for row in range(1, sheet.max_row + 1):
        if "Extended Data Record Num" in compact_text(sheet.cell(row, 3).value):
            ext_header = row
            break
    snapshot_end = ext_header if ext_header else sheet.max_row + 1
    snapshot_record_header = compact_text(sheet.cell(6, 3).value)

    for row in range(8, snapshot_end):
        if row_is_deleted(sheet, row):
            continue
        did = parse_hex_cell(sheet.cell(row, 4).value, max_value=0xFFFF)
        if did is None:
            continue
        record_num = parse_record_number(sheet.cell(row, 3).value)
        if record_num is not None:
            label = record_label_from_number_cell(
                sheet.cell(row, 3).value,
                record_num,
                snapshot_record_header,
                "Snapshot Record",
            )
            snapshot_record_names.setdefault(record_num, label)
        size = parse_int_cell(sheet.cell(row, 7).value, default=0)
        desc_en, desc_long = dual_name(sheet.cell(row, 5).value, sheet.cell(row, 6).value, hex_short("SnapshotDID", did))
        snapshot = snapshots_by_did.get(did)
        if snapshot is None:
            snapshot = SnapshotDef(record_num=record_num, did=did, desc=canonical_name(desc_en, desc_long, hex_short("SnapshotDID", did)), size=size)
            snapshots_by_did[did] = snapshot
            snapshot_param_keys[did] = set()
        snapshot.size = max(snapshot.size, size)
        conversion = sheet.cell(row, 13).value or sheet.cell(row, 14).value or sheet.cell(row, 18).value
        param = make_param_from_cells(
            name_value=sheet.cell(row, 10).value,
            chinese_name_value=sheet.cell(row, 11).value,
            byte_value=sheet.cell(row, 8).value,
            bit_value=sheet.cell(row, 9).value,
            data_type_value=sheet.cell(row, 12).value,
            unit_value=sheet.cell(row, 17).value,
            conversion_value=conversion,
            min_value=sheet.cell(row, 15).value,
            max_value=sheet.cell(row, 16).value,
            fallback_name=desc_en or hex_short("SnapshotDID", did),
            size=size,
        )
        if param and param_key(param) not in snapshot_param_keys[did]:
            snapshot.params.append(param)
            snapshot_param_keys[did].add(param_key(param))

    extended_records: list[ExtendedRecordDef] = []
    current: ExtendedRecordDef | None = None
    if ext_header:
        for row in range(ext_header + 1, sheet.max_row + 1):
            if compact_text(sheet.cell(row, 2).value).startswith("#"):
                break
            record_num = parse_record_number(sheet.cell(row, 3).value)
            if row_is_deleted(sheet, row):
                if record_num is not None:
                    current = None
                continue
            if record_num is not None:
                desc_en, desc_long = dual_name(sheet.cell(row, 4).value, sheet.cell(row, 5).value, "")
                current = ExtendedRecordDef(
                    record_num=record_num,
                    desc=canonical_name(desc_en, desc_long, f"Extended Data Record 0x{record_num:02X}"),
                    size=parse_int_cell(sheet.cell(row, 6).value, default=0),
                )
                extended_records.append(current)
            if current is None:
                continue
            size = current.size
            conversion = sheet.cell(row, 11).value or sheet.cell(row, 12).value or sheet.cell(row, 16).value
            param = make_param_from_cells(
                name_value=sheet.cell(row, 9).value,
                byte_value=sheet.cell(row, 7).value,
                bit_value=sheet.cell(row, 8).value,
                data_type_value=sheet.cell(row, 10).value,
                unit_value=sheet.cell(row, 15).value,
                conversion_value=conversion,
                min_value=sheet.cell(row, 13).value,
                max_value=sheet.cell(row, 14).value,
                fallback_name=f"ExtendedRecord_{current.record_num:02X}",
                size=size,
            )
            if param:
                current.params.append(param)

    return [snapshot for snapshot in snapshots_by_did.values() if snapshot.params], extended_records, snapshot_record_names


def update_template(template_pdx: Path, output_pdx: Path, survey: SurveyData, validate: bool = True) -> None:
    with tempfile.TemporaryDirectory(prefix="seres_pdx_") as tmp_name:
        tmp_dir = Path(tmp_name)
        with zipfile.ZipFile(template_pdx, "r") as archive:
            archive.extractall(tmp_dir)

        odx_path = tmp_dir / "SERES_ECU_CAN_v15.odx-d"
        if not odx_path.exists():
            candidates = list(tmp_dir.glob("*.odx-d"))
            if not candidates:
                raise FileNotFoundError("No .odx-d file found inside template PDX")
            odx_path = candidates[0]

        parser = etree.XMLParser(remove_blank_text=False)
        tree = etree.parse(str(odx_path), parser)
        root = tree.getroot()
        id_gen = BASE.IdGenerator(root)
        update_odx_seres(root, id_gen, survey)
        validate_diag_service_child_order(root, odx_path.name)
        BASE.validate_can_dela_odx_structure(root, odx_path.name)
        tree.write(str(odx_path), encoding="UTF-8", xml_declaration=True, pretty_print=True, standalone=False)

        keep_files = set(SERES_KEEP_FILES)
        keep_files.add(odx_path.name)
        BASE.patch_pdx_catalog_for_can_only(tmp_dir / "index.xml", keep_files)

        output_pdx.parent.mkdir(parents=True, exist_ok=True)
        if output_pdx.exists():
            output_pdx.unlink()
        with zipfile.ZipFile(output_pdx, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for path in sorted(tmp_dir.rglob("*")):
                rel_path = path.relative_to(tmp_dir).as_posix()
                if path.is_file() and rel_path in keep_files:
                    archive.write(path, rel_path)

    if validate:
        validate_with_odxtools(output_pdx)


def update_odx_seres(root: etree._Element, id_gen: Any, survey: SurveyData) -> None:
    update_layer_names(root, survey.cover)
    ddds = find_dictionary_spec(root)
    if ddds is None:
        raise RuntimeError("Template ODX has no DIAG-DATA-DICTIONARY-SPEC")
    data_object_props = ddds.find("DATA-OBJECT-PROPS")
    structures = ddds.find("STRUCTURES")
    if data_object_props is None or structures is None:
        raise RuntimeError("Template ODX is missing DOP/STRUCTURE containers")

    unit_ids = BASE.ensure_units(ddds, survey)
    generated_dop_cache: dict[tuple[str, int, str, str, str], str] = {}

    for did in survey.dids:
        BASE.prepare_data_structure(
            id_gen=id_gen,
            data_object_props=data_object_props,
            structures=structures,
            item=did,
            prefix="DID",
            unit_ids=unit_ids,
            generated_dop_cache=generated_dop_cache,
        )
        if direct_did_payload_param(did) is not None:
            remove_structure_by_id(structures, did.structure_id)
            did.structure_id = ""

    for io_did in survey.io_dids:
        BASE.prepare_data_structure(
            id_gen=id_gen,
            data_object_props=data_object_props,
            structures=structures,
            item=io_did,
            prefix="IODID",
            unit_ids=unit_ids,
            generated_dop_cache=generated_dop_cache,
        )

    for routine in survey.routines:
        english, long_name = BASE.split_name(routine.desc)
        routine.short_name = sanitize_short_name(english or hex_short("RID", routine.rid), hex_short("RID", routine.rid))
        routine.long_name = long_name or routine.short_name
        for subfn in routine.subfunctions.values():
            if subfn.option_params:
                subfn.option_structure_id = BASE.make_param_structure(
                    id_gen,
                    data_object_props,
                    structures,
                    f"STR_RoutineOption_{routine.short_name}_{subfn.control_type:02X}",
                    f"Routine Option {routine.long_name}",
                    subfn.option_params,
                    unit_ids,
                    generated_dop_cache,
                )
            if subfn.status_params:
                subfn.status_structure_id = BASE.make_param_structure(
                    id_gen,
                    data_object_props,
                    structures,
                    f"STR_RoutineStatus_{routine.short_name}_{subfn.control_type:02X}",
                    f"Routine Status {routine.long_name}",
                    subfn.status_params,
                    unit_ids,
                    generated_dop_cache,
                )

    for snapshot in survey.snapshots:
        if snapshot.params:
            snapshot.structure_id = BASE.make_param_structure(
                id_gen,
                data_object_props,
                structures,
                f"STR_Snapshot_{sanitize_short_name(snapshot.desc, hex_short('Snapshot', snapshot.did))}",
                snapshot.desc,
                snapshot.params,
                unit_ids,
                generated_dop_cache,
                byte_size=snapshot.size or None,
            )

    for record in survey.extended_records:
        if record.params:
            record.structure_id = BASE.make_param_structure(
                id_gen,
                data_object_props,
                structures,
                f"STR_ExtendedRecord_{record.record_num:02X}",
                record.desc,
                record.params,
                unit_ids,
                generated_dop_cache,
                byte_size=record.size or None,
            )

    generate_flat_did_services(root, id_gen, survey.dids)
    generate_flat_io_services(root, id_gen, survey.io_dids)
    generate_flat_routine_services(root, id_gen, survey.routines)
    for old_name in ("z_7_Read", "z_Read", "z_Control", "z_ReturnControl"):
        BASE.remove_service_and_messages(root, old_name)
    BASE.remove_service_and_messages(root, "Upload_Download_RequestDownload")
    update_dtc_dop_seres(root, id_gen, survey.dtcs)
    update_snapshot_and_extended_data_seres(
        root,
        id_gen,
        survey.snapshots,
        survey.extended_records,
        getattr(survey, "snapshot_record_names", {}),
    )
    update_base_variant_comparams(root, survey.cover)
    update_session_timing(root, survey.cover)
    ensure_core_response_services(root, id_gen)
    generate_boot_security_access_services(root, id_gen, survey)
    update_core_service_preconditions(root, survey)
    shorten_dictionary_short_names(root)
    BASE.prefix_doc_revision_labels(root)


def update_layer_names(root: etree._Element, cover: CoverInfo) -> None:
    ecu_name = sanitize_short_name(cover.ecu_name, "SERES_ECU_CAN", max_len=CANDELA_SHORT_NAME_MAX_LEN)
    container = root.find(".//DIAG-LAYER-CONTAINER")
    if container is not None:
        long_name = container.find("LONG-NAME")
        if long_name is not None:
            long_name.text = ecu_name
    base_variant = get_base_variant(root)
    if base_variant is not None:
        short_name = base_variant.find("SHORT-NAME")
        if short_name is not None:
            short_name.text = ecu_name
        long_name = base_variant.find("LONG-NAME")
        if long_name is not None:
            long_name.text = ecu_name


def get_base_variant(root: etree._Element) -> etree._Element | None:
    node = BASE.first_by_short(root, "BASE-VARIANT", "SERES_ECU_CAN")
    if node is not None:
        return node
    nodes = root.xpath("//*[local-name()='BASE-VARIANT']")
    return nodes[0] if nodes else None


def find_dictionary_spec(root: etree._Element) -> etree._Element | None:
    for ddds in root.xpath("//*[local-name()='DIAG-DATA-DICTIONARY-SPEC']"):
        if ddds.find("DATA-OBJECT-PROPS") is not None and ddds.find("STRUCTURES") is not None:
            return ddds
    nodes = root.xpath("//*[local-name()='DIAG-DATA-DICTIONARY-SPEC']")
    return nodes[0] if nodes else None


def candela_short_name(value: str, used: set[str] | None = None, max_len: int = CANDELA_SHORT_NAME_MAX_LEN) -> str:
    if len(value) <= max_len and (used is None or value not in used):
        if used is not None:
            used.add(value)
        return value

    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:8].upper()
    suffix = f"_{digest}"
    head_len = max(1, max_len - len(suffix))
    base = value[:head_len].rstrip("_") or value[:head_len]
    candidate = f"{base}{suffix}"
    if used is not None:
        index = 2
        while candidate in used:
            index_suffix = f"_{index:02d}"
            candidate = f"{candidate[:max_len - len(index_suffix)].rstrip('_')}{index_suffix}"
            index += 1
        used.add(candidate)
    return candidate


def shorten_dictionary_short_names(root: etree._Element) -> None:
    for container_name, child_name in (
        ("DATA-OBJECT-PROPS", "DATA-OBJECT-PROP"),
        ("STRUCTURES", "STRUCTURE"),
    ):
        for container in root.xpath(f"//*[local-name()='{container_name}']"):
            used: set[str] = set()
            for node in container.xpath(f"./*[local-name()='{child_name}']"):
                short = node.find("SHORT-NAME")
                if short is None or not short.text:
                    continue
                short.text = candela_short_name(short.text, used)


def validate_diag_service_child_order(root: etree._Element, source_name: str) -> None:
    order = {
        "SHORT-NAME": 0,
        "LONG-NAME": 1,
        "DESC": 2,
        "ADMIN-DATA": 3,
        "SDGS": 4,
        "FUNCT-CLASS-REFS": 5,
        "AUDIENCE": 6,
        "PROTOCOL-SNREFS": 7,
        "RELATED-DIAG-COMM-REFS": 8,
        "PRE-CONDITION-STATE-REFS": 9,
        "STATE-TRANSITION-REFS": 10,
        "COMPARAM-REFS": 11,
        "REQUEST-REF": 12,
        "POS-RESPONSE-REFS": 13,
        "NEG-RESPONSE-REFS": 14,
        "POS-RESPONSE-SUPPRESSABLE": 15,
    }
    errors: list[str] = []
    for service in root.xpath("//*[local-name()='DIAG-SERVICE']"):
        service_name = service.findtext("SHORT-NAME") or service.get("ID") or "<unnamed>"
        last_index = -1
        for child in service:
            if not isinstance(child.tag, str):
                continue
            child_name = etree.QName(child).localname
            index = order.get(child_name)
            if index is None:
                continue
            if index < last_index:
                errors.append(f"DIAG-SERVICE '{service_name}' has '{child_name}' out of ODX order")
                break
            last_index = index
    if errors:
        details = "\n".join(f"- {message}" for message in errors[:20])
        raise RuntimeError(f"Generated {source_name} has invalid DIAG-SERVICE child order:\n{details}")


def get_short_name(node: etree._Element) -> str:
    return node.findtext("SHORT-NAME") or ""


def find_container(root: etree._Element, name: str) -> etree._Element:
    node = root.find(f".//{name}")
    if node is None:
        raise RuntimeError(f"Template ODX is missing {name}")
    return node


def find_by_id(root: etree._Element, node_id: str | None) -> etree._Element | None:
    if not node_id:
        return None
    nodes = root.xpath("//*[@ID=$node_id]", node_id=node_id)
    return nodes[0] if nodes else None


def service_messages(root: etree._Element, service_short_name: str) -> tuple[etree._Element, etree._Element, etree._Element | None, etree._Element | None]:
    service = BASE.first_by_short(root, "DIAG-SERVICE", service_short_name)
    if service is None:
        raise RuntimeError(f"Template service {service_short_name} was not found")
    request_ref = service.find("REQUEST-REF")
    request = find_by_id(root, request_ref.get("ID-REF") if request_ref is not None else None)
    if request is None:
        raise RuntimeError(f"Template service {service_short_name} has no request")
    pos_ref = service.find(".//POS-RESPONSE-REF")
    neg_ref = service.find(".//NEG-RESPONSE-REF")
    positive = find_by_id(root, pos_ref.get("ID-REF") if pos_ref is not None else None)
    negative = find_by_id(root, neg_ref.get("ID-REF") if neg_ref is not None else None)
    return service, request, positive, negative


def clone_service_bundle(
    root: etree._Element,
    id_gen: Any,
    base_service_short_name: str,
    new_short_name: str,
    long_name: str,
    service_qualifier: str,
    service_name: str,
    instance_qualifier: str,
    instance_name: str,
    precondition_state_ids: Iterable[str] = (),
    static_value: int | None = None,
) -> tuple[etree._Element, etree._Element, etree._Element | None, etree._Element | None]:
    base_service, base_request, base_positive, base_negative = service_messages(root, base_service_short_name)
    request = copy.deepcopy(base_request)
    positive = copy.deepcopy(base_positive) if base_positive is not None else None
    negative = copy.deepcopy(base_negative) if base_negative is not None else None
    service = copy.deepcopy(base_service)

    request_id = id_gen.new("RQ")
    positive_id = id_gen.new("PR") if positive is not None else None
    negative_id = id_gen.new("NR") if negative is not None else None
    service_id = id_gen.new("SVC")

    request.set("ID", request_id)
    BASE.refresh_internal_ids(request, id_gen)
    BASE.set_short_long(request, f"RQ_{new_short_name}", f"RQ {long_name}")

    if positive is not None and positive_id is not None:
        positive.set("ID", positive_id)
        BASE.refresh_internal_ids(positive, id_gen)
        BASE.set_short_long(positive, f"PR_{new_short_name}", f"PR {long_name}")

    if negative is not None and negative_id is not None:
        negative.set("ID", negative_id)
        BASE.refresh_internal_ids(negative, id_gen)
        BASE.set_short_long(negative, f"NR_{new_short_name}", f"NR {long_name}")

    service.set("ID", service_id)
    BASE.refresh_internal_ids(service, id_gen)
    BASE.set_short_long(service, new_short_name, long_name)
    request_ref = service.find("REQUEST-REF")
    if request_ref is not None:
        request_ref.set("ID-REF", request_id)
    pos_ref = service.find(".//POS-RESPONSE-REF")
    if pos_ref is not None and positive_id is not None:
        pos_ref.set("ID-REF", positive_id)
    neg_ref = service.find(".//NEG-RESPONSE-REF")
    if neg_ref is not None and negative_id is not None:
        neg_ref.set("ID-REF", negative_id)
    update_service_sdg(service, id_gen, service_qualifier, service_name, instance_qualifier, instance_name, static_value)
    update_flat_preconditions(service, precondition_state_ids)

    find_container(root, "REQUESTS").append(request)
    if positive is not None:
        find_container(root, "POS-RESPONSES").append(positive)
    if negative is not None:
        find_container(root, "NEG-RESPONSES").append(negative)
    find_container(root, "DIAG-COMMS").append(service)
    return service, request, positive, negative


def update_service_sdg(
    service: etree._Element,
    id_gen: Any,
    service_qualifier: str,
    service_name: str,
    instance_qualifier: str,
    instance_name: str,
    static_value: int | None = None,
) -> None:
    sdgs = BASE.element("SDGS")
    sdg = BASE.sub(sdgs, "SDG")
    caption = BASE.sub(sdg, "SDG-CAPTION", attrib={"ID": id_gen.new("CAP")})
    BASE.sub(caption, "SHORT-NAME", "CANdelaServiceInformation")
    if static_value is not None:
        BASE.sub(sdg, "SD", static_value, attrib={"SI": "DiagInstanceStaticValue"})
    BASE.sub(sdg, "SD", sanitize_short_name(instance_qualifier, "Instance"), attrib={"SI": "DiagInstanceQualifier"})
    BASE.sub(sdg, "SD", instance_name or instance_qualifier, attrib={"SI": "DiagInstanceName"})
    BASE.sub(sdg, "SD", service_qualifier, attrib={"SI": "ServiceQualifier"})
    BASE.sub(sdg, "SD", service_name, attrib={"SI": "ServiceName"})
    BASE.sub(sdg, "SD", "no", attrib={"SI": "PositiveResponseSuppressed"})
    BASE.replace_child(service, "SDGS", sdgs, before_tags={"FUNCT-CLASS-REFS", "AUDIENCE", "PRE-CONDITION-STATE-REFS", "REQUEST-REF"})


def update_flat_preconditions(service: etree._Element, state_ids: Iterable[str]) -> None:
    state_ids = list(dict.fromkeys(state_ids))
    if not state_ids:
        existing = service.find("PRE-CONDITION-STATE-REFS")
        if existing is not None:
            service.remove(existing)
        return
    pc = BASE.element("PRE-CONDITION-STATE-REFS")
    for state_id in state_ids:
        BASE.sub(pc, "PRE-CONDITION-STATE-REF", attrib={"ID-REF": state_id})
        BASE.replace_child(
            service,
            "PRE-CONDITION-STATE-REFS",
            pc,
            before_tags={"STATE-TRANSITION-REFS", "COMPARAM-REFS", "REQUEST-REF"},
        )


def state_id_map(root: etree._Element) -> dict[str, str]:
    result: dict[str, str] = {}
    for state in root.xpath("//*[local-name()='STATE']"):
        short_name = state.findtext("SHORT-NAME")
        if short_name and state.get("ID"):
            result[short_name] = state.get("ID")
    return result


def did_state_ids(root: etree._Element, did: DidDef, mode: str) -> list[str]:
    states = state_id_map(root)
    default = states.get("Default")
    programming = states.get("Programming")
    extended = states.get("Extended")
    locked = states.get("Locked")
    unlocked_l1 = states.get("UnlockedL1")
    unlocked_fbl = states.get("Unlocked_FBL")
    level_9 = states.get("Level_9")

    result: list[str] = []
    session_map = [default, extended, default, programming, extended]
    wanted = "R" if mode == "read" else "W"
    for value, state_id in zip(did.sessions, session_map, strict=False):
        if state_id and wanted in normalize_access(value):
            result.append(state_id)
    if mode == "read":
        result.extend(state for state in (locked, unlocked_l1, level_9, unlocked_fbl) if state)
    else:
        security = normalize_access(did.write_security)
        if "FBL" in security and unlocked_fbl:
            result.append(unlocked_fbl)
        elif ("LEVEL_1" in security or "LEVEL1" in security) and unlocked_l1:
            result.append(unlocked_l1)
        elif normalize_access(did.write_security) in {"", "N", "NO", "NONE"}:
            result.extend(state for state in (locked, unlocked_l1, level_9, unlocked_fbl) if state)
        else:
            result.extend(state for state in (unlocked_l1, level_9, unlocked_fbl) if state)
    if not result:
        result = [state for state in (default, extended, unlocked_l1, level_9) if state]
    return list(dict.fromkeys(result))


CORE_RESPONSE_SERVICE_SOURCES: tuple[str, ...] = (
    "DefaultSession_Start_NoResponse",
    "ExtendedDiagnosticSession_Start_NoResponse",
    "Hard_Reset_Reset_NoResponse",
    "EnableRxAndEnableTx_Control_NoResponse",
    "DisableRxAndDisableTx_Control_NoResponse",
    "TesterPresent_Send_NoResponse",
    "ControlDTCSetting_On_NoResponse",
    "ControlDTCSetting_Off_NoResponse",
)

CORE_SERVICE_PRECONDITION_TARGETS: dict[str, tuple[int, int | None, set[str]]] = {
    "DefaultSession_Start": (0x10, 0x01, {"Default", "Programming", "Extended"}),
    "DefaultSession_Start_NoResponse": (0x10, 0x01, {"Default", "Programming", "Extended"}),
    "ExtendedDiagnosticSession_Start": (0x10, 0x03, {"Default", "Programming", "Extended"}),
    "ExtendedDiagnosticSession_Start_NoResponse": (0x10, 0x03, {"Default", "Programming", "Extended"}),
    "Hard_Reset_Reset": (0x11, 0x01, {"Default", "Programming", "Extended"}),
    "Hard_Reset_Reset_NoResponse": (0x11, 0x01, {"Default", "Programming", "Extended"}),
    "EnableRxAndEnableTx_Control": (0x28, 0x00, {"Default", "Programming", "Extended"}),
    "EnableRxAndEnableTx_Control_NoResponse": (0x28, 0x00, {"Default", "Programming", "Extended"}),
    "DisableRxAndDisableTx_Control": (0x28, 0x03, {"Default", "Programming", "Extended"}),
    "DisableRxAndDisableTx_Control_NoResponse": (0x28, 0x03, {"Default", "Programming", "Extended"}),
    "TesterPresent_Send": (0x3E, 0x00, {"Default", "Programming", "Extended"}),
    "TesterPresent_Send_NoResponse": (0x3E, 0x00, {"Default", "Programming", "Extended"}),
    "ControlDTCSetting_On": (0x85, 0x01, {"Extended"}),
    "ControlDTCSetting_On_NoResponse": (0x85, 0x01, {"Extended"}),
    "ControlDTCSetting_Off": (0x85, 0x02, {"Extended"}),
    "ControlDTCSetting_Off_NoResponse": (0x85, 0x02, {"Extended"}),
}


def service_access_state_ids(root: etree._Element, sessions: Iterable[str], security: str) -> list[str]:
    states = state_id_map(root)
    result: list[str] = []
    session_order = ("Default", "Programming", "Extended")
    session_names = {compact_text(session) for session in sessions if compact_text(session)}
    for name in session_order:
        if name in session_names and states.get(name):
            result.append(states[name])

    security_text = normalize_access(security)
    if security_text in {"", "N", "NO", "NONE"}:
        security_names = ("Locked", "UnlockedL1", "Level_9", "Unlocked_FBL")
    else:
        security_names_list: list[str] = []
        if "LEVEL_1" in security_text or "LEVEL1" in security_text:
            security_names_list.append("UnlockedL1")
        if "LEVEL_9" in security_text or "LEVEL9" in security_text:
            security_names_list.append("Level_9")
        if "FBL" in security_text:
            security_names_list.append("Unlocked_FBL")
        security_names = tuple(security_names_list or ["UnlockedL1", "Level_9", "Unlocked_FBL"])

    for name in security_names:
        if states.get(name):
            result.append(states[name])
    return list(dict.fromkeys(result))


def sdg_value(service: etree._Element, si: str, default: str = "") -> str:
    node = service.find(f'.//SD[@SI="{si}"]')
    return node.text if node is not None and node.text is not None else default


def first_coded_value_by_semantic(message: etree._Element, semantic: str) -> int | None:
    for param in message.xpath('.//*[local-name()="PARAM" and @SEMANTIC=$semantic]', semantic=semantic):
        value = parse_int_cell(param.findtext("CODED-VALUE"), default=-1)
        if value >= 0:
            return value
    return None


def first_coded_param_by_semantic(message: etree._Element, semantic: str) -> etree._Element | None:
    params = message.xpath('.//*[local-name()="PARAM" and @SEMANTIC=$semantic]', semantic=semantic)
    return params[0] if params else None


def first_coded_value_by_semantics(message: etree._Element, semantics: Iterable[str]) -> int | None:
    for semantic in semantics:
        value = first_coded_value_by_semantic(message, semantic)
        if value is not None:
            return value
    return None


def diag_service_request_key(root: etree._Element, service: etree._Element) -> tuple[int, int | None] | None:
    request_ref = service.find("REQUEST-REF")
    request = find_by_id(root, request_ref.get("ID-REF") if request_ref is not None else None)
    if request is None:
        return None
    service_id = first_coded_value_by_semantic(request, "SERVICE-ID")
    if service_id is None:
        return None
    subfunction = first_coded_value_by_semantics(request, ("SUBFUNCTION", "ACCESSMODE"))
    return service_id, subfunction


def existing_diag_service_keys(root: etree._Element) -> set[tuple[int, int | None]]:
    result: set[tuple[int, int | None]] = set()
    for service in root.xpath("//*[local-name()='DIAG-SERVICE']"):
        key = diag_service_request_key(root, service)
        if key is not None:
            result.add(key)
    return result


def strip_subfunction_prefix(value: Any) -> str:
    text = usable_text(value)
    text = re.sub(r"^\s*(?:0[xX][0-9A-Fa-f]{1,2}|[0-9A-Fa-f]{2})\s*", "", text)
    return text.strip()


def security_access_instance_name(subfunction: int, access: dict[str, Any]) -> str:
    raw_name = strip_subfunction_prefix(access.get("subfunction_name", ""))
    if raw_name:
        return sanitize_short_name(raw_name, f"SecurityLevel{subfunction:02X}")
    if subfunction == 0x09:
        return "RequstSeedOfSecurityLevelFBL"
    if subfunction == 0x0A:
        return "SendKeyOfSecurityLevelFBL"
    prefix = "RequstSeed" if subfunction % 2 else "SendKey"
    return sanitize_short_name(f"{prefix}OfSecurityLevel{subfunction:02X}", f"SecurityLevel{subfunction:02X}")


def ensure_state_transition(root: etree._Element, id_gen: Any, source_state: str, target_state: str) -> str | None:
    chart = BASE.first_by_short(root, "STATE-CHART", "SecurityAccess")
    if chart is None:
        return None
    transitions = chart.find("STATE-TRANSITIONS")
    if transitions is None:
        transitions = BASE.element("STATE-TRANSITIONS")
        BASE.replace_child(chart, "STATE-TRANSITIONS", transitions, before_tags={"START-STATE-SNREF", "STATES"})

    for transition in transitions.findall("STATE-TRANSITION"):
        source = transition.find("SOURCE-SNREF")
        target = transition.find("TARGET-SNREF")
        if (
            source is not None
            and target is not None
            and source.get("SHORT-NAME") == source_state
            and target.get("SHORT-NAME") == target_state
            and transition.get("ID")
        ):
            return transition.get("ID")

    transition = BASE.sub(transitions, "STATE-TRANSITION", attrib={"ID": id_gen.new("ST")})
    BASE.sub(transition, "SHORT-NAME", sanitize_short_name(f"{source_state}_{target_state}", "SecurityTransition"))
    BASE.sub(transition, "LONG-NAME", f"{source_state} {target_state}")
    BASE.sub(transition, "SOURCE-SNREF", attrib={"SHORT-NAME": source_state})
    BASE.sub(transition, "TARGET-SNREF", attrib={"SHORT-NAME": target_state})
    return transition.get("ID")


def set_service_transition_refs(service: etree._Element, transition_id: str | None) -> None:
    existing = service.find("STATE-TRANSITION-REFS")
    if transition_id is None:
        if existing is not None:
            service.remove(existing)
        return
    refs = BASE.element("STATE-TRANSITION-REFS")
    BASE.sub(refs, "STATE-TRANSITION-REF", attrib={"ID-REF": transition_id})
    BASE.replace_child(service, "STATE-TRANSITION-REFS", refs, before_tags={"COMPARAM-REFS", "REQUEST-REF"})


def security_access_target_state(subfunction: int, access: dict[str, Any]) -> str:
    text = normalize_access(access.get("subfunction_name", ""))
    if "FBL" in text:
        return "Unlocked_FBL"
    if "LEVEL_9" in text or "LEVEL9" in text:
        return "Level_9"
    if subfunction == 0x0A:
        return "Unlocked_FBL"
    return "UnlockedL1"


def generate_boot_security_access_services(root: etree._Element, id_gen: Any, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    existing_keys = existing_diag_service_keys(root)
    for (service_id, subfunction), access in sorted(access_map.items()):
        if service_id != 0x27 or subfunction is None:
            continue
        if "Boot" not in (access.get("sources") or set()):
            continue
        if (service_id, subfunction) in existing_keys:
            continue

        is_seed = bool(subfunction % 2)
        base_service = "SeedLevel1_Request" if is_seed else "KeyLevel1_Send"
        if BASE.first_by_short(root, "DIAG-SERVICE", base_service) is None:
            continue

        instance_name = security_access_instance_name(subfunction, access)
        service_qualifier = "Request" if is_seed else "Send"
        service_name = f"{instance_name}_{service_qualifier}"
        service, request, positive, negative = clone_service_bundle(
            root,
            id_gen,
            base_service,
            service_name,
            f"{instance_name} {service_qualifier}",
            service_qualifier,
            service_qualifier,
            instance_name,
            instance_name,
            service_access_state_ids(root, access.get("sessions") or {"Programming"}, access.get("security") or "N"),
        )
        set_service_semantic_and_funct_class(root, service, "SECURITY", "SecurityAccess")

        subfunction_semantics = ("ACCESSMODE", "SUBFUNCTION")
        for message in (request, positive):
            if message is not None:
                for semantic in subfunction_semantics:
                    set_first_coded_by_semantic(message, semantic, subfunction)

        if negative is not None:
            set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x27)

        if not is_seed:
            target_state = security_access_target_state(subfunction, access)
            transition_id = ensure_state_transition(root, id_gen, "Locked", target_state)
            set_service_transition_refs(service, transition_id)

        existing_keys.add((service_id, subfunction))


def data_object_prop_id(root: etree._Element, short_name: str) -> str | None:
    dop = BASE.first_by_short(root, "DATA-OBJECT-PROP", short_name)
    return dop.get("ID") if dop is not None else None


def make_core_positive_response(
    root: etree._Element,
    node_id: str,
    short_name: str,
    long_name: str,
    request_sid: int,
    subfunction_param: etree._Element | None,
    subfunction_value: int,
) -> etree._Element:
    response = BASE.element("POS-RESPONSE", attrib={"ID": node_id})
    BASE.sub(response, "SHORT-NAME", sanitize_short_name(short_name, "PR"))
    BASE.sub(response, "LONG-NAME", long_name)
    params = BASE.sub(response, "PARAMS")
    params.append(coded_const_param("SID_PR", "SID-PR", 0, request_sid + 0x40, "SERVICE-ID"))

    sub_short_name = subfunction_param.findtext("SHORT-NAME") if subfunction_param is not None else "SubFunction"
    sub_long_name = subfunction_param.findtext("LONG-NAME") if subfunction_param is not None else sub_short_name
    params.append(
        coded_const_param(
            sub_short_name or "SubFunction",
            sub_long_name or sub_short_name or "SubFunction",
            1,
            subfunction_value,
            "SUBFUNCTION",
        )
    )

    if request_sid == 0x10:
        p2_id = data_object_prop_id(root, "P2")
        p2ex_id = data_object_prop_id(root, "P2Ex")
        if p2_id:
            params.append(BASE.make_value_param("P2", "P2", 2, 0, p2_id))
        if p2ex_id:
            params.append(BASE.make_value_param("P2Ex", "P2Ex", 4, 0, p2ex_id))
    return response


def create_core_response_service(root: etree._Element, id_gen: Any, no_response_name: str) -> None:
    response_name = no_response_name.removesuffix("_NoResponse")
    if response_name == no_response_name or BASE.first_by_short(root, "DIAG-SERVICE", response_name) is not None:
        return

    base_service, base_request, _, base_negative = service_messages(root, no_response_name)
    request_sid = first_coded_value_by_semantic(base_request, "SERVICE-ID")
    no_response_subfunction = first_coded_value_by_semantic(base_request, "SUBFUNCTION")
    if request_sid is None or no_response_subfunction is None:
        return
    response_subfunction = no_response_subfunction & 0x7F

    request = copy.deepcopy(base_request)
    negative = copy.deepcopy(base_negative) if base_negative is not None else None
    service = copy.deepcopy(base_service)

    request_id = id_gen.new("RQ")
    positive_id = id_gen.new("PR")
    negative_id = id_gen.new("NR") if negative is not None else None
    service_id = id_gen.new("SVC")

    request.set("ID", request_id)
    BASE.refresh_internal_ids(request, id_gen)
    BASE.set_short_long(request, f"RQ_{response_name}", f"RQ {response_name}")
    set_first_coded_by_semantic(request, "SUBFUNCTION", response_subfunction)

    subfunction_param = first_coded_param_by_semantic(request, "SUBFUNCTION")
    positive = make_core_positive_response(
        root,
        positive_id,
        f"PR_{response_name}",
        f"PR {response_name}",
        request_sid,
        subfunction_param,
        response_subfunction,
    )

    if negative is not None and negative_id is not None:
        negative.set("ID", negative_id)
        BASE.refresh_internal_ids(negative, id_gen)
        BASE.set_short_long(negative, f"NR_{response_name}", f"NR {response_name}")

    service.set("ID", service_id)
    service.attrib.pop("TRANSMISSION-MODE", None)
    BASE.refresh_internal_ids(service, id_gen)
    BASE.set_short_long(service, response_name, (base_service.findtext("LONG-NAME") or response_name).replace("_NoResponse", ""))

    request_ref = service.find("REQUEST-REF")
    if request_ref is not None:
        request_ref.set("ID-REF", request_id)

    pos_refs = BASE.element("POS-RESPONSE-REFS")
    BASE.sub(pos_refs, "POS-RESPONSE-REF", attrib={"ID-REF": positive_id})
    BASE.replace_child(service, "POS-RESPONSE-REFS", pos_refs, before_tags={"NEG-RESPONSE-REFS"})

    neg_ref = service.find(".//NEG-RESPONSE-REF")
    if neg_ref is not None and negative_id is not None:
        neg_ref.set("ID-REF", negative_id)

    update_service_sdg(
        service,
        id_gen,
        sdg_value(base_service, "ServiceQualifier", "Send"),
        sdg_value(base_service, "ServiceName", sdg_value(base_service, "ServiceQualifier", "Send")),
        sdg_value(base_service, "DiagInstanceQualifier", response_name),
        sdg_value(base_service, "DiagInstanceName", response_name),
    )

    find_container(root, "REQUESTS").append(request)
    find_container(root, "POS-RESPONSES").append(positive)
    if negative is not None:
        find_container(root, "NEG-RESPONSES").append(negative)
    find_container(root, "DIAG-COMMS").append(service)


def ensure_core_response_services(root: etree._Element, id_gen: Any) -> None:
    for no_response_name in CORE_RESPONSE_SERVICE_SOURCES:
        if BASE.first_by_short(root, "DIAG-SERVICE", no_response_name) is not None:
            create_core_response_service(root, id_gen, no_response_name)


def update_core_service_preconditions(root: etree._Element, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    for service_name, (service_id, subfunction, fallback_sessions) in CORE_SERVICE_PRECONDITION_TARGETS.items():
        service = BASE.first_by_short(root, "DIAG-SERVICE", service_name)
        if service is None:
            continue
        access = access_map.get((service_id, subfunction)) or access_map.get((service_id, None))
        if access:
            sessions = access.get("sessions") or fallback_sessions
            security = access.get("security") or "N"
        else:
            sessions = fallback_sessions
            security = "N"
        update_flat_preconditions(service, service_access_state_ids(root, sessions, security))


def set_coded_value(message: etree._Element, param_short_name: str, value: int) -> None:
    for param in message.xpath('.//*[local-name()="PARAM" and SHORT-NAME=$name]', name=param_short_name):
        coded = param.find("CODED-VALUE")
        if coded is not None:
            coded.text = str(value)


def set_first_coded_by_semantic(message: etree._Element, semantic: str, value: int) -> None:
    for param in message.xpath('.//*[local-name()="PARAM" and @SEMANTIC=$semantic]', semantic=semantic):
        coded = param.find("CODED-VALUE")
        if coded is not None:
            coded.text = str(value)
            return


def set_value_param_dop(message: etree._Element, dop_id: str, short_name: str, long_name: str, byte_position: int | None = None) -> None:
    value_params = message.xpath('./*[local-name()="PARAMS"]/*[local-name()="PARAM" and @xsi:type="VALUE"]', namespaces={"xsi": BASE.XSI_NS})
    if not value_params:
        return
    param = value_params[-1]
    short = param.find("SHORT-NAME")
    if short is not None:
        short.text = sanitize_short_name(short_name, "Data")
    long = param.find("LONG-NAME")
    if long is not None:
        long.text = long_name or short_name
    if byte_position is not None:
        byte = param.find("BYTE-POSITION")
        if byte is not None:
            byte.text = str(byte_position)
    dop_ref = param.find("DOP-REF")
    if dop_ref is None:
        dop_ref = BASE.sub(param, "DOP-REF")
    dop_ref.set("ID-REF", dop_id)


def generate_flat_did_services(root: etree._Element, id_gen: Any, dids: list[DidDef]) -> None:
    for did in dids:
        if did.readable:
            service_name = f"{did.short_name}_Read"
            service, request, positive, negative = clone_service_bundle(
                root,
                id_gen,
                "z_7_Read",
                service_name,
                f"{did.long_name} Read",
                "Read",
                "Read",
                did.short_name,
                did.long_name,
                did_state_ids(root, did, "read"),
                static_value=did.did,
            )
            set_service_semantic_and_funct_class(root, service, "IDENTIFICATION", "ECU_Identification")
            set_coded_value(request, "RecordDataIdentifier", did.did)
            if positive is not None:
                set_coded_value(positive, "RecordDataIdentifier", did.did)
                set_value_param_dop(positive, did_payload_dop_id(did), did.short_name, did.long_name, byte_position=3)
            if negative is not None:
                set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x22)

        if did.writable:
            service_name = f"{did.short_name}_Write"
            service, _, _, _ = clone_service_bundle(
                root,
                id_gen,
                "z_7_Read",
                service_name,
                f"{did.long_name} Write",
                "Write",
                "Write",
                did.short_name,
                did.long_name,
                did_state_ids(root, did, "write"),
                static_value=did.did,
            )
            set_service_semantic_and_funct_class(root, service, "IDENTIFICATION", "ECU_Identification")
            request = make_write_request(id_gen.new("RQ"), f"RQ_{service_name}", did)
            positive = make_write_positive(id_gen.new("PR"), f"PR_{service_name}", did)
            negative = make_negative_response(id_gen.new("NR"), f"NR_{service_name}", f"NR {did.long_name} Write", 0x2E, "Write")
            replace_service_refs(root, service, request, positive, negative)


def replace_service_refs(
    root: etree._Element,
    service: etree._Element,
    request: etree._Element,
    positive: etree._Element | None,
    negative: etree._Element | None,
) -> None:
    old_request_ref = service.find("REQUEST-REF")
    old_pos_ref = service.find(".//POS-RESPONSE-REF")
    old_neg_ref = service.find(".//NEG-RESPONSE-REF")
    old_ids = [
        old_request_ref.get("ID-REF") if old_request_ref is not None else None,
        old_pos_ref.get("ID-REF") if old_pos_ref is not None else None,
        old_neg_ref.get("ID-REF") if old_neg_ref is not None else None,
    ]
    for old_id in old_ids:
        old_node = find_by_id(root, old_id)
        if old_node is not None and old_node.getparent() is not None:
            old_node.getparent().remove(old_node)
    find_container(root, "REQUESTS").append(request)
    if positive is not None:
        find_container(root, "POS-RESPONSES").append(positive)
    if negative is not None:
        find_container(root, "NEG-RESPONSES").append(negative)
    if old_request_ref is not None:
        old_request_ref.set("ID-REF", request.get("ID"))
    if old_pos_ref is not None and positive is not None:
        old_pos_ref.set("ID-REF", positive.get("ID"))
    if old_neg_ref is not None and negative is not None:
        old_neg_ref.set("ID-REF", negative.get("ID"))


def coded_const_param(short_name: str, long_name: str, byte_position: int, coded_value: int, semantic: str, bit_length: int = 8) -> etree._Element:
    param = BASE.element("PARAM", attrib={"SEMANTIC": semantic})
    BASE.set_xsi_type(param, "CODED-CONST")
    BASE.sub(param, "SHORT-NAME", short_name)
    BASE.sub(param, "LONG-NAME", long_name)
    BASE.sub(param, "BYTE-POSITION", byte_position)
    BASE.sub(param, "CODED-VALUE", coded_value)
    coded_type = BASE.sub(param, "DIAG-CODED-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32"})
    BASE.set_xsi_type(coded_type, "STANDARD-LENGTH-TYPE")
    BASE.sub(coded_type, "BIT-LENGTH", bit_length)
    return param


def make_write_request(node_id: str, short_name: str, did: DidDef) -> etree._Element:
    request = BASE.element("REQUEST", attrib={"ID": node_id})
    BASE.sub(request, "SHORT-NAME", sanitize_short_name(short_name, "RQ"))
    BASE.sub(request, "LONG-NAME", f"RQ {did.long_name} Write")
    params = BASE.sub(request, "PARAMS")
    params.append(coded_const_param("SID_RQ", "SID-RQ", 0, 0x2E, "SERVICE-ID"))
    params.append(coded_const_param("RecordDataIdentifier", "RecordDataIdentifier", 1, did.did, "ID", bit_length=16))
    params.append(BASE.make_value_param(did.short_name, did.long_name, 3, 0, did_payload_dop_id(did)))
    return request


def make_write_positive(node_id: str, short_name: str, did: DidDef) -> etree._Element:
    response = BASE.element("POS-RESPONSE", attrib={"ID": node_id})
    BASE.sub(response, "SHORT-NAME", sanitize_short_name(short_name, "PR"))
    BASE.sub(response, "LONG-NAME", f"PR {did.long_name} Write")
    params = BASE.sub(response, "PARAMS")
    params.append(coded_const_param("SID_PR", "SID-PR", 0, 0x6E, "SERVICE-ID"))
    params.append(coded_const_param("RecordDataIdentifier", "RecordDataIdentifier", 1, did.did, "ID", bit_length=16))
    return response


def make_negative_response(node_id: str, short_name: str, long_name: str, request_sid: int, qualifier: str) -> etree._Element:
    response = BASE.element("NEG-RESPONSE", attrib={"ID": node_id})
    BASE.sub(response, "SHORT-NAME", sanitize_short_name(short_name, "NR"))
    BASE.sub(response, "LONG-NAME", long_name)
    params = BASE.sub(response, "PARAMS")
    params.append(coded_const_param("SID_NR", "SID-NR", 0, 0x7F, "SERVICE-ID"))
    params.append(coded_const_param("SIDRQ_NR", "SIDRQ-NR", 1, request_sid, "SERVICEIDRQ"))
    value = BASE.make_value_param(qualifier, qualifier, 2, 0, "_81")
    params.append(value)
    nrc = BASE.element("PARAM", attrib={"SEMANTIC": "DATA"})
    BASE.set_xsi_type(nrc, "NRC-CONST")
    BASE.sub(nrc, "SHORT-NAME", sanitize_short_name(f"NRCConst_{qualifier}", "NRCConst"))
    BASE.sub(nrc, "LONG-NAME", qualifier)
    BASE.sub(nrc, "BYTE-POSITION", 2)
    coded_values = BASE.sub(nrc, "CODED-VALUES")
    for code in (0x13, 0x22, 0x31, 0x33, 0x72, 0x7F):
        BASE.sub(coded_values, "CODED-VALUE", code)
    coded_type = BASE.sub(nrc, "DIAG-CODED-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32"})
    BASE.set_xsi_type(coded_type, "STANDARD-LENGTH-TYPE")
    BASE.sub(coded_type, "BIT-LENGTH", 8)
    params.append(nrc)
    return response


def set_service_semantic_and_funct_class(
    root: etree._Element,
    service: etree._Element,
    semantic: str,
    funct_class_short_name: str,
) -> None:
    service.set("SEMANTIC", semantic)
    class_node = BASE.first_by_short(root, "FUNCT-CLASS", funct_class_short_name)
    if class_node is None or not class_node.get("ID"):
        return
    refs = service.find("FUNCT-CLASS-REFS")
    if refs is None:
        refs = BASE.element("FUNCT-CLASS-REFS")
        BASE.replace_child(service, "FUNCT-CLASS-REFS", refs, before_tags={"AUDIENCE", "PRE-CONDITION-STATE-REFS", "REQUEST-REF"})
    for child in list(refs):
        refs.remove(child)
    BASE.sub(refs, "FUNCT-CLASS-REF", attrib={"ID-REF": class_node.get("ID")})


def generate_flat_io_services(root: etree._Element, id_gen: Any, io_dids: list[IoDidDef]) -> None:
    control_templates = {
        0: ("z_ReturnControl", "ReturnControl", "ReturnControl"),
        1: ("z_ReturnControl", "Reset", "Reset"),
        2: ("z_ReturnControl", "Freeze", "Freeze"),
        3: ("z_Control", "Control", "Control"),
    }
    default_states = [state_id for name, state_id in state_id_map(root).items() if name in {"Extended", "UnlockedL1", "Level_9"}]
    for io_did in io_dids:
        for control in sorted(io_did.controls):
            template, qualifier, service_label = control_templates.get(control, control_templates[3])
            service_name = f"{io_did.short_name}_{qualifier}"
            _, request, positive, negative = clone_service_bundle(
                root,
                id_gen,
                template,
                service_name,
                f"{io_did.long_name} {service_label}",
                qualifier,
                service_label,
                io_did.short_name,
                io_did.long_name,
                default_states,
            )
            set_coded_value(request, "DataIdentifier", io_did.did)
            set_coded_value(request, "ControlOptionRecord_InputOutputControlParameter", control)
            if control == 3:
                set_value_param_dop(request, io_did.structure_id, "ControlOptionRecord", "ControlOptionRecord", byte_position=4)
            if positive is not None:
                set_coded_value(positive, "DataIdentifier", io_did.did)
                set_coded_value(positive, "ControlStatusRecord_InputOutputControlParameter", control)
                set_value_param_dop(positive, io_did.structure_id, "ControlStatusRecord", "ControlStatusRecord", byte_position=4)
            if negative is not None:
                set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x2F)

def generate_flat_routine_services(root: etree._Element, id_gen: Any, routines: list[RoutineDef]) -> None:
    states = state_id_map(root)
    default_states = [state for state in (states.get("Extended"), states.get("Programming"), states.get("UnlockedL1"), states.get("Unlocked_FBL")) if state]
    labels = {1: ("Start", "Start"), 2: ("Stop", "Stop"), 3: ("RequestResults", "RequestResults")}
    for routine in routines:
        for control_type, subfn in sorted(routine.subfunctions.items()):
            if not subfn.supported:
                continue
            qualifier, service_label = labels.get(control_type, (f"Control_{control_type:02X}", f"Control {control_type:02X}"))
            service_name = f"{routine.short_name}_{qualifier}"
            service, _, _, _ = clone_service_bundle(
                root,
                id_gen,
                "z_Read",
                service_name,
                f"{routine.long_name} {service_label}",
                qualifier,
                service_label,
                routine.short_name,
                routine.long_name,
                default_states,
            )
            request = make_routine_request(id_gen.new("RQ"), f"RQ_{service_name}", routine, subfn)
            positive = make_routine_positive(id_gen.new("PR"), f"PR_{service_name}", routine, subfn)
            negative = make_negative_response(id_gen.new("NR"), f"NR_{service_name}", f"NR {routine.long_name} {service_label}", 0x31, qualifier)
            replace_service_refs(root, service, request, positive, negative)


def make_routine_request(node_id: str, short_name: str, routine: RoutineDef, subfn: RoutineSubFunction) -> etree._Element:
    request = BASE.element("REQUEST", attrib={"ID": node_id})
    BASE.sub(request, "SHORT-NAME", sanitize_short_name(short_name, "RQ"))
    BASE.sub(request, "LONG-NAME", f"RQ {routine.long_name}")
    params = BASE.sub(request, "PARAMS")
    params.append(coded_const_param("SID_RQ", "SID-RQ", 0, 0x31, "SERVICE-ID"))
    params.append(coded_const_param("RoutineControlType", "RoutineControlType", 1, subfn.control_type, "SUBFUNCTION"))
    params.append(coded_const_param("RoutineIdentifier", "RoutineIdentifier", 2, routine.rid, "ID", bit_length=16))
    if subfn.option_structure_id:
        params.append(BASE.make_value_param("RoutineControlOptionRecord", "RoutineControlOptionRecord", 4, 0, subfn.option_structure_id))
    return request


def make_routine_positive(node_id: str, short_name: str, routine: RoutineDef, subfn: RoutineSubFunction) -> etree._Element:
    response = BASE.element("POS-RESPONSE", attrib={"ID": node_id})
    BASE.sub(response, "SHORT-NAME", sanitize_short_name(short_name, "PR"))
    BASE.sub(response, "LONG-NAME", f"PR {routine.long_name}")
    params = BASE.sub(response, "PARAMS")
    params.append(coded_const_param("SID_PR", "SID-PR", 0, 0x71, "SERVICE-ID"))
    params.append(coded_const_param("RoutineControlType", "RoutineControlType", 1, subfn.control_type, "SUBFUNCTION"))
    params.append(coded_const_param("RoutineIdentifier", "RoutineIdentifier", 2, routine.rid, "ID", bit_length=16))
    if subfn.status_structure_id:
        params.append(BASE.make_value_param("RoutineStatusRecord", "RoutineStatusRecord", 4, 0, subfn.status_structure_id))
    return response


def update_dtc_dop_seres(root: etree._Element, id_gen: Any, dtcs: list[DtcDef]) -> None:
    dtc_dop = BASE.first_by_short(root, "DTC-DOP", "RecordDataType")
    if dtc_dop is None:
        return
    dtcs_node = dtc_dop.find("DTCS")
    if dtcs_node is None:
        dtcs_node = BASE.sub(dtc_dop, "DTCS")
    for child in list(dtcs_node):
        dtcs_node.remove(child)
    caption_ids: dict[str, str] = {}
    for index, dtc in enumerate(dtcs):
        dtc_node = BASE.sub(dtcs_node, "DTC", attrib={"ID": id_gen.new("DTC")})
        BASE.sub(dtc_node, "SHORT-NAME", sanitize_short_name(f"DTC_{dtc.byte_code:06X}", "DTC"))
        BASE.sub(dtc_node, "TROUBLE-CODE", dtc.byte_code)
        BASE.sub(dtc_node, "DISPLAY-TROUBLE-CODE", dtc.display_code)
        BASE.sub(dtc_node, "TEXT", dtc.text or dtc.display_code)
        sdgs = BASE.sub(dtc_node, "SDGS")
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_SHORTNAME", f"DTC_0X{dtc.byte_code:06X}", first=index == 0)
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_PRIORITY_VALUE", dtc.priority or "2", first=index == 0)
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_SUPPORTED", "supported", first=index == 0)
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_CYCLE", "DEM_POWER", first=index == 0)
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_COUNTER", "40", first=index == 0)
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_SEVERITY_VALUE", "noSeverity", first=index == 0)
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_INIT_MONITOR_REQUIRED", "not required", first=index == 0)
        BASE.add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_FUNCTIONAL_UNIT_VALUE", "0x00", first=index == 0)

    BASE.update_dtc_text_table(root, dtcs)

    ext_table = BASE.first_by_short(root, "TABLE", "DTCExtendedDataRecordNumber")
    if ext_table is None:
        return
    default_structure = first_table_structure_ref(ext_table) or create_extended_record_number_structure(root, id_gen)
    BASE.clear_children(ext_table, "TABLE-ROW")
    for dtc in dtcs:
        key = BASE.dtc_table_key(dtc)
        row = BASE.append_table_row(ext_table, id_gen, f"TR_DTC_{dtc.byte_code:06X}", key, key, default_structure)
        sdgs = BASE.sub(row, "SDGS")
        sdg = BASE.sub(sdgs, "SDG")
        caption = BASE.sub(sdg, "SDG-CAPTION", attrib={"ID": id_gen.new("CAP")})
        BASE.sub(caption, "SHORT-NAME", "IsDefaultCase")
        BASE.sub(sdg, "SD", "Yes")


def first_table_structure_ref(table: etree._Element) -> str | None:
    ref = table.find(".//STRUCTURE-REF")
    return ref.get("ID-REF") if ref is not None else None


def create_extended_record_number_structure(root: etree._Element, id_gen: Any) -> str:
    structures = find_container(root, "STRUCTURES")
    structure_id = id_gen.new("STR")
    structure = BASE.sub(structures, "STRUCTURE", attrib={"ID": structure_id, "IS-VISIBLE": "false"})
    BASE.sub(structure, "SHORT-NAME", "STRUC_DTCExtendedDataRecordNumbers")
    params = BASE.sub(structure, "PARAMS")
    params.append(BASE.make_value_param("Record_Numbers", "Record Numbers", 0, 0, dop_id_by_short(root, "DATA-OBJECT-PROP", "DTCExtendedDataRecordNumbers_All") or "_62"))
    return structure_id


def dop_id_by_short(root: etree._Element, tag: str, short_name: str) -> str | None:
    node = BASE.first_by_short(root, tag, short_name)
    return node.get("ID") if node is not None else None


def update_snapshot_and_extended_data_seres(
    root: etree._Element,
    id_gen: Any,
    snapshots: list[SnapshotDef],
    extended_records: list[ExtendedRecordDef],
    snapshot_record_names: dict[int, str],
) -> None:
    update_snapshot_record_number_dop(root, "DTCSnapshotRecordNumbers_All", snapshot_record_names, include_all=True)
    update_snapshot_record_number_dop(root, "DTCSnapshotRecordNumbers_All_except_FF", snapshot_record_names, include_all=False)

    env_data = BASE.first_by_short(root, "ENV-DATA", "ENVDATA_ALLDTCS")
    did_dop_id = dop_id_by_short(root, "DATA-OBJECT-PROP", "HexDump_2Byte") or "_2"
    if env_data is not None:
        params = env_data.find("PARAMS")
        if params is None:
            params = BASE.sub(env_data, "PARAMS")
        for child in list(params):
            params.remove(child)
        byte_position = 0
        for snapshot in snapshots:
            if not snapshot.structure_id:
                continue
            did_param = BASE.element("PARAM", attrib={"SEMANTIC": "DATA"})
            BASE.set_xsi_type(did_param, "PHYS-CONST")
            snapshot_short_name = sanitize_short_name(snapshot.desc, hex_short("DID", snapshot.did))
            BASE.sub(did_param, "SHORT-NAME", snapshot_short_name)
            BASE.sub(did_param, "LONG-NAME", snapshot.desc)
            BASE.sub(did_param, "BYTE-POSITION", byte_position)
            BASE.sub(did_param, "PHYS-CONSTANT-VALUE", snapshot.did)
            BASE.sub(did_param, "DOP-REF", attrib={"ID-REF": did_dop_id})
            params.append(did_param)
            byte_position += 2
            value_param = BASE.make_value_param(
                sanitize_short_name(f"{snapshot_short_name}_Data", "SnapshotData"),
                snapshot.desc,
                byte_position,
                0,
                snapshot.structure_id,
            )
            params.append(value_param)
            byte_position += max(1, snapshot.size)

    if extended_records:
        update_extended_record_number_dop(root, "DTCExtendedDataRecordNumbers_All", extended_records, include_all=True)
        update_extended_record_number_dop(root, "DTCExtendedDataRecordNumbers_All_except_FF", extended_records, include_all=False)
        mux = BASE.first_by_short(root, "MUX", "DTCExtendedDataRecordData")
        if mux is not None:
            cases = mux.find("CASES")
            if cases is None:
                cases = BASE.sub(mux, "CASES")
            for child in list(cases):
                cases.remove(child)
            for record in extended_records:
                if not record.structure_id:
                    continue
                case = BASE.sub(cases, "CASE")
                label = extended_record_label(record)
                BASE.sub(case, "SHORT-NAME", sanitize_short_name(f"Case_0x{record.record_num:02X}_{label}", f"Case_0x{record.record_num:02X}"))
                BASE.sub(case, "STRUCTURE-REF", attrib={"ID-REF": record.structure_id})
                BASE.sub(case, "LOWER-LIMIT", label)
                BASE.sub(case, "UPPER-LIMIT", label)


def collect_env_data_groups(params: etree._Element) -> dict[int, list[etree._Element]]:
    groups: dict[int, list[etree._Element]] = {}
    current_did: int | None = None
    for param in params.findall("PARAM"):
        if param.get(XSI_TYPE) == "PHYS-CONST":
            try:
                current_did = int(compact_text(param.findtext("PHYS-CONSTANT-VALUE")), 0)
            except ValueError:
                current_did = None
            if current_did is not None:
                groups[current_did] = [param]
            continue
        if current_did is not None:
            groups[current_did].append(param)
    return groups


def existing_env_data_dids(root: etree._Element) -> set[int]:
    env_data = BASE.first_by_short(root, "ENV-DATA", "ENVDATA_ALLDTCS")
    if env_data is None:
        return set()
    params = env_data.find("PARAMS")
    if params is None:
        return set()
    return set(collect_env_data_groups(params))


def env_param_end_byte(root: etree._Element, param: etree._Element) -> int:
    byte_position = parse_int_cell(param.findtext("BYTE-POSITION"), 0)
    dop_ref = param.find("DOP-REF")
    if dop_ref is None:
        return byte_position + 1
    dop = find_by_id(root, dop_ref.get("ID-REF"))
    if dop is None:
        return byte_position + 1
    bit_lengths = dop.xpath('.//*[local-name()="BIT-LENGTH"]/text()')
    if not bit_lengths:
        return byte_position + 1
    try:
        bit_length = int(bit_lengths[0])
    except ValueError:
        return byte_position + 1
    return byte_position + max(1, (bit_length + 7) // 8)


def update_extended_record_number_dop(
    root: etree._Element,
    short_name: str,
    records: list[ExtendedRecordDef],
    *,
    include_all: bool,
) -> None:
    dop = BASE.first_by_short(root, "DATA-OBJECT-PROP", short_name)
    if dop is None:
        return
    scales = dop.find(".//COMPU-SCALES")
    if scales is None:
        return
    for child in list(scales):
        scales.remove(child)
    for record in records:
        scale = BASE.sub(scales, "COMPU-SCALE")
        BASE.sub(scale, "LOWER-LIMIT", record.record_num)
        BASE.sub(scale, "UPPER-LIMIT", record.record_num)
        const = BASE.sub(scale, "COMPU-CONST")
        BASE.sub(const, "VT", extended_record_label(record))
    if include_all:
        scale = BASE.sub(scales, "COMPU-SCALE")
        BASE.sub(scale, "LOWER-LIMIT", 255)
        BASE.sub(scale, "UPPER-LIMIT", 255)
        const = BASE.sub(scale, "COMPU-CONST")
        BASE.sub(const, "VT", "All")
    internal_constr = dop.find("INTERNAL-CONSTR")
    if internal_constr is not None:
        dop.remove(internal_constr)


def update_snapshot_record_number_dop(root: etree._Element, short_name: str, record_names: dict[int, str], *, include_all: bool) -> None:
    dop = BASE.first_by_short(root, "DATA-OBJECT-PROP", short_name)
    if dop is None:
        return
    scales = dop.find(".//COMPU-SCALES")
    if scales is None:
        return
    for child in list(scales):
        scales.remove(child)
    for record_num in sorted(record_names):
        scale = BASE.sub(scales, "COMPU-SCALE")
        BASE.sub(scale, "LOWER-LIMIT", record_num)
        BASE.sub(scale, "UPPER-LIMIT", record_num)
        const = BASE.sub(scale, "COMPU-CONST")
        BASE.sub(const, "VT", record_names.get(record_num) or f"Snapshot Record 0x{record_num:02X}")
    if include_all:
        scale = BASE.sub(scales, "COMPU-SCALE")
        BASE.sub(scale, "LOWER-LIMIT", 255)
        BASE.sub(scale, "UPPER-LIMIT", 255)
        const = BASE.sub(scale, "COMPU-CONST")
        BASE.sub(const, "VT", "All")
    internal_constr = dop.find("INTERNAL-CONSTR")
    if internal_constr is not None:
        dop.remove(internal_constr)


def update_base_variant_comparams(root: etree._Element, cover: CoverInfo) -> None:
    base_variant = get_base_variant(root)
    if base_variant is None:
        return
    base_short_name = base_variant.findtext("SHORT-NAME") or "SERES_ECU_CAN"
    comp_refs = base_variant.find("COMPARAM-REFS")
    if comp_refs is None:
        comp_refs = BASE.sub(base_variant, "COMPARAM-REFS")
    values = getattr(cover, "comm_params", {})
    for ref in comp_refs.findall("COMPARAM-REF"):
        id_ref = ref.get("ID-REF")
        if id_ref not in values:
            continue
        simple = ref.find("SIMPLE-VALUE")
        if simple is not None:
            simple.text = str(values[id_ref])
    unique_ref = None
    for ref in comp_refs.findall("COMPARAM-REF"):
        if ref.get("ID-REF") == "ISO_15765_2.CP_UniqueRespIdTable":
            unique_ref = ref
            break
    if unique_ref is None:
        unique_ref = BASE.sub(
            comp_refs,
            "COMPARAM-REF",
            attrib={"ID-REF": "ISO_15765_2.CP_UniqueRespIdTable", "DOCREF": "ISO_15765_2", "DOCTYPE": "COMPARAM-SUBSET"},
        )
    for child in list(unique_ref):
        unique_ref.remove(child)
    rx_phy = cover.rx_phy_id if cover.rx_phy_id is not None else 0x705
    tx = cover.tx_id if cover.tx_id is not None else 0x785
    rx_fun = cover.rx_fun_id if cover.rx_fun_id is not None else 0x7DF
    uudt_resp_id = UUDT_DISABLED_CAN_ID
    complex_value = BASE.sub(unique_ref, "COMPLEX-VALUE")
    for value in (
        "0",
        "normal segmented 11-bit transmit with FC",
        str(rx_phy),
        "0",
        "normal segmented 11-bit receive with FC",
        str(tx),
        "0",
        "normal unsegmented 11-bit receive",
        str(uudt_resp_id),
        base_short_name,
    ):
        BASE.sub(complex_value, "SIMPLE-VALUE", value)
    BASE.sub(unique_ref, "PROTOCOL-SNREF", attrib={"SHORT-NAME": "CAN"})


def update_session_timing(root: etree._Element, cover: CoverInfo) -> None:
    timing = getattr(cover, "session_timing", {"P2": 50, "P2Ex": 5000})
    text = f"{{P2={timing.get('P2', 50)}, P2Ex={timing.get('P2Ex', 5000)}}}"
    for state in root.xpath('//*[local-name()="STATE" and (SHORT-NAME="Default" or SHORT-NAME="Programming" or SHORT-NAME="Extended")]'):
        desc = state.find("DESC")
        if desc is None:
            desc = BASE.sub(state, "DESC")
        for child in list(desc):
            desc.remove(child)
        BASE.sub(desc, "p", text)


def validate_with_odxtools(pdx_path: Path) -> None:
    cmd = [sys.executable, "-m", "odxtools", "list", str(pdx_path), "--all"]
    result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
    if result.returncode != 0:
        details = (result.stderr or result.stdout).strip()
        raise RuntimeError(f"odxtools validation failed for {pdx_path}:\n{details}")


def find_default_xlsx(base_dir: Path) -> Path:
    candidates = sorted(path for path in base_dir.glob("*.xlsx") if not path.name.startswith("~$"))
    if not candidates:
        raise FileNotFoundError("No .xlsx survey file found in the current directory")
    return candidates[0]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Generate SERES PDX from a SERES diagnosis survey Excel file.")
    parser.add_argument("xlsx", nargs="?", type=Path, help="Input SERES diagnosis survey .xlsx file")
    parser.add_argument("--template", type=Path, default=SCRIPT_DIR / "templates" / "SERES_ECU_CAN_v15.pdx", help="Template PDX")
    parser.add_argument("--output-dir", type=Path, default=SCRIPT_DIR / "output", help="Output directory")
    parser.add_argument("--no-validate", action="store_true", help="Skip odxtools validation")
    args = parser.parse_args(argv)

    xlsx_path = args.xlsx or find_default_xlsx(SCRIPT_DIR)
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)
    if not args.template.exists():
        raise FileNotFoundError(args.template)

    survey = parse_seres_survey(xlsx_path)
    output_pdx = args.output_dir / f"{xlsx_path.stem}.pdx"
    update_template(args.template, output_pdx, survey, validate=not args.no_validate)
    did_data_objects = sum(len(did.params) for did in survey.dids)
    converted_did_objects = sum(
        1
        for did in survey.dids
        for param in did.params
        if getattr(param.conversion, "kind", "identity") in {"enum", "linear"}
    )
    print(f"Generated: {output_pdx}")
    print(
        "Parsed: "
        f"{len(survey.dids)} DID identifiers, "
        f"{did_data_objects} DID data objects, "
        f"{converted_did_objects} converted DID data objects, "
        f"{len(survey.io_dids)} IO DIDs, "
        f"{len(survey.routines)} routines, "
        f"{len(survey.dtcs)} DTCs, "
        f"{len(survey.snapshots)} snapshot DIDs, "
        f"{len(getattr(survey, 'snapshot_record_nums', []))} snapshot records, "
        f"{len(survey.extended_records)} extended records"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

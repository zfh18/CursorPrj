#!/usr/bin/env python3
"""Generate a CANdelaStudio-compatible VF PDX from a VF diagnostic survey workbook.

This script is self-contained inside the VF project. It parses the VF Excel
survey schema into the canonical diagnostic model and updates the VF
CANdelaStudio 15 PDX template with DID, IOControl, RoutineControl, DTC,
Snapshot, ExtendedData, and communication-parameter content.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import re
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Iterable
from lxml import etree
from openpyxl import load_workbook
from dataclasses import dataclass, field


#!/usr/bin/env python3
"""Shared VF ODX/PDX writer helpers.

This module intentionally has no command-line entry point. ``pdxGen_VF.py``
parses the VF workbook schema and calls these dataclasses/XML helpers to update
the CANdelaStudio 15 PDX template.
"""


XSI_NS = "http://www.w3.org/2001/XMLSchema-instance"
XSI_TYPE = f"{{{XSI_NS}}}type"
etree.register_namespace("xsi", XSI_NS)
SCRIPT_DIR = Path(__file__).resolve().parent

@dataclass
class Conversion:
    kind: str = "identity"
    enum: list[tuple[int, int, str]] = field(default_factory=list)
    a: float = 1.0
    b: float = 0.0
    precision: int | None = None


@dataclass
class ParamDef:
    name: str
    long_name: str
    byte_pos: int
    bit_pos: int
    bit_len: int
    data_type: str = "Hex(Unsigned)"
    unit: str = ""
    conversion: Conversion = field(default_factory=Conversion)
    min_value: str = ""
    max_value: str = ""
    dop_id: str = ""


@dataclass
class DidDef:
    did: int
    desc: str
    size: int
    params: list[ParamDef] = field(default_factory=list)
    write_security: str = "N"
    sessions: list[str] = field(default_factory=list)
    structure_id: str = ""
    wrapper_id: str = ""
    short_name: str = ""
    long_name: str = ""

    @property
    def readable(self) -> bool:
        return any("R" in normalize_access(v) for v in self.sessions)

    @property
    def writable(self) -> bool:
        return any("W" in normalize_access(v) for v in self.sessions)


@dataclass
class IoDidDef:
    did: int
    desc: str
    size: int
    controls: set[int] = field(default_factory=set)
    params: list[ParamDef] = field(default_factory=list)
    structure_id: str = ""
    status_wrapper_id: str = ""
    control_request_wrapper_id: str = ""
    short_name: str = ""
    long_name: str = ""


@dataclass
class RoutineSubFunction:
    control_type: int
    supported: bool
    option_params: list[ParamDef] = field(default_factory=list)
    status_params: list[ParamDef] = field(default_factory=list)
    option_structure_id: str = ""
    status_structure_id: str = ""


@dataclass
class RoutineDef:
    rid: int
    desc: str
    security: str = "N"
    sessions: list[str] = field(default_factory=list)
    subfunctions: dict[int, RoutineSubFunction] = field(default_factory=dict)
    short_name: str = ""
    long_name: str = ""


@dataclass
class DtcDef:
    display_code: str
    byte_code: int
    text: str
    priority: str = ""


@dataclass
class SnapshotDef:
    record_num: int | None
    did: int
    desc: str
    size: int
    params: list[ParamDef] = field(default_factory=list)
    structure_id: str = ""


@dataclass
class ExtendedRecordDef:
    record_num: int
    desc: str
    size: int
    params: list[ParamDef] = field(default_factory=list)
    structure_id: str = ""


@dataclass
class CoverInfo:
    ecu_name: str = "VF_ECU_CAN"
    vehicle: str = ""
    supplier: str = ""
    tx_id: int | None = None
    rx_phy_id: int | None = None
    rx_fun_id: int | None = None
    bus_type: str = "CAN"


@dataclass
class SurveyData:
    cover: CoverInfo
    dids: list[DidDef]
    io_dids: list[IoDidDef]
    routines: list[RoutineDef]
    dtcs: list[DtcDef]
    snapshots: list[SnapshotDef]
    extended_records: list[ExtendedRecordDef]


def extended_record_label(record: ExtendedRecordDef) -> str:
    return compact_text(record.desc) or f"Extended Data Record 0x{record.record_num:02X}"


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip().lstrip("'")


def compact_text(value: Any) -> str:
    return re.sub(r"\s+", " ", cell_text(value).replace("\u3000", " ")).strip()


def normalize_access(value: Any) -> str:
    return compact_text(value).upper().replace(" ", "")


def has_chinese_text(value: Any) -> bool:
    return bool(re.search(r"[\u3400-\u4DBF\u4E00-\u9FFF\uF900-\uFAFF]", cell_text(value)))


def name_parts(value: Any) -> list[str]:
    text = cell_text(value).replace("\u3000", " ")
    parts = [p.strip() for p in re.split(r"[\r\n]+", text) if p.strip()]
    result: list[str] = []
    for part in parts:
        slash_parts = [piece.strip() for piece in re.split(r"\s*/\s*", part) if piece.strip()]
        if len(slash_parts) > 1 and any(has_chinese_text(piece) for piece in slash_parts):
            result.extend(slash_parts)
        else:
            result.append(part)
    return result


def canonical_name(identifier: str, display: str, fallback: str = "") -> str:
    identifier = compact_text(identifier)
    display = compact_text(display)
    if identifier and display and identifier != display:
        return f"{identifier}\n{display}"
    return display or identifier or fallback


def split_name(value: str) -> tuple[str, str]:
    text = cell_text(value).replace("\r\n", "\n").strip()
    parts = name_parts(text)
    if not parts:
        return "", ""
    english = next((p for p in parts if re.search(r"[A-Za-z]", p) and not has_chinese_text(p)), "")
    chinese = next((p for p in parts if has_chinese_text(p)), "")
    identifier = english or ("" if chinese else parts[0])
    return identifier, chinese or text


def parse_hex_cell(value: Any, *, max_value: int = 0xFFFFFFFF) -> int | None:
    text = compact_text(value).replace(" ", "")
    if not re.fullmatch(r"(?:0[xX])?[0-9A-Fa-f]{1,8}", text):
        return None
    result = int(text[2:] if text.lower().startswith("0x") else text, 16)
    if result > max_value:
        return None
    return result


def parse_int_cell(value: Any, default: int = 0) -> int:
    text = compact_text(value)
    if not text:
        return default
    match = re.search(r"-?\d+", text)
    if not match:
        return default
    return int(match.group(0))


def parse_byte_range(value: Any, default: int = 0) -> tuple[int, int]:
    text = compact_text(value)
    if not text:
        return default, default
    nums = [int(n) for n in re.findall(r"\d+", text)]
    if not nums:
        return default, default
    start = nums[0]
    end = nums[1] if len(nums) > 1 else start
    if end < start:
        end = start
    return start, end


LINEAR_NUMBER_PATTERN = r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)"


def _parse_float_assignment(text: str, key: str, default: float) -> float:
    match = re.search(rf"\b{re.escape(key)}\s*=\s*(-?\d+(?:\.\d+)?)", text, flags=re.IGNORECASE)
    return float(match.group(1)) if match else default


def parse_structured_params(value: Any, fallback_prefix: str) -> list[ParamDef]:
    text = cell_text(value)
    if not text or text.strip().upper() in {"N/A", "NA", "N"}:
        return []
    normalized = text.replace("：", ":").replace("；", ";")
    if "Name:" not in normalized:
        # Some surveys put a single conversion-like field in this cell.
        p = make_param_from_cells(
            name_value=fallback_prefix,
            byte_value="0",
            bit_value="0-7",
            data_type_value="Hex(Unsigned)",
            conversion_value=normalized,
            fallback_name=fallback_prefix,
        )
        return [p] if p else []

    params: list[ParamDef] = []
    chunks = re.split(r"(?=Name\s*:)", normalized, flags=re.IGNORECASE)
    for idx, chunk in enumerate(chunks):
        if "Name" not in chunk:
            continue
        name_match = re.search(r"Name\s*:\s*([^,\n;]+)", chunk, flags=re.IGNORECASE)
        byte_match = re.search(r"Byte\s*:\s*([^,\n;]+)", chunk, flags=re.IGNORECASE)
        bit_match = re.search(r"Bit\s*:\s*([^,\n;]+)", chunk, flags=re.IGNORECASE)
        conv_match = re.search(r"Conversion\s*:\s*(.*)", chunk, flags=re.IGNORECASE | re.DOTALL)
        name = name_match.group(1).strip() if name_match else f"{fallback_prefix}_{idx + 1}"
        byte = byte_match.group(1).strip() if byte_match else "0"
        bit = bit_match.group(1).strip() if bit_match else "0-7"
        conv = conv_match.group(1).strip() if conv_match else ""
        param = make_param_from_cells(
            name_value=name,
            byte_value=byte,
            bit_value=bit,
            data_type_value="Hex(Unsigned)",
            conversion_value=conv,
            fallback_name=f"{fallback_prefix}_{idx + 1}",
        )
        if param:
            params.append(param)
    return params


def sanitize_short_name(value: str, fallback: str, used: set[str] | None = None, max_len: int = 120) -> str:
    base = value.strip() if value else fallback
    base = base.encode("ascii", "ignore").decode("ascii")
    base = re.sub(r"[^A-Za-z0-9_]+", "_", base).strip("_")
    if not base:
        base = fallback
    if not re.match(r"[A-Za-z_]", base):
        base = f"X_{base}"
    base = base[:max_len].rstrip("_") or fallback
    if used is None:
        return base
    candidate = base
    index = 2
    while candidate in used:
        suffix = f"_{index}"
        candidate = f"{base[: max_len - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def hex_short(prefix: str, value: int, width: int = 4) -> str:
    return f"{prefix}_0x{value:0{width}X}"


def find_sheet(workbook: Any, predicate: Any) -> Any | None:
    for sheet_name in workbook.sheetnames:
        if predicate(sheet_name.strip()):
            return workbook[sheet_name]
    return None


class IdGenerator:
    def __init__(self, root: etree._Element) -> None:
        self.used = {node.get("ID") for node in root.xpath("//*[@ID]") if node.get("ID")}
        self.index = 1

    def new(self, prefix: str = "VF") -> str:
        while True:
            candidate = f"_{prefix}_{self.index:06d}"
            self.index += 1
            if candidate not in self.used:
                self.used.add(candidate)
                return candidate


def element(tag: str, text: Any | None = None, attrib: dict[str, str] | None = None) -> etree._Element:
    node = etree.Element(tag, attrib=attrib or {})
    if text is not None:
        node.text = str(text)
    return node


def sub(parent: etree._Element, tag: str, text: Any | None = None, attrib: dict[str, str] | None = None) -> etree._Element:
    child = element(tag, text, attrib)
    parent.append(child)
    return child


def set_xsi_type(node: etree._Element, value: str) -> None:
    node.set(XSI_TYPE, value)


def first_by_short(root: etree._Element, tag: str, short_name: str) -> etree._Element | None:
    nodes = root.xpath(f'//*[local-name()="{tag}" and SHORT-NAME=$name]', name=short_name)
    return nodes[0] if nodes else None


def child_text(node: etree._Element, tag: str) -> str:
    value = node.findtext(tag)
    return value or ""


def clear_children(parent: etree._Element, tag: str) -> None:
    for child in list(parent):
        if child.tag == tag:
            parent.remove(child)


def replace_child(parent: etree._Element, old_tag: str, new_child: etree._Element, before_tags: set[str] | None = None) -> None:
    old = parent.find(old_tag)
    if old is not None:
        index = parent.index(old)
        parent.remove(old)
        parent.insert(index, new_child)
        return
    if before_tags:
        for index, child in enumerate(parent):
            if child.tag in before_tags:
                parent.insert(index, new_child)
                return
    parent.append(new_child)


def xml_local_name(node: etree._Element) -> str:
    return etree.QName(node).localname if isinstance(node.tag, str) else ""


def child_text_by_local_name(parent: etree._Element, name: str) -> str | None:
    matches = parent.xpath("./*[local-name()=$name]", name=name)
    if not matches:
        return None
    return matches[0].text


def validate_can_dela_odx_structure(root: etree._Element, source_name: str) -> None:
    """Catch ODX structures that CANdelaStudio rejects before packaging.

    odxtools is useful as a parser-level smoke test, but CANdelaStudio also
    enforces the ODX 2.2.0 content model strictly. TABLE is the most sensitive
    object for this generator because unsupported template service variants can
    otherwise leave empty tables behind.
    """

    errors: list[str] = []
    table_order = {
        "SHORT-NAME": 0,
        "LONG-NAME": 1,
        "DESC": 2,
        "KEY-LABEL": 3,
        "STRUCT-LABEL": 4,
        "ADMIN-DATA": 5,
        "KEY-DOP-REF": 6,
        "TABLE-ROW-REF": 7,
        "TABLE-ROW": 7,
        "TABLE-DIAG-COMM-CONNECTORS": 8,
        "SDGS": 9,
    }

    seen_ids: dict[str, int] = {}
    duplicate_ids: list[tuple[str, int, int]] = []
    for node in root.xpath("//*[@ID]"):
        node_id = node.get("ID")
        if not node_id:
            continue
        if node_id in seen_ids:
            duplicate_ids.append((node_id, seen_ids[node_id], node.sourceline or 0))
        else:
            seen_ids[node_id] = node.sourceline or 0
    for node_id, first_line, second_line in duplicate_ids[:20]:
        errors.append(f"duplicate ID '{node_id}' at line {second_line}, first seen at line {first_line}")

    for ref_node in root.xpath("//*[@ID-REF]"):
        id_ref = ref_node.get("ID-REF")
        if not id_ref or ref_node.get("DOCREF"):
            continue
        if id_ref not in seen_ids:
            errors.append(
                f"dangling local ID-REF '{id_ref}' at line {ref_node.sourceline or '?'} "
                f"({xml_local_name(ref_node)})"
            )

    for table in root.xpath("//*[local-name()='TABLE']"):
        short_name = child_text_by_local_name(table, "SHORT-NAME") or table.get("ID") or "<unnamed>"
        children = [xml_local_name(child) for child in table if isinstance(child.tag, str)]
        has_row = any(child_name in {"TABLE-ROW", "TABLE-ROW-REF"} for child_name in children)
        if not has_row:
            errors.append(
                f"TABLE '{short_name}' at line {table.sourceline or '?'} has no TABLE-ROW/TABLE-ROW-REF"
            )

        if not children or children[0] != "SHORT-NAME":
            errors.append(f"TABLE '{short_name}' at line {table.sourceline or '?'} does not start with SHORT-NAME")
            continue

        last_index = -1
        for child_name in children:
            order_index = table_order.get(child_name)
            if order_index is None:
                errors.append(
                    f"TABLE '{short_name}' at line {table.sourceline or '?'} has unexpected child '{child_name}'"
                )
                continue
            if order_index < last_index:
                errors.append(
                    f"TABLE '{short_name}' at line {table.sourceline or '?'} has '{child_name}' out of ODX order"
                )
                break
            last_index = order_index

    if errors:
        details = "\n".join(f"- {message}" for message in errors[:60])
        if len(errors) > 60:
            details += f"\n- ... {len(errors) - 60} more issue(s)"
        raise RuntimeError(f"Generated {source_name} is not CANdelaStudio-compatible:\n{details}")


def patch_pdx_catalog_for_can_only(index_path: Path, keep_files: set[str]) -> None:
    if not index_path.exists():
        return
    parser = etree.XMLParser(remove_blank_text=False)
    tree = etree.parse(str(index_path), parser)
    root = tree.getroot()
    for file_node in list(root.xpath("//*[local-name()='FILE']")):
        file_name = compact_text(file_node.text)
        if file_name not in keep_files:
            parent = file_node.getparent()
            if parent is not None:
                parent.remove(file_node)
    tree.write(str(index_path), encoding="UTF-8", xml_declaration=False, pretty_print=True, standalone=False)


def prefix_doc_revision_labels(root: etree._Element) -> None:
    for label in root.xpath("//*[local-name()='DOC-REVISION']/*[local-name()='REVISION-LABEL']"):
        value = compact_text(label.text)
        if value and not value.startswith("PDX_"):
            label.text = f"PDX_{value}"


def ensure_units(ddds: etree._Element, survey: SurveyData) -> dict[str, str]:
    unit_spec = ddds.find("UNIT-SPEC")
    if unit_spec is None:
        unit_spec = sub(ddds, "UNIT-SPEC")
    units_node = unit_spec.find("UNITS")
    if units_node is None:
        units_node = sub(unit_spec, "UNITS")

    existing = {
        compact_text(unit.findtext("DISPLAY-NAME") or unit.findtext("SHORT-NAME")): unit.get("ID")
        for unit in units_node.findall("UNIT")
        if unit.get("ID")
    }
    all_params: list[ParamDef] = []
    for item in [*survey.dids, *survey.io_dids, *survey.snapshots, *survey.extended_records]:
        all_params.extend(item.params)
    for routine in survey.routines:
        for subfn in routine.subfunctions.values():
            all_params.extend(subfn.option_params)
            all_params.extend(subfn.status_params)

    for param in all_params:
        unit = compact_text(param.unit)
        if not unit or unit in {"-", "/", "N/A", "NA"} or unit in existing:
            continue
        unit_id = f"_UNIT_{sanitize_short_name(unit, 'UNIT')}"
        index = 2
        while unit_id in existing.values():
            unit_id = f"_UNIT_{sanitize_short_name(unit, 'UNIT')}_{index}"
            index += 1
        unit_node = sub(units_node, "UNIT", attrib={"ID": unit_id})
        sub(unit_node, "SHORT-NAME", sanitize_short_name(unit, "UNIT"))
        sub(unit_node, "DISPLAY-NAME", unit)
        existing[unit] = unit_id
    return {k: v for k, v in existing.items() if v}


def prepare_data_structure(
    *,
    id_gen: IdGenerator,
    data_object_props: etree._Element,
    structures: etree._Element,
    item: DidDef | IoDidDef,
    prefix: str,
    unit_ids: dict[str, str],
    generated_dop_cache: dict[tuple[str, int, str, str, str], str],
    direct_single_param: bool = False,
) -> None:
    used_param_names: set[str] = set()
    english, long_name = split_name(item.desc)
    item.short_name = sanitize_short_name(english or hex_short(prefix, item.did), hex_short(prefix, item.did))
    item.long_name = long_name or item.short_name
    for param in item.params:
        param.name = sanitize_short_name(param.name, "Data", used_param_names)
        param.dop_id = ensure_param_dop(
            id_gen,
            data_object_props,
            param,
            unit_ids,
            generated_dop_cache,
            f"DOP_{item.short_name}_{param.name}",
        )
    if not item.params:
        raw = ParamDef(
            name="Data",
            long_name="Data",
            byte_pos=0,
            bit_pos=0,
            bit_len=max(8, (item.size or 1) * 8),
            data_type="Hex(Unsigned)",
        )
        raw.name = sanitize_short_name(raw.name, "Data", used_param_names)
        raw.dop_id = ensure_param_dop(
            id_gen, data_object_props, raw, unit_ids, generated_dop_cache, f"DOP_{item.short_name}_Data"
        )
        item.params.append(raw)

    direct_param = direct_did_payload_param(item) if direct_single_param and isinstance(item, DidDef) else None
    if direct_param is not None:
        item.structure_id = ""
        return

    item.structure_id = id_gen.new("DIDSTR")
    structure = sub(structures, "STRUCTURE", attrib={"ID": item.structure_id})
    sub(structure, "SHORT-NAME", item.short_name)
    sub(structure, "LONG-NAME", item.long_name)
    if item.size:
        sub(structure, "BYTE-SIZE", item.size)
    params_node = sub(structure, "PARAMS")
    for param in item.params:
        params_node.append(make_value_param(param.name, param.long_name, param.byte_pos, param.bit_pos, param.dop_id))


def make_param_structure(
    id_gen: IdGenerator,
    data_object_props: etree._Element,
    structures: etree._Element,
    short_name: str,
    long_name: str,
    params: list[ParamDef],
    unit_ids: dict[str, str],
    generated_dop_cache: dict[tuple[str, int, str, str, str], str],
    byte_size: int | None = None,
) -> str:
    used_names: set[str] = set()
    structure_id = id_gen.new("STR")
    structure = sub(structures, "STRUCTURE", attrib={"ID": structure_id})
    short_name = sanitize_short_name(short_name, "Structure")
    sub(structure, "SHORT-NAME", short_name)
    sub(structure, "LONG-NAME", long_name or short_name)
    if byte_size:
        sub(structure, "BYTE-SIZE", byte_size)
    params_node = sub(structure, "PARAMS")
    for param in params:
        param.name = sanitize_short_name(param.name, "Data", used_names)
        param.dop_id = ensure_param_dop(
            id_gen, data_object_props, param, unit_ids, generated_dop_cache, f"DOP_{short_name}_{param.name}"
        )
        params_node.append(make_value_param(param.name, param.long_name, param.byte_pos, param.bit_pos, param.dop_id))
    return structure_id


def make_wrapper_structure(
    id_gen: IdGenerator,
    structures: etree._Element,
    short_name: str,
    long_name: str,
    param_short_name: str,
    param_long_name: str,
    dop_or_structure_id: str,
) -> str:
    wrapper_id = id_gen.new("WRAP")
    wrapper = sub(structures, "STRUCTURE", attrib={"ID": wrapper_id})
    sub(wrapper, "SHORT-NAME", sanitize_short_name(short_name, "Wrapper"))
    sub(wrapper, "LONG-NAME", long_name)
    params_node = sub(wrapper, "PARAMS")
    param = make_value_param(sanitize_short_name(param_short_name, "Data"), param_long_name, 0, 0, dop_or_structure_id)
    params_node.append(param)
    return wrapper_id


def direct_did_payload_param(did: DidDef) -> ParamDef | None:
    if len(did.params) != 1:
        return None
    param = did.params[0]
    if param.byte_pos != 0 or param.bit_pos != 0 or not param.dop_id:
        return None
    expected_size = did.size if did.size > 0 else vf_param_size(did.params)
    expected_bit_len = max(8, expected_size * 8)
    return param if param.bit_len == expected_bit_len else None


def ensure_param_dop(
    id_gen: IdGenerator,
    data_object_props: etree._Element,
    param: ParamDef,
    unit_ids: dict[str, str],
    cache: dict[tuple[str, int, str, str, str], str],
    preferred_short_name: str,
) -> str:
    conv_key = conversion_cache_key(param.conversion)
    key = (param.data_type.upper(), param.bit_len, conv_key, compact_text(param.unit), param.name)
    if key in cache:
        return cache[key]
    dop_id = id_gen.new("DOP")
    dop = make_data_object_prop(dop_id, preferred_short_name, param, unit_ids)
    data_object_props.append(dop)
    cache[key] = dop_id
    return dop_id


def conversion_cache_key(conversion: Conversion) -> str:
    if conversion.kind == "enum":
        return "enum:" + "|".join(f"{lo}-{hi}:{label}" for lo, hi, label in conversion.enum)
    if conversion.kind == "linear":
        return f"linear:{conversion.a}:{conversion.b}:{conversion.precision}"
    if conversion.kind == "bcd":
        return "bcd"
    return "identity"


def make_data_object_prop(dop_id: str, preferred_short_name: str, param: ParamDef, unit_ids: dict[str, str]) -> etree._Element:
    bit_len = max(1, int(param.bit_len or 8))
    dop = element("DATA-OBJECT-PROP", attrib={"ID": dop_id})
    short_name = sanitize_short_name(preferred_short_name, "DOP")
    sub(dop, "SHORT-NAME", short_name)
    sub(dop, "LONG-NAME", param.long_name or short_name)

    conversion = param.conversion
    if conversion.kind == "enum" and conversion.enum:
        compu = sub(dop, "COMPU-METHOD")
        sub(compu, "CATEGORY", "TEXTTABLE")
        internal_to_phys = sub(compu, "COMPU-INTERNAL-TO-PHYS")
        scales = sub(internal_to_phys, "COMPU-SCALES")
        for lo, hi, label in conversion.enum:
            scale = sub(scales, "COMPU-SCALE")
            sub(scale, "LOWER-LIMIT", lo)
            sub(scale, "UPPER-LIMIT", hi)
            const = sub(scale, "COMPU-CONST")
            sub(const, "VT", label)
        coded = sub(
            dop,
            "DIAG-CODED-TYPE",
            attrib={"BASE-TYPE-ENCODING": "NONE", "BASE-DATA-TYPE": "A_UINT32"},
        )
        set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
        sub(coded, "BIT-LENGTH", bit_len)
        sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UNICODE2STRING"})
    elif conversion.kind == "linear":
        compu = sub(dop, "COMPU-METHOD")
        sub(compu, "CATEGORY", "LINEAR")
        internal_to_phys = sub(compu, "COMPU-INTERNAL-TO-PHYS")
        scales = sub(internal_to_phys, "COMPU-SCALES")
        scale = sub(scales, "COMPU-SCALE")
        coeffs = sub(scale, "COMPU-RATIONAL-COEFFS")
        numerator = sub(coeffs, "COMPU-NUMERATOR")
        sub(numerator, "V", clean_float(conversion.b))
        sub(numerator, "V", clean_float(conversion.a))
        coded = sub(
            dop,
            "DIAG-CODED-TYPE",
            attrib={"BASE-TYPE-ENCODING": "NONE", "BASE-DATA-TYPE": "A_UINT32"},
        )
        set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
        sub(coded, "BIT-LENGTH", min(bit_len, 32))
        physical = sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_FLOAT64"})
        if conversion.precision is not None:
            sub(physical, "PRECISION", conversion.precision)
    elif conversion.kind == "bcd":
        compu = sub(dop, "COMPU-METHOD")
        sub(compu, "CATEGORY", "IDENTICAL")
        coded = sub(
            dop,
            "DIAG-CODED-TYPE",
            attrib={"BASE-TYPE-ENCODING": "BCD-P", "BASE-DATA-TYPE": "A_UINT32"},
        )
        set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
        sub(coded, "BIT-LENGTH", min(bit_len, 32))
        sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32", "DISPLAY-RADIX": "DEC"})
    else:
        compu = sub(dop, "COMPU-METHOD")
        sub(compu, "CATEGORY", "IDENTICAL")
        if "ASCII" in param.data_type.upper():
            coded = sub(
                dop,
                "DIAG-CODED-TYPE",
                attrib={"BASE-TYPE-ENCODING": "ISO-8859-1", "BASE-DATA-TYPE": "A_ASCIISTRING"},
            )
            set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
            sub(coded, "BIT-LENGTH", bit_len)
            sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UNICODE2STRING"})
        elif bit_len <= 32:
            coded = sub(
                dop,
                "DIAG-CODED-TYPE",
                attrib={"BASE-TYPE-ENCODING": "NONE", "BASE-DATA-TYPE": "A_UINT32"},
            )
            set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
            sub(coded, "BIT-LENGTH", bit_len)
            sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32", "DISPLAY-RADIX": "HEX"})
        else:
            coded = sub(dop, "DIAG-CODED-TYPE", attrib={"BASE-DATA-TYPE": "A_BYTEFIELD"})
            set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
            sub(coded, "BIT-LENGTH", bit_len)
            sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_BYTEFIELD"})

    unit = compact_text(param.unit)
    if unit and unit in unit_ids and conversion.kind != "enum":
        sub(dop, "UNIT-REF", attrib={"ID-REF": unit_ids[unit]})
    return dop


def clean_float(value: float) -> str:
    return str(int(value)) if float(value).is_integer() else str(value)


def make_value_param(short_name: str, long_name: str, byte_position: int, bit_position: int, dop_id: str) -> etree._Element:
    param = element("PARAM", attrib={"SEMANTIC": "DATA"})
    set_xsi_type(param, "VALUE")
    sub(param, "SHORT-NAME", short_name)
    sub(param, "LONG-NAME", long_name or short_name)
    sub(param, "BYTE-POSITION", byte_position)
    if bit_position:
        sub(param, "BIT-POSITION", bit_position)
    sub(param, "DOP-REF", attrib={"ID-REF": dop_id})
    return param


def update_did_tables_and_services(root: etree._Element, id_gen: IdGenerator, dids: list[DidDef]) -> None:
    read_table = first_by_short(root, "TABLE", "Identification_Read_PR")
    write_table = first_by_short(root, "TABLE", "Identification_Write_RQ")
    if read_table is None or write_table is None:
        raise RuntimeError("Template identification tables were not found")
    clear_children(read_table, "TABLE-ROW")
    clear_children(write_table, "TABLE-ROW")

    readable_instances: list[tuple[int, str, str]] = []
    writable_instances: list[tuple[int, str, str]] = []
    for did in dids:
        if did.readable:
            append_table_row(read_table, id_gen, did.short_name, did.long_name, did.did, did.wrapper_id)
            readable_instances.append((did.did, did.short_name, did.long_name))
        if did.writable:
            append_table_row(write_table, id_gen, did.short_name, did.long_name, did.did, did.wrapper_id)
            writable_instances.append((did.did, did.short_name, did.long_name))

    update_service_instances(root, id_gen, "Identification_Read", "Read", "Read", readable_instances)
    update_service_instances(root, id_gen, "Identification_Write", "Write", "Write", writable_instances)
    update_preconditions_for_values(root, "Identification_Read", readable_instances, "RecordDataIdentifier", did_state_refs(dids, "read"))
    update_preconditions_for_values(root, "Identification_Write", writable_instances, "RecordDataIdentifier", did_state_refs(dids, "write"))

    eop_read_rq = first_by_short(root, "END-OF-PDU-FIELD", "EOP_DIDs_Identification_ReadRDBI_RQ")
    eop_read_pr = first_by_short(root, "END-OF-PDU-FIELD", "EOP_DIDs_Identification_ReadRDBI_PR")
    for eop in (eop_read_rq, eop_read_pr):
        if eop is not None:
            max_items = eop.find("MAX-NUMBER-OF-ITEMS")
            if max_items is not None:
                max_items.text = str(max(1, len(readable_instances)))


def did_state_refs(dids: list[DidDef], mode: str) -> dict[int, list[str]]:
    result: dict[int, list[str]] = {}
    for did in dids:
        if mode == "read":
            if not did.readable:
                continue
            states = sessions_to_state_ids(did.sessions, mode="read")
            states.extend(["_6", "_7", "_8", "_10", "_11", "_12", "_13", "_14", "_15"])
        else:
            if not did.writable:
                continue
            states = sessions_to_state_ids(did.sessions, mode="write")
            states.extend(security_to_state_ids(did.write_security))
            states.extend(["_10", "_11", "_12", "_13", "_14", "_15"])
        result[did.did] = unique_list(states)
    return result


def sessions_to_state_ids(sessions: list[str], mode: str) -> list[str]:
    wants = "R" if mode == "read" else "W"
    mapping = ["_2", "_4", "_2", "_3", "_4"]
    states: list[str] = []
    for idx, value in enumerate(sessions[:5]):
        access = normalize_access(value)
        if wants in access:
            states.append(mapping[idx])
    if not states:
        states = ["_2", "_3", "_4"]
    return unique_list(states)


def security_to_state_ids(level: str) -> list[str]:
    normalized = compact_text(level).upper()
    if "FBL" in normalized:
        return ["_8"]
    if "LEVEL1" in normalized or normalized in {"1", "L1"}:
        return ["_7"]
    return ["_6", "_7", "_8"]


def unique_list(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def append_table_row(
    table: etree._Element,
    id_gen: IdGenerator,
    short_name: str,
    long_name: str,
    key: int | str,
    structure_ref: str | None = None,
) -> etree._Element:
    row = sub(table, "TABLE-ROW", attrib={"ID": id_gen.new("ROW")})
    sub(row, "SHORT-NAME", sanitize_short_name(short_name, "Row"))
    sub(row, "LONG-NAME", long_name or short_name)
    sub(row, "KEY", key)
    if structure_ref:
        sub(row, "STRUCTURE-REF", attrib={"ID-REF": structure_ref})
    return row


def update_service_instances(
    root: etree._Element,
    id_gen: IdGenerator,
    service_short_name: str,
    qualifier: str,
    service_name: str,
    instances: list[tuple[int, str, str]],
) -> None:
    service = first_by_short(root, "DIAG-SERVICE", service_short_name)
    if service is None:
        return
    sdgs = element("SDGS")
    sdg = sub(sdgs, "SDG")
    caption = sub(sdg, "SDG-CAPTION", attrib={"ID": id_gen.new("CAP")})
    sub(caption, "SHORT-NAME", "CANdelaServiceInformation")
    sub(sdg, "SD", qualifier, attrib={"SI": "ServiceQualifier"})
    sub(sdg, "SD", service_name, attrib={"SI": "ServiceName"})
    sub(sdg, "SD", "no", attrib={"SI": "PositiveResponseSuppressed"})
    for value, inst_qualifier, inst_name in instances:
        inst = sub(sdg, "SDG")
        sub(inst, "SD", value, attrib={"SI": "DiagInstanceStaticValue"})
        sub(inst, "SD", sanitize_short_name(inst_qualifier, "Instance"), attrib={"SI": "DiagInstanceQualifier"})
        sub(inst, "SD", inst_name or inst_qualifier, attrib={"SI": "DiagInstanceName"})
    replace_child(service, "SDGS", sdgs, before_tags={"FUNCT-CLASS-REFS", "AUDIENCE", "PRE-CONDITION-STATE-REFS", "REQUEST-REF"})


def update_preconditions_for_values(
    root: etree._Element,
    service_short_name: str,
    instances: list[tuple[int, str, str]],
    in_param: str,
    state_refs_by_value: dict[int, list[str]],
) -> None:
    service = first_by_short(root, "DIAG-SERVICE", service_short_name)
    if service is None:
        return
    pc = element("PRE-CONDITION-STATE-REFS")
    for value, _, _ in instances:
        for state_id in state_refs_by_value.get(value, ["_2", "_3", "_4"]):
            ref = sub(pc, "PRE-CONDITION-STATE-REF", attrib={"ID-REF": state_id})
            sub(ref, "VALUE", value)
            sub(ref, "IN-PARAM-IF-SNREF", attrib={"SHORT-NAME": in_param})
    replace_child(service, "PRE-CONDITION-STATE-REFS", pc, before_tags={"REQUEST-REF"})


def update_flat_service_preconditions(root: etree._Element) -> None:
    all_sessions = ["_2", "_3", "_4"]
    all_security = ["_6", "_7", "_8"]
    all_authorization = ["_10", "_11", "_12", "_13", "_14", "_15"]
    default_open = [*all_sessions, *all_security, *all_authorization]
    extended_level1 = ["_4", "_7", *all_authorization]
    programming_fbl = ["_3", "_8", *all_authorization]

    for service_short_name in (
        "DefaultSession_Start",
        "ProgrammingSession_Start",
        "ExtendedDiagnosticSession_Start",
        "TesterPresent_Send",
    ):
        update_flat_preconditions(root, service_short_name, default_open)

    for service_short_name in ("Hard_Reset_Reset", "Soft_Reset_Reset"):
        update_flat_preconditions(root, service_short_name, [*all_sessions, *all_authorization])

    for service_short_name in (
        "SeedLevel1_Request",
        "KeyLevel1_Send",
        "EnableRxAndEnableTx_Control",
        "DisableRxAndDisableTx_Control",
        "ControlDTCSetting_On",
        "ControlDTCSetting_Off",
        "FaultMemory_Clear",
    ):
        update_flat_preconditions(root, service_short_name, extended_level1)

    for service_short_name in ("RequestSeedOfSecurityLevelFBL_Request", "SendKeyOfSecurityLevelFBL_Send"):
        update_flat_preconditions(root, service_short_name, programming_fbl)

    for service_short_name in (
        "FaultMemory_ReadNumber",
        "FaultMemory_ReadAllIdentified",
        "FaultMemory_Read_extended_data_record",
        "FaultMemory_Read_snapshot_record",
        "FaultMemory_ReadAllSupported",
    ):
        update_flat_preconditions(root, service_short_name, default_open)

    for service_short_name in (
        "Software_Update_Transmit",
        "Software_Update_Stop",
    ):
        update_flat_preconditions(root, service_short_name, programming_fbl)


def remove_service_and_messages(root: etree._Element, service_short_name: str) -> None:
    service = first_by_short(root, "DIAG-SERVICE", service_short_name)
    if service is None:
        return
    ids_to_remove: set[str] = set()
    request_ref = service.find("REQUEST-REF")
    if request_ref is not None and request_ref.get("ID-REF"):
        ids_to_remove.add(request_ref.get("ID-REF"))
    ids_to_remove.update(
        ref.get("ID-REF")
        for ref in service.xpath('./*[local-name()="POS-RESPONSE-REFS"]/*[local-name()="POS-RESPONSE-REF"]')
        if ref.get("ID-REF")
    )
    ids_to_remove.update(
        ref.get("ID-REF")
        for ref in service.xpath('./*[local-name()="NEG-RESPONSE-REFS"]/*[local-name()="NEG-RESPONSE-REF"]')
        if ref.get("ID-REF")
    )

    parent = service.getparent()
    if parent is not None:
        parent.remove(service)
    for element_id in ids_to_remove:
        for node in root.xpath(f'//*[@ID="{element_id}"]'):
            node_parent = node.getparent()
            if node_parent is not None:
                node_parent.remove(node)


def remove_elements_by_short_names(root: etree._Element, tag: str, short_names: set[str]) -> None:
    for short_name in short_names:
        for node in root.xpath(f'//*[local-name()="{tag}" and SHORT-NAME=$name]', name=short_name):
            parent = node.getparent()
            if parent is not None:
                parent.remove(node)


def update_io_tables_and_services(root: etree._Element, id_gen: IdGenerator, io_dids: list[IoDidDef]) -> None:
    table_names = {
        0: ("IOControl_ReturnControl_RQ", "IOControl_ReturnControl_PR", "IOControl_ReturnControl", "ReturnControl"),
        1: ("IOControl_Reset_RQ", "IOControl_Reset_PR", "IOControl_Reset", "Reset"),
        2: ("IOControl_Freeze_RQ", "IOControl_Freeze_PR", "IOControl_Freeze", "Freeze"),
        3: ("IOControl_Control_RQ", "IOControl_Control_PR", "IOControl_Control", "Control"),
    }
    # The IO DID survey section describes InputOutputControlByIdentifier (0x2F)
    # sub-functions only. Do not emit a synthetic Read service unless a future
    # schema explicitly carries read support for IO DIDs.
    remove_service_and_messages(root, "Combined_IOControl_Read")
    remove_service_and_messages(root, "IOControl_Read")
    remove_elements_by_short_names(
        root,
        "STRUCTURE",
        {"STR_EOP_DIDs_IOControl_ReadRDBI_RQ", "STR_EOP_DIDs_IOControl_ReadRDBI_PR"},
    )
    remove_elements_by_short_names(
        root,
        "END-OF-PDU-FIELD",
        {"EOP_DIDs_IOControl_ReadRDBI_RQ", "EOP_DIDs_IOControl_ReadRDBI_PR"},
    )
    remove_elements_by_short_names(root, "DATA-OBJECT-PROP", {"DOP_NRC_IOControl_Read"})
    remove_elements_by_short_names(root, "TABLE", {"IOControl_Read_PR"})

    instances_by_control: dict[int, list[tuple[int, str, str]]] = {k: [] for k in table_names}
    for control, (rq_table_name, pr_table_name, _, _) in table_names.items():
        rq_table = first_by_short(root, "TABLE", rq_table_name)
        pr_table = first_by_short(root, "TABLE", pr_table_name)
        if rq_table is not None:
            clear_children(rq_table, "TABLE-ROW")
        if pr_table is not None:
            clear_children(pr_table, "TABLE-ROW")
        for io_did in io_dids:
            if control not in io_did.controls:
                continue
            rq_structure = io_did.control_request_wrapper_id if control == 3 else None
            pr_structure = io_did.status_wrapper_id
            if rq_table is not None:
                append_table_row(rq_table, id_gen, io_did.short_name, io_did.long_name, io_did.did, rq_structure)
            if pr_table is not None:
                append_table_row(pr_table, id_gen, io_did.short_name, io_did.long_name, io_did.did, pr_structure)
            instances_by_control[control].append((io_did.did, io_did.short_name, io_did.long_name))

    for control, (_, _, service_short_name, service_label) in table_names.items():
        rq_table_name, pr_table_name, _, _ = table_names[control]
        if not instances_by_control[control]:
            remove_service_and_messages(root, service_short_name)
            remove_elements_by_short_names(root, "TABLE", {rq_table_name, pr_table_name})
            continue
        update_service_instances(root, id_gen, service_short_name, service_label, service_label, instances_by_control[control])
        state_refs = {value: ["_4", "_7", "_10", "_11", "_12", "_13", "_14", "_15"] for value, _, _ in instances_by_control[control]}
        update_preconditions_for_values(root, service_short_name, instances_by_control[control], "DataIdentifier", state_refs)


def update_routine_tables_and_services(root: etree._Element, id_gen: IdGenerator, routines: list[RoutineDef]) -> None:
    table_pairs = {
        1: ("TAB_StartOptions", "TAB_StartStatus", "Routine_Control_Start", "Start"),
        2: ("TAB_StopOptions", "TAB_StopStatus", "Routine_Control_Stop", "Stop"),
        3: ("TAB_RequestResultsOptions", "TAB_RequestResultsStatus", "Routine_Control_RequestResults", "RequestResults"),
    }
    ddds = root.find(".//DIAG-DATA-DICTIONARY-SPEC")
    tables_node = ddds.find("TABLES") if ddds is not None else None
    if tables_node is None:
        return

    key_dop_id = "_166"
    for control_type, (option_table_name, status_table_name, service_short_name, service_label) in table_pairs.items():
        option_table = ensure_table(root, tables_node, id_gen, option_table_name, f"Table Control Options ({service_label})", key_dop_id)
        status_table = ensure_table(root, tables_node, id_gen, status_table_name, f"Table Status ({service_label})", key_dop_id)
        clear_children(option_table, "TABLE-ROW")
        clear_children(status_table, "TABLE-ROW")
        instances: list[tuple[int, str, str]] = []
        state_refs: dict[int, list[str]] = {}
        for routine in routines:
            subfn = routine.subfunctions.get(control_type)
            if subfn is None or not subfn.supported:
                continue
            append_table_row(option_table, id_gen, routine.short_name, routine.long_name, routine.rid, subfn.option_structure_id or None)
            append_table_row(status_table, id_gen, routine.short_name, routine.long_name, routine.rid, subfn.status_structure_id or None)
            instances.append((routine.rid, routine.short_name, routine.long_name))
            states = sessions_to_state_ids([""] + routine.sessions, mode="write")
            states.extend(security_to_state_ids(routine.security))
            states.extend(["_10", "_11", "_12", "_13", "_14", "_15"])
            state_refs[routine.rid] = unique_list(states)

        ensure_routine_service(root, id_gen, service_short_name, service_label, control_type, option_table, status_table)
        update_service_instances(root, id_gen, service_short_name, service_label, service_label, instances)
        update_preconditions_for_values(root, service_short_name, instances, "RoutineIdentifier", state_refs)


def ensure_table(
    root: etree._Element,
    tables_node: etree._Element,
    id_gen: IdGenerator,
    short_name: str,
    long_name: str,
    key_dop_id: str,
) -> etree._Element:
    table = first_by_short(root, "TABLE", short_name)
    if table is not None:
        return table
    table = sub(tables_node, "TABLE", attrib={"ID": id_gen.new("TAB")})
    sub(table, "SHORT-NAME", short_name)
    sub(table, "LONG-NAME", long_name)
    sub(table, "KEY-DOP-REF", attrib={"ID-REF": key_dop_id})
    return table


def ensure_routine_service(
    root: etree._Element,
    id_gen: IdGenerator,
    service_short_name: str,
    service_label: str,
    control_type: int,
    option_table: etree._Element,
    status_table: etree._Element,
) -> None:
    if first_by_short(root, "DIAG-SERVICE", service_short_name) is not None:
        update_routine_request_response_tables(root, service_short_name, option_table, status_table)
        return

    base_service = first_by_short(root, "DIAG-SERVICE", "Routine_Control_Start")
    base_request = first_by_short(root, "REQUEST", "RQ_Routine_Control_Start")
    base_response = first_by_short(root, "POS-RESPONSE", "PR_Routine_Control_Start")
    base_negative = first_by_short(root, "NEG-RESPONSE", "NR_Routine_Control_Start")
    diag_comms = root.find(".//DIAG-COMMS")
    requests = root.find(".//REQUESTS")
    responses = root.find(".//POS-RESPONSES")
    negatives = root.find(".//NEG-RESPONSES")
    required_nodes = [base_service, base_request, base_response, base_negative, diag_comms, requests, responses, negatives]
    if any(node is None for node in required_nodes):
        return

    rq_id = id_gen.new("RQ")
    pr_id = id_gen.new("PR")
    nr_id = id_gen.new("NR")
    service_id = id_gen.new("SVC")

    request = copy.deepcopy(base_request)
    request.set("ID", rq_id)
    set_short_long(request, f"RQ_{service_short_name}", f"RQ {service_short_name.replace('_', ' ')}")
    set_routine_message_control_type(request, control_type)
    set_table_ref_in_message(request, option_table.get("ID"))
    refresh_internal_ids(request, id_gen)
    requests.append(request)

    response = copy.deepcopy(base_response)
    response.set("ID", pr_id)
    set_short_long(response, f"PR_{service_short_name}", f"PR {service_short_name.replace('_', ' ')}")
    set_routine_message_control_type(response, control_type)
    set_table_ref_in_message(response, status_table.get("ID"))
    refresh_internal_ids(response, id_gen)
    responses.append(response)

    negative = copy.deepcopy(base_negative)
    negative.set("ID", nr_id)
    set_short_long(negative, f"NR_{service_short_name}", f"NR {service_short_name.replace('_', ' ')}")
    refresh_internal_ids(negative, id_gen)
    negatives.append(negative)

    service = copy.deepcopy(base_service)
    service.set("ID", service_id)
    set_short_long(service, service_short_name, service_short_name.replace("_", " "))
    refresh_internal_ids(service, id_gen)
    request_ref = service.find("REQUEST-REF")
    if request_ref is not None:
        request_ref.set("ID-REF", rq_id)
    pos_ref = service.find(".//POS-RESPONSE-REF")
    if pos_ref is not None:
        pos_ref.set("ID-REF", pr_id)
    neg_ref = service.find(".//NEG-RESPONSE-REF")
    if neg_ref is not None:
        neg_ref.set("ID-REF", nr_id)
    diag_comms.append(service)
    update_routine_request_response_tables(root, service_short_name, option_table, status_table)


def set_short_long(node: etree._Element, short_name: str, long_name: str) -> None:
    short = node.find("SHORT-NAME")
    if short is not None:
        short.text = sanitize_short_name(short_name, "Node")
    long = node.find("LONG-NAME")
    if long is not None:
        long.text = long_name


def refresh_internal_ids(node: etree._Element, id_gen: IdGenerator) -> None:
    id_map: dict[str, str] = {}
    for child in node.xpath(".//*[@ID]"):
        old_id = child.get("ID")
        if not old_id:
            continue
        new_id = id_gen.new("IN")
        id_map[old_id] = new_id
        child.set("ID", new_id)
    if not id_map:
        return
    for ref_node in node.xpath(".//*[@ID-REF]"):
        old_ref = ref_node.get("ID-REF")
        if old_ref in id_map and not ref_node.get("DOCREF"):
            ref_node.set("ID-REF", id_map[old_ref])


def set_routine_message_control_type(message: etree._Element, control_type: int) -> None:
    for param in message.xpath('.//*[local-name()="PARAM" and SHORT-NAME="RoutineControlType"]'):
        coded = param.find("CODED-VALUE")
        if coded is not None:
            coded.text = str(control_type)


def set_table_ref_in_message(message: etree._Element, table_id: str | None) -> None:
    if not table_id:
        return
    table_ref = message.find(".//TABLE-REF")
    if table_ref is not None:
        table_ref.set("ID-REF", table_id)


def update_routine_request_response_tables(
    root: etree._Element,
    service_short_name: str,
    option_table: etree._Element,
    status_table: etree._Element,
) -> None:
    service = first_by_short(root, "DIAG-SERVICE", service_short_name)
    if service is None:
        return
    request_ref = service.find("REQUEST-REF")
    if request_ref is not None:
        request = root.xpath(f'//*[@ID="{request_ref.get("ID-REF")}"]')
        if request:
            set_table_ref_in_message(request[0], option_table.get("ID"))
    pos_ref = service.find(".//POS-RESPONSE-REF")
    if pos_ref is not None:
        response = root.xpath(f'//*[@ID="{pos_ref.get("ID-REF")}"]')
        if response:
            set_table_ref_in_message(response[0], status_table.get("ID"))


def update_dtc_dop(root: etree._Element, id_gen: IdGenerator, dtcs: list[DtcDef]) -> None:
    dtc_dop = first_by_short(root, "DTC-DOP", "RecordDataType")
    if dtc_dop is None:
        return
    dtcs_node = dtc_dop.find("DTCS")
    if dtcs_node is None:
        dtcs_node = sub(dtc_dop, "DTCS")
    for child in list(dtcs_node):
        dtcs_node.remove(child)
    caption_ids: dict[str, str] = {}
    for index, dtc in enumerate(dtcs):
        dtc_node = sub(dtcs_node, "DTC", attrib={"ID": id_gen.new("DTC")})
        sub(dtc_node, "SHORT-NAME", sanitize_short_name(f"DTC_{dtc.byte_code:06X}", "DTC"))
        sub(dtc_node, "TROUBLE-CODE", dtc.byte_code)
        sub(dtc_node, "DISPLAY-TROUBLE-CODE", dtc.display_code)
        sub(dtc_node, "TEXT", dtc.text or dtc.display_code)
        sdgs = sub(dtc_node, "SDGS")
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_SHORTNAME", f"DTC_0X{dtc.byte_code:06X}", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_PRIORITY_VALUE", dtc.priority or "2", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_SUPPORTED", "supported", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_CYCLE", "DEM_POWER", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_COUNTER", "40", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_SEVERITY_VALUE", "noSeverity", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_INIT_MONITOR_REQUIRED", "not required", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_FUNCTIONAL_UNIT_VALUE", "0x00", first=index == 0)

    update_dtc_text_table(root, dtcs)

    ext_table = first_by_short(root, "TABLE", "DTCExtendedDataRecordNumber")
    if ext_table is not None:
        clear_children(ext_table, "TABLE-ROW")
        default_structure = "_179"
        for dtc in dtcs:
            key = dtc_table_key(dtc)
            row = append_table_row(
                ext_table,
                id_gen,
                f"TR_DTC_{dtc.byte_code:06X}",
                key,
                key,
                default_structure,
            )
            sdgs = sub(row, "SDGS")
            sdg = sub(sdgs, "SDG")
            caption = sub(sdg, "SDG-CAPTION", attrib={"ID": id_gen.new("CAP")})
            sub(caption, "SHORT-NAME", "IsDefaultCase")
            sub(sdg, "SD", "Yes")


def dtc_table_key(dtc: DtcDef) -> str:
    text = compact_text(dtc.text or dtc.display_code)
    suffix = f"_{text[:80]}" if text else ""
    return f"0x{dtc.byte_code:06X}{suffix}"


def update_dtc_text_table(root: etree._Element, dtcs: list[DtcDef]) -> None:
    dop = first_by_short(root, "DATA-OBJECT-PROP", "TextTable_DTC")
    if dop is None:
        return
    scales = dop.find(".//COMPU-SCALES")
    if scales is None:
        return
    for child in list(scales):
        scales.remove(child)
    for dtc in dtcs:
        scale = sub(scales, "COMPU-SCALE")
        sub(scale, "LOWER-LIMIT", dtc.byte_code)
        sub(scale, "UPPER-LIMIT", dtc.byte_code)
        const = sub(scale, "COMPU-CONST")
        sub(const, "VT", dtc_table_key(dtc))

    internal_constr = dop.find("INTERNAL-CONSTR")
    if internal_constr is not None:
        dop.remove(internal_constr)


def add_dtc_sdg(
    sdgs: etree._Element,
    id_gen: IdGenerator,
    caption_ids: dict[str, str],
    short_name: str,
    value: str,
    *,
    first: bool,
) -> None:
    sdg = sub(sdgs, "SDG")
    if first or short_name not in caption_ids:
        caption_id = id_gen.new("CAP")
        caption_ids[short_name] = caption_id
        caption = sub(sdg, "SDG-CAPTION", attrib={"ID": caption_id})
        sub(caption, "SHORT-NAME", short_name)
    else:
        sub(sdg, "SDG-CAPTION-REF", attrib={"ID-REF": caption_ids[short_name]})
    sub(sdg, "SD", value)


def update_snapshot_and_extended_data(
    root: etree._Element,
    id_gen: IdGenerator,
    snapshots: list[SnapshotDef],
    extended_records: list[ExtendedRecordDef],
) -> None:
    snapshots = [snapshot for snapshot in snapshots if snapshot.structure_id]
    env_data = first_by_short(root, "ENV-DATA", "ENVDATA_ALLDTCS")
    if env_data is not None and snapshots:
        params = env_data.find("PARAMS")
        if params is None:
            params = sub(env_data, "PARAMS")
        for child in list(params):
            params.remove(child)
        byte_position = 0
        for snapshot in snapshots:
            did_param = element("PARAM", attrib={"SEMANTIC": "DATA"})
            set_xsi_type(did_param, "PHYS-CONST")
            sub(did_param, "SHORT-NAME", sanitize_short_name(hex_short("DID", snapshot.did), "SnapshotDID"))
            sub(did_param, "LONG-NAME", snapshot.desc)
            sub(did_param, "BYTE-POSITION", byte_position)
            sub(did_param, "PHYS-CONSTANT-VALUE", snapshot.did)
            sub(did_param, "DOP-REF", attrib={"ID-REF": "_17"})
            params.append(did_param)
            byte_position += 2
            if snapshot.structure_id:
                value_param = make_value_param(
                    sanitize_short_name(f"{hex_short('DID', snapshot.did)}_Data", "SnapshotData"),
                    snapshot.desc,
                    byte_position,
                    0,
                    snapshot.structure_id,
                )
                params.append(value_param)
                byte_position += max(1, snapshot.size)

    if extended_records:
        update_record_number_dop(root, "DTCExtendedDataRecordNumbers_All", extended_records, include_all=True)
        update_record_number_dop(root, "DTCExtendedDataRecordNumbers_All_except_FF", extended_records, include_all=False)
        mux = first_by_short(root, "MUX", "DTCExtendedDataRecordData")
        if mux is not None:
            cases = mux.find("CASES")
            if cases is None:
                cases = sub(mux, "CASES")
            for child in list(cases):
                cases.remove(child)
            for record in extended_records:
                if not record.structure_id:
                    continue
                case = sub(cases, "CASE")
                label = extended_record_label(record)
                sub(case, "SHORT-NAME", sanitize_short_name(f"Case_0x{record.record_num:02X}", "Case"))
                sub(case, "STRUCTURE-REF", attrib={"ID-REF": record.structure_id})
                sub(case, "LOWER-LIMIT", label)
                sub(case, "UPPER-LIMIT", label)


def make_coded_const_param(
    short_name: str,
    long_name: str,
    byte_position: int,
    coded_value: int,
    semantic: str,
) -> etree._Element:
    param = element("PARAM", attrib={"SEMANTIC": semantic})
    set_xsi_type(param, "CODED-CONST")
    sub(param, "SHORT-NAME", short_name)
    sub(param, "LONG-NAME", long_name)
    sub(param, "BYTE-POSITION", byte_position)
    sub(param, "CODED-VALUE", coded_value)
    coded_type = sub(param, "DIAG-CODED-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32"})
    set_xsi_type(coded_type, "STANDARD-LENGTH-TYPE")
    sub(coded_type, "BIT-LENGTH", 8)
    return param


def make_table_key_param(
    short_name: str,
    long_name: str,
    byte_position: int,
    table_id: str,
    param_id: str,
) -> etree._Element:
    param = element("PARAM", attrib={"SEMANTIC": "DATA", "ID": param_id})
    set_xsi_type(param, "TABLE-KEY")
    param.append(etree.ProcessingInstruction("GeneratedFromMux", "yes"))
    sub(param, "SHORT-NAME", short_name)
    sub(param, "LONG-NAME", long_name)
    sub(param, "BYTE-POSITION", byte_position)
    sub(param, "TABLE-REF", attrib={"ID-REF": table_id})
    return param


def make_table_struct_param(
    short_name: str,
    long_name: str,
    byte_position: int,
    table_key_id: str,
) -> etree._Element:
    param = element("PARAM", attrib={"SEMANTIC": "DATA"})
    set_xsi_type(param, "TABLE-STRUCT")
    sub(param, "SHORT-NAME", short_name)
    sub(param, "LONG-NAME", long_name)
    sub(param, "BYTE-POSITION", byte_position)
    sub(param, "TABLE-KEY-REF", attrib={"ID-REF": table_key_id})
    return param


def normalize_extended_data_service_for_candela(root: etree._Element, id_gen: IdGenerator) -> None:
    request = first_by_short(root, "REQUEST", "RQ_FaultMemory_Read_extended_data_record")
    if request is None:
        return
    ext_table = first_by_short(root, "TABLE", "DTCExtendedDataRecordNumber")
    if ext_table is None or not ext_table.get("ID"):
        return
    params = request.find("PARAMS")
    if params is None:
        params = sub(request, "PARAMS")
    table_key_id = ""
    for param in params.findall("PARAM"):
        if param.findtext("SHORT-NAME") != "DTC":
            continue
        if param.get("ID"):
            table_key_id = param.get("ID") or ""
            break
    if not table_key_id:
        table_key_id = id_gen.new("TK")
    for child in list(params):
        params.remove(child)
    params.append(make_coded_const_param("SID_RQ", "SID-RQ", 0, 0x19, "SERVICE-ID"))
    params.append(
        make_coded_const_param(
            "ReportDTCExtendedDataRecordByDtcNumber",
            "ReportDTCExtendedDataRecordByDtcNumber",
            1,
            0x06,
            "SUBFUNCTION",
        )
    )
    params.append(make_table_key_param("DTC", "DTC", 2, ext_table.get("ID"), table_key_id))
    params.append(
        make_table_struct_param(
            "DTCExtendedDataRecordNumber",
            "DTCExtendedDataRecordNumber",
            5,
            table_key_id,
        )
    )


def validate_edr_mapping(root: etree._Element, dtcs: list[DtcDef], extended_records: list[ExtendedRecordDef]) -> None:
    if not dtcs or not extended_records:
        return

    errors: list[str] = []

    def add_error(message: str) -> None:
        errors.append(message)

    def short_sample(values: Iterable[str], limit: int = 5) -> str:
        items = list(values)
        if len(items) <= limit:
            return ", ".join(items)
        return ", ".join(items[:limit]) + f", ... (+{len(items) - limit})"

    service = first_by_short(root, "DIAG-SERVICE", "FaultMemory_Read_extended_data_record")
    if service is None:
        add_error("missing DIAG-SERVICE FaultMemory_Read_extended_data_record")

    request = first_by_short(root, "REQUEST", "RQ_FaultMemory_Read_extended_data_record")
    if service is not None:
        request_ref = service.find("REQUEST-REF")
        request_id = request_ref.get("ID-REF") if request_ref is not None else ""
        if request_id:
            request_nodes = root.xpath(f'//*[@ID="{request_id}"]')
            if request_nodes:
                request = request_nodes[0]
            else:
                add_error(f"service references missing request ID {request_id}")

    ext_table = first_by_short(root, "TABLE", "DTCExtendedDataRecordNumber")
    if ext_table is None or not ext_table.get("ID"):
        add_error("missing TABLE DTCExtendedDataRecordNumber")

    if request is None:
        add_error("missing REQUEST RQ_FaultMemory_Read_extended_data_record")
    else:
        params = request.xpath('./*[local-name()="PARAMS"]/*[local-name()="PARAM"]')
        by_name = {compact_text(param.findtext("SHORT-NAME")): param for param in params}
        dtc_param = by_name.get("DTC")
        record_param = by_name.get("DTCExtendedDataRecordNumber")
        if dtc_param is None:
            add_error("0x19 0x06 request missing TABLE-KEY param DTC")
        else:
            if dtc_param.get(XSI_TYPE) != "TABLE-KEY":
                add_error("0x19 0x06 DTC param is not TABLE-KEY")
            if not dtc_param.get("ID"):
                add_error("0x19 0x06 DTC TABLE-KEY has no ID")
            table_ref = dtc_param.find("TABLE-REF")
            table_id = table_ref.get("ID-REF") if table_ref is not None else ""
            expected_table_id = ext_table.get("ID") if ext_table is not None else ""
            if expected_table_id and table_id != expected_table_id:
                add_error("0x19 0x06 DTC TABLE-KEY does not reference DTCExtendedDataRecordNumber")

        if record_param is None:
            add_error("0x19 0x06 request missing TABLE-STRUCT param DTCExtendedDataRecordNumber")
        else:
            if record_param.get(XSI_TYPE) != "TABLE-STRUCT":
                add_error("0x19 0x06 DTCExtendedDataRecordNumber param is not TABLE-STRUCT")
            key_ref = record_param.find("TABLE-KEY-REF")
            key_id = key_ref.get("ID-REF") if key_ref is not None else ""
            dtc_id = dtc_param.get("ID") if dtc_param is not None else ""
            if dtc_id and key_id != dtc_id:
                add_error("0x19 0x06 TABLE-STRUCT does not reference the DTC TABLE-KEY")

    if ext_table is not None:
        rows = ext_table.xpath('./*[local-name()="TABLE-ROW"]')
        rows_by_key = {compact_text(row.findtext("KEY")): row for row in rows}
        expected_keys = [dtc_table_key(dtc) for dtc in dtcs]
        missing_keys = [key for key in expected_keys if key not in rows_by_key]
        if missing_keys:
            add_error(f"DTCExtendedDataRecordNumber missing {len(missing_keys)} DTC row(s): {short_sample(missing_keys)}")
        unsupported_keys = [
            key
            for key in expected_keys
            if key in rows_by_key and rows_by_key[key].find("STRUCTURE-REF") is None
        ]
        if unsupported_keys:
            add_error(f"{len(unsupported_keys)} DTC row(s) have no EDR STRUCTURE-REF: {short_sample(unsupported_keys)}")

    expected_record_labels = [extended_record_label(record) for record in extended_records]
    for dop_name, require_all in (
        ("DTCExtendedDataRecordNumbers_All", True),
        ("DTCExtendedDataRecordNumbers_All_except_FF", False),
    ):
        dop = first_by_short(root, "DATA-OBJECT-PROP", dop_name)
        if dop is None:
            add_error(f"missing DATA-OBJECT-PROP {dop_name}")
            continue
        scales = dop.xpath('.//*[local-name()="COMPU-SCALE"]')
        labels = {compact_text(scale.findtext(".//VT")) for scale in scales}
        missing_labels = [label for label in expected_record_labels if label not in labels]
        if missing_labels:
            add_error(f"{dop_name} missing record label(s): {short_sample(missing_labels)}")
        has_all = "All" in labels
        if require_all and not has_all:
            add_error(f"{dop_name} missing All/0xFF scale")
        if not require_all and has_all:
            add_error(f"{dop_name} unexpectedly contains All/0xFF scale")

    mux = first_by_short(root, "MUX", "DTCExtendedDataRecordData")
    if mux is None:
        add_error("missing MUX DTCExtendedDataRecordData")
    else:
        cases = mux.xpath('./*[local-name()="CASES"]/*[local-name()="CASE"]')
        case_by_label = {compact_text(case.findtext("LOWER-LIMIT")): case for case in cases}
        for record, label in zip(extended_records, expected_record_labels, strict=True):
            if not record.structure_id:
                add_error(f"{label} has no generated STRUCTURE")
                continue
            case = case_by_label.get(label)
            if case is None:
                add_error(f"DTCExtendedDataRecordData missing MUX case for {label}")
                continue
            structure_ref = case.find("STRUCTURE-REF")
            structure_id = structure_ref.get("ID-REF") if structure_ref is not None else ""
            if structure_id != record.structure_id:
                add_error(f"DTCExtendedDataRecordData MUX case for {label} references {structure_id or '<none>'}")

    if errors:
        raise RuntimeError("Invalid DTC extended-data mapping:\n- " + "\n- ".join(errors))


def update_record_number_dop(
    root: etree._Element,
    short_name: str,
    records: list[ExtendedRecordDef],
    *,
    include_all: bool,
) -> None:
    dop = first_by_short(root, "DATA-OBJECT-PROP", short_name)
    if dop is None:
        return

    compu = element("COMPU-METHOD")
    sub(compu, "CATEGORY", "TEXTTABLE")
    internal = sub(compu, "COMPU-INTERNAL-TO-PHYS")
    scales = sub(internal, "COMPU-SCALES")
    for child in list(scales):
        scales.remove(child)
    for record in records:
        scale = sub(scales, "COMPU-SCALE")
        label = extended_record_label(record)
        sub(scale, "LOWER-LIMIT", record.record_num)
        sub(scale, "UPPER-LIMIT", record.record_num)
        const = sub(scale, "COMPU-CONST")
        sub(const, "VT", label)
    if include_all:
        scale = sub(scales, "COMPU-SCALE")
        sub(scale, "LOWER-LIMIT", 255)
        sub(scale, "UPPER-LIMIT", 255)
        const = sub(scale, "COMPU-CONST")
        sub(const, "VT", "All")

    coded = element(
        "DIAG-CODED-TYPE",
        attrib={"BASE-TYPE-ENCODING": "NONE", "BASE-DATA-TYPE": "A_UINT32"},
    )
    set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
    sub(coded, "BIT-LENGTH", 8)
    physical = element("PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UNICODE2STRING"})

    replace_child(dop, "COMPU-METHOD", compu, before_tags={"DIAG-CODED-TYPE", "PHYSICAL-TYPE"})
    replace_child(dop, "DIAG-CODED-TYPE", coded, before_tags={"PHYSICAL-TYPE"})
    replace_child(dop, "PHYSICAL-TYPE", physical)
    internal_constr = dop.find("INTERNAL-CONSTR")
    if internal_constr is not None:
        dop.remove(internal_constr)


#!/usr/bin/env python3
"""Generate a CANdelaStudio-compatible VF PDX from a VF diagnostic survey workbook.

This script is self-contained inside the VF project. It parses the VF Excel
survey schema into the canonical diagnostic model and updates the VF
CANdelaStudio 15 PDX template with DID, IOControl, RoutineControl, DTC,
Snapshot, ExtendedData, and communication-parameter content.
"""


VF_KEEP_FILES = {
    "ISO_11898_2_DWCAN.odx-cs",
    "ISO_11898_3_DWFTCAN.odx-cs",
    "ISO_15765_2.odx-cs",
    "ISO_15765_3.odx-cs",
    "ISO_15765_3_on_ISO_15765_2.odx-c",
    "SAE_J2411_SWCAN.odx-cs",
    "VF_ECU_CAN_v15.odx-d",
    "index.xml",
}

CANDELA_SHORT_NAME_MAX_LEN = 64
UUDT_DISABLED_CAN_ID = 0xFFFFFFFF


def parse_hex_in_text(value: Any, *, max_value: int = 0xFFFFFFFF) -> int | None:
    text = compact_text(value)
    match = re.search(r"0[xX]([0-9A-Fa-f]{1,8})|(?<![A-Za-z0-9])([0-9A-Fa-f]{2,8})(?![A-Za-z0-9])", text)
    if not match:
        return None
    raw = match.group(1) or match.group(2)
    result = int(raw, 16)
    return result if result <= max_value else None


def usable_text(value: Any) -> str:
    text = compact_text(value)
    if text.upper() in {"", "/", "N/A", "NA", "NONE", "NULL"}:
        return ""
    return text


def dual_name(english_value: Any, chinese_value: Any = "", fallback: str = "") -> tuple[str, str]:
    candidates = name_parts(english_value) + name_parts(chinese_value)
    candidates = [candidate for candidate in candidates if usable_text(candidate)]
    if not candidates:
        return fallback, fallback
    english = next((candidate for candidate in candidates if re.search(r"[A-Za-z]", candidate) and not has_chinese_text(candidate)), "")
    chinese = next((candidate for candidate in candidates if has_chinese_text(candidate)), "")
    identifier = english or (fallback if chinese else candidates[0])
    display = chinese or candidates[0]
    return identifier or fallback, display or identifier or fallback


def normalize_conversion_text(value: Any) -> str:
    text = cell_text(value)
    replacements = {
        "\u3000": " ",
        "?": ";",
        "?": ":",
        "?": ",",
        "?": "~",
        "?": "-",
        "?": "-",
        "\r": "\n",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.strip()


LINEAR_NUMBER_PATTERN = r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)"


def decimal_precision_from_token(token: str | None) -> int:
    if not token:
        return 0
    match = re.search(r"\.(\d+)", token)
    return len(match.group(1)) if match else 0


def parse_float_assignment_token(text: str, key: str, default: float) -> tuple[float, str | None]:
    match = re.search(rf"\b{re.escape(key)}\s*=\s*({LINEAR_NUMBER_PATTERN})", text, flags=re.IGNORECASE)
    if not match:
        return default, None
    token = match.group(1)
    return float(token), token


def is_conversion_directive_line(line: str) -> bool:
    return re.fullmatch(r"\s*(?:type|value|a|b|precision)\s*=.*", line, flags=re.IGNORECASE) is not None


def parse_conversion(value: Any) -> Conversion:
    text = normalize_conversion_text(value)
    if not text or text.upper() in {"/", "N/A", "NA"}:
        return Conversion()
    if re.search(r"\btype\s*=\s*identical\b", text, flags=re.IGNORECASE):
        return Conversion()
    if re.search(r"\btype\s*=\s*bcd\b", text, flags=re.IGNORECASE):
        return Conversion(kind="bcd")

    linear = re.search(
        rf"(?:phy|phys|physical|y)\s*=\s*(?:raw|XX|xx|X|x)"
        rf"(?:\s*\*\s*(?P<a>{LINEAR_NUMBER_PATTERN}))?"
        rf"(?:\s*(?P<sign>[+-])\s*(?P<b>{LINEAR_NUMBER_PATTERN}))?",
        text,
        flags=re.IGNORECASE,
    )
    if linear:
        a_token = linear.group("a")
        b_token = linear.group("b")
        a = float(a_token) if a_token else 1.0
        b = float(b_token) if b_token else 0.0
        if linear.group("sign") == "-":
            b = -b
        precision = max(decimal_precision_from_token(a_token), decimal_precision_from_token(b_token))
        return Conversion(kind="linear", a=a, b=b, precision=precision if precision > 0 else None)

    if re.search(r"\btype\s*=\s*a\s*x\s*\+\s*b\b|\by\s*=\s*a\s*\*?\s*x\s*\+\s*b\b", text, flags=re.IGNORECASE):
        a, a_token = parse_float_assignment_token(text, "a", 1.0)
        b, b_token = parse_float_assignment_token(text, "b", 0.0)
        precision_match = re.search(r"\bprecision\s*=\s*(-?\d+)", text, flags=re.IGNORECASE)
        precision = (
            int(precision_match.group(1))
            if precision_match
            else max(decimal_precision_from_token(a_token), decimal_precision_from_token(b_token))
        )
        return Conversion(kind="linear", a=a, b=b, precision=precision if precision > 0 else None)

    enum_entries: list[tuple[int, int, str]] = []
    enum_pattern = re.compile(
        r"(?:0[xX])?([0-9A-Fa-f]+)\s*(?:[-~]\s*(?:0[xX])?([0-9A-Fa-f]+))?\s*[:=]\s*([^;\n\r]+)"
    )
    for line in re.split(r"[\r\n;]+", text):
        if is_conversion_directive_line(line):
            continue
        for match in enum_pattern.finditer(line):
            label = match.group(3).strip(" ;,")
            if not label:
                continue
            lo = int(match.group(1), 16)
            hi = int(match.group(2), 16) if match.group(2) else lo
            enum_entries.append((lo, hi, label))
    if enum_entries:
        return Conversion(kind="enum", enum=enum_entries)

    inline_type_match = re.search(r"\btype\s*=\s*\w+\b", text, flags=re.IGNORECASE)
    data_text = text[inline_type_match.end():] if inline_type_match else text
    for line in re.split(r"[\r\n;]+", data_text):
        if is_conversion_directive_line(line):
            continue
        for match in enum_pattern.finditer(line):
            label = match.group(3).strip(" ;,")
            if not label:
                continue
            lo = int(match.group(1), 16)
            hi = int(match.group(2), 16) if match.group(2) else lo
            enum_entries.append((lo, hi, label))
    if enum_entries:
        return Conversion(kind="enum", enum=enum_entries)

    return Conversion()


def choose_conversion(primary_value: Any, fallback_value: Any = "") -> Conversion:
    """Parse both conversion columns and keep the richer physical meaning.

    Legacy survey sheets normally put English expressions in column L and Chinese text
    in column M. In a few rows one column contains a pass-through formula while
    the other carries the value table, so enum > linear > identity gives CANdela
    the more useful data type without hard-coding row numbers.
    """

    primary = parse_conversion(primary_value)
    fallback = parse_conversion(fallback_value)
    rank = {"enum": 4, "linear": 3, "bcd": 2, "identity": 1}
    if rank.get(fallback.kind, 0) > rank.get(primary.kind, 0):
        return fallback
    return primary


def parse_index_range(value: Any, *, size: int = 0, default: int = 0) -> tuple[int, int]:
    text = compact_text(value).upper()
    if not text or text in {"/", "N/A", "NA"}:
        return default, default
    if text == "ALL":
        return 0, max(0, size - 1)
    nums = [int(num) for num in re.findall(r"\d+", text)]
    if not nums:
        return default, default
    start = nums[0]
    end = nums[1] if len(nums) > 1 else start
    if end < start:
        end = start
    return start, end


def parse_bit_range(value: Any, byte_start: int, byte_end: int, *, size: int = 0) -> tuple[int, int]:
    text = compact_text(value).upper()
    byte_count = max(1, byte_end - byte_start + 1)
    if not text or text in {"/", "N/A", "NA", "ALL"}:
        return 0, byte_count * 8
    nums = [int(num) for num in re.findall(r"\d+", text)]
    if not nums:
        return 0, byte_count * 8
    if len(nums) > 1 and ("-" in text or "~" in text):
        start, end = nums[0], nums[1]
        if end < start:
            end = start
        return start % 8, min(max(1, end - start + 1), byte_count * 8)
    return nums[0] % 8, 1


def make_param_from_cells(
    *,
    name_value: Any,
    chinese_name_value: Any = "",
    byte_value: Any,
    bit_value: Any,
    data_type_value: Any,
    unit_value: Any = "",
    conversion_value: Any = "",
    conversion_fallback_value: Any = "",
    min_value: Any = "",
    max_value: Any = "",
    fallback_name: str = "Data",
    size: int = 0,
) -> ParamDef | None:
    english, long_name = dual_name(name_value, chinese_name_value, fallback_name)
    if not usable_text(english) and not usable_text(long_name):
        english = fallback_name
        long_name = fallback_name
    byte_start, byte_end = parse_index_range(byte_value, size=size, default=0)
    bit_pos, bit_len = parse_bit_range(bit_value, byte_start, byte_end, size=size)
    data_type = usable_text(data_type_value) or "Hex(Unsigned)"
    conversion = choose_conversion(
        usable_text(conversion_value),
        usable_text(conversion_fallback_value),
    )
    if bit_len > 32 and conversion.kind == "enum":
        conversion = Conversion()
    return ParamDef(
        name=english or fallback_name,
        long_name=long_name or english or fallback_name,
        byte_pos=byte_start,
        bit_pos=bit_pos,
        bit_len=bit_len,
        data_type=data_type,
        unit=usable_text(unit_value),
        conversion=conversion,
        min_value=usable_text(min_value),
        max_value=usable_text(max_value),
    )


def merge_session_access(read_values: Iterable[Any], write_values: Iterable[Any]) -> list[str]:
    sessions: list[str] = []
    for read_value, write_value in zip(read_values, write_values, strict=False):
        flags = ""
        if usable_text(read_value).upper() != "N":
            flags += "R"
        if usable_text(write_value).upper() != "N":
            flags += "W"
        sessions.append(flags)
    return sessions


def first_security_level(values: Iterable[Any]) -> str:
    for value in values:
        text = usable_text(value)
        if text and text.upper() != "N":
            return text
    return "N"


def is_supported_flag(value: Any, *, default: bool = False) -> bool:
    text = compact_text(value).upper()
    if text in {"Y", "YES", "TBD"}:
        return True
    if text in {"N", "NO", "-", "/", "N/A", "NA"}:
        return False
    return default


def merge_security_text(old_value: str, new_value: str) -> str:
    old_text = usable_text(old_value) or "N"
    new_text = usable_text(new_value) or "N"
    if old_text.upper() == "N":
        return new_text
    if new_text.upper() == "N" or new_text in old_text.split("&"):
        return old_text
    parts = [part for part in old_text.split("&") if part]
    for part in new_text.split("&"):
        if part and part not in parts:
            parts.append(part)
    return "&".join(parts) if parts else "N"


def vf_security_text(value: Any) -> str:
    text = usable_text(value) or "N"
    normalized = normalize_access(text)
    if normalized in {"", "N", "NO", "NONE", "LEVEL_0", "LEVEL0"}:
        return "N"
    return text


def vf_hex_tokens(text: str) -> list[int]:
    values: list[int] = []
    token_pattern = r"(?i)(?<![0-9a-z])(?:0x[0-9a-f]{1,2}|[0-9][0-9a-f]?|[a-f][0-9a-f])(?=$|[^0-9a-z])"
    for token in re.findall(token_pattern, text or ""):
        try:
            values.append(int(token, 16) if token.lower().startswith("0x") else int(token, 16))
        except ValueError:
            continue
    return values


def vf_subfunction_values(value_text: str, meaning_text: str) -> list[int]:
    """Return base sub-functions, normalizing suppress-positive-response values."""

    search_text = meaning_text.split("=", 1)[0] if "=" in meaning_text else value_text
    values = []
    for value in vf_hex_tokens(search_text):
        if value <= 0xFF:
            values.append(value & 0x7F)
    return list(dict.fromkeys(values))


def vf_parse_diagnostics_services_access(workbook: Any) -> dict[tuple[int, int | None], dict[str, Any]]:
    """Parse VF's Diagnostics Services support matrix.

    VF keeps the core UDS service support flags in one worksheet:
    Application Default/Extended are columns G/H, Bootloader
    Default/Programming are columns I/J, and each request sub-function row
    carries the supported sessions for that sub-function.
    """

    sheet = find_sheet(workbook, lambda name: name.strip() == "Diagnostics Services")
    if sheet is None:
        return {}

    access: dict[tuple[int, int | None], dict[str, Any]] = {}
    current_service_id: int | None = None
    current_service_name = ""
    current_security = "N"
    in_request_block = False
    in_subfunction_rows = False

    for row in range(1, sheet.max_row + 1):
        first_col = usable_text(sheet.cell(row, 1).value) or ""
        first_norm = first_col.strip().lower()
        if first_col and "$" in first_col and "service" in first_norm:
            current_service_name = first_col
            current_service_id = None
            current_security = "N"
            in_request_block = False
            in_subfunction_rows = False
            continue

        if first_norm.startswith("block"):
            in_request_block = "request" in first_norm
            if in_request_block:
                current_service_id = None
                current_security = "N"
                in_subfunction_rows = False
            continue

        if not in_request_block:
            continue

        description = usable_text(sheet.cell(row, 3).value) or ""
        value_text = usable_text(sheet.cell(row, 4).value) or ""
        meaning_text = usable_text(sheet.cell(row, 5).value) or ""
        description_norm = description.strip().lower()

        if "request sid" in description_norm:
            service_id = parse_hex_in_text(value_text, max_value=0xFF)
            if service_id is not None:
                current_service_id = service_id
                current_security = vf_security_text(sheet.cell(row, 13).value)
            continue

        if current_service_id is None:
            continue
        if current_service_id not in {0x10, 0x27, 0x28}:
            continue

        is_subfunction_header = any(
            marker in description_norm
            for marker in ("session type", "sub-function", "sub function", "control type")
        )
        if description_norm and not is_subfunction_header:
            in_subfunction_rows = False
        if is_subfunction_header:
            in_subfunction_rows = True
        elif not in_subfunction_rows or "=" not in meaning_text:
            continue

        requirement = compact_text(sheet.cell(row, 6).value).upper()
        if requirement == "U":
            continue

        subfunctions = vf_subfunction_values(value_text, meaning_text)
        if not subfunctions:
            continue

        sessions = {
            state_name
            for col, state_name in ((7, "Default"), (8, "Extended"), (9, "Default"), (10, "Programming"))
            if is_supported_flag(sheet.cell(row, col).value)
        }
        if not sessions:
            continue

        sources = {
            source
            for col, source in ((7, "Application"), (8, "Application"), (9, "Boot"), (10, "Boot"))
            if is_supported_flag(sheet.cell(row, col).value)
        }
        if not sources:
            continue

        security = vf_security_text(sheet.cell(row, 13).value) if usable_text(sheet.cell(row, 13).value) else current_security
        subfunction_name = meaning_text.split("=", 1)[1].split("/", 1)[0].strip() if "=" in meaning_text else description
        for subfunction in subfunctions:
            key = (current_service_id, subfunction)
            item = access.setdefault(
                key,
                {
                    "service_id": current_service_id,
                    "subfunction": subfunction,
                    "service_name": current_service_name,
                    "subfunction_name": subfunction_name,
                    "sessions": set(),
                    "security": "N",
                    "sources": set(),
                },
            )
            item["sessions"].update(sessions)
            item["security"] = merge_security_text(item["security"], security)
            item["sources"].update(sources)

    return access


# ---------------------------------------------------------------------------
# VF workbook adapter
# ---------------------------------------------------------------------------


def vf_is_end(value: Any) -> bool:
    return compact_text(value).upper().startswith("#END")


def vf_yes(value: Any, default: bool = False) -> bool:
    text = compact_text(value).upper()
    if text in {"Y", "YES", "TRUE", "SUPPORTED", "SUPPORT", "TBD"}:
        return True
    if text in {"N", "NO", "FALSE", "-", "/", "N/A", "NA", "NOT SUPPORTED", "NOTSUPPORTED"}:
        return False
    return default


def vf_find_sheet(workbook: Any, name: str) -> Any | None:
    return find_sheet(workbook, lambda sheet_name: sheet_name.strip().casefold() == name.casefold())


def vf_parse_baudrate(value: Any, default: int = 500000) -> int:
    text = compact_text(value).lower()
    number = parse_int_cell(text, default)
    if "kbps" in text or "kbit" in text or text.endswith("k"):
        return number * 1000
    return number


def vf_record_number(value: Any) -> int | None:
    text = compact_text(value).replace("0x", "").replace("0X", "")
    match = re.search(r"[0-9A-Fa-f]{1,2}", text)
    if not match:
        return None
    return int(match.group(0), 16)


def vf_make_param(
    *,
    name_value: Any,
    byte_value: Any,
    bit_value: Any,
    bit_len_value: Any,
    data_type_value: Any,
    method_value: Any = "",
    unit_value: Any = "",
    min_value: Any = "",
    max_value: Any = "",
    fallback_name: str = "Data",
    byte_offset: int = 0,
) -> ParamDef | None:
    name, long_name = split_name(cell_text(name_value) or fallback_name)
    byte_abs = parse_int_cell(byte_value, default=-999999)
    bit_len = parse_int_cell(bit_len_value, default=0)
    if byte_abs == -999999 or bit_len <= 0:
        return None
    bit_pos = parse_int_cell(bit_value, default=0)
    conversion = parse_conversion(method_value)
    if conversion.kind == "identity" and re.fullmatch(r"(?i)\s*BCD\s*", usable_text(data_type_value)):
        conversion = Conversion(kind="bcd")
    if bit_len > 32 and conversion.kind == "enum":
        conversion = Conversion()
    return ParamDef(
        name=name or fallback_name,
        long_name=long_name or name or fallback_name,
        byte_pos=max(0, byte_abs - byte_offset),
        bit_pos=bit_pos,
        bit_len=bit_len,
        data_type=usable_text(data_type_value) or "Unsigned",
        unit=usable_text(unit_value),
        conversion=conversion,
        min_value=usable_text(min_value),
        max_value=usable_text(max_value),
    )


def vf_param_size(params: Iterable[ParamDef], default: int = 1) -> int:
    max_end = 0
    for param in params:
        max_end = max(max_end, param.byte_pos + max(1, (param.bit_pos + max(1, param.bit_len) + 7) // 8))
    return max(default, max_end)


def normalized_data_type(value: Any) -> str:
    return compact_text(value).upper()


def is_ascii_data_type(value: Any) -> bool:
    return "ASCII" in normalized_data_type(value)


def merge_split_byte_did_params(did: DidDef) -> None:
    if len(did.params) <= 1:
        return

    ordered = sorted(did.params, key=lambda param: (param.byte_pos, param.bit_pos, param.name))
    data_type = normalized_data_type(ordered[0].data_type)
    unit = compact_text(ordered[0].unit)
    if not data_type or not is_ascii_data_type(data_type):
        return

    for expected_byte, param in enumerate(ordered):
        if (
            normalized_data_type(param.data_type) != data_type
            or compact_text(param.unit) != unit
            or param.byte_pos != expected_byte
            or param.bit_pos != 0
            or param.bit_len != 8
            or param.conversion.kind != "identity"
        ):
            return

    identifier, display = split_name(did.desc)
    name = display or identifier or hex_short("DID", did.did)
    did.params = [
        ParamDef(
            name=identifier or name,
            long_name=name,
            byte_pos=0,
            bit_pos=0,
            bit_len=len(ordered) * 8,
            data_type=ordered[0].data_type,
            unit=unit,
        )
    ]


def vf_parse_cover(workbook: Any) -> CoverInfo:
    cover = CoverInfo(ecu_name="VF_ECU_CAN")
    sheet = vf_find_sheet(workbook, "Change Log & General Info")
    rows: dict[str, str] = {}
    if sheet is not None:
        for row in range(1, sheet.max_row + 1):
            key = compact_text(sheet.cell(row, 1).value)
            if key:
                rows[key.casefold()] = compact_text(sheet.cell(row, 2).value)

    cover.ecu_name = rows.get("ecu-name", "") or "VF_ECU_CAN"
    cover.vehicle = rows.get("vehicle program - project name", "")
    cover.supplier = rows.get("ecu supplier", "")
    cover.bus_type = "CAN"
    cover.rx_fun_id = parse_hex_in_text(rows.get("funcreqid", ""), max_value=0x1FFFFFFF)
    cover.rx_phy_id = parse_hex_in_text(rows.get("physreqid", ""), max_value=0x1FFFFFFF)
    cover.tx_id = parse_hex_in_text(rows.get("physrespid", ""), max_value=0x1FFFFFFF)
    baudrate = vf_parse_baudrate(rows.get("baudrate", "500kbps"))

    cover.comm_params = {
        "ISO_15765_2.CP_CanFuncReqId": cover.rx_fun_id if cover.rx_fun_id is not None else 0x6FF,
        "ISO_15765_3.CP_P2Max": 150_000,
        "ISO_15765_3.CP_P2Star": 5_100_000,
        "ISO_15765_3.CP_TesterPresentTime": 2_000_000,
        "ISO_15765_3.CP_P3Phys": 100_000,
        "ISO_15765_3.CP_P3Func": 100_000,
        "ISO_15765_2.CP_StMin": 20,
        "ISO_15765_2.CP_BlockSize": 0,
        "ISO_15765_2.CP_As": 70_000,
        "ISO_15765_2.CP_Ar": 70_000,
        "ISO_15765_2.CP_Bs": 150_000,
        "ISO_15765_2.CP_Br": 70_000,
        "ISO_15765_2.CP_Cs": 70_000,
        "ISO_15765_2.CP_Cr": 150_000,
        "ISO_11898_2_DWCAN.CP_Baudrate": baudrate,
    }
    cover.session_timing = {"P2": 50, "P2Ex": 2000}
    return cover


def vf_parse_did_sheet(sheet: Any, *, system_sheet: bool) -> list[DidDef]:
    dids: dict[int, DidDef] = {}
    current: DidDef | None = None
    current_supported = True

    for row in range(3, sheet.max_row + 1):
        if vf_is_end(sheet.cell(row, 1).value):
            break
        did_value = parse_hex_cell(sheet.cell(row, 1).value, max_value=0xFFFF)
        if did_value is not None:
            current_supported = vf_yes(sheet.cell(row, 15).value, default=False) if system_sheet else True
            if not current_supported:
                current = None
                continue
            desc_name, desc_display = split_name(cell_text(sheet.cell(row, 2).value))
            desc = canonical_name(desc_name, desc_display, hex_short("DID", did_value))
            rw_state = usable_text(sheet.cell(row, 3).value) or "R"
            app_access = usable_text(sheet.cell(row, 11).value) or rw_state
            boot_access = usable_text(sheet.cell(row, 12).value)
            security = usable_text(sheet.cell(row, 13).value)
            sessions = [app_access, app_access, boot_access, boot_access, app_access]
            current = dids.setdefault(
                did_value,
                DidDef(
                    did=did_value,
                    desc=desc,
                    size=0,
                    write_security=security or "N",
                    sessions=sessions,
                ),
            )
            current.desc = current.desc or desc
            current.write_security = current.write_security if current.write_security != "N" else (security or "N")
            current.sessions = current.sessions or sessions

        if current is None or not current_supported or not usable_text(sheet.cell(row, 4).value):
            continue

        param = vf_make_param(
            name_value=sheet.cell(row, 4).value,
            byte_value=sheet.cell(row, 5).value,
            bit_value=sheet.cell(row, 6).value,
            bit_len_value=sheet.cell(row, 7).value,
            data_type_value=sheet.cell(row, 8).value,
            method_value=sheet.cell(row, 9).value,
            unit_value=sheet.cell(row, 10).value,
            min_value=sheet.cell(row, 14).value if not system_sheet else "",
            max_value=sheet.cell(row, 15).value if not system_sheet else "",
            fallback_name=current.desc,
            byte_offset=3,
        )
        if param is not None:
            current.params.append(param)

    for did in dids.values():
        did.size = vf_param_size(did.params)
    return list(dids.values())


def vf_parse_dids(workbook: Any) -> list[DidDef]:
    result: dict[int, DidDef] = {}
    for sheet_name, system_sheet in (("System DID", True), ("ECU DID", False)):
        sheet = vf_find_sheet(workbook, sheet_name)
        if sheet is None:
            continue
        for did in vf_parse_did_sheet(sheet, system_sheet=system_sheet):
            existing = result.get(did.did)
            if existing is None:
                result[did.did] = did
            else:
                existing.params.extend(did.params)
                existing.size = max(existing.size, did.size)
                existing.write_security = existing.write_security if existing.write_security != "N" else did.write_security
                existing.sessions = existing.sessions or did.sessions
    for did in result.values():
        merge_split_byte_did_params(did)
        did.size = vf_param_size(did.params)
    return list(result.values())


def vf_parse_io_dids(workbook: Any) -> list[IoDidDef]:
    sheet = vf_find_sheet(workbook, "IO Control (2F)")
    if sheet is None:
        return []
    by_id: dict[int, IoDidDef] = {}
    current: IoDidDef | None = None
    current_control: int | None = None
    current_direction = ""
    for row in range(24, sheet.max_row + 1):
        if vf_is_end(sheet.cell(row, 1).value):
            break
        did_value = parse_hex_cell(sheet.cell(row, 2).value, max_value=0xFFFF)
        if did_value is not None:
            desc = usable_text(sheet.cell(row, 3).value) or hex_short("IODID", did_value)
            current = by_id.setdefault(did_value, IoDidDef(did=did_value, desc=desc, size=0))
        control = parse_hex_cell(sheet.cell(row, 1).value, max_value=0xFF)
        if control is not None:
            current_control = control
            if current is not None:
                current.controls.add(control)
        direction = usable_text(sheet.cell(row, 4).value)
        if direction:
            current_direction = direction
        if current is None or not usable_text(sheet.cell(row, 5).value):
            continue
        param = vf_make_param(
            name_value=sheet.cell(row, 5).value,
            byte_value=sheet.cell(row, 6).value,
            bit_value=sheet.cell(row, 7).value,
            bit_len_value=sheet.cell(row, 8).value,
            data_type_value=sheet.cell(row, 9).value,
            method_value=sheet.cell(row, 10).value,
            unit_value=sheet.cell(row, 11).value,
            fallback_name=current.desc,
            byte_offset=4,
        )
        if param is not None and not normalize_access(current_direction).startswith("RESP"):
            current.params.append(param)
        if current_control is not None:
            current.controls.add(current_control)
    for io_did in by_id.values():
        io_did.size = vf_param_size(io_did.params)
        if not io_did.controls:
            io_did.controls.add(3)
    return list(by_id.values())


def vf_parse_routine_control_type(value: Any) -> int | None:
    parsed = parse_hex_cell(value, max_value=0xFF)
    if parsed is not None and parsed in {1, 2, 3}:
        return parsed
    match = re.search(r"\b([123])\b", compact_text(value))
    return int(match.group(1)) if match else None


def vf_parse_routines(workbook: Any) -> list[RoutineDef]:
    sheet = vf_find_sheet(workbook, "Routines Data (0x31)")
    if sheet is None:
        return []
    routines: dict[int, RoutineDef] = {}
    current_routine: RoutineDef | None = None
    current_subfn: RoutineSubFunction | None = None
    current_direction = ""
    current_supported = True
    for row in range(3, sheet.max_row + 1):
        if vf_is_end(sheet.cell(row, 1).value):
            break
        rid = parse_hex_cell(sheet.cell(row, 3).value, max_value=0xFFFF)
        control_type = vf_parse_routine_control_type(sheet.cell(row, 2).value)
        if rid is not None:
            current_supported = vf_yes(sheet.cell(row, 15).value, default=True)
            desc = usable_text(sheet.cell(row, 4).value) or hex_short("RID", rid)
            current_routine = routines.setdefault(
                rid,
                RoutineDef(
                    rid=rid,
                    desc=desc,
                    security=usable_text(sheet.cell(row, 13).value) or "N",
                    sessions=[usable_text(sheet.cell(row, 1).value)],
                ),
            )
        if current_routine is not None and control_type is not None:
            current_subfn = current_routine.subfunctions.setdefault(
                control_type, RoutineSubFunction(control_type=control_type, supported=current_supported)
            )
            current_subfn.supported = current_subfn.supported or current_supported
        direction = usable_text(sheet.cell(row, 5).value)
        if direction:
            current_direction = direction
        if current_routine is None or current_subfn is None or not current_supported or not usable_text(sheet.cell(row, 6).value):
            continue
        param = vf_make_param(
            name_value=sheet.cell(row, 6).value,
            byte_value=sheet.cell(row, 7).value,
            bit_value=sheet.cell(row, 8).value,
            bit_len_value=sheet.cell(row, 9).value,
            data_type_value=sheet.cell(row, 10).value,
            method_value=sheet.cell(row, 11).value,
            unit_value=sheet.cell(row, 12).value,
            fallback_name=current_routine.desc,
            byte_offset=4,
        )
        if param is None:
            continue
        if normalize_access(current_direction).startswith("RESP"):
            current_subfn.status_params.append(param)
        else:
            current_subfn.option_params.append(param)
    return list(routines.values())


def vf_encode_dtc(value: Any) -> tuple[int, str] | None:
    text = compact_text(value).upper().replace(" ", "").replace("_", "")
    text = re.sub(r"^0X", "", text)
    match = re.fullmatch(r"([PCBU])([0-3A-F][0-9A-F]{3})([0-9A-F]{2})", text)
    if match:
        category, base, ftb = match.groups()
        category_base = {"P": 0x0, "C": 0x4, "B": 0x8, "U": 0xC}[category]
        first = int(base[0], 16) & 0x3
        encoded = f"{category_base + first:X}{base[1:]}{ftb}"
        return int(encoded, 16), f"{category}{base}{ftb}"
    match = re.fullmatch(r"([0-9A-F]{4})-?([0-9A-F]{2})", text)
    if match:
        encoded = "".join(match.groups())
        return int(encoded, 16), f"{match.group(1)}-{match.group(2)}"
    match = re.search(r"([0-9A-F]{6})", text)
    if match:
        return int(match.group(1), 16), match.group(1)
    return None


def vf_parse_dtcs(workbook: Any) -> list[DtcDef]:
    sheet = vf_find_sheet(workbook, "DTC-List")
    if sheet is None:
        return []
    result: list[DtcDef] = []
    seen: set[int] = set()
    for row in range(3, sheet.max_row + 1):
        encoded = vf_encode_dtc(sheet.cell(row, 2).value) or vf_encode_dtc(sheet.cell(row, 3).value)
        if encoded is None:
            continue
        byte_code, display = encoded
        if byte_code in seen:
            continue
        seen.add(byte_code)
        priority = re.search(r"\d+", compact_text(sheet.cell(row, 6).value))
        result.append(
            DtcDef(
                display_code=display,
                byte_code=byte_code,
                text=usable_text(sheet.cell(row, 5).value) or display,
                priority=priority.group(0) if priority else usable_text(sheet.cell(row, 6).value),
            )
        )
    return result


def vf_snapshot_record_nums(value: Any) -> list[int]:
    return sorted({int(match, 16) for match in re.findall(r"0[xX]([0-9A-Fa-f]{1,2})", compact_text(value))})


def vf_snapshot_record_names(value: Any) -> dict[int, str]:
    names: dict[int, str] = {}
    for line in re.split(r"[\r\n]+", cell_text(value)):
        match = re.search(r"0[xX]([0-9A-Fa-f]{1,2})\s*[-–—]\s*(.+)", line)
        if not match:
            continue
        record_num = int(match.group(1), 16)
        label = compact_text(match.group(2))
        if label:
            names[record_num] = label
    return names


def vf_parse_snapshots(workbook: Any) -> tuple[list[SnapshotDef], list[int], dict[int, str]]:
    sheet = vf_find_sheet(workbook, "Snapshot DIDs")
    if sheet is None:
        return [], [], {}
    record_header = sheet.cell(3, 1).value
    record_nums = vf_snapshot_record_nums(record_header) or [1, 2]
    record_names = vf_snapshot_record_names(record_header)
    snapshots: list[SnapshotDef] = []
    for row in range(5, sheet.max_row + 1):
        if vf_is_end(sheet.cell(row, 1).value):
            break
        did = parse_hex_cell(sheet.cell(row, 1).value, max_value=0xFFFF)
        if did is None or not vf_yes(sheet.cell(row, 11).value, default=False):
            continue
        size = parse_int_cell(sheet.cell(row, 6).value, default=0)
        if size <= 0:
            continue
        desc = usable_text(sheet.cell(row, 4).value) or hex_short("Snapshot", did)
        snapshots.append(
            SnapshotDef(
                record_num=None,
                did=did,
                desc=desc,
                size=size,
                params=[
                    ParamDef(
                        name=desc,
                        long_name=desc,
                        byte_pos=0,
                        bit_pos=0,
                        bit_len=max(8, size * 8),
                        data_type=usable_text(sheet.cell(row, 5).value) or "Unsigned",
                    )
                ],
            )
        )
    return snapshots, record_nums, record_names


def vf_parse_extended_records(workbook: Any) -> list[ExtendedRecordDef]:
    sheet = vf_find_sheet(workbook, "Extended Data")
    if sheet is None:
        return []
    result: list[ExtendedRecordDef] = []
    for row in range(6, sheet.max_row + 1):
        if vf_is_end(sheet.cell(row, 1).value):
            break
        if not vf_yes(sheet.cell(row, 8).value, default=False):
            continue
        record_num = vf_record_number(sheet.cell(row, 2).value)
        size = parse_int_cell(sheet.cell(row, 3).value, default=0)
        if record_num is None or size <= 0:
            continue
        desc = usable_text(sheet.cell(row, 1).value) or f"Extended Record 0x{record_num:02X}"
        result.append(
            ExtendedRecordDef(
                record_num=record_num,
                desc=desc,
                size=size,
                params=[
                    ParamDef(
                        name=desc,
                        long_name=desc,
                        byte_pos=0,
                        bit_pos=0,
                        bit_len=max(8, size * 8),
                        data_type=usable_text(sheet.cell(row, 4).value) or "INT",
                    )
                ],
            )
        )
    return result


def parse_vf_survey(xlsx_path: Path) -> SurveyData:
    workbook = load_workbook(xlsx_path, data_only=True)
    cover = vf_parse_cover(workbook)
    dids = vf_parse_dids(workbook)
    io_dids = vf_parse_io_dids(workbook)
    routines = vf_parse_routines(workbook)
    dtcs = vf_parse_dtcs(workbook)
    snapshots, snapshot_record_nums, snapshot_record_names = vf_parse_snapshots(workbook)
    extended_records = vf_parse_extended_records(workbook)
    survey = SurveyData(cover, dids, io_dids, routines, dtcs, snapshots, extended_records)
    survey.snapshot_record_nums = snapshot_record_nums
    survey.snapshot_record_names = snapshot_record_names
    survey.core_service_access = vf_parse_diagnostics_services_access(workbook)
    return survey


def update_template(template_pdx: Path, output_pdx: Path, survey: SurveyData, validate: bool = True) -> None:
    with tempfile.TemporaryDirectory(prefix="vf_pdx_") as tmp_name:
        tmp_dir = Path(tmp_name)
        with zipfile.ZipFile(template_pdx, "r") as archive:
            archive.extractall(tmp_dir)

        odx_path = tmp_dir / "VF_ECU_CAN_v15.odx-d"
        if not odx_path.exists():
            candidates = list(tmp_dir.glob("*.odx-d"))
            if not candidates:
                raise FileNotFoundError("No .odx-d file found inside template PDX")
            odx_path = candidates[0]

        parser = etree.XMLParser(remove_blank_text=False)
        tree = etree.parse(str(odx_path), parser)
        root = tree.getroot()
        id_gen = IdGenerator(root)
        update_odx_vf(root, id_gen, survey)
        validate_diag_service_child_order(root, odx_path.name)
        validate_can_dela_odx_structure(root, odx_path.name)
        tree.write(str(odx_path), encoding="UTF-8", xml_declaration=True, pretty_print=True, standalone=False)

        keep_files = set(VF_KEEP_FILES)
        keep_files.add(odx_path.name)
        patch_pdx_catalog_for_can_only(tmp_dir / "index.xml", keep_files)

        output_pdx.parent.mkdir(parents=True, exist_ok=True)
        if output_pdx.exists():
            output_pdx.unlink()
        with zipfile.ZipFile(output_pdx, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for path in sorted(tmp_dir.rglob("*")):
                rel_path = path.relative_to(tmp_dir).as_posix()
                if path.is_file() and rel_path in keep_files:
                    archive.write(path, rel_path)

    if validate:
        validate_with_odxtools(output_pdx)


def update_odx_vf(root: etree._Element, id_gen: Any, survey: SurveyData) -> None:
    update_layer_names(root, survey.cover)
    ddds = find_dictionary_spec(root)
    if ddds is None:
        raise RuntimeError("Template ODX has no DIAG-DATA-DICTIONARY-SPEC")
    data_object_props = ddds.find("DATA-OBJECT-PROPS")
    structures = ddds.find("STRUCTURES")
    if data_object_props is None or structures is None:
        raise RuntimeError("Template ODX is missing DOP/STRUCTURE containers")

    unit_ids = ensure_units(ddds, survey)
    generated_dop_cache: dict[tuple[str, int, str, str, str], str] = {}

    for did in survey.dids:
        prepare_data_structure(
            id_gen=id_gen,
            data_object_props=data_object_props,
            structures=structures,
            item=did,
            prefix="DID",
            unit_ids=unit_ids,
            generated_dop_cache=generated_dop_cache,
            direct_single_param=True,
        )

    for io_did in survey.io_dids:
        prepare_data_structure(
            id_gen=id_gen,
            data_object_props=data_object_props,
            structures=structures,
            item=io_did,
            prefix="IODID",
            unit_ids=unit_ids,
            generated_dop_cache=generated_dop_cache,
        )

    for routine in survey.routines:
        english, long_name = split_name(routine.desc)
        routine.short_name = sanitize_short_name(english or hex_short("RID", routine.rid), hex_short("RID", routine.rid))
        routine.long_name = long_name or routine.short_name
        for subfn in routine.subfunctions.values():
            if subfn.option_params:
                subfn.option_structure_id = make_param_structure(
                    id_gen,
                    data_object_props,
                    structures,
                    f"STR_RoutineOption_{routine.short_name}_{subfn.control_type:02X}",
                    f"Routine Option {routine.long_name}",
                    subfn.option_params,
                    unit_ids,
                    generated_dop_cache,
                )
            if subfn.status_params:
                subfn.status_structure_id = make_param_structure(
                    id_gen,
                    data_object_props,
                    structures,
                    f"STR_RoutineStatus_{routine.short_name}_{subfn.control_type:02X}",
                    f"Routine Status {routine.long_name}",
                    subfn.status_params,
                    unit_ids,
                    generated_dop_cache,
                )

    for snapshot in survey.snapshots:
        if snapshot.params:
            snapshot.structure_id = make_param_structure(
                id_gen,
                data_object_props,
                structures,
                f"STR_Snapshot_{sanitize_short_name(snapshot.desc, hex_short('Snapshot', snapshot.did))}",
                snapshot.desc,
                snapshot.params,
                unit_ids,
                generated_dop_cache,
                byte_size=snapshot.size or None,
            )

    for record in survey.extended_records:
        if record.params:
            record.structure_id = make_param_structure(
                id_gen,
                data_object_props,
                structures,
                f"STR_ExtendedRecord_{record.record_num:02X}",
                record.desc,
                record.params,
                unit_ids,
                generated_dop_cache,
                byte_size=record.size or None,
            )

    ensure_vf_placeholder_services(root, id_gen)
    generate_flat_did_services(root, id_gen, survey.dids)
    generate_flat_io_services(root, id_gen, survey.io_dids)
    generate_flat_routine_services(root, id_gen, survey.routines)
    for old_name in ("z_7_Read", "z_Read", "z_Control", "z_ReturnControl"):
        remove_service_and_messages(root, old_name)
    remove_service_and_messages(root, "Upload_Download_RequestDownload")
    update_vf_dtc_dop(root, id_gen, survey.dtcs)
    update_vf_snapshot_and_extended_data(
        root,
        id_gen,
        survey.snapshots,
        survey.extended_records,
        getattr(survey, "snapshot_record_nums", []),
        getattr(survey, "snapshot_record_names", {}),
    )
    update_base_variant_comparams(root, survey.cover)
    update_session_timing(root, survey.cover)
    ensure_core_response_services(root, id_gen)
    ensure_vf_session_control_services(root, id_gen, survey)
    ensure_vf_communication_control_services(root, id_gen, survey)
    ensure_vf_security_access_subfunction_scales(root, survey)
    prune_vf_unsupported_security_access_services(root, survey)
    generate_boot_security_access_services(root, id_gen, survey)
    update_vf_security_access_preconditions(root, id_gen, survey)
    update_core_service_preconditions(root, survey)
    shorten_dictionary_short_names(root)
    prefix_doc_revision_labels(root)


def update_layer_names(root: etree._Element, cover: CoverInfo) -> None:
    container = root.find(".//DIAG-LAYER-CONTAINER")
    if container is not None:
        long_name = container.find("LONG-NAME")
        if long_name is not None:
            long_name.text = cover.ecu_name
    base_variant = get_base_variant(root)
    if base_variant is not None:
        long_name = base_variant.find("LONG-NAME")
        if long_name is not None:
            long_name.text = cover.ecu_name


def get_base_variant(root: etree._Element) -> etree._Element | None:
    node = first_by_short(root, "BASE-VARIANT", "VF_ECU_CAN")
    if node is not None:
        return node
    nodes = root.xpath("//*[local-name()='BASE-VARIANT']")
    return nodes[0] if nodes else None


def find_dictionary_spec(root: etree._Element) -> etree._Element | None:
    for ddds in root.xpath("//*[local-name()='DIAG-DATA-DICTIONARY-SPEC']"):
        if ddds.find("DATA-OBJECT-PROPS") is not None and ddds.find("STRUCTURES") is not None:
            return ddds
    nodes = root.xpath("//*[local-name()='DIAG-DATA-DICTIONARY-SPEC']")
    return nodes[0] if nodes else None


def candela_short_name(value: str, used: set[str] | None = None, max_len: int = CANDELA_SHORT_NAME_MAX_LEN) -> str:
    if len(value) <= max_len and (used is None or value not in used):
        if used is not None:
            used.add(value)
        return value

    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:8].upper()
    suffix = f"_{digest}"
    head_len = max(1, max_len - len(suffix))
    base = value[:head_len].rstrip("_") or value[:head_len]
    candidate = f"{base}{suffix}"
    if used is not None:
        index = 2
        while candidate in used:
            index_suffix = f"_{index:02d}"
            candidate = f"{candidate[:max_len - len(index_suffix)].rstrip('_')}{index_suffix}"
            index += 1
        used.add(candidate)
    return candidate


def shorten_dictionary_short_names(root: etree._Element) -> None:
    for container_name, child_name in (
        ("DATA-OBJECT-PROPS", "DATA-OBJECT-PROP"),
        ("STRUCTURES", "STRUCTURE"),
    ):
        for container in root.xpath(f"//*[local-name()='{container_name}']"):
            used: set[str] = set()
            for node in container.xpath(f"./*[local-name()='{child_name}']"):
                short = node.find("SHORT-NAME")
                if short is None or not short.text:
                    continue
                short.text = candela_short_name(short.text, used)


def validate_diag_service_child_order(root: etree._Element, source_name: str) -> None:
    order = {
        "SHORT-NAME": 0,
        "LONG-NAME": 1,
        "DESC": 2,
        "ADMIN-DATA": 3,
        "SDGS": 4,
        "FUNCT-CLASS-REFS": 5,
        "AUDIENCE": 6,
        "PROTOCOL-SNREFS": 7,
        "RELATED-DIAG-COMM-REFS": 8,
        "PRE-CONDITION-STATE-REFS": 9,
        "STATE-TRANSITION-REFS": 10,
        "COMPARAM-REFS": 11,
        "REQUEST-REF": 12,
        "POS-RESPONSE-REFS": 13,
        "NEG-RESPONSE-REFS": 14,
        "POS-RESPONSE-SUPPRESSABLE": 15,
    }
    errors: list[str] = []
    for service in root.xpath("//*[local-name()='DIAG-SERVICE']"):
        service_name = service.findtext("SHORT-NAME") or service.get("ID") or "<unnamed>"
        last_index = -1
        for child in service:
            if not isinstance(child.tag, str):
                continue
            child_name = etree.QName(child).localname
            index = order.get(child_name)
            if index is None:
                continue
            if index < last_index:
                errors.append(f"DIAG-SERVICE '{service_name}' has '{child_name}' out of ODX order")
                break
            last_index = index
    if errors:
        details = "\n".join(f"- {message}" for message in errors[:20])
        raise RuntimeError(f"Generated {source_name} has invalid DIAG-SERVICE child order:\n{details}")


def get_short_name(node: etree._Element) -> str:
    return node.findtext("SHORT-NAME") or ""


def find_container(root: etree._Element, name: str) -> etree._Element:
    node = root.find(f".//{name}")
    if node is None:
        raise RuntimeError(f"Template ODX is missing {name}")
    return node


def find_by_id(root: etree._Element, node_id: str | None) -> etree._Element | None:
    if not node_id:
        return None
    nodes = root.xpath("//*[@ID=$node_id]", node_id=node_id)
    return nodes[0] if nodes else None


def service_messages(root: etree._Element, service_short_name: str) -> tuple[etree._Element, etree._Element, etree._Element | None, etree._Element | None]:
    service = first_by_short(root, "DIAG-SERVICE", service_short_name)
    if service is None:
        raise RuntimeError(f"Template service {service_short_name} was not found")
    request_ref = service.find("REQUEST-REF")
    request = find_by_id(root, request_ref.get("ID-REF") if request_ref is not None else None)
    if request is None:
        raise RuntimeError(f"Template service {service_short_name} has no request")
    pos_ref = service.find(".//POS-RESPONSE-REF")
    neg_ref = service.find(".//NEG-RESPONSE-REF")
    positive = find_by_id(root, pos_ref.get("ID-REF") if pos_ref is not None else None)
    negative = find_by_id(root, neg_ref.get("ID-REF") if neg_ref is not None else None)
    return service, request, positive, negative


def clone_service_bundle(
    root: etree._Element,
    id_gen: Any,
    base_service_short_name: str,
    new_short_name: str,
    long_name: str,
    service_qualifier: str,
    service_name: str,
    instance_qualifier: str,
    instance_name: str,
    precondition_state_ids: Iterable[str] = (),
    static_value: int | None = None,
    positive_response_suppressed: str = "no",
) -> tuple[etree._Element, etree._Element, etree._Element | None, etree._Element | None]:
    base_service, base_request, base_positive, base_negative = service_messages(root, base_service_short_name)
    request = copy.deepcopy(base_request)
    positive = copy.deepcopy(base_positive) if base_positive is not None else None
    negative = copy.deepcopy(base_negative) if base_negative is not None else None
    service = copy.deepcopy(base_service)

    request_id = id_gen.new("RQ")
    positive_id = id_gen.new("PR") if positive is not None else None
    negative_id = id_gen.new("NR") if negative is not None else None
    service_id = id_gen.new("SVC")

    request.set("ID", request_id)
    refresh_internal_ids(request, id_gen)
    set_short_long(request, f"RQ_{new_short_name}", f"RQ {long_name}")

    if positive is not None and positive_id is not None:
        positive.set("ID", positive_id)
        refresh_internal_ids(positive, id_gen)
        set_short_long(positive, f"PR_{new_short_name}", f"PR {long_name}")

    if negative is not None and negative_id is not None:
        negative.set("ID", negative_id)
        refresh_internal_ids(negative, id_gen)
        set_short_long(negative, f"NR_{new_short_name}", f"NR {long_name}")

    service.set("ID", service_id)
    refresh_internal_ids(service, id_gen)
    set_short_long(service, new_short_name, long_name)
    request_ref = service.find("REQUEST-REF")
    if request_ref is not None:
        request_ref.set("ID-REF", request_id)
    pos_ref = service.find(".//POS-RESPONSE-REF")
    if pos_ref is not None and positive_id is not None:
        pos_ref.set("ID-REF", positive_id)
    neg_ref = service.find(".//NEG-RESPONSE-REF")
    if neg_ref is not None and negative_id is not None:
        neg_ref.set("ID-REF", negative_id)
    update_service_sdg(
        service,
        id_gen,
        service_qualifier,
        service_name,
        instance_qualifier,
        instance_name,
        static_value,
        positive_response_suppressed=positive_response_suppressed,
    )
    update_flat_preconditions(service, precondition_state_ids)

    find_container(root, "REQUESTS").append(request)
    if positive is not None:
        find_container(root, "POS-RESPONSES").append(positive)
    if negative is not None:
        find_container(root, "NEG-RESPONSES").append(negative)
    find_container(root, "DIAG-COMMS").append(service)
    return service, request, positive, negative


def update_service_sdg(
    service: etree._Element,
    id_gen: Any,
    service_qualifier: str,
    service_name: str,
    instance_qualifier: str,
    instance_name: str,
    static_value: int | None = None,
    positive_response_suppressed: str = "no",
) -> None:
    sdgs = element("SDGS")
    sdg = sub(sdgs, "SDG")
    caption = sub(sdg, "SDG-CAPTION", attrib={"ID": id_gen.new("CAP")})
    sub(caption, "SHORT-NAME", "CANdelaServiceInformation")
    if static_value is not None:
        sub(sdg, "SD", static_value, attrib={"SI": "DiagInstanceStaticValue"})
    sub(sdg, "SD", sanitize_short_name(instance_qualifier, "Instance"), attrib={"SI": "DiagInstanceQualifier"})
    sub(sdg, "SD", instance_name or instance_qualifier, attrib={"SI": "DiagInstanceName"})
    sub(sdg, "SD", service_qualifier, attrib={"SI": "ServiceQualifier"})
    sub(sdg, "SD", service_name, attrib={"SI": "ServiceName"})
    sub(sdg, "SD", positive_response_suppressed, attrib={"SI": "PositiveResponseSuppressed"})
    replace_child(service, "SDGS", sdgs, before_tags={"FUNCT-CLASS-REFS", "AUDIENCE", "PRE-CONDITION-STATE-REFS", "REQUEST-REF"})


def update_flat_preconditions(service: etree._Element, state_ids: Iterable[str]) -> None:
    state_ids = list(dict.fromkeys(state_ids))
    if not state_ids:
        existing = service.find("PRE-CONDITION-STATE-REFS")
        if existing is not None:
            service.remove(existing)
        return
    pc = element("PRE-CONDITION-STATE-REFS")
    for state_id in state_ids:
        sub(pc, "PRE-CONDITION-STATE-REF", attrib={"ID-REF": state_id})
        replace_child(
            service,
            "PRE-CONDITION-STATE-REFS",
            pc,
            before_tags={"STATE-TRANSITION-REFS", "COMPARAM-REFS", "REQUEST-REF"},
        )


def state_id_map(root: etree._Element) -> dict[str, str]:
    result: dict[str, str] = {}
    for state in root.xpath("//*[local-name()='STATE']"):
        short_name = state.findtext("SHORT-NAME")
        if short_name and state.get("ID"):
            result[short_name] = state.get("ID")
    return result


def did_state_ids(root: etree._Element, did: DidDef, mode: str) -> list[str]:
    states = state_id_map(root)
    default = states.get("Default")
    programming = states.get("Programming")
    extended = states.get("Extended")
    locked = states.get("Locked")
    unlocked_l1 = states.get("UnlockedL1")
    unlocked_fbl = states.get("Unlocked_FBL")
    level_9 = states.get("Level_9")

    result: list[str] = []
    session_map = [default, extended, default, programming, extended]
    wanted = "R" if mode == "read" else "W"
    for value, state_id in zip(did.sessions, session_map, strict=False):
        if state_id and wanted in normalize_access(value):
            result.append(state_id)
    if mode == "read":
        result.extend(state for state in (locked, unlocked_l1, level_9, unlocked_fbl) if state)
    else:
        security = normalize_access(did.write_security)
        if "FBL" in security and unlocked_fbl:
            result.append(unlocked_fbl)
        elif ("LEVEL_1" in security or "LEVEL1" in security) and unlocked_l1:
            result.append(unlocked_l1)
        elif normalize_access(did.write_security) in {"", "N", "NO", "NONE"}:
            result.extend(state for state in (locked, unlocked_l1, level_9, unlocked_fbl) if state)
        else:
            result.extend(state for state in (unlocked_l1, level_9, unlocked_fbl) if state)
    if not result:
        result = [state for state in (default, extended, unlocked_l1, level_9) if state]
    return list(dict.fromkeys(result))


CORE_RESPONSE_SERVICE_SOURCES: tuple[str, ...] = (
    "DefaultSession_Start_NoResponse",
    "ExtendedDiagnosticSession_Start_NoResponse",
    "Hard_Reset_Reset_NoResponse",
    "EnableRxAndEnableTx_Control_NoResponse",
    "DisableRxAndDisableTx_Control_NoResponse",
    "TesterPresent_Send_NoResponse",
    "ControlDTCSetting_On_NoResponse",
    "ControlDTCSetting_Off_NoResponse",
)

CORE_SERVICE_PRECONDITION_TARGETS: dict[str, tuple[int, int | None, set[str]]] = {
    "DefaultSession_Start": (0x10, 0x01, {"Default", "Programming", "Extended"}),
    "DefaultSession_Start_NoResponse": (0x10, 0x01, {"Default", "Programming", "Extended"}),
    "ProgrammingSession_Start": (0x10, 0x02, {"Default", "Programming", "Extended"}),
    "ProgrammingSession_Start_NoResponse": (0x10, 0x02, {"Default", "Programming", "Extended"}),
    "ExtendedDiagnosticSession_Start": (0x10, 0x03, {"Default", "Programming", "Extended"}),
    "ExtendedDiagnosticSession_Start_NoResponse": (0x10, 0x03, {"Default", "Programming", "Extended"}),
    "CodingSession_Start": (0x10, 0x41, {"Extended"}),
    "CodingSession_Start_NoResponse": (0x10, 0x41, {"Extended"}),
    "Hard_Reset_Reset": (0x11, 0x01, {"Default", "Programming", "Extended"}),
    "Hard_Reset_Reset_NoResponse": (0x11, 0x01, {"Default", "Programming", "Extended"}),
    "EnableRxAndEnableTx_Control": (0x28, 0x00, {"Default", "Programming", "Extended"}),
    "EnableRxAndEnableTx_Control_NoResponse": (0x28, 0x00, {"Default", "Programming", "Extended"}),
    "EnableRxAndDisableTx_Control": (0x28, 0x01, {"Extended", "Programming"}),
    "EnableRxAndDisableTx_Control_NoResponse": (0x28, 0x01, {"Extended", "Programming"}),
    "DisableRxAndDisableTx_Control": (0x28, 0x03, {"Default", "Programming", "Extended"}),
    "DisableRxAndDisableTx_Control_NoResponse": (0x28, 0x03, {"Default", "Programming", "Extended"}),
    "TesterPresent_Send": (0x3E, 0x00, {"Default", "Programming", "Extended"}),
    "TesterPresent_Send_NoResponse": (0x3E, 0x00, {"Default", "Programming", "Extended"}),
    "ControlDTCSetting_On": (0x85, 0x01, {"Extended"}),
    "ControlDTCSetting_On_NoResponse": (0x85, 0x01, {"Extended"}),
    "ControlDTCSetting_Off": (0x85, 0x02, {"Extended"}),
    "ControlDTCSetting_Off_NoResponse": (0x85, 0x02, {"Extended"}),
}


def service_access_state_ids(root: etree._Element, sessions: Iterable[str], security: str) -> list[str]:
    states = state_id_map(root)
    result: list[str] = []
    session_order = ("Default", "Programming", "Extended")
    session_names = {compact_text(session) for session in sessions if compact_text(session)}
    for name in session_order:
        if name in session_names and states.get(name):
            result.append(states[name])

    security_text = normalize_access(security)
    if security_text in {"", "N", "NO", "NONE"}:
        security_names = ("Locked", "UnlockedL1", "Level_9", "Unlocked_FBL")
    else:
        security_names_list: list[str] = []
        if "LEVEL_1" in security_text or "LEVEL1" in security_text:
            security_names_list.append("UnlockedL1")
        if "LEVEL_9" in security_text or "LEVEL9" in security_text:
            security_names_list.append("Level_9")
        if "FBL" in security_text:
            security_names_list.append("Unlocked_FBL")
        security_names = tuple(security_names_list or ["UnlockedL1", "Level_9", "Unlocked_FBL"])

    for name in security_names:
        if states.get(name):
            result.append(states[name])
    return list(dict.fromkeys(result))


def ensure_vf_funct_class(root: etree._Element, id_gen: Any, short_name: str, long_name: str) -> str:
    existing = first_by_short(root, "FUNCT-CLASS", short_name)
    if existing is not None and existing.get("ID"):
        return existing.get("ID")
    base_variant = get_base_variant(root)
    classes = base_variant.find("FUNCT-CLASSS")
    if classes is None:
        classes = element("FUNCT-CLASSS")
        replace_child(base_variant, "FUNCT-CLASSS", classes, before_tags={"STATE-CHARTS", "DIAG-COMMS"})
    node = sub(classes, "FUNCT-CLASS", attrib={"ID": id_gen.new("FC")})
    sub(node, "SHORT-NAME", short_name)
    sub(node, "LONG-NAME", long_name)
    return node.get("ID")


def append_vf_placeholder_service(
    root: etree._Element,
    id_gen: Any,
    short_name: str,
    long_name: str,
    semantic: str,
    funct_class: str,
    funct_class_long: str,
    request: etree._Element,
    positive: etree._Element | None,
    negative: etree._Element | None,
    service_qualifier: str,
    service_name: str,
    instance_qualifier: str = "z",
    instance_name: str = "VF placeholder",
) -> None:
    if first_by_short(root, "DIAG-SERVICE", short_name) is not None:
        return
    service = element("DIAG-SERVICE", attrib={"ID": id_gen.new("SVC"), "SEMANTIC": semantic, "ADDRESSING": "FUNCTIONAL-OR-PHYSICAL"})
    sub(service, "SHORT-NAME", short_name)
    sub(service, "LONG-NAME", long_name)
    update_service_sdg(service, id_gen, service_qualifier, service_name, instance_qualifier, instance_name)
    class_id = ensure_vf_funct_class(root, id_gen, funct_class, funct_class_long)
    refs = element("FUNCT-CLASS-REFS")
    sub(refs, "FUNCT-CLASS-REF", attrib={"ID-REF": class_id})
    replace_child(service, "FUNCT-CLASS-REFS", refs, before_tags={"AUDIENCE", "PRE-CONDITION-STATE-REFS", "REQUEST-REF"})
    update_flat_preconditions(service, [state for name, state in state_id_map(root).items() if name in {"Default", "Extended", "Programming", "Locked", "UnlockedL1", "Unlocked_FBL"}])
    sub(service, "REQUEST-REF", attrib={"ID-REF": request.get("ID")})
    if positive is not None:
        refs = sub(service, "POS-RESPONSE-REFS")
        sub(refs, "POS-RESPONSE-REF", attrib={"ID-REF": positive.get("ID")})
    if negative is not None:
        refs = sub(service, "NEG-RESPONSE-REFS")
        sub(refs, "NEG-RESPONSE-REF", attrib={"ID-REF": negative.get("ID")})
    find_container(root, "REQUESTS").append(request)
    if positive is not None:
        find_container(root, "POS-RESPONSES").append(positive)
    if negative is not None:
        find_container(root, "NEG-RESPONSES").append(negative)
    find_container(root, "DIAG-COMMS").append(service)


def make_vf_read_placeholder_request(id_gen: Any, short_name: str, service_id: int = 0x22, did: int = 0xF190) -> etree._Element:
    request = element("REQUEST", attrib={"ID": id_gen.new("RQ")})
    sub(request, "SHORT-NAME", f"RQ_{short_name}")
    sub(request, "LONG-NAME", f"RQ {short_name}")
    params = sub(request, "PARAMS")
    params.append(coded_const_param("SID_RQ", "SID-RQ", 0, service_id, "SERVICE-ID"))
    params.append(coded_const_param("RecordDataIdentifier", "RecordDataIdentifier", 1, did, "ID", bit_length=16))
    return request


def make_vf_read_placeholder_positive(id_gen: Any, short_name: str, service_id: int = 0x62, did: int = 0xF190) -> etree._Element:
    response = element("POS-RESPONSE", attrib={"ID": id_gen.new("PR")})
    sub(response, "SHORT-NAME", f"PR_{short_name}")
    sub(response, "LONG-NAME", f"PR {short_name}")
    params = sub(response, "PARAMS")
    params.append(coded_const_param("SID_PR", "SID-PR", 0, service_id, "SERVICE-ID"))
    params.append(coded_const_param("RecordDataIdentifier", "RecordDataIdentifier", 1, did, "ID", bit_length=16))
    params.append(make_value_param("Data", "Data", 3, 0, "_1"))
    return response


def make_vf_io_placeholder_request(id_gen: Any, short_name: str, did: int = 0xFE01, control: int = 0x03) -> etree._Element:
    request = element("REQUEST", attrib={"ID": id_gen.new("RQ")})
    sub(request, "SHORT-NAME", f"RQ_{short_name}")
    sub(request, "LONG-NAME", f"RQ {short_name}")
    params = sub(request, "PARAMS")
    params.append(coded_const_param("SID_RQ", "SID-RQ", 0, 0x2F, "SERVICE-ID"))
    params.append(coded_const_param("DataIdentifier", "DataIdentifier", 1, did, "ID", bit_length=16))
    params.append(coded_const_param("ControlOptionRecord_InputOutputControlParameter", "InputOutputControlParameter", 3, control, "ID"))
    params.append(make_value_param("ControlOptionRecord", "ControlOptionRecord", 4, 0, "_1"))
    return request


def make_vf_io_placeholder_positive(id_gen: Any, short_name: str, did: int = 0xFE01, control: int = 0x03) -> etree._Element:
    response = element("POS-RESPONSE", attrib={"ID": id_gen.new("PR")})
    sub(response, "SHORT-NAME", f"PR_{short_name}")
    sub(response, "LONG-NAME", f"PR {short_name}")
    params = sub(response, "PARAMS")
    params.append(coded_const_param("SID_PR", "SID-PR", 0, 0x6F, "SERVICE-ID"))
    params.append(coded_const_param("DataIdentifier", "DataIdentifier", 1, did, "ID", bit_length=16))
    params.append(coded_const_param("ControlStatusRecord_InputOutputControlParameter", "InputOutputControlParameter", 3, control, "ID"))
    params.append(make_value_param("ControlStatusRecord", "ControlStatusRecord", 4, 0, "_1"))
    return response


def ensure_vf_placeholder_services(root: etree._Element, id_gen: Any) -> None:
    """Create the flat-service seeds expected by the shared flat-service writer.

    They are deleted again after generated DID/IO/Routine instances are cloned.
    Keeping the seed services in-memory lets the shared writer work without
    requiring permanent placeholder services in the VF template.
    """

    append_vf_placeholder_service(
        root,
        id_gen,
        "z_7_Read",
        "VF DID Read placeholder",
        "IDENTIFICATION",
        "ECU_Identification",
        "ECU Identification",
        make_vf_read_placeholder_request(id_gen, "z_7_Read"),
        make_vf_read_placeholder_positive(id_gen, "z_7_Read"),
        make_negative_response(id_gen.new("NR"), "NR_z_7_Read", "NR VF DID Read placeholder", 0x22, "Read"),
        "Read",
        "Read",
    )
    append_vf_placeholder_service(
        root,
        id_gen,
        "z_Read",
        "VF Routine placeholder",
        "ROUTINE",
        "Routine_Control",
        "Routine Control",
        make_vf_read_placeholder_request(id_gen, "z_Read", service_id=0x31, did=0x0203),
        make_vf_read_placeholder_positive(id_gen, "z_Read", service_id=0x71, did=0x0203),
        make_negative_response(id_gen.new("NR"), "NR_z_Read", "NR VF Routine placeholder", 0x31, "Routine"),
        "Read",
        "Read",
    )
    append_vf_placeholder_service(
        root,
        id_gen,
        "z_Control",
        "VF IOControl Control placeholder",
        "CONTROL",
        "IOControl",
        "IO Control",
        make_vf_io_placeholder_request(id_gen, "z_Control", control=0x03),
        make_vf_io_placeholder_positive(id_gen, "z_Control", control=0x03),
        make_negative_response(id_gen.new("NR"), "NR_z_Control", "NR VF IOControl placeholder", 0x2F, "Control"),
        "Control",
        "Control",
    )
    append_vf_placeholder_service(
        root,
        id_gen,
        "z_ReturnControl",
        "VF IOControl ReturnControl placeholder",
        "CONTROL",
        "IOControl",
        "IO Control",
        make_vf_io_placeholder_request(id_gen, "z_ReturnControl", control=0x00),
        make_vf_io_placeholder_positive(id_gen, "z_ReturnControl", control=0x00),
        make_negative_response(id_gen.new("NR"), "NR_z_ReturnControl", "NR VF IOControl placeholder", 0x2F, "ReturnControl"),
        "ReturnControl",
        "ReturnControl",
    )


def sdg_value(service: etree._Element, si: str, default: str = "") -> str:
    node = service.find(f'.//SD[@SI="{si}"]')
    return node.text if node is not None and node.text is not None else default


def first_coded_value_by_semantic(message: etree._Element, semantic: str) -> int | None:
    for param in message.xpath('.//*[local-name()="PARAM" and @SEMANTIC=$semantic]', semantic=semantic):
        value = parse_int_cell(param.findtext("CODED-VALUE"), default=-1)
        if value >= 0:
            return value
    return None


def first_coded_param_by_semantic(message: etree._Element, semantic: str) -> etree._Element | None:
    params = message.xpath('.//*[local-name()="PARAM" and @SEMANTIC=$semantic]', semantic=semantic)
    return params[0] if params else None


def first_coded_value_by_semantics(message: etree._Element, semantics: Iterable[str]) -> int | None:
    for semantic in semantics:
        value = first_coded_value_by_semantic(message, semantic)
        if value is not None:
            return value
    return None


def diag_service_request_key(root: etree._Element, service: etree._Element) -> tuple[int, int | None] | None:
    request_ref = service.find("REQUEST-REF")
    request = find_by_id(root, request_ref.get("ID-REF") if request_ref is not None else None)
    if request is None:
        return None
    service_id = first_coded_value_by_semantic(request, "SERVICE-ID")
    if service_id is None:
        return None
    subfunction = first_coded_value_by_semantics(request, ("SUBFUNCTION", "ACCESSMODE"))
    return service_id, subfunction


def existing_diag_service_keys(root: etree._Element) -> set[tuple[int, int | None]]:
    result: set[tuple[int, int | None]] = set()
    for service in root.xpath("//*[local-name()='DIAG-SERVICE']"):
        key = diag_service_request_key(root, service)
        if key is not None:
            result.add(key)
    return result


def strip_subfunction_prefix(value: Any) -> str:
    text = usable_text(value)
    text = re.sub(r"^\s*(?:0[xX][0-9A-Fa-f]{1,2}|[0-9A-Fa-f]{2})\s*", "", text)
    return text.strip()


def security_access_instance_name(subfunction: int, access: dict[str, Any]) -> str:
    if subfunction == 0x07:
        return "RequestSeedOfSecurityLevelFBL"
    if subfunction == 0x08:
        return "SendKeyOfSecurityLevelFBL"
    raw_name = strip_subfunction_prefix(access.get("subfunction_name", ""))
    if raw_name:
        return sanitize_short_name(raw_name, f"SecurityLevel{subfunction:02X}")
    prefix = "RequestSeed" if subfunction % 2 else "SendKey"
    return sanitize_short_name(f"{prefix}OfSecurityLevel{subfunction:02X}", f"SecurityLevel{subfunction:02X}")


def ensure_state_transition(root: etree._Element, id_gen: Any, source_state: str, target_state: str) -> str | None:
    chart = first_by_short(root, "STATE-CHART", "SecurityAccess")
    if chart is None:
        return None
    transitions = chart.find("STATE-TRANSITIONS")
    if transitions is None:
        transitions = element("STATE-TRANSITIONS")
        replace_child(chart, "STATE-TRANSITIONS", transitions, before_tags={"START-STATE-SNREF", "STATES"})

    for transition in transitions.findall("STATE-TRANSITION"):
        source = transition.find("SOURCE-SNREF")
        target = transition.find("TARGET-SNREF")
        if (
            source is not None
            and target is not None
            and source.get("SHORT-NAME") == source_state
            and target.get("SHORT-NAME") == target_state
            and transition.get("ID")
        ):
            return transition.get("ID")

    transition = sub(transitions, "STATE-TRANSITION", attrib={"ID": id_gen.new("ST")})
    sub(transition, "SHORT-NAME", sanitize_short_name(f"{source_state}_{target_state}", "SecurityTransition"))
    sub(transition, "LONG-NAME", f"{source_state} {target_state}")
    sub(transition, "SOURCE-SNREF", attrib={"SHORT-NAME": source_state})
    sub(transition, "TARGET-SNREF", attrib={"SHORT-NAME": target_state})
    return transition.get("ID")


def set_service_transition_refs(service: etree._Element, transition_id: str | None) -> None:
    existing = service.find("STATE-TRANSITION-REFS")
    if transition_id is None:
        if existing is not None:
            service.remove(existing)
        return
    refs = element("STATE-TRANSITION-REFS")
    sub(refs, "STATE-TRANSITION-REF", attrib={"ID-REF": transition_id})
    replace_child(service, "STATE-TRANSITION-REFS", refs, before_tags={"COMPARAM-REFS", "REQUEST-REF"})


def security_access_target_state(subfunction: int, access: dict[str, Any]) -> str:
    text = normalize_access(access.get("subfunction_name", ""))
    if "FBL" in text:
        return "Unlocked_FBL"
    if "LEVEL_9" in text or "LEVEL9" in text:
        return "Level_9"
    if subfunction == 0x08:
        return "Unlocked_FBL"
    return "UnlockedL1"


def generate_boot_security_access_services(root: etree._Element, id_gen: Any, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    existing_keys = existing_diag_service_keys(root)
    for (service_id, subfunction), access in sorted(access_map.items()):
        if service_id != 0x27 or subfunction is None:
            continue
        if "Boot" not in (access.get("sources") or set()):
            continue
        if (service_id, subfunction) in existing_keys:
            continue

        is_seed = bool(subfunction % 2)
        base_service = "SeedLevel1_Request" if is_seed else "KeyLevel1_Send"
        if first_by_short(root, "DIAG-SERVICE", base_service) is None:
            continue

        instance_name = security_access_instance_name(subfunction, access)
        service_qualifier = "Request" if is_seed else "Send"
        service_name = f"{instance_name}_{service_qualifier}"
        service, request, positive, negative = clone_service_bundle(
            root,
            id_gen,
            base_service,
            service_name,
            f"{instance_name} {service_qualifier}",
            service_qualifier,
            service_qualifier,
            instance_name,
            instance_name,
            service_access_state_ids(root, access.get("sessions") or {"Programming"}, access.get("security") or "N"),
        )
        set_service_semantic_and_funct_class(root, service, "SECURITY", "SecurityAccess")

        subfunction_semantics = ("ACCESSMODE", "SUBFUNCTION")
        for message in (request, positive):
            if message is not None:
                for semantic in subfunction_semantics:
                    set_first_coded_by_semantic(message, semantic, subfunction)

        if negative is not None:
            set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x27)

        if not is_seed:
            target_state = security_access_target_state(subfunction, access)
            transition_id = ensure_state_transition(root, id_gen, "Locked", target_state)
            set_service_transition_refs(service, transition_id)

        existing_keys.add((service_id, subfunction))


def prune_vf_unsupported_security_access_services(root: etree._Element, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    supported = {key for key in access_map if key[0] == 0x27}
    if not supported:
        return
    to_remove: list[str] = []
    for service in root.xpath("//*[local-name()='DIAG-SERVICE']"):
        key = diag_service_request_key(root, service)
        if key is None or key[0] != 0x27:
            continue
        if key not in supported:
            short_name = service.findtext("SHORT-NAME")
            if short_name:
                to_remove.append(short_name)
    for short_name in to_remove:
        remove_service_and_messages(root, short_name)


def ensure_vf_security_access_subfunction_scales(root: etree._Element, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    security_subfunctions = sorted(subfunction for service_id, subfunction in access_map if service_id == 0x27 and subfunction is not None)
    if not security_subfunctions:
        return
    dop = first_by_short(root, "DATA-OBJECT-PROP", "Subfunction_SecurityAccess")
    if dop is None:
        return
    scales = dop.find(".//COMPU-SCALES")
    if scales is None:
        internal = dop.find(".//COMPU-INTERNAL-TO-PHYS")
        if internal is None:
            compu = dop.find("COMPU-METHOD")
            if compu is None:
                compu = sub(dop, "COMPU-METHOD")
                sub(compu, "CATEGORY", "TEXTTABLE")
            internal = sub(compu, "COMPU-INTERNAL-TO-PHYS")
        scales = sub(internal, "COMPU-SCALES")

    existing: set[int] = set()
    for scale in scales.findall("COMPU-SCALE"):
        low = parse_int_cell(scale.findtext("LOWER-LIMIT"))
        high = parse_int_cell(scale.findtext("UPPER-LIMIT"), low)
        if low is not None and high == low:
            existing.add(low)

    for subfunction in security_subfunctions:
        if subfunction in existing:
            continue
        scale = sub(scales, "COMPU-SCALE")
        sub(scale, "LOWER-LIMIT", subfunction)
        sub(scale, "UPPER-LIMIT", subfunction)
        inverse = sub(scale, "COMPU-INVERSE-VALUE")
        sub(inverse, "V", 1 if subfunction % 2 else 2)
        const = sub(scale, "COMPU-CONST")
        sub(const, "VT", "requestSeed" if subfunction % 2 else "sendKey")


def update_vf_security_access_preconditions(root: etree._Element, id_gen: Any, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    for service in root.xpath("//*[local-name()='DIAG-SERVICE']"):
        key = diag_service_request_key(root, service)
        if key is None or key[0] != 0x27:
            continue
        access = access_map.get(key)
        if not access:
            continue
        update_flat_preconditions(service, service_access_state_ids(root, access.get("sessions") or {"Extended", "Programming"}, access.get("security") or "N"))
        subfunction = key[1]
        if subfunction is not None and not subfunction % 2:
            target_state = security_access_target_state(subfunction, access)
            transition_id = ensure_state_transition(root, id_gen, "Locked", target_state)
            set_service_transition_refs(service, transition_id)


def data_object_prop_id(root: etree._Element, short_name: str) -> str | None:
    dop = first_by_short(root, "DATA-OBJECT-PROP", short_name)
    return dop.get("ID") if dop is not None else None


def make_core_positive_response(
    root: etree._Element,
    node_id: str,
    short_name: str,
    long_name: str,
    request_sid: int,
    subfunction_param: etree._Element | None,
    subfunction_value: int,
) -> etree._Element:
    response = element("POS-RESPONSE", attrib={"ID": node_id})
    sub(response, "SHORT-NAME", sanitize_short_name(short_name, "PR"))
    sub(response, "LONG-NAME", long_name)
    params = sub(response, "PARAMS")
    params.append(coded_const_param("SID_PR", "SID-PR", 0, request_sid + 0x40, "SERVICE-ID"))

    sub_short_name = subfunction_param.findtext("SHORT-NAME") if subfunction_param is not None else "SubFunction"
    sub_long_name = subfunction_param.findtext("LONG-NAME") if subfunction_param is not None else sub_short_name
    params.append(
        coded_const_param(
            sub_short_name or "SubFunction",
            sub_long_name or sub_short_name or "SubFunction",
            1,
            subfunction_value,
            "SUBFUNCTION",
        )
    )

    if request_sid == 0x10:
        p2_id = data_object_prop_id(root, "P2")
        p2ex_id = data_object_prop_id(root, "P2Ex")
        if p2_id:
            params.append(make_value_param("P2", "P2", 2, 0, p2_id))
        if p2ex_id:
            params.append(make_value_param("P2Ex", "P2Ex", 4, 0, p2ex_id))
    return response


def create_core_response_service(root: etree._Element, id_gen: Any, no_response_name: str) -> None:
    response_name = no_response_name.removesuffix("_NoResponse")
    if response_name == no_response_name or first_by_short(root, "DIAG-SERVICE", response_name) is not None:
        return

    base_service, base_request, _, base_negative = service_messages(root, no_response_name)
    request_sid = first_coded_value_by_semantic(base_request, "SERVICE-ID")
    no_response_subfunction = first_coded_value_by_semantic(base_request, "SUBFUNCTION")
    if request_sid is None or no_response_subfunction is None:
        return
    response_subfunction = no_response_subfunction & 0x7F

    request = copy.deepcopy(base_request)
    negative = copy.deepcopy(base_negative) if base_negative is not None else None
    service = copy.deepcopy(base_service)

    request_id = id_gen.new("RQ")
    positive_id = id_gen.new("PR")
    negative_id = id_gen.new("NR") if negative is not None else None
    service_id = id_gen.new("SVC")

    request.set("ID", request_id)
    refresh_internal_ids(request, id_gen)
    set_short_long(request, f"RQ_{response_name}", f"RQ {response_name}")
    set_first_coded_by_semantic(request, "SUBFUNCTION", response_subfunction)

    subfunction_param = first_coded_param_by_semantic(request, "SUBFUNCTION")
    positive = make_core_positive_response(
        root,
        positive_id,
        f"PR_{response_name}",
        f"PR {response_name}",
        request_sid,
        subfunction_param,
        response_subfunction,
    )

    if negative is not None and negative_id is not None:
        negative.set("ID", negative_id)
        refresh_internal_ids(negative, id_gen)
        set_short_long(negative, f"NR_{response_name}", f"NR {response_name}")

    service.set("ID", service_id)
    service.attrib.pop("TRANSMISSION-MODE", None)
    refresh_internal_ids(service, id_gen)
    set_short_long(service, response_name, (base_service.findtext("LONG-NAME") or response_name).replace("_NoResponse", ""))

    request_ref = service.find("REQUEST-REF")
    if request_ref is not None:
        request_ref.set("ID-REF", request_id)

    pos_refs = element("POS-RESPONSE-REFS")
    sub(pos_refs, "POS-RESPONSE-REF", attrib={"ID-REF": positive_id})
    replace_child(service, "POS-RESPONSE-REFS", pos_refs, before_tags={"NEG-RESPONSE-REFS"})

    neg_ref = service.find(".//NEG-RESPONSE-REF")
    if neg_ref is not None and negative_id is not None:
        neg_ref.set("ID-REF", negative_id)

    update_service_sdg(
        service,
        id_gen,
        sdg_value(base_service, "ServiceQualifier", "Send"),
        sdg_value(base_service, "ServiceName", sdg_value(base_service, "ServiceQualifier", "Send")),
        sdg_value(base_service, "DiagInstanceQualifier", response_name),
        sdg_value(base_service, "DiagInstanceName", response_name),
    )

    find_container(root, "REQUESTS").append(request)
    find_container(root, "POS-RESPONSES").append(positive)
    if negative is not None:
        find_container(root, "NEG-RESPONSES").append(negative)
    find_container(root, "DIAG-COMMS").append(service)


def ensure_core_response_services(root: etree._Element, id_gen: Any) -> None:
    for no_response_name in CORE_RESPONSE_SERVICE_SOURCES:
        if first_by_short(root, "DIAG-SERVICE", no_response_name) is not None:
            create_core_response_service(root, id_gen, no_response_name)


def copy_state_transition_refs(source_service: etree._Element | None, target_service: etree._Element) -> None:
    if source_service is None:
        return
    source_refs = source_service.find("STATE-TRANSITION-REFS")
    if source_refs is None:
        existing = target_service.find("STATE-TRANSITION-REFS")
        if existing is not None:
            target_service.remove(existing)
        return
    replace_child(
        target_service,
        "STATE-TRANSITION-REFS",
        copy.deepcopy(source_refs),
        before_tags={"COMPARAM-REFS", "REQUEST-REF"},
    )


def ensure_vf_session_control_services(root: etree._Element, id_gen: Any, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    session_defs = {
        0x01: ("DefaultSession", "Default Session", "DefaultSession_Start", "DefaultSession_Start_NoResponse"),
        0x02: ("ProgrammingSession", "Programming Session", "ProgrammingSession_Start", "ProgrammingSession_Start_NoResponse"),
        0x03: (
            "ExtendedDiagnosticSession",
            "Extended Diagnostic Session",
            "ExtendedDiagnosticSession_Start",
            "ExtendedDiagnosticSession_Start_NoResponse",
        ),
        0x41: ("CodingSession", "Coding Session"),
    }
    for subfunction, definition in session_defs.items():
        instance_qualifier = definition[0]
        instance_name = definition[1]
        service_short_name = definition[2] if len(definition) > 2 else f"{instance_qualifier}_Start"
        no_response_short_name = definition[3] if len(definition) > 3 else f"{service_short_name}_NoResponse"
        access = access_map.get((0x10, subfunction))
        if not access or not access.get("sessions"):
            continue

        preconditions = service_access_state_ids(root, access.get("sessions") or {"Extended"}, access.get("security") or "N")
        if first_by_short(root, "DIAG-SERVICE", service_short_name) is None:
            service, request, positive, negative = clone_service_bundle(
                root,
                id_gen,
                "ExtendedDiagnosticSession_Start",
                service_short_name,
                f"{instance_name} Start",
                "Start",
                "Start",
                instance_qualifier,
                instance_name,
                preconditions,
            )
            for message in (request, positive):
                if message is not None:
                    set_first_coded_by_semantic(message, "SUBFUNCTION", subfunction)
            if negative is not None:
                set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x10)

        if first_by_short(root, "DIAG-SERVICE", no_response_short_name) is None:
            no_response_base = (
                "DefaultSession_Start_NoResponse"
                if subfunction == 0x01
                else "ExtendedDiagnosticSession_Start_NoResponse"
            )
            service, request, _, negative = clone_service_bundle(
                root,
                id_gen,
                no_response_base,
                no_response_short_name,
                f"{instance_name} Start_NoResponse",
                "Start",
                "Start",
                instance_qualifier,
                instance_name,
                preconditions,
                positive_response_suppressed="yes",
            )
            set_first_coded_by_semantic(request, "SUBFUNCTION", subfunction | 0x80)
            if negative is not None:
                set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x10)
            copy_state_transition_refs(first_by_short(root, "DIAG-SERVICE", service_short_name), service)


def ensure_vf_communication_control_services(root: etree._Element, id_gen: Any, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    service_defs = {
        0x01: ("EnableRxAndDisableTx", "Enable Rx And Disable Tx"),
    }
    for subfunction, (instance_qualifier, instance_name) in service_defs.items():
        access = access_map.get((0x28, subfunction))
        if not access or not access.get("sessions"):
            continue

        preconditions = service_access_state_ids(root, access.get("sessions") or {"Extended", "Programming"}, access.get("security") or "N")
        service_short_name = f"{instance_qualifier}_Control"
        if first_by_short(root, "DIAG-SERVICE", service_short_name) is None:
            service, request, positive, negative = clone_service_bundle(
                root,
                id_gen,
                "EnableRxAndEnableTx_Control",
                service_short_name,
                f"{instance_name} Control",
                "Control",
                "Control",
                instance_qualifier,
                instance_name,
                preconditions,
            )
            for message in (request, positive):
                if message is not None:
                    set_first_coded_by_semantic(message, "SUBFUNCTION", subfunction)
            if negative is not None:
                set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x28)

        no_response_short_name = f"{service_short_name}_NoResponse"
        if first_by_short(root, "DIAG-SERVICE", no_response_short_name) is None:
            service, request, _, negative = clone_service_bundle(
                root,
                id_gen,
                "EnableRxAndEnableTx_Control_NoResponse",
                no_response_short_name,
                f"{instance_name} Control_NoResponse",
                "Control",
                "Control",
                instance_qualifier,
                instance_name,
                preconditions,
                positive_response_suppressed="yes",
            )
            set_first_coded_by_semantic(request, "SUBFUNCTION", subfunction | 0x80)
            if negative is not None:
                set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x28)


def update_core_service_preconditions(root: etree._Element, survey: SurveyData) -> None:
    access_map = getattr(survey, "core_service_access", {}) or {}
    for service_name, (service_id, subfunction, fallback_sessions) in CORE_SERVICE_PRECONDITION_TARGETS.items():
        service = first_by_short(root, "DIAG-SERVICE", service_name)
        if service is None:
            continue
        access = access_map.get((service_id, subfunction)) or access_map.get((service_id, None))
        if access:
            sessions = access.get("sessions") or fallback_sessions
            security = access.get("security") or "N"
        else:
            sessions = fallback_sessions
            security = "N"
        update_flat_preconditions(service, service_access_state_ids(root, sessions, security))


def set_coded_value(message: etree._Element, param_short_name: str, value: int) -> None:
    for param in message.xpath('.//*[local-name()="PARAM" and SHORT-NAME=$name]', name=param_short_name):
        coded = param.find("CODED-VALUE")
        if coded is not None:
            coded.text = str(value)


def set_first_coded_by_semantic(message: etree._Element, semantic: str, value: int) -> None:
    for param in message.xpath('.//*[local-name()="PARAM" and @SEMANTIC=$semantic]', semantic=semantic):
        coded = param.find("CODED-VALUE")
        if coded is not None:
            coded.text = str(value)
            return


def set_value_param_dop(message: etree._Element, dop_id: str, short_name: str, long_name: str, byte_position: int | None = None) -> None:
    value_params = message.xpath('./*[local-name()="PARAMS"]/*[local-name()="PARAM" and @xsi:type="VALUE"]', namespaces={"xsi": XSI_NS})
    if not value_params:
        return
    param = value_params[-1]
    short = param.find("SHORT-NAME")
    if short is not None:
        short.text = sanitize_short_name(short_name, "Data")
    long = param.find("LONG-NAME")
    if long is not None:
        long.text = long_name or short_name
    if byte_position is not None:
        byte = param.find("BYTE-POSITION")
        if byte is not None:
            byte.text = str(byte_position)
    dop_ref = param.find("DOP-REF")
    if dop_ref is None:
        dop_ref = sub(param, "DOP-REF")
    dop_ref.set("ID-REF", dop_id)


def did_payload_dop_id(did: DidDef) -> str:
    direct_param = direct_did_payload_param(did)
    if direct_param is not None:
        return direct_param.dop_id
    return did.structure_id


def remove_param_by_short_name(message: etree._Element, short_name: str) -> None:
    params = message.find("PARAMS")
    if params is None:
        return
    for param in list(params):
        if param.findtext("SHORT-NAME") == short_name:
            params.remove(param)


def generate_flat_did_services(root: etree._Element, id_gen: Any, dids: list[DidDef]) -> None:
    for did in dids:
        if did.readable:
            service_name = f"{did.short_name}_Read"
            service, request, positive, negative = clone_service_bundle(
                root,
                id_gen,
                "z_7_Read",
                service_name,
                f"{did.long_name} Read",
                "Read",
                "Read",
                did.short_name,
                did.long_name,
                did_state_ids(root, did, "read"),
                static_value=did.did,
            )
            set_service_semantic_and_funct_class(root, service, "IDENTIFICATION", "ECU_Identification")
            set_coded_value(request, "RecordDataIdentifier", did.did)
            if positive is not None:
                set_coded_value(positive, "RecordDataIdentifier", did.did)
                set_value_param_dop(positive, did_payload_dop_id(did), did.short_name, did.long_name, byte_position=3)
            if negative is not None:
                set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x22)

        if did.writable:
            service_name = f"{did.short_name}_Write"
            service, _, _, _ = clone_service_bundle(
                root,
                id_gen,
                "z_7_Read",
                service_name,
                f"{did.long_name} Write",
                "Write",
                "Write",
                did.short_name,
                did.long_name,
                did_state_ids(root, did, "write"),
                static_value=did.did,
            )
            set_service_semantic_and_funct_class(root, service, "IDENTIFICATION", "ECU_Identification")
            request = make_write_request(id_gen.new("RQ"), f"RQ_{service_name}", did)
            positive = make_write_positive(id_gen.new("PR"), f"PR_{service_name}", did)
            negative = make_negative_response(id_gen.new("NR"), f"NR_{service_name}", f"NR {did.long_name} Write", 0x2E, "Write")
            replace_service_refs(root, service, request, positive, negative)


def replace_service_refs(
    root: etree._Element,
    service: etree._Element,
    request: etree._Element,
    positive: etree._Element | None,
    negative: etree._Element | None,
) -> None:
    old_request_ref = service.find("REQUEST-REF")
    old_pos_ref = service.find(".//POS-RESPONSE-REF")
    old_neg_ref = service.find(".//NEG-RESPONSE-REF")
    old_ids = [
        old_request_ref.get("ID-REF") if old_request_ref is not None else None,
        old_pos_ref.get("ID-REF") if old_pos_ref is not None else None,
        old_neg_ref.get("ID-REF") if old_neg_ref is not None else None,
    ]
    for old_id in old_ids:
        old_node = find_by_id(root, old_id)
        if old_node is not None and old_node.getparent() is not None:
            old_node.getparent().remove(old_node)
    find_container(root, "REQUESTS").append(request)
    if positive is not None:
        find_container(root, "POS-RESPONSES").append(positive)
    if negative is not None:
        find_container(root, "NEG-RESPONSES").append(negative)
    if old_request_ref is not None:
        old_request_ref.set("ID-REF", request.get("ID"))
    if old_pos_ref is not None and positive is not None:
        old_pos_ref.set("ID-REF", positive.get("ID"))
    if old_neg_ref is not None and negative is not None:
        old_neg_ref.set("ID-REF", negative.get("ID"))


def coded_const_param(short_name: str, long_name: str, byte_position: int, coded_value: int, semantic: str, bit_length: int = 8) -> etree._Element:
    param = element("PARAM", attrib={"SEMANTIC": semantic})
    set_xsi_type(param, "CODED-CONST")
    sub(param, "SHORT-NAME", short_name)
    sub(param, "LONG-NAME", long_name)
    sub(param, "BYTE-POSITION", byte_position)
    sub(param, "CODED-VALUE", coded_value)
    coded_type = sub(param, "DIAG-CODED-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32"})
    set_xsi_type(coded_type, "STANDARD-LENGTH-TYPE")
    sub(coded_type, "BIT-LENGTH", bit_length)
    return param


def make_write_request(node_id: str, short_name: str, did: DidDef) -> etree._Element:
    request = element("REQUEST", attrib={"ID": node_id})
    sub(request, "SHORT-NAME", sanitize_short_name(short_name, "RQ"))
    sub(request, "LONG-NAME", f"RQ {did.long_name} Write")
    params = sub(request, "PARAMS")
    params.append(coded_const_param("SID_RQ", "SID-RQ", 0, 0x2E, "SERVICE-ID"))
    params.append(coded_const_param("RecordDataIdentifier", "RecordDataIdentifier", 1, did.did, "ID", bit_length=16))
    params.append(make_value_param(did.short_name, did.long_name, 3, 0, did_payload_dop_id(did)))
    return request


def make_write_positive(node_id: str, short_name: str, did: DidDef) -> etree._Element:
    response = element("POS-RESPONSE", attrib={"ID": node_id})
    sub(response, "SHORT-NAME", sanitize_short_name(short_name, "PR"))
    sub(response, "LONG-NAME", f"PR {did.long_name} Write")
    params = sub(response, "PARAMS")
    params.append(coded_const_param("SID_PR", "SID-PR", 0, 0x6E, "SERVICE-ID"))
    params.append(coded_const_param("RecordDataIdentifier", "RecordDataIdentifier", 1, did.did, "ID", bit_length=16))
    return response


def make_negative_response(node_id: str, short_name: str, long_name: str, request_sid: int, qualifier: str) -> etree._Element:
    response = element("NEG-RESPONSE", attrib={"ID": node_id})
    sub(response, "SHORT-NAME", sanitize_short_name(short_name, "NR"))
    sub(response, "LONG-NAME", long_name)
    params = sub(response, "PARAMS")
    params.append(coded_const_param("SID_NR", "SID-NR", 0, 0x7F, "SERVICE-ID"))
    params.append(coded_const_param("SIDRQ_NR", "SIDRQ-NR", 1, request_sid, "SERVICEIDRQ"))
    value = make_value_param(qualifier, qualifier, 2, 0, "_1")
    params.append(value)
    nrc = element("PARAM", attrib={"SEMANTIC": "DATA"})
    set_xsi_type(nrc, "NRC-CONST")
    sub(nrc, "SHORT-NAME", sanitize_short_name(f"NRCConst_{qualifier}", "NRCConst"))
    sub(nrc, "LONG-NAME", qualifier)
    sub(nrc, "BYTE-POSITION", 2)
    coded_values = sub(nrc, "CODED-VALUES")
    for code in (0x13, 0x22, 0x31, 0x33, 0x72, 0x7F):
        sub(coded_values, "CODED-VALUE", code)
    coded_type = sub(nrc, "DIAG-CODED-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32"})
    set_xsi_type(coded_type, "STANDARD-LENGTH-TYPE")
    sub(coded_type, "BIT-LENGTH", 8)
    params.append(nrc)
    return response


def set_service_semantic_and_funct_class(
    root: etree._Element,
    service: etree._Element,
    semantic: str,
    funct_class_short_name: str,
) -> None:
    service.set("SEMANTIC", semantic)
    class_node = first_by_short(root, "FUNCT-CLASS", funct_class_short_name)
    if class_node is None or not class_node.get("ID"):
        return
    refs = service.find("FUNCT-CLASS-REFS")
    if refs is None:
        refs = element("FUNCT-CLASS-REFS")
        replace_child(service, "FUNCT-CLASS-REFS", refs, before_tags={"AUDIENCE", "PRE-CONDITION-STATE-REFS", "REQUEST-REF"})
    for child in list(refs):
        refs.remove(child)
    sub(refs, "FUNCT-CLASS-REF", attrib={"ID-REF": class_node.get("ID")})


def generate_flat_io_services(root: etree._Element, id_gen: Any, io_dids: list[IoDidDef]) -> None:
    control_templates = {
        0: ("z_ReturnControl", "ReturnControl", "ReturnControl"),
        1: ("z_ReturnControl", "Reset", "Reset"),
        2: ("z_ReturnControl", "Freeze", "Freeze"),
        3: ("z_Control", "Control", "Control"),
    }
    default_states = [state_id for name, state_id in state_id_map(root).items() if name in {"Extended", "UnlockedL1", "Level_9"}]
    for io_did in io_dids:
        for control in sorted(io_did.controls):
            template, qualifier, service_label = control_templates.get(control, control_templates[3])
            service_name = f"{io_did.short_name}_{qualifier}"
            _, request, positive, negative = clone_service_bundle(
                root,
                id_gen,
                template,
                service_name,
                f"{io_did.long_name} {service_label}",
                qualifier,
                service_label,
                io_did.short_name,
                io_did.long_name,
                default_states,
            )
            set_coded_value(request, "DataIdentifier", io_did.did)
            set_coded_value(request, "ControlOptionRecord_InputOutputControlParameter", control)
            if control == 3:
                set_value_param_dop(request, io_did.structure_id, "ControlOptionRecord", "ControlOptionRecord", byte_position=4)
            else:
                remove_param_by_short_name(request, "ControlOptionRecord")
            if positive is not None:
                set_coded_value(positive, "DataIdentifier", io_did.did)
                set_coded_value(positive, "ControlStatusRecord_InputOutputControlParameter", control)
                set_value_param_dop(positive, io_did.structure_id, "ControlStatusRecord", "ControlStatusRecord", byte_position=4)
            if negative is not None:
                set_first_coded_by_semantic(negative, "SERVICEIDRQ", 0x2F)

def generate_flat_routine_services(root: etree._Element, id_gen: Any, routines: list[RoutineDef]) -> None:
    states = state_id_map(root)
    default_states = [state for state in (states.get("Extended"), states.get("Programming"), states.get("UnlockedL1"), states.get("Unlocked_FBL")) if state]
    labels = {1: ("Start", "Start"), 2: ("Stop", "Stop"), 3: ("RequestResults", "RequestResults")}
    for routine in routines:
        for control_type, subfn in sorted(routine.subfunctions.items()):
            if not subfn.supported:
                continue
            qualifier, service_label = labels.get(control_type, (f"Control_{control_type:02X}", f"Control {control_type:02X}"))
            service_name = f"{routine.short_name}_{qualifier}"
            service, _, _, _ = clone_service_bundle(
                root,
                id_gen,
                "z_Read",
                service_name,
                f"{routine.long_name} {service_label}",
                qualifier,
                service_label,
                routine.short_name,
                routine.long_name,
                default_states,
            )
            request = make_routine_request(id_gen.new("RQ"), f"RQ_{service_name}", routine, subfn)
            positive = make_routine_positive(id_gen.new("PR"), f"PR_{service_name}", routine, subfn)
            negative = make_negative_response(id_gen.new("NR"), f"NR_{service_name}", f"NR {routine.long_name} {service_label}", 0x31, qualifier)
            replace_service_refs(root, service, request, positive, negative)


def make_routine_request(node_id: str, short_name: str, routine: RoutineDef, subfn: RoutineSubFunction) -> etree._Element:
    request = element("REQUEST", attrib={"ID": node_id})
    sub(request, "SHORT-NAME", sanitize_short_name(short_name, "RQ"))
    sub(request, "LONG-NAME", f"RQ {routine.long_name}")
    params = sub(request, "PARAMS")
    params.append(coded_const_param("SID_RQ", "SID-RQ", 0, 0x31, "SERVICE-ID"))
    params.append(coded_const_param("RoutineControlType", "RoutineControlType", 1, subfn.control_type, "SUBFUNCTION"))
    params.append(coded_const_param("RoutineIdentifier", "RoutineIdentifier", 2, routine.rid, "ID", bit_length=16))
    if subfn.option_structure_id:
        params.append(make_value_param("RoutineControlOptionRecord", "RoutineControlOptionRecord", 4, 0, subfn.option_structure_id))
    return request


def make_routine_positive(node_id: str, short_name: str, routine: RoutineDef, subfn: RoutineSubFunction) -> etree._Element:
    response = element("POS-RESPONSE", attrib={"ID": node_id})
    sub(response, "SHORT-NAME", sanitize_short_name(short_name, "PR"))
    sub(response, "LONG-NAME", f"PR {routine.long_name}")
    params = sub(response, "PARAMS")
    params.append(coded_const_param("SID_PR", "SID-PR", 0, 0x71, "SERVICE-ID"))
    params.append(coded_const_param("RoutineControlType", "RoutineControlType", 1, subfn.control_type, "SUBFUNCTION"))
    params.append(coded_const_param("RoutineIdentifier", "RoutineIdentifier", 2, routine.rid, "ID", bit_length=16))
    if subfn.status_structure_id:
        params.append(make_value_param("RoutineStatusRecord", "RoutineStatusRecord", 4, 0, subfn.status_structure_id))
    return response


def normalize_vf_texttable_dop(dop: etree._Element, bit_length: int, labels: list[tuple[int, str]]) -> None:
    compu = element("COMPU-METHOD")
    sub(compu, "CATEGORY", "TEXTTABLE")
    internal = sub(compu, "COMPU-INTERNAL-TO-PHYS")
    scales = sub(internal, "COMPU-SCALES")
    for value, label in labels:
        scale = sub(scales, "COMPU-SCALE")
        sub(scale, "LOWER-LIMIT", value)
        sub(scale, "UPPER-LIMIT", value)
        const = sub(scale, "COMPU-CONST")
        sub(const, "VT", label)

    coded = element(
        "DIAG-CODED-TYPE",
        attrib={"BASE-TYPE-ENCODING": "NONE", "BASE-DATA-TYPE": "A_UINT32"},
    )
    set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
    sub(coded, "BIT-LENGTH", bit_length)
    physical = element("PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UNICODE2STRING"})

    replace_child(dop, "COMPU-METHOD", compu, before_tags={"DIAG-CODED-TYPE", "PHYSICAL-TYPE"})
    replace_child(dop, "DIAG-CODED-TYPE", coded, before_tags={"PHYSICAL-TYPE"})
    replace_child(dop, "PHYSICAL-TYPE", physical)
    internal_constr = dop.find("INTERNAL-CONSTR")
    if internal_constr is not None:
        dop.remove(internal_constr)


def ensure_vf_texttable_dtc(root: etree._Element, id_gen: Any) -> str:
    existing = first_by_short(root, "DATA-OBJECT-PROP", "TextTable_DTC")
    if existing is not None and existing.get("ID"):
        normalize_vf_texttable_dop(existing, 24, [])
        return existing.get("ID")
    ddds = find_dictionary_spec(root)
    if ddds is None:
        raise RuntimeError("Template ODX has no DIAG-DATA-DICTIONARY-SPEC")
    data_object_props = ddds.find("DATA-OBJECT-PROPS")
    if data_object_props is None:
        data_object_props = sub(ddds, "DATA-OBJECT-PROPS")
    dop = sub(data_object_props, "DATA-OBJECT-PROP", attrib={"ID": id_gen.new("DOP")})
    sub(dop, "SHORT-NAME", "TextTable_DTC")
    sub(dop, "LONG-NAME", "Text Table DTC")
    normalize_vf_texttable_dop(dop, 24, [])
    return dop.get("ID")


def ensure_vf_tables_container(root: etree._Element) -> etree._Element:
    ddds = find_dictionary_spec(root)
    if ddds is None:
        raise RuntimeError("Template ODX has no DIAG-DATA-DICTIONARY-SPEC")
    tables = ddds.find("TABLES")
    if tables is None:
        tables = sub(ddds, "TABLES")
    return tables


def make_vf_ext_record_table_structure(root: etree._Element, id_gen: Any) -> str:
    existing = first_by_short(root, "STRUCTURE", "STRUC_DTCExtendedDataRecordNumbers")
    if existing is not None and existing.get("ID"):
        return existing.get("ID")
    structures = find_container(root, "STRUCTURES")
    structure = sub(structures, "STRUCTURE", attrib={"ID": id_gen.new("STR"), "IS-VISIBLE": "false"})
    sub(structure, "SHORT-NAME", "STRUC_DTCExtendedDataRecordNumbers")
    params = sub(structure, "PARAMS")
    params.append(
        make_value_param(
            "Record_Numbers",
            "Record Numbers",
            0,
            0,
            dop_id_by_short(root, "DATA-OBJECT-PROP", "DTCExtendedDataRecordNumbers_All") or "_54",
        )
    )
    return structure.get("ID")


def normalize_vf_extended_data_request(root: etree._Element, id_gen: Any, table_id: str) -> None:
    service = first_by_short(root, "DIAG-SERVICE", "FaultMemory_Read_DTC_extended_data_record_by_DTC_number")
    if service is None:
        return
    request_ref = service.find("REQUEST-REF")
    request = find_by_id(root, request_ref.get("ID-REF") if request_ref is not None else None)
    if request is None:
        return
    params = request.find("PARAMS")
    if params is None:
        params = sub(request, "PARAMS")
    table_key_id = id_gen.new("TK")
    for child in list(params):
        params.remove(child)
    params.append(coded_const_param("SID_RQ", "SID-RQ", 0, 0x19, "SERVICE-ID"))
    params.append(
        coded_const_param(
            "ReportDTCExtendedDataRecordByDtcNumber",
            "ReportDTCExtendedDataRecordByDtcNumber",
            1,
            0x06,
            "SUBFUNCTION",
        )
    )
    params.append(make_table_key_param("DTC", "DTC", 2, table_id, table_key_id))
    params.append(make_table_struct_param("DTCExtendedDataRecordNumber", "DTCExtendedDataRecordNumber", 5, table_key_id))


def normalize_vf_snapshot_request(root: etree._Element) -> None:
    service = first_by_short(root, "DIAG-SERVICE", "FaultMemory_Read_DTC_snapshot_record_by_DTC_number")
    if service is None:
        return
    request_ref = service.find("REQUEST-REF")
    request = find_by_id(root, request_ref.get("ID-REF") if request_ref is not None else None)
    if request is None:
        return
    for param in request.xpath('./*[local-name()="PARAMS"]/*[local-name()="PARAM"]'):
        if param.findtext("SHORT-NAME") == "Record_Numbers":
            dop_ref = param.find("DOP-REF")
            if dop_ref is not None:
                dop_ref.set("ID-REF", dop_id_by_short(root, "DATA-OBJECT-PROP", "DTCSnapshotRecordNumbers_All") or "_53")


def ensure_vf_dtc_extended_table(root: etree._Element, id_gen: Any, dtcs: list[DtcDef]) -> None:
    texttable_id = ensure_vf_texttable_dtc(root, id_gen)
    tables = ensure_vf_tables_container(root)
    table = first_by_short(root, "TABLE", "DTCExtendedDataRecordNumber")
    if table is None:
        table = sub(tables, "TABLE", attrib={"ID": id_gen.new("TAB")})
        sub(table, "SHORT-NAME", "DTCExtendedDataRecordNumber")
        sub(table, "LONG-NAME", "DTCExtendedDataRecordNumber")
        sub(table, "KEY-DOP-REF", attrib={"ID-REF": texttable_id})
    else:
        key_ref = table.find("KEY-DOP-REF")
        if key_ref is None:
            key_ref = element("KEY-DOP-REF", attrib={"ID-REF": texttable_id})
            table.insert(2, key_ref)
        key_ref.set("ID-REF", texttable_id)
    for row in table.findall("TABLE-ROW"):
        table.remove(row)
    structure_id = make_vf_ext_record_table_structure(root, id_gen)
    for dtc in dtcs:
        label = dtc_table_key(dtc)
        row = append_table_row(table, id_gen, f"TR_DTC_{dtc.byte_code:06X}", label, label, structure_id)
        sdgs = sub(row, "SDGS")
        sdg = sub(sdgs, "SDG")
        caption = sub(sdg, "SDG-CAPTION", attrib={"ID": id_gen.new("CAP")})
        sub(caption, "SHORT-NAME", "IsDefaultCase")
        sub(sdg, "SD", "Yes")
    normalize_vf_extended_data_request(root, id_gen, table.get("ID"))
    normalize_vf_snapshot_request(root)


def update_vf_dtc_dop(root: etree._Element, id_gen: Any, dtcs: list[DtcDef]) -> None:
    dtc_dop = first_by_short(root, "DTC-DOP", "RecordDataType")
    if dtc_dop is None:
        data_dop = first_by_short(root, "DATA-OBJECT-PROP", "RecordDataType")
        record_id = data_dop.get("ID") if data_dop is not None and data_dop.get("ID") else id_gen.new("DTC_DOP")
        if data_dop is not None and data_dop.getparent() is not None:
            data_dop.getparent().remove(data_dop)
        ddds = find_dictionary_spec(root)
        if ddds is None:
            return
        dtc_dops = ddds.find("DTC-DOPS")
        if dtc_dops is None:
            dtc_dops = element("DTC-DOPS")
            ddds.insert(0, dtc_dops)
        dtc_dop = sub(dtc_dops, "DTC-DOP", attrib={"ID": record_id})
        sub(dtc_dop, "SHORT-NAME", "RecordDataType")
        sub(dtc_dop, "LONG-NAME", "RecordDataType")
        coded = sub(
            dtc_dop,
            "DIAG-CODED-TYPE",
            attrib={"BASE-TYPE-ENCODING": "NONE", "BASE-DATA-TYPE": "A_UINT32"},
        )
        set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
        sub(coded, "BIT-LENGTH", 24)
        sub(dtc_dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32", "DISPLAY-RADIX": "HEX"})
        compu = sub(dtc_dop, "COMPU-METHOD")
        sub(compu, "CATEGORY", "IDENTICAL")
    dtcs_node = dtc_dop.find("DTCS")
    if dtcs_node is None:
        dtcs_node = sub(dtc_dop, "DTCS")
    for child in list(dtcs_node):
        dtcs_node.remove(child)
    caption_ids: dict[str, str] = {}
    for index, dtc in enumerate(dtcs):
        dtc_node = sub(dtcs_node, "DTC", attrib={"ID": id_gen.new("DTC")})
        sub(dtc_node, "SHORT-NAME", sanitize_short_name(f"DTC_{dtc.byte_code:06X}", "DTC"))
        sub(dtc_node, "TROUBLE-CODE", dtc.byte_code)
        sub(dtc_node, "DISPLAY-TROUBLE-CODE", dtc.display_code)
        sub(dtc_node, "TEXT", dtc.text or dtc.display_code)
        sdgs = sub(dtc_node, "SDGS")
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_SHORTNAME", f"DTC_0X{dtc.byte_code:06X}", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_PRIORITY_VALUE", dtc.priority or "2", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_SUPPORTED", "supported", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_CYCLE", "DEM_POWER", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_COUNTER", "40", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_SEVERITY_VALUE", "noSeverity", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_INIT_MONITOR_REQUIRED", "not required", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_FUNCTIONAL_UNIT_VALUE", "0x00", first=index == 0)

    ensure_vf_dtc_extended_table(root, id_gen, dtcs)
    update_dtc_text_table(root, dtcs)

    ext_table = first_by_short(root, "TABLE", "DTCExtendedDataRecordNumber")
    if ext_table is None:
        return
    default_structure = first_table_structure_ref(ext_table) or create_extended_record_number_structure(root, id_gen)
    clear_children(ext_table, "TABLE-ROW")
    for dtc in dtcs:
        label = dtc_table_key(dtc)
        row = append_table_row(ext_table, id_gen, f"TR_DTC_{dtc.byte_code:06X}", label, label, default_structure)
        sdgs = sub(row, "SDGS")
        sdg = sub(sdgs, "SDG")
        caption = sub(sdg, "SDG-CAPTION", attrib={"ID": id_gen.new("CAP")})
        sub(caption, "SHORT-NAME", "IsDefaultCase")
        sub(sdg, "SD", "Yes")


def first_table_structure_ref(table: etree._Element) -> str | None:
    ref = table.find(".//STRUCTURE-REF")
    return ref.get("ID-REF") if ref is not None else None


def create_extended_record_number_structure(root: etree._Element, id_gen: Any) -> str:
    structures = find_container(root, "STRUCTURES")
    structure_id = id_gen.new("STR")
    structure = sub(structures, "STRUCTURE", attrib={"ID": structure_id, "IS-VISIBLE": "false"})
    sub(structure, "SHORT-NAME", "STRUC_DTCExtendedDataRecordNumbers")
    params = sub(structure, "PARAMS")
    params.append(make_value_param("Record_Numbers", "Record Numbers", 0, 0, dop_id_by_short(root, "DATA-OBJECT-PROP", "DTCExtendedDataRecordNumbers_All") or "_62"))
    return structure_id


def dop_id_by_short(root: etree._Element, tag: str, short_name: str) -> str | None:
    node = first_by_short(root, tag, short_name)
    return node.get("ID") if node is not None else None


def snapshot_number_of_ids_param(
    root: etree._Element,
    short_name: str,
    long_name: str,
    byte_position: int,
) -> etree._Element:
    param = element("PARAM")
    set_xsi_type(param, "VALUE")
    sub(param, "SHORT-NAME", short_name)
    sub(param, "LONG-NAME", long_name)
    sub(param, "BYTE-POSITION", byte_position)
    sub(
        param,
        "DOP-REF",
        attrib={
            "ID-REF": (
                dop_id_by_short(root, "DATA-OBJECT-PROP", "DTCSnapshotRecordNumberOfIdentifiers")
                or dop_id_by_short(root, "DATA-OBJECT-PROP", "HexDump_1Byte")
                or "_73"
            )
        },
    )
    return param


def ensure_vf_ddds_container(root: etree._Element, tag: str, before_tags: set[str]) -> etree._Element:
    ddds = find_dictionary_spec(root)
    if ddds is None:
        raise RuntimeError("Template ODX has no DIAG-DATA-DICTIONARY-SPEC")
    existing = ddds.find(tag)
    if existing is not None:
        return existing
    container = element(tag)
    for index, child in enumerate(ddds):
        if child.tag in before_tags:
            ddds.insert(index, container)
            return container
    ddds.append(container)
    return container


def ensure_vf_snapshot_env_data(root: etree._Element, id_gen: Any) -> etree._Element:
    env_data = first_by_short(root, "ENV-DATA", "ENVDATA_ALLDTCS")
    if env_data is not None:
        return env_data
    env_datas = ensure_vf_ddds_container(root, "ENV-DATAS", {"UNIT-SPEC", "TABLES"})
    env_data = sub(env_datas, "ENV-DATA", attrib={"ID": id_gen.new("ENV")})
    sub(env_data, "SHORT-NAME", "ENVDATA_ALLDTCS")
    sub(env_data, "LONG-NAME", "ENVDATA_ALLDTCS")
    sub(env_data, "PARAMS")
    sub(env_data, "ALL-VALUE")
    return env_data


def ensure_vf_snapshot_env_data_desc(root: etree._Element, id_gen: Any, env_data_id: str) -> str:
    desc = first_by_short(root, "ENV-DATA-DESC", "DTCSnapshotRecordData")
    if desc is None:
        descs = ensure_vf_ddds_container(root, "ENV-DATA-DESCS", {"DATA-OBJECT-PROPS", "STRUCTURES"})
        desc = sub(descs, "ENV-DATA-DESC", attrib={"ID": id_gen.new("EDD")})
        sub(desc, "SHORT-NAME", "DTCSnapshotRecordData")
        sub(desc, "LONG-NAME", "DTCSnapshotRecordData")
    param_ref = desc.find("PARAM-SNREF")
    if param_ref is None:
        param_ref = element("PARAM-SNREF", attrib={"SHORT-NAME": "DTC"})
        env_refs = desc.find("ENV-DATA-REFS")
        insert_at = desc.index(env_refs) if env_refs is not None else len(desc)
        desc.insert(insert_at, param_ref)
    else:
        param_ref.set("SHORT-NAME", "DTC")
    env_refs = desc.find("ENV-DATA-REFS")
    if env_refs is None:
        env_refs = sub(desc, "ENV-DATA-REFS")
    for child in list(env_refs):
        env_refs.remove(child)
    sub(env_refs, "ENV-DATA-REF", attrib={"ID-REF": env_data_id})
    return desc.get("ID")


def update_vf_snapshot_and_extended_data(
    root: etree._Element,
    id_gen: Any,
    snapshots: list[SnapshotDef],
    extended_records: list[ExtendedRecordDef],
    snapshot_record_nums: list[int],
    snapshot_record_names: dict[int, str],
) -> None:
    update_snapshot_record_number_dop(
        root,
        "DTCSnapshotRecordNumbers_All",
        snapshot_record_nums,
        snapshot_record_names,
        include_all=True,
    )
    update_snapshot_record_number_dop(
        root,
        "DTCSnapshotRecordNumbers_All_except_FF",
        snapshot_record_nums,
        snapshot_record_names,
        include_all=False,
    )

    env_data = first_by_short(root, "ENV-DATA", "ENVDATA_ALLDTCS")
    if env_data is None and snapshots:
        env_data = ensure_vf_snapshot_env_data(root, id_gen)
    did_dop_id = dop_id_by_short(root, "DATA-OBJECT-PROP", "HexDump_2Byte") or "_2"
    if env_data is not None:
        params = env_data.find("PARAMS")
        if params is None:
            params = sub(env_data, "PARAMS")
        for child in list(params):
            params.remove(child)
        byte_position = 0
        for snapshot in snapshots:
            if not snapshot.structure_id:
                continue
            did_param = element("PARAM", attrib={"SEMANTIC": "DATA"})
            set_xsi_type(did_param, "PHYS-CONST")
            snapshot_short_name = sanitize_short_name(snapshot.desc, hex_short("DID", snapshot.did))
            sub(did_param, "SHORT-NAME", snapshot_short_name)
            sub(did_param, "LONG-NAME", snapshot.desc)
            sub(did_param, "BYTE-POSITION", byte_position)
            sub(did_param, "PHYS-CONSTANT-VALUE", snapshot.did)
            sub(did_param, "DOP-REF", attrib={"ID-REF": did_dop_id})
            params.append(did_param)
            byte_position += 2
            value_param = make_value_param(
                sanitize_short_name(f"{snapshot_short_name}_Data", "SnapshotData"),
                snapshot.desc,
                byte_position,
                0,
                snapshot.structure_id,
            )
            params.append(value_param)
            byte_position += max(1, snapshot.size)
        snapshot_env_desc_id = ensure_vf_snapshot_env_data_desc(root, id_gen, env_data.get("ID"))
        list_struct = first_by_short(root, "STRUCTURE", "ListOfDTCSnapshotRecord")
        if list_struct is not None and snapshot_env_desc_id:
            params = list_struct.find("PARAMS")
            if params is None:
                params = sub(list_struct, "PARAMS")
            for child in list(params):
                params.remove(child)
            params.append(
                make_value_param(
                    "Record_Numbers",
                    "Record Numbers",
                    0,
                    0,
                    dop_id_by_short(root, "DATA-OBJECT-PROP", "DTCSnapshotRecordNumbers_All") or "_53",
                )
            )
            params.append(
                snapshot_number_of_ids_param(
                    root,
                    "DTCSnapshotRecordNumberOfIdentifiers",
                    "DTCSnapshotRecordNumberOfIdentifiers",
                    1,
                )
            )
            params.append(
                make_value_param(
                    "DTCSnapshotRecordData",
                    "DTCSnapshotRecordData",
                    2,
                    0,
                    snapshot_env_desc_id,
                )
            )

    if extended_records:
        update_record_number_dop(root, "DTCExtendedDataRecordNumbers_All", extended_records, include_all=True)
        update_record_number_dop(root, "DTCExtendedDataRecordNumbers_All_except_FF", extended_records, include_all=False)
        mux = first_by_short(root, "MUX", "DTCExtendedDataRecordData")
        if mux is not None:
            cases = mux.find("CASES")
            if cases is None:
                cases = sub(mux, "CASES")
            for child in list(cases):
                cases.remove(child)
            for record in extended_records:
                if not record.structure_id:
                    continue
                case = sub(cases, "CASE")
                label = extended_record_label(record)
                sub(case, "SHORT-NAME", sanitize_short_name(f"Case_0x{record.record_num:02X}", "Case"))
                sub(case, "STRUCTURE-REF", attrib={"ID-REF": record.structure_id})
                sub(case, "LOWER-LIMIT", record.record_num)
                sub(case, "UPPER-LIMIT", record.record_num)


def collect_env_data_groups(params: etree._Element) -> dict[int, list[etree._Element]]:
    groups: dict[int, list[etree._Element]] = {}
    current_did: int | None = None
    for param in params.findall("PARAM"):
        if param.get(XSI_TYPE) == "PHYS-CONST":
            try:
                current_did = int(compact_text(param.findtext("PHYS-CONSTANT-VALUE")), 0)
            except ValueError:
                current_did = None
            if current_did is not None:
                groups[current_did] = [param]
            continue
        if current_did is not None:
            groups[current_did].append(param)
    return groups


def existing_env_data_dids(root: etree._Element) -> set[int]:
    env_data = first_by_short(root, "ENV-DATA", "ENVDATA_ALLDTCS")
    if env_data is None:
        return set()
    params = env_data.find("PARAMS")
    if params is None:
        return set()
    return set(collect_env_data_groups(params))


def env_param_end_byte(root: etree._Element, param: etree._Element) -> int:
    byte_position = parse_int_cell(param.findtext("BYTE-POSITION"), 0)
    dop_ref = param.find("DOP-REF")
    if dop_ref is None:
        return byte_position + 1
    dop = find_by_id(root, dop_ref.get("ID-REF"))
    if dop is None:
        return byte_position + 1
    bit_lengths = dop.xpath('.//*[local-name()="BIT-LENGTH"]/text()')
    if not bit_lengths:
        return byte_position + 1
    try:
        bit_length = int(bit_lengths[0])
    except ValueError:
        return byte_position + 1
    return byte_position + max(1, (bit_length + 7) // 8)


def update_snapshot_record_number_dop(
    root: etree._Element,
    short_name: str,
    record_nums: list[int],
    record_names: dict[int, str],
    *,
    include_all: bool,
) -> None:
    dop = first_by_short(root, "DATA-OBJECT-PROP", short_name)
    if dop is None:
        return
    labels: list[tuple[int, str]] = []
    seen: set[int] = set()
    for record_num in record_nums:
        if record_num == 0xFF:
            continue
        if record_num in seen:
            continue
        seen.add(record_num)
        labels.append((record_num, record_names.get(record_num, f"Snapshot Record 0x{record_num:02X}")))
    if include_all:
        labels.append((0xFF, record_names.get(0xFF, "All")))
    normalize_vf_texttable_dop(dop, 8, labels)


def update_base_variant_comparams(root: etree._Element, cover: CoverInfo) -> None:
    base_variant = get_base_variant(root)
    if base_variant is None:
        return
    base_short_name = base_variant.findtext("SHORT-NAME") or "VF_ECU_CAN"
    comp_refs = base_variant.find("COMPARAM-REFS")
    if comp_refs is None:
        comp_refs = sub(base_variant, "COMPARAM-REFS")
    values = getattr(cover, "comm_params", {})
    for ref in comp_refs.findall("COMPARAM-REF"):
        id_ref = ref.get("ID-REF")
        if id_ref not in values:
            continue
        simple = ref.find("SIMPLE-VALUE")
        if simple is not None:
            simple.text = str(values[id_ref])
    unique_ref = None
    for ref in comp_refs.findall("COMPARAM-REF"):
        if ref.get("ID-REF") == "ISO_15765_2.CP_UniqueRespIdTable":
            unique_ref = ref
            break
    if unique_ref is None:
        unique_ref = sub(
            comp_refs,
            "COMPARAM-REF",
            attrib={"ID-REF": "ISO_15765_2.CP_UniqueRespIdTable", "DOCREF": "ISO_15765_2", "DOCTYPE": "COMPARAM-SUBSET"},
        )
    for child in list(unique_ref):
        unique_ref.remove(child)
    rx_phy = cover.rx_phy_id if cover.rx_phy_id is not None else 0x705
    tx = cover.tx_id if cover.tx_id is not None else 0x785
    rx_fun = cover.rx_fun_id if cover.rx_fun_id is not None else 0x7DF
    uudt_resp_id = UUDT_DISABLED_CAN_ID
    complex_value = sub(unique_ref, "COMPLEX-VALUE")
    for value in (
        "0",
        "normal segmented 11-bit transmit with FC",
        str(rx_phy),
        "0",
        "normal segmented 11-bit receive with FC",
        str(tx),
        "0",
        "normal unsegmented 11-bit receive",
        str(uudt_resp_id),
        base_short_name,
    ):
        sub(complex_value, "SIMPLE-VALUE", value)
    sub(unique_ref, "PROTOCOL-SNREF", attrib={"SHORT-NAME": "CAN"})


def update_session_timing(root: etree._Element, cover: CoverInfo) -> None:
    timing = getattr(cover, "session_timing", {"P2": 50, "P2Ex": 5000})
    text = f"{{P2={timing.get('P2', 50)}, P2Ex={timing.get('P2Ex', 5000)}}}"
    for state in root.xpath('//*[local-name()="STATE" and (SHORT-NAME="Default" or SHORT-NAME="Programming" or SHORT-NAME="Extended")]'):
        desc = state.find("DESC")
        if desc is None:
            desc = sub(state, "DESC")
        for child in list(desc):
            desc.remove(child)
        sub(desc, "p", text)


def validate_with_odxtools(pdx_path: Path) -> None:
    cmd = [sys.executable, "-m", "odxtools", "list", str(pdx_path), "--all"]
    result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
    if result.returncode != 0:
        details = (result.stderr or result.stdout).strip()
        raise RuntimeError(f"odxtools validation failed for {pdx_path}:\n{details}")


def find_default_xlsx(base_dir: Path) -> Path:
    candidates = sorted(path for path in base_dir.glob("*.xlsx") if not path.name.startswith("~$"))
    if not candidates:
        raise FileNotFoundError("No .xlsx survey file found in the current directory")
    return candidates[0]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Generate VF PDX from a VF diagnosis survey Excel file.")
    parser.add_argument("xlsx", nargs="?", type=Path, help="Input VF diagnosis survey .xlsx file")
    parser.add_argument("--template", type=Path, default=SCRIPT_DIR / "templates" / "VF_ECU_CAN_v15.pdx", help="Template PDX")
    parser.add_argument("--output-dir", type=Path, default=SCRIPT_DIR / "output", help="Output directory")
    parser.add_argument("--no-validate", action="store_true", help="Skip odxtools validation")
    args = parser.parse_args(argv)

    xlsx_path = args.xlsx or find_default_xlsx(SCRIPT_DIR)
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)
    if not args.template.exists():
        raise FileNotFoundError(args.template)

    survey = parse_vf_survey(xlsx_path)
    output_pdx = args.output_dir / f"{xlsx_path.stem}.pdx"
    update_template(args.template, output_pdx, survey, validate=not args.no_validate)
    did_data_objects = sum(len(did.params) for did in survey.dids)
    converted_did_objects = sum(
        1
        for did in survey.dids
        for param in did.params
        if getattr(param.conversion, "kind", "identity") in {"enum", "linear"}
    )
    print(f"Generated: {output_pdx}")
    print(
        "Parsed: "
        f"{len(survey.dids)} DID identifiers, "
        f"{did_data_objects} DID data objects, "
        f"{converted_did_objects} converted DID data objects, "
        f"{len(survey.io_dids)} IO DIDs, "
        f"{len(survey.routines)} routines, "
        f"{len(survey.dtcs)} DTCs, "
        f"{len(survey.snapshots)} snapshot DIDs, "
        f"{len(getattr(survey, 'snapshot_record_nums', []))} snapshot records, "
        f"{len(survey.extended_records)} extended records"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

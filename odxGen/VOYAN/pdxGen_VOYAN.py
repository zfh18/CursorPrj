#!/usr/bin/env python3
"""Generate a VOYAH/VOYAN CANdelaStudio-compatible PDX from a diagnosis survey.

The generator intentionally uses the Vector CANdelaStudio-exported PDX as a
structural template. It parses the OEM Excel survey into an ECU data model, then
updates ODX DTC/DID/IO/Routine tables and communication parameters in the
template before repackaging the result as a .pdx.
"""

from __future__ import annotations

import argparse
import copy
import re
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from lxml import etree
from openpyxl import load_workbook


XSI_NS = "http://www.w3.org/2001/XMLSchema-instance"
XSI_TYPE = f"{{{XSI_NS}}}type"
etree.register_namespace("xsi", XSI_NS)

CAN_ONLY_PDX_FILES = {
    "FGL_UDS.odx-d",
    "ISO_11898_2_DWCAN.odx-cs",
    "ISO_15765_2.odx-cs",
    "ISO_15765_3.odx-cs",
    "ISO_15765_3_on_ISO_15765_2.odx-c",
    "index.xml",
    "VOYAN_ECU_CAN_v15.odx-d",
}


@dataclass
class Conversion:
    kind: str = "identity"
    enum: list[tuple[int, int, str]] = field(default_factory=list)
    a: float = 1.0
    b: float = 0.0
    precision: int | None = None


@dataclass
class ParamDef:
    name: str
    long_name: str
    byte_pos: int
    bit_pos: int
    bit_len: int
    data_type: str = "Hex(Unsigned)"
    unit: str = ""
    conversion: Conversion = field(default_factory=Conversion)
    min_value: str = ""
    max_value: str = ""
    dop_id: str = ""


@dataclass
class DidDef:
    did: int
    desc: str
    size: int
    params: list[ParamDef] = field(default_factory=list)
    write_security: str = "N"
    sessions: list[str] = field(default_factory=list)
    structure_id: str = ""
    wrapper_id: str = ""
    short_name: str = ""
    long_name: str = ""

    @property
    def readable(self) -> bool:
        return any("R" in normalize_access(v) for v in self.sessions)

    @property
    def writable(self) -> bool:
        return any("W" in normalize_access(v) for v in self.sessions)


@dataclass
class IoDidDef:
    did: int
    desc: str
    size: int
    controls: set[int] = field(default_factory=set)
    params: list[ParamDef] = field(default_factory=list)
    structure_id: str = ""
    status_wrapper_id: str = ""
    control_request_wrapper_id: str = ""
    short_name: str = ""
    long_name: str = ""


@dataclass
class RoutineSubFunction:
    control_type: int
    supported: bool
    option_params: list[ParamDef] = field(default_factory=list)
    status_params: list[ParamDef] = field(default_factory=list)
    option_structure_id: str = ""
    status_structure_id: str = ""


@dataclass
class RoutineDef:
    rid: int
    desc: str
    security: str = "N"
    sessions: list[str] = field(default_factory=list)
    subfunctions: dict[int, RoutineSubFunction] = field(default_factory=dict)
    short_name: str = ""
    long_name: str = ""


@dataclass
class DtcDef:
    display_code: str
    byte_code: int
    text: str
    priority: str = ""


@dataclass
class SnapshotDef:
    record_num: int | None
    did: int
    desc: str
    size: int
    params: list[ParamDef] = field(default_factory=list)
    structure_id: str = ""


@dataclass
class ExtendedRecordDef:
    record_num: int
    desc: str
    size: int
    params: list[ParamDef] = field(default_factory=list)
    structure_id: str = ""


@dataclass
class CoverInfo:
    ecu_name: str = "VOYAN_ECU_CAN"
    vehicle: str = ""
    supplier: str = ""
    tx_id: int | None = None
    rx_phy_id: int | None = None
    rx_fun_id: int | None = None
    bus_type: str = "CAN"


@dataclass
class SurveyData:
    cover: CoverInfo
    dids: list[DidDef]
    io_dids: list[IoDidDef]
    routines: list[RoutineDef]
    dtcs: list[DtcDef]
    snapshots: list[SnapshotDef]
    extended_records: list[ExtendedRecordDef]


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip().lstrip("'")


def compact_text(value: Any) -> str:
    return re.sub(r"\s+", " ", cell_text(value).replace("\u3000", " ")).strip()


def normalize_access(value: Any) -> str:
    return compact_text(value).upper().replace(" ", "")


def split_name(value: str) -> tuple[str, str]:
    text = cell_text(value)
    parts = [p.strip() for p in re.split(r"[\r\n]+", text) if p.strip()]
    if not parts:
        return "", ""
    english = next((p for p in parts if re.search(r"[A-Za-z]", p)), parts[0])
    return english, text.replace("\r\n", "\n").strip()


def parse_hex_cell(value: Any, *, max_value: int = 0xFFFFFFFF) -> int | None:
    text = compact_text(value).replace(" ", "")
    if not re.fullmatch(r"(?:0[xX])?[0-9A-Fa-f]{1,8}", text):
        return None
    result = int(text[2:] if text.lower().startswith("0x") else text, 16)
    if result > max_value:
        return None
    return result


def parse_hex_in_text(value: Any) -> int | None:
    text = cell_text(value)
    match = re.search(r"0[xX]([0-9A-Fa-f]+)", text)
    if match:
        return int(match.group(1), 16)
    return None


def parse_int_cell(value: Any, default: int = 0) -> int:
    text = compact_text(value)
    if not text:
        return default
    match = re.search(r"-?\d+", text)
    if not match:
        return default
    return int(match.group(0))


def parse_byte_range(value: Any, default: int = 0) -> tuple[int, int]:
    text = compact_text(value)
    if not text:
        return default, default
    nums = [int(n) for n in re.findall(r"\d+", text)]
    if not nums:
        return default, default
    start = nums[0]
    end = nums[1] if len(nums) > 1 else start
    if end < start:
        end = start
    return start, end


def parse_bit_range(value: Any, byte_start: int, byte_end: int) -> tuple[int, int]:
    text = compact_text(value)
    if not text:
        return 0, max(1, (byte_end - byte_start + 1) * 8)
    nums = [int(n) for n in re.findall(r"\d+", text)]
    if not nums:
        return 0, max(1, (byte_end - byte_start + 1) * 8)
    if "-" in text and len(nums) > 1:
        start, end = nums[0], nums[1]
        return start % 8, max(1, end - start + 1)
    return nums[0] % 8, 1


def parse_conversion(value: Any) -> Conversion:
    text = cell_text(value)
    normalized = (
        text.replace("：", ":")
        .replace("；", ";")
        .replace("–", "-")
        .replace("—", "-")
    )
    enum_entries: list[tuple[int, int, str]] = []
    enum_pattern = re.compile(
        r"(?:0[xX])?([0-9A-Fa-f]+)\s*(?:-\s*(?:0[xX])?([0-9A-Fa-f]+))?\s*:\s*([^;\n\r]+)"
    )
    for match in enum_pattern.finditer(normalized):
        lo = int(match.group(1), 16)
        hi = int(match.group(2), 16) if match.group(2) else lo
        label = match.group(3).strip()
        if label:
            enum_entries.append((lo, hi, label))
    if enum_entries and re.search(r"0[xX]", normalized):
        return Conversion(kind="enum", enum=enum_entries)

    if re.search(r"y\s*=\s*a\s*x\s*\+\s*b", normalized, flags=re.IGNORECASE):
        a = _parse_float_assignment(normalized, "a", 1.0)
        b = _parse_float_assignment(normalized, "b", 0.0)
        precision_match = re.search(r"precision\s*=\s*(-?\d+)", normalized, flags=re.IGNORECASE)
        precision = int(precision_match.group(1)) if precision_match else None
        return Conversion(kind="linear", a=a, b=b, precision=precision)

    return Conversion()


def _parse_float_assignment(text: str, key: str, default: float) -> float:
    match = re.search(rf"\b{re.escape(key)}\s*=\s*(-?\d+(?:\.\d+)?)", text, flags=re.IGNORECASE)
    return float(match.group(1)) if match else default


def make_param_from_cells(
    *,
    name_value: Any,
    byte_value: Any,
    bit_value: Any,
    data_type_value: Any,
    unit_value: Any = "",
    conversion_value: Any = "",
    min_value: Any = "",
    max_value: Any = "",
    fallback_name: str = "Data",
    default_byte: int = 0,
) -> ParamDef | None:
    english, long_name = split_name(cell_text(name_value) or fallback_name)
    if not english and not long_name:
        return None
    byte_start, byte_end = parse_byte_range(byte_value, default_byte)
    bit_pos, bit_len = parse_bit_range(bit_value, byte_start, byte_end)
    data_type = compact_text(data_type_value) or "Hex(Unsigned)"
    return ParamDef(
        name=english or fallback_name,
        long_name=long_name or english or fallback_name,
        byte_pos=byte_start,
        bit_pos=bit_pos,
        bit_len=bit_len,
        data_type=data_type,
        unit=compact_text(unit_value),
        conversion=parse_conversion(conversion_value),
        min_value=compact_text(min_value),
        max_value=compact_text(max_value),
    )


def parse_structured_params(value: Any, fallback_prefix: str) -> list[ParamDef]:
    text = cell_text(value)
    if not text or text.strip().upper() in {"N/A", "NA", "N"}:
        return []
    normalized = text.replace("：", ":").replace("；", ";")
    if "Name:" not in normalized:
        # Some surveys put a single conversion-like field in this cell.
        p = make_param_from_cells(
            name_value=fallback_prefix,
            byte_value="0",
            bit_value="0-7",
            data_type_value="Hex(Unsigned)",
            conversion_value=normalized,
            fallback_name=fallback_prefix,
        )
        return [p] if p else []

    params: list[ParamDef] = []
    chunks = re.split(r"(?=Name\s*:)", normalized, flags=re.IGNORECASE)
    for idx, chunk in enumerate(chunks):
        if "Name" not in chunk:
            continue
        name_match = re.search(r"Name\s*:\s*([^,\n;]+)", chunk, flags=re.IGNORECASE)
        byte_match = re.search(r"Byte\s*:\s*([^,\n;]+)", chunk, flags=re.IGNORECASE)
        bit_match = re.search(r"Bit\s*:\s*([^,\n;]+)", chunk, flags=re.IGNORECASE)
        conv_match = re.search(r"Conversion\s*:\s*(.*)", chunk, flags=re.IGNORECASE | re.DOTALL)
        name = name_match.group(1).strip() if name_match else f"{fallback_prefix}_{idx + 1}"
        byte = byte_match.group(1).strip() if byte_match else "0"
        bit = bit_match.group(1).strip() if bit_match else "0-7"
        conv = conv_match.group(1).strip() if conv_match else ""
        param = make_param_from_cells(
            name_value=name,
            byte_value=byte,
            bit_value=bit,
            data_type_value="Hex(Unsigned)",
            conversion_value=conv,
            fallback_name=f"{fallback_prefix}_{idx + 1}",
        )
        if param:
            params.append(param)
    return params


def sanitize_short_name(value: str, fallback: str, used: set[str] | None = None, max_len: int = 120) -> str:
    base = value.strip() if value else fallback
    base = base.encode("ascii", "ignore").decode("ascii")
    base = re.sub(r"[^A-Za-z0-9_]+", "_", base).strip("_")
    if not base:
        base = fallback
    if not re.match(r"[A-Za-z_]", base):
        base = f"X_{base}"
    base = base[:max_len].rstrip("_") or fallback
    if used is None:
        return base
    candidate = base
    index = 2
    while candidate in used:
        suffix = f"_{index}"
        candidate = f"{base[: max_len - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def hex_short(prefix: str, value: int, width: int = 4) -> str:
    return f"{prefix}_0x{value:0{width}X}"


def find_sheet(workbook: Any, predicate: Any) -> Any | None:
    for sheet_name in workbook.sheetnames:
        if predicate(sheet_name.strip()):
            return workbook[sheet_name]
    return None


def parse_survey(xlsx_path: Path) -> SurveyData:
    workbook = load_workbook(xlsx_path, data_only=True)
    cover = parse_cover(workbook)
    dids = parse_read_write_dids(workbook)
    io_dids = parse_io_dids(workbook)
    routines = parse_routines(workbook)
    dtcs = parse_dtcs(workbook)
    snapshots, extended_records = parse_snapshot_extended(workbook)
    return SurveyData(cover, dids, io_dids, routines, dtcs, snapshots, extended_records)


def parse_cover(workbook: Any) -> CoverInfo:
    sheet = find_sheet(workbook, lambda name: name.startswith("0_1"))
    if sheet is None:
        return CoverInfo()

    cover = CoverInfo()
    cover.ecu_name = compact_text(sheet["B12"].value) or compact_text(sheet["B11"].value) or cover.ecu_name
    cover.vehicle = compact_text(sheet["E9"].value).replace("车型:", "").replace("车型：", "").strip()
    cover.supplier = compact_text(sheet["F12"].value)
    cover.tx_id = parse_hex_in_text(sheet["C12"].value)
    cover.rx_phy_id = parse_hex_in_text(sheet["C13"].value)
    cover.rx_fun_id = parse_hex_in_text(sheet["C14"].value)
    cover.bus_type = compact_text(sheet["D12"].value) or compact_text(sheet["F10"].value) or "CAN"
    return cover


def locate_header_row(sheet: Any, checks: Iterable[tuple[int, str]], max_scan: int = 30) -> int:
    for row in range(1, min(sheet.max_row, max_scan) + 1):
        ok = True
        for col, needle in checks:
            if needle.lower() not in compact_text(sheet.cell(row, col).value).lower():
                ok = False
                break
        if ok:
            return row
    return 1


def parse_read_write_dids(workbook: Any) -> list[DidDef]:
    sheets = []
    for sheet_name in workbook.sheetnames:
        stripped = sheet_name.strip()
        if "Read&Write DID" in stripped or (
            "DID" in stripped and "IO" not in stripped and "Routine" not in stripped
        ):
            sheet = workbook[sheet_name]
            header = locate_header_row(sheet, [(3, "DID")])
            if header > 1:
                sheets.append(sheet)

    dids: list[DidDef] = []
    for sheet in sheets:
        header = locate_header_row(sheet, [(3, "DID")])
        current: DidDef | None = None
        for row in range(header + 1, sheet.max_row + 1):
            did_value = parse_hex_cell(sheet.cell(row, 3).value, max_value=0xFFFF)
            note = compact_text(sheet.cell(row, 2).value).lower()
            if did_value is not None and not any(x in note for x in ("these", "please", "note")):
                desc_english, desc_long = split_name(cell_text(sheet.cell(row, 4).value))
                size = parse_int_cell(sheet.cell(row, 5).value, default=0)
                current = DidDef(
                    did=did_value,
                    desc=desc_long or desc_english or hex_short("DID", did_value),
                    size=size,
                    write_security=compact_text(sheet.cell(row, 14).value) or "N",
                    sessions=[compact_text(sheet.cell(row, c).value) for c in range(15, 20)],
                )
                first_param = make_param_from_cells(
                    name_value=sheet.cell(row, 8).value,
                    byte_value=sheet.cell(row, 6).value,
                    bit_value=sheet.cell(row, 7).value,
                    data_type_value=sheet.cell(row, 13).value,
                    unit_value=sheet.cell(row, 11).value,
                    conversion_value=sheet.cell(row, 12).value,
                    min_value=sheet.cell(row, 9).value,
                    max_value=sheet.cell(row, 10).value,
                    fallback_name=desc_english or hex_short("DID", did_value),
                )
                if first_param:
                    current.params.append(first_param)
                dids.append(current)
                continue

            if current is None:
                continue
            has_param = any(compact_text(sheet.cell(row, c).value) for c in (6, 7, 8, 12, 13))
            if not has_param:
                continue
            param = make_param_from_cells(
                name_value=sheet.cell(row, 8).value,
                byte_value=sheet.cell(row, 6).value,
                bit_value=sheet.cell(row, 7).value,
                data_type_value=sheet.cell(row, 13).value,
                unit_value=sheet.cell(row, 11).value,
                conversion_value=sheet.cell(row, 12).value,
                min_value=sheet.cell(row, 9).value,
                max_value=sheet.cell(row, 10).value,
                fallback_name=f"{hex_short('DID', current.did)}_Data",
            )
            if param:
                current.params.append(param)

    dedup: dict[int, DidDef] = {}
    for did in dids:
        dedup[did.did] = did
    return list(dedup.values())


def parse_io_dids(workbook: Any) -> list[IoDidDef]:
    sheet = find_sheet(workbook, lambda name: "IO DID" in name)
    if sheet is None:
        return []

    header = locate_header_row(sheet, [(2, "DID")], max_scan=20)
    io_by_id: dict[int, IoDidDef] = {}
    current: IoDidDef | None = None
    for row in range(header + 1, sheet.max_row + 1):
        did_value = parse_hex_cell(sheet.cell(row, 2).value, max_value=0xFFFF)
        control = parse_hex_cell(sheet.cell(row, 4).value, max_value=0x03)
        if did_value is not None:
            desc_english, desc_long = split_name(cell_text(sheet.cell(row, 3).value))
            current = io_by_id.get(did_value)
            if current is None:
                current = IoDidDef(
                    did=did_value,
                    desc=desc_long or desc_english or hex_short("DID", did_value),
                    size=parse_int_cell(sheet.cell(row, 5).value, default=0),
                )
                io_by_id[did_value] = current

        if current is not None and control is not None:
            current.controls.add(control)

        if did_value is not None and control is not None:
            param = make_param_from_cells(
                name_value=sheet.cell(row, 8).value,
                byte_value=sheet.cell(row, 6).value,
                bit_value=sheet.cell(row, 7).value,
                data_type_value=sheet.cell(row, 13).value,
                unit_value=sheet.cell(row, 11).value,
                conversion_value=sheet.cell(row, 12).value,
                min_value=sheet.cell(row, 9).value,
                max_value=sheet.cell(row, 10).value,
                fallback_name=desc_english or hex_short("IODID", did_value),
            )
            if param:
                current.params.append(param)
            continue

        if current is None:
            continue
        has_param = any(compact_text(sheet.cell(row, c).value) for c in (6, 7, 8, 12, 13))
        if not has_param:
            continue
        param = make_param_from_cells(
            name_value=sheet.cell(row, 8).value,
            byte_value=sheet.cell(row, 6).value,
            bit_value=sheet.cell(row, 7).value,
            data_type_value=sheet.cell(row, 13).value,
            unit_value=sheet.cell(row, 11).value,
            conversion_value=sheet.cell(row, 12).value,
            min_value=sheet.cell(row, 9).value,
            max_value=sheet.cell(row, 10).value,
            fallback_name=f"{hex_short('IODID', current.did)}_Data",
        )
        if param:
            current.params.append(param)
    return list(io_by_id.values())


def parse_routines(workbook: Any) -> list[RoutineDef]:
    sheet = find_sheet(workbook, lambda name: "Routine DID" in name)
    if sheet is None:
        return []

    header = locate_header_row(sheet, [(2, "Routin")], max_scan=20)
    routines: list[RoutineDef] = []
    current: RoutineDef | None = None
    for row in range(header + 1, sheet.max_row + 1):
        rid = parse_hex_cell(sheet.cell(row, 2).value, max_value=0xFFFF)
        control_type = parse_hex_cell(str(sheet.cell(row, 4).value).split()[0] if sheet.cell(row, 4).value else "", max_value=0x03)
        supported = compact_text(sheet.cell(row, 5).value).upper() == "Y"
        if rid is not None and control_type is not None:
            desc_english, desc_long = split_name(cell_text(sheet.cell(row, 3).value))
            current = RoutineDef(
                rid=rid,
                desc=desc_long or desc_english or hex_short("RID", rid),
                security=compact_text(sheet.cell(row, 8).value) or "N",
                sessions=[compact_text(sheet.cell(row, c).value) for c in range(9, 14)],
            )
            current.subfunctions[control_type] = RoutineSubFunction(
                control_type=control_type,
                supported=supported,
                option_params=parse_structured_params(sheet.cell(row, 6).value, "RoutineControlOptionRecord"),
                status_params=parse_structured_params(sheet.cell(row, 7).value, "RoutineStatusRecord"),
            )
            routines.append(current)
            continue

        if current is not None and control_type is not None:
            current.subfunctions[control_type] = RoutineSubFunction(
                control_type=control_type,
                supported=supported,
                option_params=parse_structured_params(sheet.cell(row, 6).value, "RoutineControlOptionRecord"),
                status_params=parse_structured_params(sheet.cell(row, 7).value, "RoutineStatusRecord"),
            )
    return routines


def parse_dtcs(workbook: Any) -> list[DtcDef]:
    sheet = find_sheet(workbook, lambda name: "DTC Information" in name)
    if sheet is None:
        return []
    header = locate_header_row(sheet, [(2, "DTC Display")], max_scan=30)
    dtcs: list[DtcDef] = []
    for row in range(header + 1, sheet.max_row + 1):
        display = compact_text(sheet.cell(row, 2).value).upper()
        byte_text = compact_text(sheet.cell(row, 3).value).upper().replace("0X", "")
        if not re.fullmatch(r"[PCBU][0-9A-F]{6}", display):
            continue
        if not re.fullmatch(r"[0-9A-F]{6}", byte_text):
            continue
        _, text = split_name(cell_text(sheet.cell(row, 4).value))
        dtcs.append(
            DtcDef(
                display_code=display,
                byte_code=int(byte_text, 16),
                text=text or display,
                priority=compact_text(sheet.cell(row, 5).value),
            )
        )
    return dtcs


def parse_snapshot_extended(workbook: Any) -> tuple[list[SnapshotDef], list[ExtendedRecordDef]]:
    sheet = find_sheet(workbook, lambda name: "Snapshot&Extended" in name)
    if sheet is None:
        return [], []

    snapshots: list[SnapshotDef] = []
    current_snapshot: SnapshotDef | None = None
    ext_header = locate_header_row(sheet, [(3, "Extended Data Record Num")], max_scan=40)
    snapshot_end_row = ext_header if ext_header > 1 else sheet.max_row + 1
    for row in range(1, snapshot_end_row):
        did = parse_hex_cell(sheet.cell(row, 4).value, max_value=0xFFFF)
        if did is not None and 0x0B00 <= did <= 0x0BFF:
            record = parse_hex_cell(sheet.cell(row, 3).value, max_value=0xFF)
            _, desc = split_name(cell_text(sheet.cell(row, 5).value))
            current_snapshot = SnapshotDef(
                record_num=record,
                did=did,
                desc=desc or hex_short("SnapshotDID", did),
                size=parse_int_cell(sheet.cell(row, 6).value, default=0),
            )
            has_param = any(compact_text(sheet.cell(row, c).value) for c in (7, 8, 9, 13, 14))
            if has_param:
                param = make_param_from_cells(
                    name_value=sheet.cell(row, 9).value,
                    byte_value=sheet.cell(row, 7).value,
                    bit_value=sheet.cell(row, 8).value,
                    data_type_value=sheet.cell(row, 14).value,
                    unit_value=sheet.cell(row, 12).value,
                    conversion_value=sheet.cell(row, 13).value,
                    min_value=sheet.cell(row, 10).value,
                    max_value=sheet.cell(row, 11).value,
                    fallback_name=hex_short("SnapshotDID", did),
                )
                if param:
                    current_snapshot.params.append(param)
            snapshots.append(current_snapshot)
            continue
        if current_snapshot and not compact_text(sheet.cell(row, 4).value):
            has_param = any(compact_text(sheet.cell(row, c).value) for c in (7, 8, 9, 13, 14))
            if has_param:
                param = make_param_from_cells(
                    name_value=sheet.cell(row, 9).value,
                    byte_value=sheet.cell(row, 7).value,
                    bit_value=sheet.cell(row, 8).value,
                    data_type_value=sheet.cell(row, 14).value,
                    unit_value=sheet.cell(row, 12).value,
                    conversion_value=sheet.cell(row, 13).value,
                    min_value=sheet.cell(row, 10).value,
                    max_value=sheet.cell(row, 11).value,
                    fallback_name=f"{hex_short('SnapshotDID', current_snapshot.did)}_Data",
                )
                if param:
                    current_snapshot.params.append(param)

    snapshots = [snapshot for snapshot in snapshots if snapshot.params]

    extended_records: list[ExtendedRecordDef] = []
    current_ext: ExtendedRecordDef | None = None
    for row in range(ext_header + 1, sheet.max_row + 1):
        record = parse_int_cell(sheet.cell(row, 3).value, default=-1)
        if compact_text(sheet.cell(row, 2).value) and record >= 0 and compact_text(sheet.cell(row, 4).value):
            _, desc = split_name(cell_text(sheet.cell(row, 4).value))
            current_ext = ExtendedRecordDef(
                record_num=record,
                desc=desc or f"Extended Data Record 0x{record:02X}",
                size=parse_int_cell(sheet.cell(row, 5).value, default=0),
            )
            param = make_param_from_cells(
                name_value=sheet.cell(row, 8).value,
                byte_value=sheet.cell(row, 6).value,
                bit_value=sheet.cell(row, 7).value,
                data_type_value=sheet.cell(row, 13).value,
                unit_value=sheet.cell(row, 11).value,
                conversion_value=sheet.cell(row, 12).value,
                min_value=sheet.cell(row, 9).value,
                max_value=sheet.cell(row, 10).value,
                fallback_name=f"ExtendedRecord_{record:02X}",
            )
            if param:
                current_ext.params.append(param)
            extended_records.append(current_ext)
            continue
        if current_ext is None:
            continue
        has_param = any(compact_text(sheet.cell(row, c).value) for c in (6, 7, 8, 12, 13))
        if not has_param:
            continue
        param = make_param_from_cells(
            name_value=sheet.cell(row, 8).value,
            byte_value=sheet.cell(row, 6).value,
            bit_value=sheet.cell(row, 7).value,
            data_type_value=sheet.cell(row, 13).value,
            unit_value=sheet.cell(row, 11).value,
            conversion_value=sheet.cell(row, 12).value,
            min_value=sheet.cell(row, 9).value,
            max_value=sheet.cell(row, 10).value,
            fallback_name=f"ExtendedRecord_{current_ext.record_num:02X}_Data",
        )
        if param:
            current_ext.params.append(param)

    return snapshots, extended_records


class IdGenerator:
    def __init__(self, root: etree._Element) -> None:
        self.used = {node.get("ID") for node in root.xpath("//*[@ID]") if node.get("ID")}
        self.index = 1

    def new(self, prefix: str = "VOYAN") -> str:
        while True:
            candidate = f"_{prefix}_{self.index:06d}"
            self.index += 1
            if candidate not in self.used:
                self.used.add(candidate)
                return candidate


def element(tag: str, text: Any | None = None, attrib: dict[str, str] | None = None) -> etree._Element:
    node = etree.Element(tag, attrib=attrib or {})
    if text is not None:
        node.text = str(text)
    return node


def sub(parent: etree._Element, tag: str, text: Any | None = None, attrib: dict[str, str] | None = None) -> etree._Element:
    child = element(tag, text, attrib)
    parent.append(child)
    return child


def set_xsi_type(node: etree._Element, value: str) -> None:
    node.set(XSI_TYPE, value)


def first_by_short(root: etree._Element, tag: str, short_name: str) -> etree._Element | None:
    nodes = root.xpath(f'//*[local-name()="{tag}" and SHORT-NAME=$name]', name=short_name)
    return nodes[0] if nodes else None


def child_text(node: etree._Element, tag: str) -> str:
    value = node.findtext(tag)
    return value or ""


def clear_children(parent: etree._Element, tag: str) -> None:
    for child in list(parent):
        if child.tag == tag:
            parent.remove(child)


def replace_child(parent: etree._Element, old_tag: str, new_child: etree._Element, before_tags: set[str] | None = None) -> None:
    old = parent.find(old_tag)
    if old is not None:
        index = parent.index(old)
        parent.remove(old)
        parent.insert(index, new_child)
        return
    if before_tags:
        for index, child in enumerate(parent):
            if child.tag in before_tags:
                parent.insert(index, new_child)
                return
    parent.append(new_child)


def update_template(template_pdx: Path, output_pdx: Path, survey: SurveyData, validate: bool = True) -> None:
    with tempfile.TemporaryDirectory(prefix="voyan_pdx_") as tmp_name:
        tmp_dir = Path(tmp_name)
        with zipfile.ZipFile(template_pdx, "r") as archive:
            archive.extractall(tmp_dir)

        odx_path = tmp_dir / "VOYAN_ECU_CAN_v15.odx-d"
        if not odx_path.exists():
            candidates = list(tmp_dir.glob("*.odx-d"))
            if not candidates:
                raise FileNotFoundError("No .odx-d file found inside template PDX")
            odx_path = candidates[0]

        parser = etree.XMLParser(remove_blank_text=False)
        tree = etree.parse(str(odx_path), parser)
        root = tree.getroot()
        id_gen = IdGenerator(root)
        update_odx(root, id_gen, survey)
        validate_can_dela_odx_structure(root, odx_path.name)
        tree.write(str(odx_path), encoding="UTF-8", xml_declaration=True, pretty_print=True, standalone=False)

        fgl_path = tmp_dir / "FGL_UDS.odx-d"
        if fgl_path.exists():
            patch_fgl_communication_params(fgl_path, survey.cover)
        can_stack_path = tmp_dir / "ISO_15765_3_on_ISO_15765_2.odx-c"
        if can_stack_path.exists():
            prune_can_comparam_spec(can_stack_path)
        keep_files = set(CAN_ONLY_PDX_FILES)
        keep_files.add(odx_path.name)
        patch_pdx_catalog_for_can_only(tmp_dir / "index.xml", keep_files)

        output_pdx.parent.mkdir(parents=True, exist_ok=True)
        if output_pdx.exists():
            output_pdx.unlink()
        with zipfile.ZipFile(output_pdx, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for path in sorted(tmp_dir.rglob("*")):
                rel_path = path.relative_to(tmp_dir).as_posix()
                if path.is_file() and rel_path in keep_files:
                    archive.write(path, rel_path)

    if validate:
        validate_with_odxtools(output_pdx)


def xml_local_name(node: etree._Element) -> str:
    return etree.QName(node).localname if isinstance(node.tag, str) else ""


def child_text_by_local_name(parent: etree._Element, name: str) -> str | None:
    matches = parent.xpath("./*[local-name()=$name]", name=name)
    if not matches:
        return None
    return matches[0].text


def validate_can_dela_odx_structure(root: etree._Element, source_name: str) -> None:
    """Catch ODX structures that CANdelaStudio rejects before packaging.

    odxtools is useful as a parser-level smoke test, but CANdelaStudio also
    enforces the ODX 2.2.0 content model strictly. TABLE is the most sensitive
    object for this generator because unsupported template service variants can
    otherwise leave empty tables behind.
    """

    errors: list[str] = []
    table_order = {
        "SHORT-NAME": 0,
        "LONG-NAME": 1,
        "DESC": 2,
        "KEY-LABEL": 3,
        "STRUCT-LABEL": 4,
        "ADMIN-DATA": 5,
        "KEY-DOP-REF": 6,
        "TABLE-ROW-REF": 7,
        "TABLE-ROW": 7,
        "TABLE-DIAG-COMM-CONNECTORS": 8,
        "SDGS": 9,
    }

    seen_ids: dict[str, int] = {}
    duplicate_ids: list[tuple[str, int, int]] = []
    for node in root.xpath("//*[@ID]"):
        node_id = node.get("ID")
        if not node_id:
            continue
        if node_id in seen_ids:
            duplicate_ids.append((node_id, seen_ids[node_id], node.sourceline or 0))
        else:
            seen_ids[node_id] = node.sourceline or 0
    for node_id, first_line, second_line in duplicate_ids[:20]:
        errors.append(f"duplicate ID '{node_id}' at line {second_line}, first seen at line {first_line}")

    for ref_node in root.xpath("//*[@ID-REF]"):
        id_ref = ref_node.get("ID-REF")
        if not id_ref or ref_node.get("DOCREF"):
            continue
        if id_ref not in seen_ids:
            errors.append(
                f"dangling local ID-REF '{id_ref}' at line {ref_node.sourceline or '?'} "
                f"({xml_local_name(ref_node)})"
            )

    for table in root.xpath("//*[local-name()='TABLE']"):
        short_name = child_text_by_local_name(table, "SHORT-NAME") or table.get("ID") or "<unnamed>"
        children = [xml_local_name(child) for child in table if isinstance(child.tag, str)]
        has_row = any(child_name in {"TABLE-ROW", "TABLE-ROW-REF"} for child_name in children)
        if not has_row:
            errors.append(
                f"TABLE '{short_name}' at line {table.sourceline or '?'} has no TABLE-ROW/TABLE-ROW-REF"
            )

        if not children or children[0] != "SHORT-NAME":
            errors.append(f"TABLE '{short_name}' at line {table.sourceline or '?'} does not start with SHORT-NAME")
            continue

        last_index = -1
        for child_name in children:
            order_index = table_order.get(child_name)
            if order_index is None:
                errors.append(
                    f"TABLE '{short_name}' at line {table.sourceline or '?'} has unexpected child '{child_name}'"
                )
                continue
            if order_index < last_index:
                errors.append(
                    f"TABLE '{short_name}' at line {table.sourceline or '?'} has '{child_name}' out of ODX order"
                )
                break
            last_index = order_index

    if errors:
        details = "\n".join(f"- {message}" for message in errors[:60])
        if len(errors) > 60:
            details += f"\n- ... {len(errors) - 60} more issue(s)"
        raise RuntimeError(f"Generated {source_name} is not CANdelaStudio-compatible:\n{details}")


def patch_pdx_catalog_for_can_only(index_path: Path, keep_files: set[str]) -> None:
    if not index_path.exists():
        return
    parser = etree.XMLParser(remove_blank_text=False)
    tree = etree.parse(str(index_path), parser)
    root = tree.getroot()
    for file_node in list(root.xpath("//*[local-name()='FILE']")):
        file_name = compact_text(file_node.text)
        if file_name not in keep_files:
            parent = file_node.getparent()
            if parent is not None:
                parent.remove(file_node)
    tree.write(str(index_path), encoding="UTF-8", xml_declaration=False, pretty_print=True, standalone=False)


def update_odx(root: etree._Element, id_gen: IdGenerator, survey: SurveyData) -> None:
    base_variant = first_by_short(root, "BASE-VARIANT", "VOYAN_ECU_CAN")
    container = root.find(".//DIAG-LAYER-CONTAINER")
    if container is not None:
        long_name = container.find("LONG-NAME")
        if long_name is not None:
            long_name.text = survey.cover.ecu_name
    if base_variant is not None:
        long_name = base_variant.find("LONG-NAME")
        if long_name is not None:
            long_name.text = survey.cover.ecu_name

    ddds = root.find(".//DIAG-DATA-DICTIONARY-SPEC")
    if ddds is None:
        raise RuntimeError("Template ODX has no DIAG-DATA-DICTIONARY-SPEC")
    data_object_props = ddds.find("DATA-OBJECT-PROPS")
    structures = ddds.find("STRUCTURES")
    tables = ddds.find("TABLES")
    if data_object_props is None or structures is None or tables is None:
        raise RuntimeError("Template ODX is missing DOP/STRUCTURE/TABLE containers")

    unit_ids = ensure_units(ddds, survey)
    generated_dop_cache: dict[tuple[str, int, str, str, str], str] = {}

    for did in survey.dids:
        prepare_data_structure(
            id_gen=id_gen,
            data_object_props=data_object_props,
            structures=structures,
            item=did,
            prefix="DID",
            unit_ids=unit_ids,
            generated_dop_cache=generated_dop_cache,
        )
        did.wrapper_id = make_wrapper_structure(
            id_gen,
            structures,
            f"TR_Identification_{did.short_name}",
            f"TR Identification {did.long_name}",
            did.short_name,
            did.long_name,
            did.structure_id,
        )

    for io_did in survey.io_dids:
        prepare_data_structure(
            id_gen=id_gen,
            data_object_props=data_object_props,
            structures=structures,
            item=io_did,
            prefix="IODID",
            unit_ids=unit_ids,
            generated_dop_cache=generated_dop_cache,
        )
        io_did.status_wrapper_id = make_wrapper_structure(
            id_gen,
            structures,
            f"TR_ControlStatusRecord_{io_did.short_name}",
            f"TR ControlStatusRecord {io_did.long_name}",
            io_did.short_name,
            io_did.long_name,
            io_did.structure_id,
        )
        io_did.control_request_wrapper_id = make_wrapper_structure(
            id_gen,
            structures,
            f"TR_IOControl_Control_RQ_{io_did.short_name}",
            f"TR IOControl Control RQ {io_did.long_name}",
            "ControlOptionRecord",
            "ControlOptionRecord",
            io_did.status_wrapper_id,
        )

    for routine in survey.routines:
        english, long_name = split_name(routine.desc)
        routine.short_name = sanitize_short_name(english or hex_short("RID", routine.rid), hex_short("RID", routine.rid))
        routine.long_name = long_name or routine.short_name
        for subfn in routine.subfunctions.values():
            if subfn.option_params:
                subfn.option_structure_id = make_param_structure(
                    id_gen,
                    data_object_props,
                    structures,
                    f"STR_RoutineOption_{routine.short_name}_{subfn.control_type:02X}",
                    f"Routine Option {routine.long_name}",
                    subfn.option_params,
                    unit_ids,
                    generated_dop_cache,
                )
            if subfn.status_params:
                subfn.status_structure_id = make_param_structure(
                    id_gen,
                    data_object_props,
                    structures,
                    f"STR_RoutineStatus_{routine.short_name}_{subfn.control_type:02X}",
                    f"Routine Status {routine.long_name}",
                    subfn.status_params,
                    unit_ids,
                    generated_dop_cache,
                )

    for snapshot in survey.snapshots:
        if snapshot.params:
            snapshot.structure_id = make_param_structure(
                id_gen,
                data_object_props,
                structures,
                f"STR_Snapshot_{hex_short('DID', snapshot.did)}",
                snapshot.desc,
                snapshot.params,
                unit_ids,
                generated_dop_cache,
                byte_size=snapshot.size or None,
            )

    for record in survey.extended_records:
        if record.params:
            record.structure_id = make_param_structure(
                id_gen,
                data_object_props,
                structures,
                f"STR_ExtendedRecord_{record.record_num:02X}",
                record.desc,
                record.params,
                unit_ids,
                generated_dop_cache,
                byte_size=record.size or None,
            )

    update_did_tables_and_services(root, id_gen, survey.dids)
    update_io_tables_and_services(root, id_gen, survey.io_dids)
    update_routine_tables_and_services(root, id_gen, survey.routines)
    update_dtc_dop(root, id_gen, survey.dtcs)
    update_snapshot_and_extended_data(root, id_gen, survey.snapshots, survey.extended_records)
    normalize_extended_data_service_for_candela(root, id_gen)
    validate_edr_mapping(root, survey.dtcs, survey.extended_records)
    remove_service_and_messages(root, "Software_Update_RequestDownload")
    update_flat_service_preconditions(root)
    remove_service_and_messages(root, "Combined_Identification_Read")
    update_base_variant_comparams(root, survey.cover)
    ensure_base_variant_protocol_parent(root)
    update_session_timing(root)
    prefix_doc_revision_labels(root)


def prefix_doc_revision_labels(root: etree._Element) -> None:
    for label in root.xpath("//*[local-name()='DOC-REVISION']/*[local-name()='REVISION-LABEL']"):
        value = compact_text(label.text)
        if value and not value.startswith("PDX_"):
            label.text = f"PDX_{value}"


def ensure_units(ddds: etree._Element, survey: SurveyData) -> dict[str, str]:
    unit_spec = ddds.find("UNIT-SPEC")
    if unit_spec is None:
        unit_spec = sub(ddds, "UNIT-SPEC")
    units_node = unit_spec.find("UNITS")
    if units_node is None:
        units_node = sub(unit_spec, "UNITS")

    existing = {
        compact_text(unit.findtext("DISPLAY-NAME") or unit.findtext("SHORT-NAME")): unit.get("ID")
        for unit in units_node.findall("UNIT")
        if unit.get("ID")
    }
    all_params: list[ParamDef] = []
    for item in [*survey.dids, *survey.io_dids, *survey.snapshots, *survey.extended_records]:
        all_params.extend(item.params)
    for routine in survey.routines:
        for subfn in routine.subfunctions.values():
            all_params.extend(subfn.option_params)
            all_params.extend(subfn.status_params)

    for param in all_params:
        unit = compact_text(param.unit)
        if not unit or unit in {"-", "/", "N/A", "NA"} or unit in existing:
            continue
        unit_id = f"_UNIT_{sanitize_short_name(unit, 'UNIT')}"
        index = 2
        while unit_id in existing.values():
            unit_id = f"_UNIT_{sanitize_short_name(unit, 'UNIT')}_{index}"
            index += 1
        unit_node = sub(units_node, "UNIT", attrib={"ID": unit_id})
        sub(unit_node, "SHORT-NAME", sanitize_short_name(unit, "UNIT"))
        sub(unit_node, "DISPLAY-NAME", unit)
        existing[unit] = unit_id
    return {k: v for k, v in existing.items() if v}


def prepare_data_structure(
    *,
    id_gen: IdGenerator,
    data_object_props: etree._Element,
    structures: etree._Element,
    item: DidDef | IoDidDef,
    prefix: str,
    unit_ids: dict[str, str],
    generated_dop_cache: dict[tuple[str, int, str, str, str], str],
) -> None:
    used_param_names: set[str] = set()
    english, long_name = split_name(item.desc)
    item.short_name = sanitize_short_name(english or hex_short(prefix, item.did), hex_short(prefix, item.did))
    item.long_name = long_name or item.short_name
    for param in item.params:
        param.name = sanitize_short_name(param.name, "Data", used_param_names)
        param.dop_id = ensure_param_dop(
            id_gen,
            data_object_props,
            param,
            unit_ids,
            generated_dop_cache,
            f"DOP_{item.short_name}_{param.name}",
        )
    if not item.params:
        raw = ParamDef(
            name="Data",
            long_name="Data",
            byte_pos=0,
            bit_pos=0,
            bit_len=max(8, (item.size or 1) * 8),
            data_type="Hex(Unsigned)",
        )
        raw.name = sanitize_short_name(raw.name, "Data", used_param_names)
        raw.dop_id = ensure_param_dop(
            id_gen, data_object_props, raw, unit_ids, generated_dop_cache, f"DOP_{item.short_name}_Data"
        )
        item.params.append(raw)

    item.structure_id = id_gen.new("DIDSTR")
    structure = sub(structures, "STRUCTURE", attrib={"ID": item.structure_id})
    sub(structure, "SHORT-NAME", item.short_name)
    sub(structure, "LONG-NAME", item.long_name)
    if item.size:
        sub(structure, "BYTE-SIZE", item.size)
    params_node = sub(structure, "PARAMS")
    for param in item.params:
        params_node.append(make_value_param(param.name, param.long_name, param.byte_pos, param.bit_pos, param.dop_id))


def make_param_structure(
    id_gen: IdGenerator,
    data_object_props: etree._Element,
    structures: etree._Element,
    short_name: str,
    long_name: str,
    params: list[ParamDef],
    unit_ids: dict[str, str],
    generated_dop_cache: dict[tuple[str, int, str, str, str], str],
    byte_size: int | None = None,
) -> str:
    used_names: set[str] = set()
    structure_id = id_gen.new("STR")
    structure = sub(structures, "STRUCTURE", attrib={"ID": structure_id})
    short_name = sanitize_short_name(short_name, "Structure")
    sub(structure, "SHORT-NAME", short_name)
    sub(structure, "LONG-NAME", long_name or short_name)
    if byte_size:
        sub(structure, "BYTE-SIZE", byte_size)
    params_node = sub(structure, "PARAMS")
    for param in params:
        param.name = sanitize_short_name(param.name, "Data", used_names)
        param.dop_id = ensure_param_dop(
            id_gen, data_object_props, param, unit_ids, generated_dop_cache, f"DOP_{short_name}_{param.name}"
        )
        params_node.append(make_value_param(param.name, param.long_name, param.byte_pos, param.bit_pos, param.dop_id))
    return structure_id


def make_wrapper_structure(
    id_gen: IdGenerator,
    structures: etree._Element,
    short_name: str,
    long_name: str,
    param_short_name: str,
    param_long_name: str,
    dop_or_structure_id: str,
) -> str:
    wrapper_id = id_gen.new("WRAP")
    wrapper = sub(structures, "STRUCTURE", attrib={"ID": wrapper_id})
    sub(wrapper, "SHORT-NAME", sanitize_short_name(short_name, "Wrapper"))
    sub(wrapper, "LONG-NAME", long_name)
    params_node = sub(wrapper, "PARAMS")
    param = make_value_param(sanitize_short_name(param_short_name, "Data"), param_long_name, 0, 0, dop_or_structure_id)
    params_node.append(param)
    return wrapper_id


def ensure_param_dop(
    id_gen: IdGenerator,
    data_object_props: etree._Element,
    param: ParamDef,
    unit_ids: dict[str, str],
    cache: dict[tuple[str, int, str, str, str], str],
    preferred_short_name: str,
) -> str:
    conv_key = conversion_cache_key(param.conversion)
    key = (param.data_type.upper(), param.bit_len, conv_key, compact_text(param.unit), param.name)
    if key in cache:
        return cache[key]
    dop_id = id_gen.new("DOP")
    dop = make_data_object_prop(dop_id, preferred_short_name, param, unit_ids)
    data_object_props.append(dop)
    cache[key] = dop_id
    return dop_id


def conversion_cache_key(conversion: Conversion) -> str:
    if conversion.kind == "enum":
        return "enum:" + "|".join(f"{lo}-{hi}:{label}" for lo, hi, label in conversion.enum)
    if conversion.kind == "linear":
        return f"linear:{conversion.a}:{conversion.b}:{conversion.precision}"
    return "identity"


def make_data_object_prop(dop_id: str, preferred_short_name: str, param: ParamDef, unit_ids: dict[str, str]) -> etree._Element:
    bit_len = max(1, int(param.bit_len or 8))
    dop = element("DATA-OBJECT-PROP", attrib={"ID": dop_id})
    short_name = sanitize_short_name(preferred_short_name, "DOP")
    sub(dop, "SHORT-NAME", short_name)
    sub(dop, "LONG-NAME", param.long_name or short_name)

    conversion = param.conversion
    if conversion.kind == "enum" and conversion.enum:
        compu = sub(dop, "COMPU-METHOD")
        sub(compu, "CATEGORY", "TEXTTABLE")
        internal_to_phys = sub(compu, "COMPU-INTERNAL-TO-PHYS")
        scales = sub(internal_to_phys, "COMPU-SCALES")
        for lo, hi, label in conversion.enum:
            scale = sub(scales, "COMPU-SCALE")
            sub(scale, "LOWER-LIMIT", lo)
            sub(scale, "UPPER-LIMIT", hi)
            const = sub(scale, "COMPU-CONST")
            sub(const, "VT", label)
        coded = sub(
            dop,
            "DIAG-CODED-TYPE",
            attrib={"BASE-TYPE-ENCODING": "NONE", "BASE-DATA-TYPE": "A_UINT32"},
        )
        set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
        sub(coded, "BIT-LENGTH", bit_len)
        sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UNICODE2STRING"})
    elif conversion.kind == "linear":
        compu = sub(dop, "COMPU-METHOD")
        sub(compu, "CATEGORY", "LINEAR")
        internal_to_phys = sub(compu, "COMPU-INTERNAL-TO-PHYS")
        scales = sub(internal_to_phys, "COMPU-SCALES")
        scale = sub(scales, "COMPU-SCALE")
        coeffs = sub(scale, "COMPU-RATIONAL-COEFFS")
        numerator = sub(coeffs, "COMPU-NUMERATOR")
        sub(numerator, "V", clean_float(conversion.b))
        sub(numerator, "V", clean_float(conversion.a))
        coded = sub(
            dop,
            "DIAG-CODED-TYPE",
            attrib={"BASE-TYPE-ENCODING": "NONE", "BASE-DATA-TYPE": "A_UINT32"},
        )
        set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
        sub(coded, "BIT-LENGTH", min(bit_len, 32))
        physical = sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_FLOAT64"})
        if conversion.precision is not None:
            sub(physical, "PRECISION", conversion.precision)
    else:
        compu = sub(dop, "COMPU-METHOD")
        sub(compu, "CATEGORY", "IDENTICAL")
        if "ASCII" in param.data_type.upper():
            coded = sub(
                dop,
                "DIAG-CODED-TYPE",
                attrib={"BASE-TYPE-ENCODING": "ISO-8859-1", "BASE-DATA-TYPE": "A_ASCIISTRING"},
            )
            set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
            sub(coded, "BIT-LENGTH", bit_len)
            sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UNICODE2STRING"})
        elif bit_len <= 32:
            coded = sub(
                dop,
                "DIAG-CODED-TYPE",
                attrib={"BASE-TYPE-ENCODING": "NONE", "BASE-DATA-TYPE": "A_UINT32"},
            )
            set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
            sub(coded, "BIT-LENGTH", bit_len)
            sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32", "DISPLAY-RADIX": "HEX"})
        else:
            coded = sub(dop, "DIAG-CODED-TYPE", attrib={"BASE-DATA-TYPE": "A_BYTEFIELD"})
            set_xsi_type(coded, "STANDARD-LENGTH-TYPE")
            sub(coded, "BIT-LENGTH", bit_len)
            sub(dop, "PHYSICAL-TYPE", attrib={"BASE-DATA-TYPE": "A_BYTEFIELD"})

    unit = compact_text(param.unit)
    if unit and unit in unit_ids and conversion.kind != "enum":
        sub(dop, "UNIT-REF", attrib={"ID-REF": unit_ids[unit]})
    return dop


def clean_float(value: float) -> str:
    return str(int(value)) if float(value).is_integer() else str(value)


def make_value_param(short_name: str, long_name: str, byte_position: int, bit_position: int, dop_id: str) -> etree._Element:
    param = element("PARAM", attrib={"SEMANTIC": "DATA"})
    set_xsi_type(param, "VALUE")
    sub(param, "SHORT-NAME", short_name)
    sub(param, "LONG-NAME", long_name or short_name)
    sub(param, "BYTE-POSITION", byte_position)
    if bit_position:
        sub(param, "BIT-POSITION", bit_position)
    sub(param, "DOP-REF", attrib={"ID-REF": dop_id})
    return param


def update_did_tables_and_services(root: etree._Element, id_gen: IdGenerator, dids: list[DidDef]) -> None:
    read_table = first_by_short(root, "TABLE", "Identification_Read_PR")
    write_table = first_by_short(root, "TABLE", "Identification_Write_RQ")
    if read_table is None or write_table is None:
        raise RuntimeError("Template identification tables were not found")
    clear_children(read_table, "TABLE-ROW")
    clear_children(write_table, "TABLE-ROW")

    readable_instances: list[tuple[int, str, str]] = []
    writable_instances: list[tuple[int, str, str]] = []
    for did in dids:
        if did.readable:
            append_table_row(read_table, id_gen, did.short_name, did.long_name, did.did, did.wrapper_id)
            readable_instances.append((did.did, did.short_name, did.long_name))
        if did.writable:
            append_table_row(write_table, id_gen, did.short_name, did.long_name, did.did, did.wrapper_id)
            writable_instances.append((did.did, did.short_name, did.long_name))

    update_service_instances(root, id_gen, "Identification_Read", "Read", "Read", readable_instances)
    update_service_instances(root, id_gen, "Identification_Write", "Write", "Write", writable_instances)
    update_preconditions_for_values(root, "Identification_Read", readable_instances, "RecordDataIdentifier", did_state_refs(dids, "read"))
    update_preconditions_for_values(root, "Identification_Write", writable_instances, "RecordDataIdentifier", did_state_refs(dids, "write"))

    eop_read_rq = first_by_short(root, "END-OF-PDU-FIELD", "EOP_DIDs_Identification_ReadRDBI_RQ")
    eop_read_pr = first_by_short(root, "END-OF-PDU-FIELD", "EOP_DIDs_Identification_ReadRDBI_PR")
    for eop in (eop_read_rq, eop_read_pr):
        if eop is not None:
            max_items = eop.find("MAX-NUMBER-OF-ITEMS")
            if max_items is not None:
                max_items.text = str(max(1, len(readable_instances)))


def did_state_refs(dids: list[DidDef], mode: str) -> dict[int, list[str]]:
    result: dict[int, list[str]] = {}
    for did in dids:
        if mode == "read":
            if not did.readable:
                continue
            states = sessions_to_state_ids(did.sessions, mode="read")
            states.extend(["_6", "_7", "_8", "_10", "_11", "_12", "_13", "_14", "_15"])
        else:
            if not did.writable:
                continue
            states = sessions_to_state_ids(did.sessions, mode="write")
            states.extend(security_to_state_ids(did.write_security))
            states.extend(["_10", "_11", "_12", "_13", "_14", "_15"])
        result[did.did] = unique_list(states)
    return result


def sessions_to_state_ids(sessions: list[str], mode: str) -> list[str]:
    wants = "R" if mode == "read" else "W"
    mapping = ["_2", "_4", "_2", "_3", "_4"]
    states: list[str] = []
    for idx, value in enumerate(sessions[:5]):
        access = normalize_access(value)
        if wants in access:
            states.append(mapping[idx])
    if not states:
        states = ["_2", "_3", "_4"]
    return unique_list(states)


def security_to_state_ids(level: str) -> list[str]:
    normalized = compact_text(level).upper()
    if "FBL" in normalized:
        return ["_8"]
    if "LEVEL1" in normalized or normalized in {"1", "L1"}:
        return ["_7"]
    return ["_6", "_7", "_8"]


def unique_list(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def append_table_row(
    table: etree._Element,
    id_gen: IdGenerator,
    short_name: str,
    long_name: str,
    key: int | str,
    structure_ref: str | None = None,
) -> etree._Element:
    row = sub(table, "TABLE-ROW", attrib={"ID": id_gen.new("ROW")})
    sub(row, "SHORT-NAME", sanitize_short_name(short_name, "Row"))
    sub(row, "LONG-NAME", long_name or short_name)
    sub(row, "KEY", key)
    if structure_ref:
        sub(row, "STRUCTURE-REF", attrib={"ID-REF": structure_ref})
    return row


def update_service_instances(
    root: etree._Element,
    id_gen: IdGenerator,
    service_short_name: str,
    qualifier: str,
    service_name: str,
    instances: list[tuple[int, str, str]],
) -> None:
    service = first_by_short(root, "DIAG-SERVICE", service_short_name)
    if service is None:
        return
    sdgs = element("SDGS")
    sdg = sub(sdgs, "SDG")
    caption = sub(sdg, "SDG-CAPTION", attrib={"ID": id_gen.new("CAP")})
    sub(caption, "SHORT-NAME", "CANdelaServiceInformation")
    sub(sdg, "SD", qualifier, attrib={"SI": "ServiceQualifier"})
    sub(sdg, "SD", service_name, attrib={"SI": "ServiceName"})
    sub(sdg, "SD", "no", attrib={"SI": "PositiveResponseSuppressed"})
    for value, inst_qualifier, inst_name in instances:
        inst = sub(sdg, "SDG")
        sub(inst, "SD", value, attrib={"SI": "DiagInstanceStaticValue"})
        sub(inst, "SD", sanitize_short_name(inst_qualifier, "Instance"), attrib={"SI": "DiagInstanceQualifier"})
        sub(inst, "SD", inst_name or inst_qualifier, attrib={"SI": "DiagInstanceName"})
    replace_child(service, "SDGS", sdgs, before_tags={"FUNCT-CLASS-REFS", "AUDIENCE", "PRE-CONDITION-STATE-REFS", "REQUEST-REF"})


def update_preconditions_for_values(
    root: etree._Element,
    service_short_name: str,
    instances: list[tuple[int, str, str]],
    in_param: str,
    state_refs_by_value: dict[int, list[str]],
) -> None:
    service = first_by_short(root, "DIAG-SERVICE", service_short_name)
    if service is None:
        return
    pc = element("PRE-CONDITION-STATE-REFS")
    for value, _, _ in instances:
        for state_id in state_refs_by_value.get(value, ["_2", "_3", "_4"]):
            ref = sub(pc, "PRE-CONDITION-STATE-REF", attrib={"ID-REF": state_id})
            sub(ref, "VALUE", value)
            sub(ref, "IN-PARAM-IF-SNREF", attrib={"SHORT-NAME": in_param})
    replace_child(service, "PRE-CONDITION-STATE-REFS", pc, before_tags={"REQUEST-REF"})


def update_flat_preconditions(root: etree._Element, service_short_name: str, state_refs: Iterable[str]) -> None:
    service = first_by_short(root, "DIAG-SERVICE", service_short_name)
    if service is None:
        return
    pc = element("PRE-CONDITION-STATE-REFS")
    for state_id in unique_list(state_refs):
        sub(pc, "PRE-CONDITION-STATE-REF", attrib={"ID-REF": state_id})
    replace_child(service, "PRE-CONDITION-STATE-REFS", pc, before_tags={"REQUEST-REF"})


def update_flat_service_preconditions(root: etree._Element) -> None:
    all_sessions = ["_2", "_3", "_4"]
    all_security = ["_6", "_7", "_8"]
    all_authorization = ["_10", "_11", "_12", "_13", "_14", "_15"]
    default_open = [*all_sessions, *all_security, *all_authorization]
    extended_level1 = ["_4", "_7", *all_authorization]
    programming_fbl = ["_3", "_8", *all_authorization]

    for service_short_name in (
        "DefaultSession_Start",
        "ProgrammingSession_Start",
        "ExtendedDiagnosticSession_Start",
        "TesterPresent_Send",
    ):
        update_flat_preconditions(root, service_short_name, default_open)

    for service_short_name in ("Hard_Reset_Reset", "Soft_Reset_Reset"):
        update_flat_preconditions(root, service_short_name, [*all_sessions, *all_authorization])

    for service_short_name in (
        "SeedLevel1_Request",
        "KeyLevel1_Send",
        "EnableRxAndEnableTx_Control",
        "DisableRxAndDisableTx_Control",
        "ControlDTCSetting_On",
        "ControlDTCSetting_Off",
        "FaultMemory_Clear",
    ):
        update_flat_preconditions(root, service_short_name, extended_level1)

    for service_short_name in ("RequstSeedOfSecurityLevelFBL_Request", "SendKeyOfSecurityLevelFBL_Send"):
        update_flat_preconditions(root, service_short_name, programming_fbl)

    for service_short_name in (
        "FaultMemory_ReadNumber",
        "FaultMemory_ReadAllIdentified",
        "FaultMemory_Read_extended_data_record",
        "FaultMemory_Read_snapshot_record",
        "FaultMemory_ReadAllSupported",
    ):
        update_flat_preconditions(root, service_short_name, default_open)

    for service_short_name in (
        "Software_Update_Transmit",
        "Software_Update_Stop",
    ):
        update_flat_preconditions(root, service_short_name, programming_fbl)


def remove_service_and_messages(root: etree._Element, service_short_name: str) -> None:
    service = first_by_short(root, "DIAG-SERVICE", service_short_name)
    if service is None:
        return
    ids_to_remove: set[str] = set()
    request_ref = service.find("REQUEST-REF")
    if request_ref is not None and request_ref.get("ID-REF"):
        ids_to_remove.add(request_ref.get("ID-REF"))
    ids_to_remove.update(
        ref.get("ID-REF")
        for ref in service.xpath('./*[local-name()="POS-RESPONSE-REFS"]/*[local-name()="POS-RESPONSE-REF"]')
        if ref.get("ID-REF")
    )
    ids_to_remove.update(
        ref.get("ID-REF")
        for ref in service.xpath('./*[local-name()="NEG-RESPONSE-REFS"]/*[local-name()="NEG-RESPONSE-REF"]')
        if ref.get("ID-REF")
    )

    parent = service.getparent()
    if parent is not None:
        parent.remove(service)
    for element_id in ids_to_remove:
        for node in root.xpath(f'//*[@ID="{element_id}"]'):
            node_parent = node.getparent()
            if node_parent is not None:
                node_parent.remove(node)


def remove_elements_by_short_names(root: etree._Element, tag: str, short_names: set[str]) -> None:
    for short_name in short_names:
        for node in root.xpath(f'//*[local-name()="{tag}" and SHORT-NAME=$name]', name=short_name):
            parent = node.getparent()
            if parent is not None:
                parent.remove(node)


def update_io_tables_and_services(root: etree._Element, id_gen: IdGenerator, io_dids: list[IoDidDef]) -> None:
    table_names = {
        0: ("IOControl_ReturnControl_RQ", "IOControl_ReturnControl_PR", "IOControl_ReturnControl", "ReturnControl"),
        1: ("IOControl_Reset_RQ", "IOControl_Reset_PR", "IOControl_Reset", "Reset"),
        2: ("IOControl_Freeze_RQ", "IOControl_Freeze_PR", "IOControl_Freeze", "Freeze"),
        3: ("IOControl_Control_RQ", "IOControl_Control_PR", "IOControl_Control", "Control"),
    }
    # The IO DID survey section describes InputOutputControlByIdentifier (0x2F)
    # sub-functions only. Do not emit a synthetic Read service unless a future
    # schema explicitly carries read support for IO DIDs.
    remove_service_and_messages(root, "Combined_IOControl_Read")
    remove_service_and_messages(root, "IOControl_Read")
    remove_elements_by_short_names(
        root,
        "STRUCTURE",
        {"STR_EOP_DIDs_IOControl_ReadRDBI_RQ", "STR_EOP_DIDs_IOControl_ReadRDBI_PR"},
    )
    remove_elements_by_short_names(
        root,
        "END-OF-PDU-FIELD",
        {"EOP_DIDs_IOControl_ReadRDBI_RQ", "EOP_DIDs_IOControl_ReadRDBI_PR"},
    )
    remove_elements_by_short_names(root, "DATA-OBJECT-PROP", {"DOP_NRC_IOControl_Read"})
    remove_elements_by_short_names(root, "TABLE", {"IOControl_Read_PR"})

    instances_by_control: dict[int, list[tuple[int, str, str]]] = {k: [] for k in table_names}
    for control, (rq_table_name, pr_table_name, _, _) in table_names.items():
        rq_table = first_by_short(root, "TABLE", rq_table_name)
        pr_table = first_by_short(root, "TABLE", pr_table_name)
        if rq_table is not None:
            clear_children(rq_table, "TABLE-ROW")
        if pr_table is not None:
            clear_children(pr_table, "TABLE-ROW")
        for io_did in io_dids:
            if control not in io_did.controls:
                continue
            rq_structure = io_did.control_request_wrapper_id if control == 3 else None
            pr_structure = io_did.status_wrapper_id
            if rq_table is not None:
                append_table_row(rq_table, id_gen, io_did.short_name, io_did.long_name, io_did.did, rq_structure)
            if pr_table is not None:
                append_table_row(pr_table, id_gen, io_did.short_name, io_did.long_name, io_did.did, pr_structure)
            instances_by_control[control].append((io_did.did, io_did.short_name, io_did.long_name))

    for control, (_, _, service_short_name, service_label) in table_names.items():
        rq_table_name, pr_table_name, _, _ = table_names[control]
        if not instances_by_control[control]:
            remove_service_and_messages(root, service_short_name)
            remove_elements_by_short_names(root, "TABLE", {rq_table_name, pr_table_name})
            continue
        update_service_instances(root, id_gen, service_short_name, service_label, service_label, instances_by_control[control])
        state_refs = {value: ["_4", "_7", "_10", "_11", "_12", "_13", "_14", "_15"] for value, _, _ in instances_by_control[control]}
        update_preconditions_for_values(root, service_short_name, instances_by_control[control], "DataIdentifier", state_refs)


def update_routine_tables_and_services(root: etree._Element, id_gen: IdGenerator, routines: list[RoutineDef]) -> None:
    table_pairs = {
        1: ("TAB_StartOptions", "TAB_StartStatus", "Routine_Control_Start", "Start"),
        2: ("TAB_StopOptions", "TAB_StopStatus", "Routine_Control_Stop", "Stop"),
        3: ("TAB_RequestResultsOptions", "TAB_RequestResultsStatus", "Routine_Control_RequestResults", "RequestResults"),
    }
    ddds = root.find(".//DIAG-DATA-DICTIONARY-SPEC")
    tables_node = ddds.find("TABLES") if ddds is not None else None
    if tables_node is None:
        return

    key_dop_id = "_166"
    for control_type, (option_table_name, status_table_name, service_short_name, service_label) in table_pairs.items():
        option_table = ensure_table(root, tables_node, id_gen, option_table_name, f"Table Control Options ({service_label})", key_dop_id)
        status_table = ensure_table(root, tables_node, id_gen, status_table_name, f"Table Status ({service_label})", key_dop_id)
        clear_children(option_table, "TABLE-ROW")
        clear_children(status_table, "TABLE-ROW")
        instances: list[tuple[int, str, str]] = []
        state_refs: dict[int, list[str]] = {}
        for routine in routines:
            subfn = routine.subfunctions.get(control_type)
            if subfn is None or not subfn.supported:
                continue
            append_table_row(option_table, id_gen, routine.short_name, routine.long_name, routine.rid, subfn.option_structure_id or None)
            append_table_row(status_table, id_gen, routine.short_name, routine.long_name, routine.rid, subfn.status_structure_id or None)
            instances.append((routine.rid, routine.short_name, routine.long_name))
            states = sessions_to_state_ids([""] + routine.sessions, mode="write")
            states.extend(security_to_state_ids(routine.security))
            states.extend(["_10", "_11", "_12", "_13", "_14", "_15"])
            state_refs[routine.rid] = unique_list(states)

        ensure_routine_service(root, id_gen, service_short_name, service_label, control_type, option_table, status_table)
        update_service_instances(root, id_gen, service_short_name, service_label, service_label, instances)
        update_preconditions_for_values(root, service_short_name, instances, "RoutineIdentifier", state_refs)


def ensure_table(
    root: etree._Element,
    tables_node: etree._Element,
    id_gen: IdGenerator,
    short_name: str,
    long_name: str,
    key_dop_id: str,
) -> etree._Element:
    table = first_by_short(root, "TABLE", short_name)
    if table is not None:
        return table
    table = sub(tables_node, "TABLE", attrib={"ID": id_gen.new("TAB")})
    sub(table, "SHORT-NAME", short_name)
    sub(table, "LONG-NAME", long_name)
    sub(table, "KEY-DOP-REF", attrib={"ID-REF": key_dop_id})
    return table


def ensure_routine_service(
    root: etree._Element,
    id_gen: IdGenerator,
    service_short_name: str,
    service_label: str,
    control_type: int,
    option_table: etree._Element,
    status_table: etree._Element,
) -> None:
    if first_by_short(root, "DIAG-SERVICE", service_short_name) is not None:
        update_routine_request_response_tables(root, service_short_name, option_table, status_table)
        return

    base_service = first_by_short(root, "DIAG-SERVICE", "Routine_Control_Start")
    base_request = first_by_short(root, "REQUEST", "RQ_Routine_Control_Start")
    base_response = first_by_short(root, "POS-RESPONSE", "PR_Routine_Control_Start")
    base_negative = first_by_short(root, "NEG-RESPONSE", "NR_Routine_Control_Start")
    diag_comms = root.find(".//DIAG-COMMS")
    requests = root.find(".//REQUESTS")
    responses = root.find(".//POS-RESPONSES")
    negatives = root.find(".//NEG-RESPONSES")
    required_nodes = [base_service, base_request, base_response, base_negative, diag_comms, requests, responses, negatives]
    if any(node is None for node in required_nodes):
        return

    rq_id = id_gen.new("RQ")
    pr_id = id_gen.new("PR")
    nr_id = id_gen.new("NR")
    service_id = id_gen.new("SVC")

    request = copy.deepcopy(base_request)
    request.set("ID", rq_id)
    set_short_long(request, f"RQ_{service_short_name}", f"RQ {service_short_name.replace('_', ' ')}")
    set_routine_message_control_type(request, control_type)
    set_table_ref_in_message(request, option_table.get("ID"))
    refresh_internal_ids(request, id_gen)
    requests.append(request)

    response = copy.deepcopy(base_response)
    response.set("ID", pr_id)
    set_short_long(response, f"PR_{service_short_name}", f"PR {service_short_name.replace('_', ' ')}")
    set_routine_message_control_type(response, control_type)
    set_table_ref_in_message(response, status_table.get("ID"))
    refresh_internal_ids(response, id_gen)
    responses.append(response)

    negative = copy.deepcopy(base_negative)
    negative.set("ID", nr_id)
    set_short_long(negative, f"NR_{service_short_name}", f"NR {service_short_name.replace('_', ' ')}")
    refresh_internal_ids(negative, id_gen)
    negatives.append(negative)

    service = copy.deepcopy(base_service)
    service.set("ID", service_id)
    set_short_long(service, service_short_name, service_short_name.replace("_", " "))
    refresh_internal_ids(service, id_gen)
    request_ref = service.find("REQUEST-REF")
    if request_ref is not None:
        request_ref.set("ID-REF", rq_id)
    pos_ref = service.find(".//POS-RESPONSE-REF")
    if pos_ref is not None:
        pos_ref.set("ID-REF", pr_id)
    neg_ref = service.find(".//NEG-RESPONSE-REF")
    if neg_ref is not None:
        neg_ref.set("ID-REF", nr_id)
    diag_comms.append(service)
    update_routine_request_response_tables(root, service_short_name, option_table, status_table)


def set_short_long(node: etree._Element, short_name: str, long_name: str) -> None:
    short = node.find("SHORT-NAME")
    if short is not None:
        short.text = sanitize_short_name(short_name, "Node")
    long = node.find("LONG-NAME")
    if long is not None:
        long.text = long_name


def refresh_internal_ids(node: etree._Element, id_gen: IdGenerator) -> None:
    id_map: dict[str, str] = {}
    for child in node.xpath(".//*[@ID]"):
        old_id = child.get("ID")
        if not old_id:
            continue
        new_id = id_gen.new("IN")
        id_map[old_id] = new_id
        child.set("ID", new_id)
    if not id_map:
        return
    for ref_node in node.xpath(".//*[@ID-REF]"):
        old_ref = ref_node.get("ID-REF")
        if old_ref in id_map and not ref_node.get("DOCREF"):
            ref_node.set("ID-REF", id_map[old_ref])


def set_routine_message_control_type(message: etree._Element, control_type: int) -> None:
    for param in message.xpath('.//*[local-name()="PARAM" and SHORT-NAME="RoutineControlType"]'):
        coded = param.find("CODED-VALUE")
        if coded is not None:
            coded.text = str(control_type)


def set_table_ref_in_message(message: etree._Element, table_id: str | None) -> None:
    if not table_id:
        return
    table_ref = message.find(".//TABLE-REF")
    if table_ref is not None:
        table_ref.set("ID-REF", table_id)


def update_routine_request_response_tables(
    root: etree._Element,
    service_short_name: str,
    option_table: etree._Element,
    status_table: etree._Element,
) -> None:
    service = first_by_short(root, "DIAG-SERVICE", service_short_name)
    if service is None:
        return
    request_ref = service.find("REQUEST-REF")
    if request_ref is not None:
        request = root.xpath(f'//*[@ID="{request_ref.get("ID-REF")}"]')
        if request:
            set_table_ref_in_message(request[0], option_table.get("ID"))
    pos_ref = service.find(".//POS-RESPONSE-REF")
    if pos_ref is not None:
        response = root.xpath(f'//*[@ID="{pos_ref.get("ID-REF")}"]')
        if response:
            set_table_ref_in_message(response[0], status_table.get("ID"))


def update_dtc_dop(root: etree._Element, id_gen: IdGenerator, dtcs: list[DtcDef]) -> None:
    dtc_dop = first_by_short(root, "DTC-DOP", "RecordDataType")
    if dtc_dop is None:
        return
    dtcs_node = dtc_dop.find("DTCS")
    if dtcs_node is None:
        dtcs_node = sub(dtc_dop, "DTCS")
    for child in list(dtcs_node):
        dtcs_node.remove(child)
    caption_ids: dict[str, str] = {}
    for index, dtc in enumerate(dtcs):
        dtc_node = sub(dtcs_node, "DTC", attrib={"ID": id_gen.new("DTC")})
        sub(dtc_node, "SHORT-NAME", sanitize_short_name(f"DTC_{dtc.byte_code:06X}", "DTC"))
        sub(dtc_node, "TROUBLE-CODE", dtc.byte_code)
        sub(dtc_node, "DISPLAY-TROUBLE-CODE", dtc.display_code)
        sub(dtc_node, "TEXT", dtc.text or dtc.display_code)
        sdgs = sub(dtc_node, "SDGS")
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_SHORTNAME", f"DTC_0X{dtc.byte_code:06X}", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_PRIORITY_VALUE", dtc.priority or "2", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_SUPPORTED", "supported", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_CYCLE", "DEM_POWER", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_AGING_COUNTER", "40", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_SEVERITY_VALUE", "noSeverity", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_INIT_MONITOR_REQUIRED", "not required", first=index == 0)
        add_dtc_sdg(sdgs, id_gen, caption_ids, "DTC_FUNCTIONAL_UNIT_VALUE", "0x00", first=index == 0)

    update_dtc_text_table(root, dtcs)

    ext_table = first_by_short(root, "TABLE", "DTCExtendedDataRecordNumber")
    if ext_table is not None:
        clear_children(ext_table, "TABLE-ROW")
        default_structure = "_179"
        for dtc in dtcs:
            key = dtc_table_key(dtc)
            row = append_table_row(
                ext_table,
                id_gen,
                f"TR_DTC_{dtc.byte_code:06X}",
                key,
                key,
                default_structure,
            )
            sdgs = sub(row, "SDGS")
            sdg = sub(sdgs, "SDG")
            caption = sub(sdg, "SDG-CAPTION", attrib={"ID": id_gen.new("CAP")})
            sub(caption, "SHORT-NAME", "IsDefaultCase")
            sub(sdg, "SD", "Yes")


def dtc_table_key(dtc: DtcDef) -> str:
    text = compact_text(dtc.text or dtc.display_code)
    suffix = f"_{text[:80]}" if text else ""
    return f"0x{dtc.byte_code:06X}{suffix}"


def update_dtc_text_table(root: etree._Element, dtcs: list[DtcDef]) -> None:
    dop = first_by_short(root, "DATA-OBJECT-PROP", "TextTable_DTC")
    if dop is None:
        return
    scales = dop.find(".//COMPU-SCALES")
    if scales is None:
        return
    for child in list(scales):
        scales.remove(child)
    for dtc in dtcs:
        scale = sub(scales, "COMPU-SCALE")
        sub(scale, "LOWER-LIMIT", dtc.byte_code)
        sub(scale, "UPPER-LIMIT", dtc.byte_code)
        const = sub(scale, "COMPU-CONST")
        sub(const, "VT", dtc_table_key(dtc))

    internal_constr = dop.find("INTERNAL-CONSTR")
    if internal_constr is not None:
        dop.remove(internal_constr)


def add_dtc_sdg(
    sdgs: etree._Element,
    id_gen: IdGenerator,
    caption_ids: dict[str, str],
    short_name: str,
    value: str,
    *,
    first: bool,
) -> None:
    sdg = sub(sdgs, "SDG")
    if first or short_name not in caption_ids:
        caption_id = id_gen.new("CAP")
        caption_ids[short_name] = caption_id
        caption = sub(sdg, "SDG-CAPTION", attrib={"ID": caption_id})
        sub(caption, "SHORT-NAME", short_name)
    else:
        sub(sdg, "SDG-CAPTION-REF", attrib={"ID-REF": caption_ids[short_name]})
    sub(sdg, "SD", value)


def update_snapshot_and_extended_data(
    root: etree._Element,
    id_gen: IdGenerator,
    snapshots: list[SnapshotDef],
    extended_records: list[ExtendedRecordDef],
) -> None:
    snapshots = [snapshot for snapshot in snapshots if snapshot.structure_id]
    env_data = first_by_short(root, "ENV-DATA", "ENVDATA_ALLDTCS")
    if env_data is not None and snapshots:
        params = env_data.find("PARAMS")
        if params is None:
            params = sub(env_data, "PARAMS")
        for child in list(params):
            params.remove(child)
        byte_position = 0
        for snapshot in snapshots:
            did_param = element("PARAM", attrib={"SEMANTIC": "DATA"})
            set_xsi_type(did_param, "PHYS-CONST")
            sub(did_param, "SHORT-NAME", sanitize_short_name(hex_short("DID", snapshot.did), "SnapshotDID"))
            sub(did_param, "LONG-NAME", snapshot.desc)
            sub(did_param, "BYTE-POSITION", byte_position)
            sub(did_param, "PHYS-CONSTANT-VALUE", snapshot.did)
            sub(did_param, "DOP-REF", attrib={"ID-REF": "_17"})
            params.append(did_param)
            byte_position += 2
            if snapshot.structure_id:
                value_param = make_value_param(
                    sanitize_short_name(f"{hex_short('DID', snapshot.did)}_Data", "SnapshotData"),
                    snapshot.desc,
                    byte_position,
                    0,
                    snapshot.structure_id,
                )
                params.append(value_param)
                byte_position += max(1, snapshot.size)

    if extended_records:
        update_record_number_dop(root, "DTCExtendedDataRecordNumbers_All", extended_records, include_all=True)
        update_record_number_dop(root, "DTCExtendedDataRecordNumbers_All_except_FF", extended_records, include_all=False)
        mux = first_by_short(root, "MUX", "DTCExtendedDataRecordData")
        if mux is not None:
            cases = mux.find("CASES")
            if cases is None:
                cases = sub(mux, "CASES")
            for child in list(cases):
                cases.remove(child)
            for record in extended_records:
                if not record.structure_id:
                    continue
                case = sub(cases, "CASE")
                label = f"Extended Data Record 0x{record.record_num:02X}"
                sub(case, "SHORT-NAME", sanitize_short_name(f"Case_0x{record.record_num:02X}", "Case"))
                sub(case, "STRUCTURE-REF", attrib={"ID-REF": record.structure_id})
                sub(case, "LOWER-LIMIT", label)
                sub(case, "UPPER-LIMIT", label)


def make_coded_const_param(
    short_name: str,
    long_name: str,
    byte_position: int,
    coded_value: int,
    semantic: str,
) -> etree._Element:
    param = element("PARAM", attrib={"SEMANTIC": semantic})
    set_xsi_type(param, "CODED-CONST")
    sub(param, "SHORT-NAME", short_name)
    sub(param, "LONG-NAME", long_name)
    sub(param, "BYTE-POSITION", byte_position)
    sub(param, "CODED-VALUE", coded_value)
    coded_type = sub(param, "DIAG-CODED-TYPE", attrib={"BASE-DATA-TYPE": "A_UINT32"})
    set_xsi_type(coded_type, "STANDARD-LENGTH-TYPE")
    sub(coded_type, "BIT-LENGTH", 8)
    return param


def make_table_key_param(
    short_name: str,
    long_name: str,
    byte_position: int,
    table_id: str,
    param_id: str,
) -> etree._Element:
    param = element("PARAM", attrib={"SEMANTIC": "DATA", "ID": param_id})
    set_xsi_type(param, "TABLE-KEY")
    param.append(etree.ProcessingInstruction("GeneratedFromMux", "yes"))
    sub(param, "SHORT-NAME", short_name)
    sub(param, "LONG-NAME", long_name)
    sub(param, "BYTE-POSITION", byte_position)
    sub(param, "TABLE-REF", attrib={"ID-REF": table_id})
    return param


def make_table_struct_param(
    short_name: str,
    long_name: str,
    byte_position: int,
    table_key_id: str,
) -> etree._Element:
    param = element("PARAM", attrib={"SEMANTIC": "DATA"})
    set_xsi_type(param, "TABLE-STRUCT")
    sub(param, "SHORT-NAME", short_name)
    sub(param, "LONG-NAME", long_name)
    sub(param, "BYTE-POSITION", byte_position)
    sub(param, "TABLE-KEY-REF", attrib={"ID-REF": table_key_id})
    return param


def normalize_extended_data_service_for_candela(root: etree._Element, id_gen: IdGenerator) -> None:
    request = first_by_short(root, "REQUEST", "RQ_FaultMemory_Read_extended_data_record")
    if request is None:
        return
    ext_table = first_by_short(root, "TABLE", "DTCExtendedDataRecordNumber")
    if ext_table is None or not ext_table.get("ID"):
        return
    params = request.find("PARAMS")
    if params is None:
        params = sub(request, "PARAMS")
    table_key_id = ""
    for param in params.findall("PARAM"):
        if param.findtext("SHORT-NAME") != "DTC":
            continue
        if param.get("ID"):
            table_key_id = param.get("ID") or ""
            break
    if not table_key_id:
        table_key_id = id_gen.new("TK")
    for child in list(params):
        params.remove(child)
    params.append(make_coded_const_param("SID_RQ", "SID-RQ", 0, 0x19, "SERVICE-ID"))
    params.append(
        make_coded_const_param(
            "ReportDTCExtendedDataRecordByDtcNumber",
            "ReportDTCExtendedDataRecordByDtcNumber",
            1,
            0x06,
            "SUBFUNCTION",
        )
    )
    params.append(make_table_key_param("DTC", "DTC", 2, ext_table.get("ID"), table_key_id))
    params.append(
        make_table_struct_param(
            "DTCExtendedDataRecordNumber",
            "DTCExtendedDataRecordNumber",
            5,
            table_key_id,
        )
    )


def validate_edr_mapping(root: etree._Element, dtcs: list[DtcDef], extended_records: list[ExtendedRecordDef]) -> None:
    if not dtcs or not extended_records:
        return

    errors: list[str] = []

    def add_error(message: str) -> None:
        errors.append(message)

    def short_sample(values: Iterable[str], limit: int = 5) -> str:
        items = list(values)
        if len(items) <= limit:
            return ", ".join(items)
        return ", ".join(items[:limit]) + f", ... (+{len(items) - limit})"

    service = first_by_short(root, "DIAG-SERVICE", "FaultMemory_Read_extended_data_record")
    if service is None:
        add_error("missing DIAG-SERVICE FaultMemory_Read_extended_data_record")

    request = first_by_short(root, "REQUEST", "RQ_FaultMemory_Read_extended_data_record")
    if service is not None:
        request_ref = service.find("REQUEST-REF")
        request_id = request_ref.get("ID-REF") if request_ref is not None else ""
        if request_id:
            request_nodes = root.xpath(f'//*[@ID="{request_id}"]')
            if request_nodes:
                request = request_nodes[0]
            else:
                add_error(f"service references missing request ID {request_id}")

    ext_table = first_by_short(root, "TABLE", "DTCExtendedDataRecordNumber")
    if ext_table is None or not ext_table.get("ID"):
        add_error("missing TABLE DTCExtendedDataRecordNumber")

    if request is None:
        add_error("missing REQUEST RQ_FaultMemory_Read_extended_data_record")
    else:
        params = request.xpath('./*[local-name()="PARAMS"]/*[local-name()="PARAM"]')
        by_name = {compact_text(param.findtext("SHORT-NAME")): param for param in params}
        dtc_param = by_name.get("DTC")
        record_param = by_name.get("DTCExtendedDataRecordNumber")
        if dtc_param is None:
            add_error("0x19 0x06 request missing TABLE-KEY param DTC")
        else:
            if dtc_param.get(XSI_TYPE) != "TABLE-KEY":
                add_error("0x19 0x06 DTC param is not TABLE-KEY")
            if not dtc_param.get("ID"):
                add_error("0x19 0x06 DTC TABLE-KEY has no ID")
            table_ref = dtc_param.find("TABLE-REF")
            table_id = table_ref.get("ID-REF") if table_ref is not None else ""
            expected_table_id = ext_table.get("ID") if ext_table is not None else ""
            if expected_table_id and table_id != expected_table_id:
                add_error("0x19 0x06 DTC TABLE-KEY does not reference DTCExtendedDataRecordNumber")

        if record_param is None:
            add_error("0x19 0x06 request missing TABLE-STRUCT param DTCExtendedDataRecordNumber")
        else:
            if record_param.get(XSI_TYPE) != "TABLE-STRUCT":
                add_error("0x19 0x06 DTCExtendedDataRecordNumber param is not TABLE-STRUCT")
            key_ref = record_param.find("TABLE-KEY-REF")
            key_id = key_ref.get("ID-REF") if key_ref is not None else ""
            dtc_id = dtc_param.get("ID") if dtc_param is not None else ""
            if dtc_id and key_id != dtc_id:
                add_error("0x19 0x06 TABLE-STRUCT does not reference the DTC TABLE-KEY")

    if ext_table is not None:
        rows = ext_table.xpath('./*[local-name()="TABLE-ROW"]')
        rows_by_key = {compact_text(row.findtext("KEY")): row for row in rows}
        expected_keys = [dtc_table_key(dtc) for dtc in dtcs]
        missing_keys = [key for key in expected_keys if key not in rows_by_key]
        if missing_keys:
            add_error(f"DTCExtendedDataRecordNumber missing {len(missing_keys)} DTC row(s): {short_sample(missing_keys)}")
        unsupported_keys = [
            key
            for key in expected_keys
            if key in rows_by_key and rows_by_key[key].find("STRUCTURE-REF") is None
        ]
        if unsupported_keys:
            add_error(f"{len(unsupported_keys)} DTC row(s) have no EDR STRUCTURE-REF: {short_sample(unsupported_keys)}")

    expected_record_labels = [f"Extended Data Record 0x{record.record_num:02X}" for record in extended_records]
    for dop_name, require_all in (
        ("DTCExtendedDataRecordNumbers_All", True),
        ("DTCExtendedDataRecordNumbers_All_except_FF", False),
    ):
        dop = first_by_short(root, "DATA-OBJECT-PROP", dop_name)
        if dop is None:
            add_error(f"missing DATA-OBJECT-PROP {dop_name}")
            continue
        scales = dop.xpath('.//*[local-name()="COMPU-SCALE"]')
        labels = {compact_text(scale.findtext(".//VT")) for scale in scales}
        missing_labels = [label for label in expected_record_labels if label not in labels]
        if missing_labels:
            add_error(f"{dop_name} missing record label(s): {short_sample(missing_labels)}")
        has_all = "All" in labels
        if require_all and not has_all:
            add_error(f"{dop_name} missing All/0xFF scale")
        if not require_all and has_all:
            add_error(f"{dop_name} unexpectedly contains All/0xFF scale")

    mux = first_by_short(root, "MUX", "DTCExtendedDataRecordData")
    if mux is None:
        add_error("missing MUX DTCExtendedDataRecordData")
    else:
        cases = mux.xpath('./*[local-name()="CASES"]/*[local-name()="CASE"]')
        case_by_label = {compact_text(case.findtext("LOWER-LIMIT")): case for case in cases}
        for record, label in zip(extended_records, expected_record_labels, strict=True):
            if not record.structure_id:
                add_error(f"{label} has no generated STRUCTURE")
                continue
            case = case_by_label.get(label)
            if case is None:
                add_error(f"DTCExtendedDataRecordData missing MUX case for {label}")
                continue
            structure_ref = case.find("STRUCTURE-REF")
            structure_id = structure_ref.get("ID-REF") if structure_ref is not None else ""
            if structure_id != record.structure_id:
                add_error(f"DTCExtendedDataRecordData MUX case for {label} references {structure_id or '<none>'}")

    if errors:
        raise RuntimeError("Invalid DTC extended-data mapping:\n- " + "\n- ".join(errors))


def update_record_number_dop(
    root: etree._Element,
    short_name: str,
    records: list[ExtendedRecordDef],
    *,
    include_all: bool,
) -> None:
    dop = first_by_short(root, "DATA-OBJECT-PROP", short_name)
    if dop is None:
        return
    scales = dop.find(".//COMPU-SCALES")
    if scales is None:
        return
    for child in list(scales):
        scales.remove(child)
    for record in records:
        scale = sub(scales, "COMPU-SCALE")
        label = f"Extended Data Record 0x{record.record_num:02X}"
        sub(scale, "LOWER-LIMIT", record.record_num)
        sub(scale, "UPPER-LIMIT", record.record_num)
        const = sub(scale, "COMPU-CONST")
        sub(const, "VT", label)
    if include_all:
        scale = sub(scales, "COMPU-SCALE")
        sub(scale, "LOWER-LIMIT", 255)
        sub(scale, "UPPER-LIMIT", 255)
        const = sub(scale, "COMPU-CONST")
        sub(const, "VT", "All")
    internal_constr = dop.find("INTERNAL-CONSTR")
    if internal_constr is not None:
        dop.remove(internal_constr)


def update_base_variant_comparams(root: etree._Element, cover: CoverInfo) -> None:
    base_variant = first_by_short(root, "BASE-VARIANT", "VOYAN_ECU_CAN")
    if base_variant is None:
        return
    comp_refs = base_variant.find("COMPARAM-REFS")
    if comp_refs is None:
        comp_refs = sub(base_variant, "COMPARAM-REFS")
    unique_ref = None
    for ref in comp_refs.findall("COMPARAM-REF"):
        if ref.get("ID-REF") == "ISO_15765_2.CP_UniqueRespIdTable":
            unique_ref = ref
            break
    if unique_ref is None:
        unique_ref = sub(
            comp_refs,
            "COMPARAM-REF",
            attrib={"ID-REF": "ISO_15765_2.CP_UniqueRespIdTable", "DOCREF": "ISO_15765_2", "DOCTYPE": "COMPARAM-SUBSET"},
        )
    for child in list(unique_ref):
        unique_ref.remove(child)
    complex_value = sub(unique_ref, "COMPLEX-VALUE")
    rx_phy = cover.rx_phy_id if cover.rx_phy_id is not None else 0x700
    tx = cover.tx_id if cover.tx_id is not None else 0x600
    values = [
        "0",
        "normal segmented 11-bit transmit with FC",
        str(rx_phy),
        "0",
        "normal segmented 11-bit receive with FC",
        str(tx),
        "0",
        "normal unsegmented 11-bit receive",
        "4294967295",
        "VOYAN_ECU_CAN",
    ]
    for value in values:
        sub(complex_value, "SIMPLE-VALUE", value)
    sub(unique_ref, "PROTOCOL-SNREF", attrib={"SHORT-NAME": "CAN"})


def ensure_base_variant_protocol_parent(root: etree._Element) -> None:
    base_variant = first_by_short(root, "BASE-VARIANT", "VOYAN_ECU_CAN")
    if base_variant is None:
        return
    parent_refs = base_variant.find("PARENT-REFS")
    if parent_refs is None:
        parent_refs = etree.Element("PARENT-REFS")
        comp_refs = base_variant.find("COMPARAM-REFS")
        if comp_refs is not None:
            base_variant.insert(base_variant.index(comp_refs), parent_refs)
        else:
            base_variant.append(parent_refs)

    for parent_ref in parent_refs.findall("PARENT-REF"):
        if parent_ref.get("ID-REF") == "CAN" and parent_ref.get(XSI_TYPE) == "PROTOCOL-REF":
            parent_ref.set("DOCREF", "DLC_FGL_UDS")
            parent_ref.set("DOCTYPE", "CONTAINER")
            return

    sub(
        parent_refs,
        "PARENT-REF",
        attrib={
            "ID-REF": "CAN",
            "DOCREF": "DLC_FGL_UDS",
            "DOCTYPE": "CONTAINER",
            XSI_TYPE: "PROTOCOL-REF",
        },
    )


def update_session_timing(root: etree._Element) -> None:
    for state in root.xpath('//*[local-name()="STATE" and (SHORT-NAME="Default" or SHORT-NAME="Programming" or SHORT-NAME="Extended")]'):
        desc = state.find("DESC")
        if desc is None:
            desc = sub(state, "DESC")
        for child in list(desc):
            desc.remove(child)
        sub(desc, "p", "{P2=50, P2Ex=5000}")


def patch_fgl_communication_params(fgl_path: Path, cover: CoverInfo) -> None:
    parser = etree.XMLParser(remove_blank_text=False)
    tree = etree.parse(str(fgl_path), parser)
    root = tree.getroot()
    prune_fgl_to_can(root)
    values = {
        "ISO_15765_2.CP_CanFuncReqId": cover.rx_fun_id if cover.rx_fun_id is not None else 0x7DF,
        "ISO_15765_3.CP_P2Max": 150000,
        "ISO_15765_3.CP_P2Star": 5100000,
        "ISO_15765_3.CP_TesterPresentTime": 4000000,
        "ISO_15765_3.CP_P3Phys": 50000,
        "ISO_15765_3.CP_P3Func": 50000,
        "ISO_15765_2.CP_StMin": 20_000,
        "ISO_15765_2.CP_As": 70_000,
        "ISO_15765_2.CP_Ar": 70_000,
        "ISO_15765_2.CP_Bs": 150_000,
        "ISO_15765_2.CP_Br": 70_000,
        "ISO_15765_2.CP_Cs": 70_000,
        "ISO_15765_2.CP_Cr": 150_000,
    }
    for ref in root.xpath('//*[local-name()="COMPARAM-REF"]'):
        id_ref = ref.get("ID-REF")
        if id_ref not in values:
            continue
        simple = ref.find("SIMPLE-VALUE")
        if simple is not None:
            simple.text = str(values[id_ref])
    tree.write(str(fgl_path), encoding="UTF-8", xml_declaration=True, pretty_print=True, standalone=False)


def prune_can_comparam_spec(stack_path: Path) -> None:
    parser = etree.XMLParser(remove_blank_text=False)
    tree = etree.parse(str(stack_path), parser)
    root = tree.getroot()
    keep_stack = "ISO_15765_3_on_ISO_15765_2_on_ISO_11898_2_DWCAN"
    for prot_stack in list(root.xpath("//*[local-name()='PROT-STACK']")):
        if prot_stack.get("ID") != keep_stack:
            parent = prot_stack.getparent()
            if parent is not None:
                parent.remove(prot_stack)
    tree.write(str(stack_path), encoding="UTF-8", xml_declaration=True, pretty_print=True, standalone=False)


def prune_fgl_to_can(root: etree._Element) -> None:
    for protocol in list(root.xpath("//*[local-name()='PROTOCOL']")):
        if child_text_by_local_name(protocol, "SHORT-NAME") != "CAN":
            parent = protocol.getparent()
            if parent is not None:
                parent.remove(protocol)

    for functional_group in root.xpath("//*[local-name()='FUNCTIONAL-GROUP']"):
        comp_refs = functional_group.find("COMPARAM-REFS")
        if comp_refs is not None:
            functional_group.remove(comp_refs)

        parent_refs = functional_group.find("PARENT-REFS")
        if parent_refs is not None:
            functional_group.remove(parent_refs)


def validate_with_odxtools(pdx_path: Path) -> None:
    cmd = [sys.executable, "-m", "odxtools", "list", str(pdx_path), "--all"]
    result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
    if result.returncode != 0:
        details = (result.stderr or result.stdout).strip()
        raise RuntimeError(f"odxtools validation failed for {pdx_path}:\n{details}")


def find_default_xlsx(base_dir: Path) -> Path:
    candidates = sorted(path for path in base_dir.glob("*.xlsx") if not path.name.startswith("~$"))
    if not candidates:
        raise FileNotFoundError("No .xlsx survey file found in the current directory")
    return candidates[0]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Generate VOYAN PDX from a VOYAH diagnosis survey Excel file.")
    parser.add_argument("xlsx", nargs="?", type=Path, help="Input diagnosis survey .xlsx file")
    parser.add_argument("--template", type=Path, default=Path("templates") / "VOYAN_ECU_CAN_v15.pdx", help="Template PDX")
    parser.add_argument("--output-dir", type=Path, default=Path("output"), help="Output directory")
    parser.add_argument("--no-validate", action="store_true", help="Skip odxtools validation")
    args = parser.parse_args(argv)

    xlsx_path = args.xlsx or find_default_xlsx(Path.cwd())
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)
    if not args.template.exists():
        raise FileNotFoundError(args.template)

    survey = parse_survey(xlsx_path)
    output_pdx = args.output_dir / f"{xlsx_path.stem}.pdx"
    update_template(args.template, output_pdx, survey, validate=not args.no_validate)
    print(f"Generated: {output_pdx}")
    print(
        "Parsed: "
        f"{len(survey.dids)} DIDs, "
        f"{len(survey.io_dids)} IO DIDs, "
        f"{len(survey.routines)} routines, "
        f"{len(survey.dtcs)} DTCs, "
        f"{len(survey.snapshots)} snapshots, "
        f"{len(survey.extended_records)} extended records"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
